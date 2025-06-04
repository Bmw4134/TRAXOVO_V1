"""
Driver Report Generator

This module processes driving history data to generate daily reports for:
1. Prior Day Report - Late Starts, Early Ends, Not on Job
2. Current Day Report - Late Starts, Not on Job

These reports are saved to a dated folder and can be accessed through the web interface.
"""

import os
import csv
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Initialize logger
logger = logging.getLogger(__name__)

def process_driving_history(file_path):
    """
    Process the driving history CSV file to extract driver location data
    
    Args:
        file_path (str): Path to the driving history CSV file
        
    Returns:
        dict: Processed driving history data organized by driver and date
    """
    try:
        # Read the CSV file
        logger.info(f"Processing driving history file: {file_path}")
        driving_data = []
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            csv_reader = csv.reader(f)
            
            # Skip header rows (first 7 rows)
            for _ in range(7):
                next(csv_reader)
            
            # Process the data rows
            current_driver = None
            for row in csv_reader:
                if len(row) < 11:  # Skip rows with insufficient data
                    continue
                    
                # Check if this is a new driver header row
                if row[0].startswith('#') and len(row[10].strip()) == 0:
                    current_driver = row[0].strip()
                    continue
                
                # Regular event row
                if current_driver and row[10].strip() and row[6].strip():
                    try:
                        event_datetime = datetime.strptime(row[6].strip(), '%m/%d/%Y %I:%M:%S %p')
                        
                        driving_data.append({
                            'driver_id': current_driver,
                            'event_datetime': event_datetime,
                            'event_date': event_datetime.date(),
                            'event_type': row[7].strip(),
                            'location': row[10].strip(),
                            'latitude': float(row[8]) if row[8].strip() else None,
                            'longitude': float(row[9]) if row[9].strip() else None,
                            'contact': row[4].strip(),
                            'phone': row[5].strip()
                        })
                    except Exception as e:
                        logger.warning(f"Error processing row for driver {current_driver}: {e}")
                        continue
        
        # Convert to DataFrame for easier processing
        df = pd.DataFrame(driving_data)
        
        if df.empty:
            logger.warning("No valid driving data found in file")
            return None
            
        logger.info(f"Processed {len(df)} driving events from {df['driver_id'].nunique()} drivers")
        return df
        
    except Exception as e:
        logger.error(f"Error processing driving history file: {e}")
        return None

def map_job_sites(driving_history_df, job_sites_mapping=None):
    """
    Map GPS coordinates to known job sites
    
    Args:
        driving_history_df (DataFrame): Processed driving history data
        job_sites_mapping (dict): Optional mapping of job sites with coords
        
    Returns:
        DataFrame: Driving history with job site assignments
    """
    # If no job sites mapping provided, use a simple algorithm based on location name
    if not job_sites_mapping:
        # Extract potential job site identifiers from location strings
        driving_history_df['job_site'] = driving_history_df['location'].apply(
            lambda x: x.split(',')[0] if ',' in x else x.split(' ')[0]
        )
    else:
        # For a more sophisticated approach, we would calculate distance to known job coordinates
        # and assign the closest job site to each location
        # This is a placeholder for when we have proper job site data
        pass
    
    return driving_history_df

def determine_shifts(driving_history_df, shift_start_time=datetime.strptime('07:00:00', '%H:%M:%S').time(), 
                    shift_end_time=datetime.strptime('16:00:00', '%H:%M:%S').time()):
    """
    Determine shift information (expected vs. actual start/end times)
    
    Args:
        driving_history_df (DataFrame): Processed driving history data
        shift_start_time (time): Expected shift start time
        shift_end_time (time): Expected shift end time
        
    Returns:
        dict: Shift analysis by driver and date
    """
    driver_dates = driving_history_df.groupby(['driver_id', 'event_date'])
    
    shift_data = {}
    
    for (driver_id, event_date), events in driver_dates:
        # Sort events by time
        day_events = events.sort_values('event_datetime')
        
        # Find first "Key On" event of the day (actual start)
        first_key_on = day_events[day_events['event_type'] == 'Key On'].iloc[0] if not day_events[day_events['event_type'] == 'Key On'].empty else None
        
        # Find last "Key Off" event of the day (actual end)
        last_key_off = day_events[day_events['event_type'] == 'Key Off'].iloc[-1] if not day_events[day_events['event_type'] == 'Key Off'].empty else None
        
        # Calculate shift metrics
        if first_key_on is not None:
            # Create expected start time for comparison
            expected_start = datetime.combine(event_date, shift_start_time)
            actual_start = first_key_on['event_datetime']
            
            # Calculate minutes late (positive means late)
            minutes_late = (actual_start - expected_start).total_seconds() / 60
            
            # Determine if late start (more than 5 minutes late)
            is_late_start = minutes_late > 5
            
            # Determine starting job site
            start_job_site = first_key_on['job_site'] if 'job_site' in first_key_on else None
            
            # Find the first job arrival
            first_arrival = day_events[day_events['event_type'] == 'Arrived'].iloc[0] if not day_events[day_events['event_type'] == 'Arrived'].empty else None
            first_job_site = first_arrival['job_site'] if first_arrival is not None and 'job_site' in first_arrival else None
        else:
            expected_start = None
            actual_start = None
            minutes_late = None
            is_late_start = False
            start_job_site = None
            first_job_site = None
            
        if last_key_off is not None:
            # Create expected end time for comparison
            expected_end = datetime.combine(event_date, shift_end_time)
            actual_end = last_key_off['event_datetime']
            
            # Calculate minutes early (positive means early departure)
            minutes_early = (expected_end - actual_end).total_seconds() / 60
            
            # Determine if early end (more than 15 minutes early)
            is_early_end = minutes_early > 15
            
            # Determine ending job site
            end_job_site = last_key_off['job_site'] if 'job_site' in last_key_off else None
        else:
            expected_end = None
            actual_end = None
            minutes_early = None
            is_early_end = False
            end_job_site = None
            
        # Store the shift data
        if (driver_id, event_date) not in shift_data:
            shift_data[(driver_id, event_date)] = {
                'driver_id': driver_id,
                'date': event_date,
                'expected_start': expected_start,
                'actual_start': actual_start,
                'minutes_late': minutes_late,
                'is_late_start': is_late_start,
                'expected_end': expected_end,
                'actual_end': actual_end,
                'minutes_early': minutes_early,
                'is_early_end': is_early_end,
                'start_job_site': start_job_site,
                'end_job_site': end_job_site,
                'first_job_site': first_job_site
            }
    
    # Convert to DataFrame
    shift_df = pd.DataFrame(list(shift_data.values()))
    
    return shift_df

def identify_not_on_job(shift_df, job_assignments=None):
    """
    Identify drivers who are not at their assigned job sites
    
    Args:
        shift_df (DataFrame): Shift data
        job_assignments (dict): Map of drivers to expected job sites
        
    Returns:
        DataFrame: Updated shift data with not_on_job flag
    """
    # If no job assignments are provided, we'll use a simple approach
    # where we assume a driver should be at the same job site consistently
    if not job_assignments:
        # Group by driver to find their most common job site
        driver_groups = shift_df.groupby('driver_id')
        
        for driver_id, group in driver_groups:
            most_common_site = None
            
            # Find the most frequent job site for this driver
            if 'first_job_site' in group.columns and not group['first_job_site'].isnull().all():
                most_common_site = group['first_job_site'].mode()[0] if not group['first_job_site'].mode().empty else None
            
            # Update the not_on_job flag for each shift
            for idx in group.index:
                expected_job = most_common_site
                actual_job = shift_df.loc[idx, 'first_job_site']
                
                # Mark as not on job if the job sites don't match
                if expected_job and actual_job and expected_job != actual_job:
                    shift_df.loc[idx, 'not_on_job'] = True
                    shift_df.loc[idx, 'expected_job_site'] = expected_job
                else:
                    shift_df.loc[idx, 'not_on_job'] = False
                    shift_df.loc[idx, 'expected_job_site'] = expected_job
    else:
        # Use the provided job assignments
        for idx in shift_df.index:
            driver_id = shift_df.loc[idx, 'driver_id']
            
            if driver_id in job_assignments:
                expected_job = job_assignments[driver_id]
                actual_job = shift_df.loc[idx, 'first_job_site']
                
                shift_df.loc[idx, 'expected_job_site'] = expected_job
                shift_df.loc[idx, 'not_on_job'] = expected_job != actual_job
            else:
                shift_df.loc[idx, 'not_on_job'] = False
                
    return shift_df

def generate_prior_day_report(driving_history_df=None, report_date=None, job_assignments=None):
    """
    Generate prior day attendance report
    
    Args:
        driving_history_df (DataFrame): Processed driving history
        report_date (date): Date for which to generate the report (defaults to yesterday)
        job_assignments (dict): Optional driver-to-job assignments
        
    Returns:
        tuple: (late_starts_df, early_ends_df, not_on_job_df) for the prior day
    """
    try:
        # If no report date specified, use yesterday
        if report_date is None:
            report_date = datetime.now().date() - timedelta(days=1)
            
        logger.info(f"Generating prior day report for {report_date}")
        
        # If no driving history provided, try to load from default location
        if driving_history_df is None:
            history_file = 'attached_assets/DrivingHistory.csv'
            if os.path.exists(history_file):
                driving_history_df = process_driving_history(history_file)
            else:
                logger.error(f"No driving history file found at {history_file}")
                return None, None, None
        
        # Filter data for the report date
        day_data = driving_history_df[driving_history_df['event_date'] == report_date]
        
        if day_data.empty:
            logger.warning(f"No driving data found for {report_date}")
            return None, None, None
            
        # Map to job sites
        day_data = map_job_sites(day_data, job_assignments)
        
        # Analyze shifts
        shift_df = determine_shifts(day_data)
        
        if shift_df.empty:
            logger.warning(f"No shift data could be generated for {report_date}")
            return None, None, None
            
        # Identify not on job
        shift_df = identify_not_on_job(shift_df, job_assignments)
        
        # Create report dataframes
        late_starts_df = shift_df[shift_df['is_late_start'] == True].copy() if 'is_late_start' in shift_df.columns else pd.DataFrame()
        early_ends_df = shift_df[shift_df['is_early_end'] == True].copy() if 'is_early_end' in shift_df.columns else pd.DataFrame()
        not_on_job_df = shift_df[shift_df['not_on_job'] == True].copy() if 'not_on_job' in shift_df.columns else pd.DataFrame()
        
        # Generate Excel report
        reports_dir = f'reports/{datetime.now().strftime("%Y-%m-%d")}'
        os.makedirs(reports_dir, exist_ok=True)
        
        report_path = f"{reports_dir}/prior_day_report_{report_date.strftime('%Y-%m-%d')}.xlsx"
        
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            # Late Starts Sheet
            if not late_starts_df.empty:
                late_starts_df.to_excel(writer, sheet_name='Late Starts', index=False)
            else:
                pd.DataFrame({'Message': ['No late starts recorded for this date']}).to_excel(
                    writer, sheet_name='Late Starts', index=False)
            
            # Early Ends Sheet
            if not early_ends_df.empty:
                early_ends_df.to_excel(writer, sheet_name='Early Ends', index=False)
            else:
                pd.DataFrame({'Message': ['No early ends recorded for this date']}).to_excel(
                    writer, sheet_name='Early Ends', index=False)
            
            # Not On Job Sheet
            if not not_on_job_df.empty:
                not_on_job_df.to_excel(writer, sheet_name='Not On Job', index=False)
            else:
                pd.DataFrame({'Message': ['No drivers were at incorrect job sites for this date']}).to_excel(
                    writer, sheet_name='Not On Job', index=False)
            
            # Format the workbook
            workbook = writer.book
            
            # Apply formatting to each sheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format header row
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        logger.info(f"Prior day report generated: {report_path}")
        
        return late_starts_df, early_ends_df, not_on_job_df, report_path
        
    except Exception as e:
        logger.error(f"Error generating prior day report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None, None, None, None

def generate_current_day_report(driving_history_df=None, report_date=None, job_assignments=None):
    """
    Generate current day attendance report
    
    Args:
        driving_history_df (DataFrame): Processed driving history
        report_date (date): Date for which to generate the report (defaults to today)
        job_assignments (dict): Optional driver-to-job assignments
        
    Returns:
        tuple: (late_starts_df, not_on_job_df) for the current day
    """
    try:
        # If no report date specified, use today
        if report_date is None:
            report_date = datetime.now().date()
            
        logger.info(f"Generating current day report for {report_date}")
        
        # If no driving history provided, try to load from default location
        if driving_history_df is None:
            history_file = 'attached_assets/DrivingHistory.csv'
            if os.path.exists(history_file):
                driving_history_df = process_driving_history(history_file)
            else:
                logger.error(f"No driving history file found at {history_file}")
                return None, None, None
        
        # Filter data for the report date
        day_data = driving_history_df[driving_history_df['event_date'] == report_date]
        
        if day_data.empty:
            logger.warning(f"No driving data found for {report_date}")
            return None, None, None
            
        # Map to job sites
        day_data = map_job_sites(day_data, job_assignments)
        
        # Analyze shifts
        shift_df = determine_shifts(day_data)
        
        if shift_df.empty:
            logger.warning(f"No shift data could be generated for {report_date}")
            return None, None, None
            
        # Identify not on job
        shift_df = identify_not_on_job(shift_df, job_assignments)
        
        # Create report dataframes
        late_starts_df = shift_df[shift_df['is_late_start'] == True].copy() if 'is_late_start' in shift_df.columns else pd.DataFrame()
        not_on_job_df = shift_df[shift_df['not_on_job'] == True].copy() if 'not_on_job' in shift_df.columns else pd.DataFrame()
        
        # Generate Excel report
        reports_dir = f'reports/{datetime.now().strftime("%Y-%m-%d")}'
        os.makedirs(reports_dir, exist_ok=True)
        
        report_path = f"{reports_dir}/current_day_report_{report_date.strftime('%Y-%m-%d')}.xlsx"
        
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            # Late Starts Sheet
            if not late_starts_df.empty:
                late_starts_df.to_excel(writer, sheet_name='Late Starts', index=False)
            else:
                pd.DataFrame({'Message': ['No late starts recorded for this date']}).to_excel(
                    writer, sheet_name='Late Starts', index=False)
            
            # Not On Job Sheet
            if not not_on_job_df.empty:
                not_on_job_df.to_excel(writer, sheet_name='Not On Job', index=False)
            else:
                pd.DataFrame({'Message': ['No drivers were at incorrect job sites for this date']}).to_excel(
                    writer, sheet_name='Not On Job', index=False)
            
            # Format the workbook
            workbook = writer.book
            
            # Apply formatting to each sheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format header row
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        logger.info(f"Current day report generated: {report_path}")
        
        return late_starts_df, not_on_job_df, report_path
        
    except Exception as e:
        logger.error(f"Error generating current day report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None, None, None

def generate_all_reports():
    """
    Generate both prior day and current day reports
    
    Returns:
        dict: Dictionary with report paths and status
    """
    try:
        # Load driving history once for both reports
        history_file = 'attached_assets/DrivingHistory.csv'
        if os.path.exists(history_file):
            driving_history_df = process_driving_history(history_file)
        else:
            logger.error(f"No driving history file found at {history_file}")
            return {
                'success': False,
                'message': f"Driving history file not found: {history_file}"
            }
            
        if driving_history_df is None or driving_history_df.empty:
            return {
                'success': False,
                'message': "Error processing driving history data"
            }
            
        # Generate prior day report
        prior_day = datetime.now().date() - timedelta(days=1)
        _, _, _, prior_day_path = generate_prior_day_report(
            driving_history_df=driving_history_df,
            report_date=prior_day
        )
        
        # Generate current day report
        current_day = datetime.now().date()
        _, _, current_day_path = generate_current_day_report(
            driving_history_df=driving_history_df,
            report_date=current_day
        )
        
        # Return results
        return {
            'success': True,
            'prior_day_report': prior_day_path,
            'current_day_report': current_day_path,
            'message': "Reports generated successfully"
        }
        
    except Exception as e:
        logger.error(f"Error generating reports: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'message': f"Error generating reports: {str(e)}"
        }

if __name__ == "__main__":
    # Test functionality
    result = generate_all_reports()
    print(result)
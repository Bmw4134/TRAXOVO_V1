"""
Dynamic Timecard Processor

This module provides functionality for processing timecard data from Excel files,
with formula evaluation capabilities to accurately extract driver hours.
"""

import os
import re
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from formulas import Parser
from openpyxl import load_workbook
from fuzzywuzzy import fuzz, process

# Configure logging
logger = logging.getLogger(__name__)

def process_dynamic_timecard(file_path, start_date=None, end_date=None):
    """
    Process timecard data from an Excel file with formula support.
    
    Args:
        file_path: Path to the Excel file
        start_date: Start date for filtering (optional)
        end_date: End date for filtering (optional)
        
    Returns:
        DataFrame: Processed timecard data
    """
    try:
        logger.info(f"Processing timecard file: {file_path}")
        
        # Load Excel file with openpyxl for formula extraction
        wb = load_workbook(file_path, data_only=False)
        
        # Get all sheet names
        sheet_names = wb.sheetnames
        
        # Try to find the relevant sheet
        timecard_sheet = None
        for name in sheet_names:
            name_lower = name.lower()
            if any(term in name_lower for term in ['timecard', 'time card', 'hours', 'daily']):
                timecard_sheet = name
                break
        
        # If no specific sheet found, use the first one
        if not timecard_sheet and sheet_names:
            timecard_sheet = sheet_names[0]
        
        if not timecard_sheet:
            logger.error("No suitable sheet found in the timecard file")
            return None
        
        # Load the sheet
        sheet = wb[timecard_sheet]
        
        # Try to detect the structure of the timecard
        # First, find headers
        header_row = detect_header_row(sheet)
        
        if header_row is None:
            logger.warning("Could not detect header row in timecard")
            # Try using pandas to read the file
            return process_with_pandas(file_path, start_date, end_date)
        
        # Get column headers
        columns = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            columns.append(cell_value if cell_value else f"Column{col}")
        
        # Clean and normalize column names
        columns = [clean_column_name(col) for col in columns]
        
        # Try to identify key columns
        date_col = employee_col = hours_col = job_col = None
        
        for i, col in enumerate(columns):
            col_lower = col.lower()
            if any(term in col_lower for term in ['date', 'day', 'workday']):
                date_col = i
            elif any(term in col_lower for term in ['employee', 'name', 'driver', 'operator']):
                employee_col = i
            elif any(term in col_lower for term in ['hours', 'time', 'duration']):
                hours_col = i
            elif any(term in col_lower for term in ['job', 'project', 'site', 'location']):
                job_col = i
        
        # If key columns not found, try using pandas
        if date_col is None or employee_col is None:
            logger.warning("Could not identify key columns in timecard")
            return process_with_pandas(file_path, start_date, end_date)
        
        # Extract data
        data = []
        
        for row in range(header_row + 1, sheet.max_row + 1):
            # Skip empty rows
            if all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1)):
                continue
            
            row_data = {}
            
            # Extract values for each column
            for i, col_name in enumerate(columns):
                col = i + 1
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                # Handle formulas
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_value = evaluate_formula(cell_value, wb, sheet, row, col)
                
                row_data[col_name] = cell_value
            
            data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Clean up the DataFrame
        df = clean_timecard_dataframe(df, date_col, employee_col, hours_col, job_col, columns)
        
        # Filter by date if requested
        if start_date or end_date:
            df = filter_by_date(df, start_date, end_date)
        
        return df
    
    except Exception as e:
        logger.error(f"Error processing timecard file: {str(e)}")
        # Fall back to pandas
        return process_with_pandas(file_path, start_date, end_date)

def detect_header_row(sheet):
    """
    Detect the header row in the sheet.
    
    Args:
        sheet: The Excel sheet
        
    Returns:
        int: Header row index or None if not found
    """
    # Check the first 10 rows
    for row in range(1, min(11, sheet.max_row + 1)):
        # Count non-empty cells
        non_empty = sum(1 for col in range(1, min(20, sheet.max_column + 1))
                        if sheet.cell(row=row, column=col).value is not None)
        
        # Check if this row has several non-empty cells
        if non_empty >= 3:
            # Check if cells contain header-like text
            header_like = 0
            for col in range(1, min(20, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value).lower()
                    if any(term in cell_str for term in ['date', 'name', 'employee', 'hours', 'job', 'time', 'driver']):
                        header_like += 1
            
            if header_like >= 2:
                return row
    
    # If no clear header found, use the first row with data
    for row in range(1, min(11, sheet.max_row + 1)):
        non_empty = sum(1 for col in range(1, min(20, sheet.max_column + 1))
                        if sheet.cell(row=row, column=col).value is not None)
        if non_empty >= 3:
            return row
    
    return None

def clean_column_name(name):
    """
    Clean and normalize column name.
    
    Args:
        name: Column name
        
    Returns:
        str: Cleaned column name
    """
    if name is None:
        return "unnamed"
    
    # Convert to string
    name = str(name).strip()
    
    # Remove special characters and replace spaces with underscores
    name = re.sub(r'[^\w\s]', '', name)
    name = name.replace(' ', '_')
    
    # Convert to lowercase
    name = name.lower()
    
    # Ensure the name is not empty
    if not name:
        return "unnamed"
    
    return name

def evaluate_formula(formula, workbook, sheet, row, col):
    """
    Evaluate an Excel formula.
    
    Args:
        formula: The formula to evaluate
        workbook: The Excel workbook
        sheet: The current sheet
        row: Current row
        col: Current column
        
    Returns:
        The evaluated formula result
    """
    try:
        # Try using the formulas library
        parser = Parser()
        parsed = parser.parse(formula)
        
        # Get the cell address
        cell_address = f"{chr(64 + col)}{row}"
        sheet_name = sheet.title
        
        # Set up the formula in the parser context
        parser.set_cell_value(f"{sheet_name}!{cell_address}", formula)
        
        # Evaluate
        result = parser.evaluate(f"{sheet_name}!{cell_address}")
        
        # If result is a number, return it
        if isinstance(result, (int, float)):
            return result
        
        # Otherwise, try to convert to number
        try:
            return float(result)
        except (ValueError, TypeError):
            pass
        
        return result
    
    except Exception as e:
        logger.debug(f"Formula evaluation failed: {str(e)}")
        
        # Fall back to manual evaluation for simple formulas
        try:
            # Handle SUM
            if "SUM" in formula:
                # Extract the range
                match = re.search(r'SUM\((.*?)\)', formula)
                if match:
                    range_str = match.group(1)
                    # Very simple case: SUM(X1:X10)
                    if ":" in range_str:
                        start, end = range_str.split(":")
                        # For now, only handle same column
                        if start[0] == end[0]:
                            col_letter = start[0]
                            start_row = int(start[1:])
                            end_row = int(end[1:])
                            col_idx = ord(col_letter) - 64
                            
                            # Sum the values
                            total = 0
                            for r in range(start_row, end_row + 1):
                                val = sheet.cell(row=r, column=col_idx).value
                                if isinstance(val, (int, float)):
                                    total += val
                            
                            return total
            
            # Handle basic arithmetic
            if "+" in formula or "-" in formula or "*" in formula or "/" in formula:
                # Remove the = sign
                expr = formula[1:]
                # Replace cell references with their values
                for cell_ref in re.findall(r'[A-Z]+[0-9]+', expr):
                    col_letter = ''.join(c for c in cell_ref if c.isalpha())
                    row_number = int(''.join(c for c in cell_ref if c.isdigit()))
                    col_idx = 0
                    for i, c in enumerate(col_letter):
                        col_idx = col_idx * 26 + (ord(c) - 64)
                    
                    cell_value = sheet.cell(row=row_number, column=col_idx).value
                    if cell_value is None:
                        cell_value = 0
                    
                    expr = expr.replace(cell_ref, str(cell_value))
                
                # Evaluate the expression
                return eval(expr)
        
        except Exception:
            pass
        
        # If all else fails, return None
        return None

def clean_timecard_dataframe(df, date_col, employee_col, hours_col, job_col, columns):
    """
    Clean and normalize the timecard DataFrame.
    
    Args:
        df: The DataFrame to clean
        date_col: Index of the date column
        employee_col: Index of the employee column
        hours_col: Index of the hours column
        job_col: Index of the job column
        columns: List of column names
        
    Returns:
        DataFrame: Cleaned DataFrame
    """
    # Rename columns to standard names
    column_mapping = {}
    
    if date_col is not None:
        column_mapping[columns[date_col]] = 'date'
    
    if employee_col is not None:
        column_mapping[columns[employee_col]] = 'driver_name'
    
    if hours_col is not None:
        column_mapping[columns[hours_col]] = 'hours'
    
    if job_col is not None:
        column_mapping[columns[job_col]] = 'job_number'
    
    # Apply renaming
    df = df.rename(columns=column_mapping)
    
    # Convert date column to datetime
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
    
    # Ensure hours is numeric
    if 'hours' in df.columns:
        df['hours'] = pd.to_numeric(df['hours'], errors='coerce')
    
    # Drop rows with missing employee or date
    if 'driver_name' in df.columns and 'date' in df.columns:
        df = df.dropna(subset=['driver_name', 'date'])
    
    # Reset index
    df = df.reset_index(drop=True)
    
    return df

def filter_by_date(df, start_date=None, end_date=None):
    """
    Filter DataFrame by date range.
    
    Args:
        df: The DataFrame to filter
        start_date: Start date (inclusive)
        end_date: End date (inclusive)
        
    Returns:
        DataFrame: Filtered DataFrame
    """
    if 'date' not in df.columns:
        return df
    
    # Ensure dates column is datetime
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    
    # Filter by start date
    if start_date:
        start_dt = pd.to_datetime(start_date).normalize()
        df = df[df['date'] >= start_dt]
    
    # Filter by end date
    if end_date:
        end_dt = pd.to_datetime(end_date).normalize()
        df = df[df['date'] <= end_dt]
    
    return df

def process_with_pandas(file_path, start_date=None, end_date=None):
    """
    Process timecard file using pandas.
    
    Args:
        file_path: Path to the Excel file
        start_date: Start date for filtering (optional)
        end_date: End date for filtering (optional)
        
    Returns:
        DataFrame: Processed timecard data
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_path)
        
        # Clean column names
        df.columns = [clean_column_name(col) for col in df.columns]
        
        # Try to identify key columns
        date_col = employee_col = hours_col = job_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if any(term in col_lower for term in ['date', 'day', 'workday']):
                date_col = col
            elif any(term in col_lower for term in ['employee', 'name', 'driver', 'operator']):
                employee_col = col
            elif any(term in col_lower for term in ['hours', 'time', 'duration']):
                hours_col = col
            elif any(term in col_lower for term in ['job', 'project', 'site', 'location']):
                job_col = col
        
        # Rename columns to standard names
        column_mapping = {}
        
        if date_col:
            column_mapping[date_col] = 'date'
        
        if employee_col:
            column_mapping[employee_col] = 'driver_name'
        
        if hours_col:
            column_mapping[hours_col] = 'hours'
        
        if job_col:
            column_mapping[job_col] = 'job_number'
        
        # Apply renaming
        df = df.rename(columns=column_mapping)
        
        # Convert date column to datetime
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        # Ensure hours is numeric
        if 'hours' in df.columns:
            df['hours'] = pd.to_numeric(df['hours'], errors='coerce')
        
        # Filter by date if requested
        if start_date or end_date:
            df = filter_by_date(df, start_date, end_date)
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    except Exception as e:
        logger.error(f"Error processing with pandas: {str(e)}")
        return None
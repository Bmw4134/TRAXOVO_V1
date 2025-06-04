"""
Export Functions Utility Module

This module provides helper functions for generating and exporting various report formats,
including Excel, CSV, and PDF through a consistent interface.
"""

import os
import logging
import csv
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Set up logging
logger = logging.getLogger(__name__)

def ensure_exports_folder(subfolder=""):
    """Ensure the exports folder exists"""
    exports_path = os.path.join('exports', subfolder) if subfolder else 'exports'
    os.makedirs(exports_path, exist_ok=True)
    return exports_path
    
def generate_unique_filename(base_name, extension, subfolder=""):
    """Generate a unique filename with timestamp"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{base_name}_{timestamp}.{extension}"
    return filename

def format_column_headers(worksheet, headers):
    """Format column headers with styling"""
    # Define header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to headers
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Set column width based on header length (min 10, max 30)
        col_width = min(max(len(str(header)) + 2, 10), 30)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = col_width

def export_to_excel(data, filename, sheet_name="Data", subfolder=""):
    """
    Export data to Excel XLSX format
    
    Args:
        data (list): List of dictionaries containing row data
        filename (str): Output filename
        sheet_name (str): Name of the worksheet
        subfolder (str): Optional subfolder within exports directory
        
    Returns:
        str: Path to the created file
    """
    try:
        # Ensure exports folder exists
        exports_path = ensure_exports_folder(subfolder)
        file_path = os.path.join(exports_path, filename)
        
        # Create workbook and active sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # If data is empty, return with empty file
        if not data:
            workbook.save(file_path)
            logger.info(f"Created empty Excel file: {file_path}")
            return file_path
            
        # Get headers from first row dict keys
        headers = list(data[0].keys())
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
        
        # Format headers
        format_column_headers(worksheet, headers)
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save workbook
        workbook.save(file_path)
        logger.info(f"Exported Excel file: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        return None

def export_to_csv(data, filename, subfolder=""):
    """
    Export data to CSV format
    
    Args:
        data (list): List of dictionaries containing row data
        filename (str): Output filename
        subfolder (str): Optional subfolder within exports directory
        
    Returns:
        str: Path to the created file
    """
    try:
        # Ensure exports folder exists
        exports_path = ensure_exports_folder(subfolder)
        file_path = os.path.join(exports_path, filename)
        
        # If data is empty, return with empty file
        if not data:
            with open(file_path, 'w', newline='') as csvfile:
                pass
            logger.info(f"Created empty CSV file: {file_path}")
            return file_path
            
        # Get headers from first row dict keys
        headers = list(data[0].keys())
        
        # Write CSV file
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"Exported CSV file: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}")
        return None

def export_to_pdf(data, filename, title="Report", subfolder=""):
    """
    Export data to PDF format
    
    Args:
        data (list): List of dictionaries containing row data
        filename (str): Output filename
        title (str): Report title
        subfolder (str): Optional subfolder within exports directory
        
    Returns:
        str: Path to the created file
    """
    try:
        # Ensure exports folder exists
        exports_path = ensure_exports_folder(subfolder)
        file_path = os.path.join(exports_path, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Create elements list to build PDF
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # If data is empty, add note and return
        if not data:
            elements.append(Paragraph("No data available for this report.", styles['Normal']))
            doc.build(elements)
            logger.info(f"Created empty PDF file: {file_path}")
            return file_path
            
        # Get headers from first row dict keys
        headers = list(data[0].keys())
        
        # Prepare table data including headers
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(header, '')) for header in headers])
        
        # Create table
        table = Table(table_data)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # Add alternating row colors
        for row in range(1, len(table_data)):
            if row % 2 == 0:
                style.add('BACKGROUND', (0, row), (-1, row), colors.lightgrey)
        
        table.setStyle(style)
        elements.append(table)
        
        # Add generation timestamp
        timestamp_style = ParagraphStyle(
            name='TimestampStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.gray,
            alignment=1
        )
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Spacer(1, 0.25*inch))
        elements.append(Paragraph(f"Generated on: {timestamp}", timestamp_style))
        
        # Build PDF
        doc.build(elements)
        logger.info(f"Exported PDF file: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Error exporting to PDF: {str(e)}")
        return None

def dataframe_to_list(df):
    """
    Convert a pandas DataFrame to a list of dictionaries
    
    Args:
        df (pandas.DataFrame): Input DataFrame
        
    Returns:
        list: List of dictionaries (each representing a row)
    """
    if df is None or df.empty:
        return []
    
    return df.to_dict(orient='records')

def export_dataframe(df, filename, format_type='xlsx', subfolder="", title="Report"):
    """
    Export a pandas DataFrame to the specified format
    
    Args:
        df (pandas.DataFrame): DataFrame to export
        filename (str): Output filename (without extension)
        format_type (str): 'xlsx', 'csv', or 'pdf'
        subfolder (str): Optional subfolder within exports directory
        title (str): Report title (for PDF)
        
    Returns:
        str: Path to the created file or None if failed
    """
    # Convert DataFrame to list of dictionaries for consistent handling
    data = dataframe_to_list(df)
    
    # Add extension if not already present
    if not filename.lower().endswith(f".{format_type.lower()}"):
        filename = f"{filename}.{format_type.lower()}"
    
    # Export based on format type
    if format_type.lower() == 'xlsx':
        return export_to_excel(data, filename, subfolder=subfolder)
    elif format_type.lower() == 'csv':
        return export_to_csv(data, filename, subfolder=subfolder)
    elif format_type.lower() == 'pdf':
        return export_to_pdf(data, filename, title=title, subfolder=subfolder)
    else:
        logger.error(f"Unsupported export format: {format_type}")
        return None

def export_fsi_format(df, region, month, year, subfolder="foundation"):
    """
    Export data in Foundation System Import (FSI) format
    
    Args:
        df (pandas.DataFrame): DataFrame with the data to export
        region (str): Region code (DFW, HOU, WT)
        month (str): Month name
        year (str): Year
        subfolder (str): Optional subfolder within exports directory
        
    Returns:
        str: Path to the created file or None if failed
    """
    try:
        # Ensure exports folder exists
        exports_path = ensure_exports_folder(subfolder)
        filename = f"FSI_IMPORT_{region}_{month}_{year}.csv"
        file_path = os.path.join(exports_path, filename)
        
        # Export without header (Foundation format requirement)
        df.to_csv(file_path, index=False, header=False)
        logger.info(f"Exported FSI format file: {file_path}")
        return file_path
    
    except Exception as e:
        logger.error(f"Error exporting FSI format: {str(e)}")
        return None
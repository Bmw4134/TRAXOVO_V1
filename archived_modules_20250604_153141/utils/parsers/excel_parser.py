"""
Excel File Parser

This module provides functions for parsing Excel files with various structures, 
handling multiple sheets, detecting data areas, and normalizing content.
"""
import os
import re
import logging
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Tuple, Set, Union
from datetime import datetime

# Configure logging
logger = logging.getLogger(__name__)

# Constants for data detection
HEADER_ROW_THRESHOLD = 0.5  # Percentage of non-empty cells to consider a header row
DATA_CELL_THRESHOLD = 0.3   # Percentage of numeric cells to consider a data area
MIN_DATA_ROWS = 5           # Minimum number of rows to consider a valid data section
MAX_HEADER_ROWS = 10        # Maximum number of rows to check for headers
MIN_COLUMNS = 3             # Minimum number of columns to consider a valid data section

# Patterns for identifying different types of data
EMPLOYEE_PATTERNS = [
    r'employee',
    r'operator',
    r'driver',
    r'personnel',
    r'staff'
]

ASSET_PATTERNS = [
    r'asset',
    r'equipment',
    r'vehicle',
    r'machine',
    r'unit'
]

JOB_PATTERNS = [
    r'job',
    r'project',
    r'work\s*order',
    r'site'
]

TIME_PATTERNS = [
    r'hour',
    r'time',
    r'duration',
    r'period'
]

COST_PATTERNS = [
    r'cost',
    r'price',
    r'rate',
    r'amount',
    r'charge'
]

DATE_PATTERNS = [
    r'date',
    r'day',
    r'month',
    r'week',
    r'period'
]

def process_file(file_path: str, file_type: str = 'auto') -> Dict[str, Any]:
    """
    Process an Excel file and extract structured data
    
    Args:
        file_path (str): Path to the Excel file
        file_type (str): Type of file to guide processing
        
    Returns:
        Dict[str, Any]: Processing results with extracted data
    """
    try:
        # Initialize result structure
        result = {
            'success': False,
            'message': '',
            'file_path': file_path,
            'file_type': file_type,
            'sheets': {},
            'summary': {},
            'detected_structures': {},
            'extracted_data': {},
            'metadata': {}
        }
        
        # Check if file exists
        if not os.path.exists(file_path):
            result['message'] = f"File not found: {file_path}"
            return result
        
        # Basic file info
        file_info = {
            'size': os.path.getsize(file_path),
            'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
            'filename': os.path.basename(file_path)
        }
        result['metadata']['file_info'] = file_info
        
        # Get workbook info using openpyxl (for detailed metadata)
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            sheets_info = {}
            
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                # Get basic sheet info
                sheets_info[sheet_name] = {
                    'title': sheet.title,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                }
                
            result['metadata']['sheets_info'] = sheets_info
            wb.close()
        except Exception as e:
            logger.warning(f"Error getting detailed sheet info: {e}")
            # Still continue with pandas
        
        # Process the file with pandas
        # Try to read all sheets
        try:
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
            
            # Process each sheet
            for sheet_name in sheet_names:
                sheet_result = process_sheet(xl, sheet_name, file_type)
                result['sheets'][sheet_name] = sheet_result
            
            # Calculate overall summary
            summary = calculate_summary(result['sheets'])
            result['summary'] = summary
            
            # Extract specific data based on file type
            extracted_data = extract_specific_data(result, file_type)
            result['extracted_data'] = extracted_data
            
            # Detect common structures across sheets
            detected_structures = detect_structures(result['sheets'])
            result['detected_structures'] = detected_structures
            
            # Success!
            result['success'] = True
            result['message'] = f"Successfully processed {len(sheet_names)} sheets"
            result['record_count'] = summary.get('total_rows', 0)
            
            return result
            
        except Exception as e:
            result['message'] = f"Error processing Excel file: {str(e)}"
            logger.exception(f"Error processing Excel file {file_path}: {e}")
            return result
            
    except Exception as e:
        logger.exception(f"Unexpected error processing file {file_path}: {e}")
        return {
            'success': False,
            'message': f"Unexpected error: {str(e)}",
            'file_path': file_path,
            'file_type': file_type
        }

def process_sheet(xl: pd.ExcelFile, sheet_name: str, file_type: str) -> Dict[str, Any]:
    """
    Process a single sheet in an Excel file
    
    Args:
        xl (pd.ExcelFile): ExcelFile object
        sheet_name (str): Name of the sheet to process
        file_type (str): Type of file to guide processing
        
    Returns:
        Dict[str, Any]: Sheet processing results
    """
    result = {
        'name': sheet_name,
        'data_sections': [],
        'summary_sections': [],
        'header_rows': [],
        'column_types': {},
        'column_summaries': {},
        'total_rows': 0,
        'total_columns': 0,
        'empty_cells_percentage': 0,
        'numeric_cells_percentage': 0,
        'value_frequencies': {}
    }
    
    try:
        # First attempt - try with default parameters
        df = xl.parse(sheet_name)
        
        # If we have less than 5 rows or 3 columns, try reading with no header
        if len(df) < 5 or len(df.columns) < 3:
            df = xl.parse(sheet_name, header=None)
        
        # Store basic stats
        result['total_rows'] = len(df)
        result['total_columns'] = len(df.columns)
        
        # Count empty cells
        empty_cells = df.isna().sum().sum()
        total_cells = len(df) * len(df.columns)
        result['empty_cells_percentage'] = (empty_cells / total_cells * 100) if total_cells > 0 else 0
        
        # Count numeric cells
        numeric_cells = df.select_dtypes(include=['number']).count().sum()
        result['numeric_cells_percentage'] = (numeric_cells / total_cells * 100) if total_cells > 0 else 0
        
        # Detect data sections
        data_sections = detect_data_sections(df)
        result['data_sections'] = data_sections
        
        # Detect summary sections
        summary_sections = detect_summary_sections(df)
        result['summary_sections'] = summary_sections
        
        # Detect header rows
        header_rows = detect_header_rows(df)
        result['header_rows'] = header_rows
        
        # Determine column types
        column_types = analyze_column_types(df)
        result['column_types'] = column_types
        
        # Calculate column summaries (mean, median, min, max for numeric columns)
        column_summaries = calculate_column_summaries(df)
        result['column_summaries'] = column_summaries
        
        # Calculate value frequencies for categorical columns
        value_frequencies = calculate_value_frequencies(df)
        result['value_frequencies'] = value_frequencies
        
        # Extract sheet structure based on file type
        structure = extract_sheet_structure(df, file_type, sheet_name)
        result['structure'] = structure
        
        return result
        
    except Exception as e:
        logger.warning(f"Error processing sheet {sheet_name}: {e}")
        # Return partial result with error
        result['error'] = str(e)
        return result

def detect_data_sections(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Detect contiguous sections of data in a dataframe
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        List[Dict[str, Any]]: List of detected data sections
    """
    data_sections = []
    
    # Skip if dataframe is too small
    if len(df) < MIN_DATA_ROWS or len(df.columns) < MIN_COLUMNS:
        return data_sections
    
    # Create a mask of non-empty cells
    non_empty_mask = ~df.isna()
    
    # Calculate density by row (percentage of non-empty cells)
    row_density = non_empty_mask.mean(axis=1)
    
    # Find potential starting points for data sections
    potential_starts = []
    in_dense_section = False
    
    for i, density in enumerate(row_density):
        if not in_dense_section and density > DATA_CELL_THRESHOLD:
            potential_starts.append(i)
            in_dense_section = True
        elif in_dense_section and density < DATA_CELL_THRESHOLD:
            in_dense_section = False
    
    # For each potential start, find the end of the section
    for start in potential_starts:
        end = start
        while end < len(row_density) and row_density[end] > DATA_CELL_THRESHOLD:
            end += 1
        
        # Only consider sections with enough rows
        if end - start >= MIN_DATA_ROWS:
            # Look for header row(s) before the data section
            header_row = max(0, start - 1)
            for i in range(start - 1, max(0, start - MAX_HEADER_ROWS), -1):
                if row_density[i] > HEADER_ROW_THRESHOLD:
                    header_row = i
                    break
            
            # Detect columns with data (non-empty)
            data_columns = []
            for col in df.columns:
                col_data = df.loc[start:end-1, col]
                if col_data.notna().mean() > DATA_CELL_THRESHOLD:
                    data_columns.append(col)
            
            # Add the section if it has enough data columns
            if len(data_columns) >= MIN_COLUMNS:
                data_sections.append({
                    'start_row': start,
                    'end_row': end,
                    'header_row': header_row,
                    'row_count': end - start,
                    'column_count': len(data_columns),
                    'data_columns': data_columns,
                    'non_empty_percentage': non_empty_mask.loc[start:end-1, data_columns].mean().mean() * 100
                })
    
    return data_sections

def detect_summary_sections(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Detect summary sections (like totals, averages) in a dataframe
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        List[Dict[str, Any]]: List of detected summary sections
    """
    summary_sections = []
    
    # Skip if dataframe is too small
    if len(df) < 3 or len(df.columns) < 3:
        return summary_sections
    
    # Keywords that often appear in summary rows
    summary_keywords = [
        'total', 'sum', 'average', 'avg', 'mean', 'subtotal', 
        'grand total', 'count', 'min', 'max'
    ]
    
    # Look for rows with summary keywords
    for i in range(len(df)):
        row = df.iloc[i]
        row_str = ' '.join(str(x).lower() for x in row if not pd.isna(x))
        
        contains_summary = False
        for keyword in summary_keywords:
            if keyword in row_str:
                contains_summary = True
                break
        
        # Check if this row has more numeric content than surrounding rows
        if i > 0 and i < len(df) - 1:
            prev_row = df.iloc[i-1]
            next_row = df.iloc[i+1]
            
            # Count numeric cells in current and surrounding rows
            row_numeric = sum(1 for x in row if isinstance(x, (int, float)) and not pd.isna(x))
            prev_numeric = sum(1 for x in prev_row if isinstance(x, (int, float)) and not pd.isna(x))
            next_numeric = sum(1 for x in next_row if isinstance(x, (int, float)) and not pd.isna(x))
            
            # Check for patterns like empty row, then summary row
            if row_numeric > 0 and (prev_numeric == 0 or next_numeric == 0):
                contains_summary = True
        
        if contains_summary:
            # Determine which columns contain summary values
            summary_columns = []
            for j, col in enumerate(df.columns):
                cell = row[j]
                if isinstance(cell, (int, float)) and not pd.isna(cell):
                    summary_columns.append(col)
            
            # Only add if we have at least one summary column
            if summary_columns:
                summary_sections.append({
                    'row': i,
                    'columns': summary_columns,
                    'text': row_str[:100] + ('...' if len(row_str) > 100 else '')
                })
    
    return summary_sections

def detect_header_rows(df: pd.DataFrame) -> List[int]:
    """
    Detect potential header rows in a dataframe
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        List[int]: List of row indices that appear to be headers
    """
    header_rows = []
    
    # Skip if dataframe is too small
    if len(df) < 3:
        return header_rows
    
    # Look at the first few rows
    for i in range(min(MAX_HEADER_ROWS, len(df))):
        row = df.iloc[i]
        
        # Header characteristics:
        # 1. Mostly string values
        # 2. Not many NaN values
        # 3. Often shorter than data rows
        str_count = sum(1 for x in row if isinstance(x, str))
        nan_count = sum(1 for x in row if pd.isna(x))
        
        # Calculate the percentage of string values in non-NaN cells
        non_nan_cells = len(row) - nan_count
        str_percentage = (str_count / non_nan_cells) if non_nan_cells > 0 else 0
        
        # Calculate the percentage of non-empty cells
        non_empty_percentage = (non_nan_cells / len(row)) if len(row) > 0 else 0
        
        # Check if the next row has more numeric values (suggesting this is a header)
        next_numeric_percentage = 0
        if i < len(df) - 1:
            next_row = df.iloc[i+1]
            next_numeric = sum(1 for x in next_row if isinstance(x, (int, float)) and not pd.isna(x))
            next_numeric_percentage = (next_numeric / len(next_row)) if len(next_row) > 0 else 0
        
        # Combine factors to decide if this is a header row
        header_score = (str_percentage * 0.5 + 
                       non_empty_percentage * 0.3 + 
                       next_numeric_percentage * 0.2)
        
        if header_score > 0.5 and non_empty_percentage > HEADER_ROW_THRESHOLD:
            header_rows.append(i)
    
    return header_rows

def analyze_column_types(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Analyze data types and characteristics of each column
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        Dict[str, Dict[str, Any]]: Dictionary of column information
    """
    column_types = {}
    
    for col in df.columns:
        column_data = df[col]
        non_null_data = column_data.dropna()
        
        # Skip if no data
        if len(non_null_data) == 0:
            column_types[str(col)] = {
                'data_type': 'unknown',
                'empty': True,
                'unique_count': 0,
                'null_percentage': 100.0
            }
            continue
        
        # Basic stats
        null_percentage = (column_data.isna().sum() / len(column_data)) * 100
        unique_count = non_null_data.nunique()
        unique_percentage = (unique_count / len(non_null_data)) * 100 if len(non_null_data) > 0 else 0
        
        # Determine column purpose
        column_purpose = 'data'
        column_str = str(col).lower()
        
        # Check for ID columns
        if ('id' in column_str or column_str.endswith('no') or column_str.endswith('number')):
            if unique_percentage > 80:
                column_purpose = 'identifier'
        
        # Check for date columns
        date_match = any(re.search(pattern, column_str) for pattern in DATE_PATTERNS)
        try_date = False
        is_date = False
        
        if date_match or (non_null_data.dtype == 'object' and unique_percentage > 0):
            try_date = True
            
        if try_date:
            # Try to convert to datetime
            try:
                pd.to_datetime(non_null_data, errors='raise')
                column_purpose = 'date'
                is_date = True
            except:
                pass
        
        # Check for numeric columns
        is_numeric = pd.api.types.is_numeric_dtype(non_null_data.dtype)
        
        # Determine more specific data type
        if is_date:
            data_type = 'date'
        elif is_numeric:
            if non_null_data.dtype == 'int64' or non_null_data.dtype == 'int32':
                data_type = 'integer'
            else:
                data_type = 'float'
                
            # Check if this is a cost/money column
            cost_match = any(re.search(pattern, column_str) for pattern in COST_PATTERNS)
            if cost_match:
                column_purpose = 'cost'
            
            # Check if this is a time/hours column
            time_match = any(re.search(pattern, column_str) for pattern in TIME_PATTERNS)
            if time_match:
                column_purpose = 'time'
        elif unique_percentage < 10 and len(non_null_data) >= 10:
            data_type = 'categorical'
        else:
            data_type = 'text'
            
            # Check if this is an employee column
            employee_match = any(re.search(pattern, column_str) for pattern in EMPLOYEE_PATTERNS)
            if employee_match:
                column_purpose = 'employee'
            
            # Check if this is an asset column
            asset_match = any(re.search(pattern, column_str) for pattern in ASSET_PATTERNS)
            if asset_match:
                column_purpose = 'asset'
            
            # Check if this is a job/project column
            job_match = any(re.search(pattern, column_str) for pattern in JOB_PATTERNS)
            if job_match:
                column_purpose = 'job'
        
        # Store column info
        column_types[str(col)] = {
            'data_type': data_type,
            'purpose': column_purpose,
            'unique_count': int(unique_count),
            'unique_percentage': float(unique_percentage),
            'null_percentage': float(null_percentage),
            'is_numeric': is_numeric,
            'is_date': is_date
        }
    
    return column_types

def calculate_column_summaries(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Calculate statistical summaries for numeric columns
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        Dict[str, Dict[str, Any]]: Dictionary of column summaries
    """
    column_summaries = {}
    
    # Process numeric columns
    numeric_df = df.select_dtypes(include=['number'])
    
    for col in numeric_df.columns:
        try:
            # Calculate basic statistics
            stats = {
                'mean': float(numeric_df[col].mean()),
                'median': float(numeric_df[col].median()),
                'min': float(numeric_df[col].min()),
                'max': float(numeric_df[col].max()),
                'sum': float(numeric_df[col].sum()),
                'std': float(numeric_df[col].std())
            }
            
            # Add percentiles
            stats['percentile_25'] = float(numeric_df[col].quantile(0.25))
            stats['percentile_75'] = float(numeric_df[col].quantile(0.75))
            
            column_summaries[str(col)] = stats
        except Exception as e:
            logger.warning(f"Error calculating summary for column {col}: {e}")
            column_summaries[str(col)] = {'error': str(e)}
    
    return column_summaries

def calculate_value_frequencies(df: pd.DataFrame) -> Dict[str, Dict[str, int]]:
    """
    Calculate value frequencies for categorical and low-cardinality columns
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        Dict[str, Dict[str, int]]: Dictionary of value frequencies
    """
    value_frequencies = {}
    
    # Limit to object, boolean and low-cardinality numeric columns
    for col in df.columns:
        # Skip if too many unique values
        if df[col].nunique() > 50 or df[col].nunique() <= 1:
            continue
        
        try:
            # Calculate value frequencies
            freq = df[col].value_counts().head(20).to_dict()
            
            # Convert keys to strings (for JSON serialization)
            freq_str = {str(k): int(v) for k, v in freq.items()}
            value_frequencies[str(col)] = freq_str
        except Exception as e:
            logger.warning(f"Error calculating frequencies for column {col}: {e}")
    
    return value_frequencies

def extract_sheet_structure(df: pd.DataFrame, file_type: str, sheet_name: str) -> Dict[str, Any]:
    """
    Extract the structure of a sheet based on the file type
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        file_type (str): Type of file to guide extraction
        sheet_name (str): Name of the sheet
        
    Returns:
        Dict[str, Any]: Structure information
    """
    structure = {
        'type': 'generic',
        'attributes': {}
    }
    
    # Determine sheet type based on name and content
    sheet_name_lower = sheet_name.lower()
    
    # Check for summary or data sheets
    if 'summary' in sheet_name_lower or 'total' in sheet_name_lower:
        structure['type'] = 'summary'
    elif 'data' in sheet_name_lower or 'detail' in sheet_name_lower:
        structure['type'] = 'detail'
    elif 'config' in sheet_name_lower or 'setup' in sheet_name_lower:
        structure['type'] = 'configuration'
    
    # Check for file type specific structures
    if file_type == 'fringe':
        # Look for employee benefit information
        employee_cols = 0
        benefit_cols = 0
        
        for col in df.columns:
            col_str = str(col).lower()
            if any(re.search(pattern, col_str) for pattern in EMPLOYEE_PATTERNS):
                employee_cols += 1
            if 'benefit' in col_str or 'fringe' in col_str:
                benefit_cols += 1
        
        if employee_cols > 0 and benefit_cols > 0:
            structure['type'] = 'employee_benefits'
            structure['attributes']['employee_columns'] = employee_cols
            structure['attributes']['benefit_columns'] = benefit_cols
    
    elif file_type == 'billing':
        # Look for billing information
        if any('cost' in str(col).lower() or 'rate' in str(col).lower() for col in df.columns):
            structure['type'] = 'billing'
            
            # Check if this is a rate sheet or charges sheet
            rate_cols = sum(1 for col in df.columns if 'rate' in str(col).lower())
            charge_cols = sum(1 for col in df.columns if 'charge' in str(col).lower() or 'amount' in str(col).lower())
            
            if rate_cols > charge_cols:
                structure['attributes']['billing_type'] = 'rates'
            else:
                structure['attributes']['billing_type'] = 'charges'
    
    elif file_type == 'activity':
        # Look for activity/usage information
        time_cols = sum(1 for col in df.columns if any(re.search(pattern, str(col).lower()) for pattern in TIME_PATTERNS))
        date_cols = sum(1 for col in df.columns if any(re.search(pattern, str(col).lower()) for pattern in DATE_PATTERNS))
        
        if time_cols > 0 and date_cols > 0:
            structure['type'] = 'activity'
            structure['attributes']['time_columns'] = time_cols
            structure['attributes']['date_columns'] = date_cols
    
    elif file_type == 'maintenance':
        # Look for work order information
        if any('work' in str(col).lower() or 'order' in str(col).lower() or 'wo' in str(col).lower() for col in df.columns):
            structure['type'] = 'maintenance'
            
            # Check if this is a work order detail or summary
            detail_indicator = any('detail' in str(col).lower() for col in df.columns)
            if detail_indicator:
                structure['attributes']['maintenance_type'] = 'detail'
            else:
                structure['attributes']['maintenance_type'] = 'summary'
    
    return structure

def calculate_summary(sheets_data: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calculate summary statistics across all sheets
    
    Args:
        sheets_data (Dict[str, Dict[str, Any]]): Data from all sheets
        
    Returns:
        Dict[str, Any]: Summary statistics
    """
    summary = {
        'total_sheets': len(sheets_data),
        'total_rows': 0,
        'total_columns': 0,
        'total_data_sections': 0,
        'sheet_types': {},
        'largest_sheet': '',
        'largest_sheet_rows': 0,
        'column_purposes': {}
    }
    
    for sheet_name, sheet_data in sheets_data.items():
        # Count rows and columns
        summary['total_rows'] += sheet_data.get('total_rows', 0)
        
        # Track largest sheet
        if sheet_data.get('total_rows', 0) > summary['largest_sheet_rows']:
            summary['largest_sheet_rows'] = sheet_data.get('total_rows', 0)
            summary['largest_sheet'] = sheet_name
        
        # Count data sections
        summary['total_data_sections'] += len(sheet_data.get('data_sections', []))
        
        # Count sheet types
        sheet_type = sheet_data.get('structure', {}).get('type', 'generic')
        if sheet_type in summary['sheet_types']:
            summary['sheet_types'][sheet_type] += 1
        else:
            summary['sheet_types'][sheet_type] = 1
        
        # Count column purposes
        for col, col_info in sheet_data.get('column_types', {}).items():
            purpose = col_info.get('purpose', 'data')
            if purpose in summary['column_purposes']:
                summary['column_purposes'][purpose] += 1
            else:
                summary['column_purposes'][purpose] = 1
    
    return summary

def extract_specific_data(result: Dict[str, Any], file_type: str) -> Dict[str, Any]:
    """
    Extract specific data based on file type
    
    Args:
        result (Dict[str, Any]): Processing results with sheet data
        file_type (str): Type of file to guide extraction
        
    Returns:
        Dict[str, Any]: Extracted specific data
    """
    extracted_data = {
        'key_values': {},
        'main_entities': [],
        'relationships': [],
        'time_periods': []
    }
    
    if file_type == 'fringe':
        return extract_fringe_data(result)
    elif file_type == 'billing':
        return extract_billing_data(result)
    elif file_type == 'activity':
        return extract_activity_data(result)
    elif file_type == 'maintenance':
        return extract_maintenance_data(result)
    else:
        # Generic extraction for unknown file types
        return extract_generic_data(result)

def extract_fringe_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract data specific to fringe benefit files
    
    Args:
        result (Dict[str, Any]): Processing results
        
    Returns:
        Dict[str, Any]: Extracted fringe data
    """
    extracted_data = {
        'employees': [],
        'benefits': [],
        'employee_benefits': [],
        'period': '',
        'totals': {}
    }
    
    # Try to determine period from filename or sheet names
    filename = result.get('metadata', {}).get('file_info', {}).get('filename', '')
    if 'jan' in filename.lower() or 'january' in filename.lower():
        extracted_data['period'] = 'January'
    elif 'feb' in filename.lower() or 'february' in filename.lower():
        extracted_data['period'] = 'February'
    elif 'mar' in filename.lower() or 'march' in filename.lower():
        extracted_data['period'] = 'March'
    elif 'apr' in filename.lower() or 'april' in filename.lower():
        extracted_data['period'] = 'April'
    elif 'may' in filename.lower():
        extracted_data['period'] = 'May'
    elif 'jun' in filename.lower() or 'june' in filename.lower():
        extracted_data['period'] = 'June'
    elif 'jul' in filename.lower() or 'july' in filename.lower():
        extracted_data['period'] = 'July'
    elif 'aug' in filename.lower() or 'august' in filename.lower():
        extracted_data['period'] = 'August'
    elif 'sep' in filename.lower() or 'september' in filename.lower():
        extracted_data['period'] = 'September'
    elif 'oct' in filename.lower() or 'october' in filename.lower():
        extracted_data['period'] = 'October'
    elif 'nov' in filename.lower() or 'november' in filename.lower():
        extracted_data['period'] = 'November'
    elif 'dec' in filename.lower() or 'december' in filename.lower():
        extracted_data['period'] = 'December'
    
    # Extract year if present in filename
    year_match = re.search(r'20\d{2}', filename)
    if year_match:
        extracted_data['year'] = year_match.group(0)
    
    # Analyze each sheet for employee and benefit information
    for sheet_name, sheet_data in result.get('sheets', {}).items():
        if sheet_data.get('structure', {}).get('type') == 'employee_benefits':
            # Extract employee columns
            employee_cols = []
            benefit_cols = []
            
            for col, col_info in sheet_data.get('column_types', {}).items():
                if col_info.get('purpose') == 'employee':
                    employee_cols.append(col)
                if 'benefit' in col.lower() or 'fringe' in col.lower():
                    benefit_cols.append(col)
            
            # Extract employees
            for col in employee_cols:
                employees = sheet_data.get('value_frequencies', {}).get(col, {})
                if employees:
                    # Add non-empty employees
                    for emp, count in employees.items():
                        if emp and emp.lower() != 'nan' and emp.lower() != 'total':
                            if emp not in extracted_data['employees']:
                                extracted_data['employees'].append(emp)
            
            # Extract benefits
            for col in benefit_cols:
                if col not in extracted_data['benefits']:
                    extracted_data['benefits'].append(col)
            
            # Extract totals
            for section in sheet_data.get('summary_sections', []):
                for col in section.get('columns', []):
                    if col in benefit_cols:
                        extracted_data['totals'][col] = True
    
    return extracted_data

def extract_billing_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract data specific to billing files
    
    Args:
        result (Dict[str, Any]): Processing results
        
    Returns:
        Dict[str, Any]: Extracted billing data
    """
    extracted_data = {
        'assets': [],
        'rates': {},
        'charges': {},
        'period': '',
        'total_billed': 0,
        'districts': []
    }
    
    # Try to determine period from filename
    filename = result.get('metadata', {}).get('file_info', {}).get('filename', '')
    
    # Extract month/period
    month_match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)', filename.lower())
    if month_match:
        month = month_match.group(0)
        if month == 'jan' or month == 'january':
            extracted_data['period'] = 'January'
        elif month == 'feb' or month == 'february':
            extracted_data['period'] = 'February'
        elif month == 'mar' or month == 'march':
            extracted_data['period'] = 'March'
        elif month == 'apr' or month == 'april':
            extracted_data['period'] = 'April'
        elif month == 'may':
            extracted_data['period'] = 'May'
        elif month == 'jun' or month == 'june':
            extracted_data['period'] = 'June'
        elif month == 'jul' or month == 'july':
            extracted_data['period'] = 'July'
        elif month == 'aug' or month == 'august':
            extracted_data['period'] = 'August'
        elif month == 'sep' or month == 'september':
            extracted_data['period'] = 'September'
        elif month == 'oct' or month == 'october':
            extracted_data['period'] = 'October'
        elif month == 'nov' or month == 'november':
            extracted_data['period'] = 'November'
        elif month == 'dec' or month == 'december':
            extracted_data['period'] = 'December'
    
    # Extract year if present
    year_match = re.search(r'20\d{2}', filename)
    if year_match:
        extracted_data['year'] = year_match.group(0)
    
    # Analyze sheets for billing information
    for sheet_name, sheet_data in result.get('sheets', {}).items():
        if sheet_data.get('structure', {}).get('type') == 'billing':
            billing_type = sheet_data.get('structure', {}).get('attributes', {}).get('billing_type', '')
            
            # Find asset columns
            asset_cols = []
            rate_cols = []
            charge_cols = []
            district_cols = []
            
            for col, col_info in sheet_data.get('column_types', {}).items():
                if col_info.get('purpose') == 'asset':
                    asset_cols.append(col)
                if 'rate' in col.lower() and col_info.get('is_numeric'):
                    rate_cols.append(col)
                if ('charge' in col.lower() or 'amount' in col.lower() or 'total' in col.lower()) and col_info.get('is_numeric'):
                    charge_cols.append(col)
                if 'district' in col.lower():
                    district_cols.append(col)
            
            # Extract assets
            for col in asset_cols:
                assets = sheet_data.get('value_frequencies', {}).get(col, {})
                if assets:
                    # Add non-empty assets
                    for asset, count in assets.items():
                        if asset and asset.lower() != 'nan' and asset.lower() != 'total':
                            if asset not in extracted_data['assets']:
                                extracted_data['assets'].append(asset)
            
            # Extract districts
            for col in district_cols:
                districts = sheet_data.get('value_frequencies', {}).get(col, {})
                if districts:
                    for district, count in districts.items():
                        if district and district.lower() != 'nan' and district.lower() != 'total':
                            if district not in extracted_data['districts']:
                                extracted_data['districts'].append(district)
            
            # Extract rate information
            for col in rate_cols:
                extracted_data['rates'][col] = sheet_data.get('column_summaries', {}).get(col, {})
            
            # Extract charge information
            for col in charge_cols:
                extracted_data['charges'][col] = sheet_data.get('column_summaries', {}).get(col, {})
                # Add to total billed if this looks like a total column
                if 'total' in col.lower() and 'sum' in sheet_data.get('column_summaries', {}).get(col, {}):
                    extracted_data['total_billed'] += sheet_data.get('column_summaries', {}).get(col, {}).get('sum', 0)
    
    return extracted_data

def extract_activity_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract data specific to activity/usage files
    
    Args:
        result (Dict[str, Any]): Processing results
        
    Returns:
        Dict[str, Any]: Extracted activity data
    """
    extracted_data = {
        'assets': [],
        'employees': [],
        'jobs': [],
        'time_periods': [],
        'hour_totals': {},
        'efficiency_metrics': {}
    }
    
    # Try to determine period from filename
    filename = result.get('metadata', {}).get('file_info', {}).get('filename', '')
    
    # Extract month/period
    month_match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)', filename.lower())
    if month_match:
        month = month_match.group(0)
        if month in ['jan', 'january']:
            extracted_data['period'] = 'January'
        elif month in ['feb', 'february']:
            extracted_data['period'] = 'February'
        elif month in ['mar', 'march']:
            extracted_data['period'] = 'March'
        elif month in ['apr', 'april']:
            extracted_data['period'] = 'April'
        elif month == 'may':
            extracted_data['period'] = 'May'
        elif month in ['jun', 'june']:
            extracted_data['period'] = 'June'
        elif month in ['jul', 'july']:
            extracted_data['period'] = 'July'
        elif month in ['aug', 'august']:
            extracted_data['period'] = 'August'
        elif month in ['sep', 'september']:
            extracted_data['period'] = 'September'
        elif month in ['oct', 'october']:
            extracted_data['period'] = 'October'
        elif month in ['nov', 'november']:
            extracted_data['period'] = 'November'
        elif month in ['dec', 'december']:
            extracted_data['period'] = 'December'
    
    # Extract year if present
    year_match = re.search(r'20\d{2}', filename)
    if year_match:
        extracted_data['year'] = year_match.group(0)
    
    # Analyze sheets for activity information
    for sheet_name, sheet_data in result.get('sheets', {}).items():
        if sheet_data.get('structure', {}).get('type') == 'activity':
            # Find relevant columns
            asset_cols = []
            employee_cols = []
            job_cols = []
            time_cols = []
            
            for col, col_info in sheet_data.get('column_types', {}).items():
                if col_info.get('purpose') == 'asset':
                    asset_cols.append(col)
                if col_info.get('purpose') == 'employee':
                    employee_cols.append(col)
                if col_info.get('purpose') == 'job':
                    job_cols.append(col)
                if col_info.get('purpose') == 'time':
                    time_cols.append(col)
            
            # Extract assets
            for col in asset_cols:
                assets = sheet_data.get('value_frequencies', {}).get(col, {})
                if assets:
                    for asset, count in assets.items():
                        if asset and asset.lower() != 'nan' and asset.lower() != 'total':
                            if asset not in extracted_data['assets']:
                                extracted_data['assets'].append(asset)
            
            # Extract employees
            for col in employee_cols:
                employees = sheet_data.get('value_frequencies', {}).get(col, {})
                if employees:
                    for emp, count in employees.items():
                        if emp and emp.lower() != 'nan' and emp.lower() != 'total':
                            if emp not in extracted_data['employees']:
                                extracted_data['employees'].append(emp)
            
            # Extract jobs
            for col in job_cols:
                jobs = sheet_data.get('value_frequencies', {}).get(col, {})
                if jobs:
                    for job, count in jobs.items():
                        if job and job.lower() != 'nan' and job.lower() != 'total':
                            if job not in extracted_data['jobs']:
                                extracted_data['jobs'].append(job)
            
            # Extract time totals
            for col in time_cols:
                extracted_data['hour_totals'][col] = sheet_data.get('column_summaries', {}).get(col, {}).get('sum', 0)
                
                # Check for efficiency metrics
                if 'efficiency' in col.lower() or 'utilized' in col.lower() or 'productive' in col.lower():
                    extracted_data['efficiency_metrics'][col] = {
                        'mean': sheet_data.get('column_summaries', {}).get(col, {}).get('mean', 0),
                        'min': sheet_data.get('column_summaries', {}).get(col, {}).get('min', 0),
                        'max': sheet_data.get('column_summaries', {}).get(col, {}).get('max', 0)
                    }
    
    return extracted_data

def extract_maintenance_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract data specific to maintenance/work order files
    
    Args:
        result (Dict[str, Any]): Processing results
        
    Returns:
        Dict[str, Any]: Extracted maintenance data
    """
    extracted_data = {
        'work_orders': [],
        'assets': [],
        'service_types': [],
        'costs': {},
        'date_range': {},
        'districts': []
    }
    
    # Try to determine period from filename
    filename = result.get('metadata', {}).get('file_info', {}).get('filename', '')
    
    # Extract date range if present
    date_matches = re.findall(r'\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}', filename)
    if len(date_matches) >= 2:
        extracted_data['date_range']['start'] = date_matches[0]
        extracted_data['date_range']['end'] = date_matches[1]
    
    # Extract year if present
    year_match = re.search(r'20\d{2}', filename)
    if year_match:
        extracted_data['year'] = year_match.group(0)
    
    # Analyze sheets for maintenance information
    for sheet_name, sheet_data in result.get('sheets', {}).items():
        if sheet_data.get('structure', {}).get('type') == 'maintenance':
            # Find relevant columns
            wo_cols = []
            asset_cols = []
            service_cols = []
            cost_cols = []
            district_cols = []
            
            for col, col_info in sheet_data.get('column_types', {}).items():
                col_lower = col.lower()
                if 'work' in col_lower or 'order' in col_lower or 'wo' in col_lower or col_info.get('purpose') == 'identifier':
                    wo_cols.append(col)
                if col_info.get('purpose') == 'asset':
                    asset_cols.append(col)
                if 'service' in col_lower or 'type' in col_lower or 'repair' in col_lower:
                    service_cols.append(col)
                if col_info.get('purpose') == 'cost':
                    cost_cols.append(col)
                if 'district' in col_lower:
                    district_cols.append(col)
            
            # Extract work orders
            for col in wo_cols:
                work_orders = sheet_data.get('value_frequencies', {}).get(col, {})
                if work_orders:
                    for wo, count in work_orders.items():
                        if wo and wo.lower() != 'nan' and wo.lower() != 'total':
                            if wo not in extracted_data['work_orders']:
                                extracted_data['work_orders'].append(wo)
            
            # Extract assets
            for col in asset_cols:
                assets = sheet_data.get('value_frequencies', {}).get(col, {})
                if assets:
                    for asset, count in assets.items():
                        if asset and asset.lower() != 'nan' and asset.lower() != 'total':
                            if asset not in extracted_data['assets']:
                                extracted_data['assets'].append(asset)
            
            # Extract service types
            for col in service_cols:
                services = sheet_data.get('value_frequencies', {}).get(col, {})
                if services:
                    for service, count in services.items():
                        if service and service.lower() != 'nan' and service.lower() != 'total':
                            if service not in extracted_data['service_types']:
                                extracted_data['service_types'].append(service)
            
            # Extract districts
            for col in district_cols:
                districts = sheet_data.get('value_frequencies', {}).get(col, {})
                if districts:
                    for district, count in districts.items():
                        if district and district.lower() != 'nan' and district.lower() != 'total':
                            if district not in extracted_data['districts']:
                                extracted_data['districts'].append(district)
            
            # Extract cost information
            for col in cost_cols:
                extracted_data['costs'][col] = sheet_data.get('column_summaries', {}).get(col, {})
    
    return extracted_data

def extract_generic_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract generic data for unknown file types
    
    Args:
        result (Dict[str, Any]): Processing results
        
    Returns:
        Dict[str, Any]: Extracted generic data
    """
    extracted_data = {
        'key_entities': [],
        'numeric_totals': {},
        'possible_date_range': {},
        'key_columns': []
    }
    
    # Go through each sheet
    for sheet_name, sheet_data in result.get('sheets', {}).items():
        # Find important columns
        for col, col_info in sheet_data.get('column_types', {}).items():
            purpose = col_info.get('purpose', '')
            
            # Track important columns
            if purpose in ['identifier', 'asset', 'employee', 'job']:
                if col not in extracted_data['key_columns']:
                    extracted_data['key_columns'].append(col)
            
            # Extract entities
            if purpose in ['asset', 'employee', 'job']:
                entities = sheet_data.get('value_frequencies', {}).get(col, {})
                if entities:
                    for entity, count in entities.items():
                        if entity and entity.lower() != 'nan' and entity.lower() != 'total':
                            entity_info = {
                                'name': entity,
                                'type': purpose,
                                'source_column': col,
                                'count': count
                            }
                            if entity_info not in extracted_data['key_entities']:
                                extracted_data['key_entities'].append(entity_info)
            
            # Track date ranges
            if col_info.get('is_date', False):
                date_summary = sheet_data.get('column_summaries', {}).get(col, {})
                if 'min' in date_summary and 'max' in date_summary:
                    extracted_data['possible_date_range'][col] = {
                        'min': date_summary['min'],
                        'max': date_summary['max']
                    }
            
            # Track numeric totals
            if col_info.get('is_numeric', False) and col_info.get('purpose') in ['cost', 'time']:
                extracted_data['numeric_totals'][col] = sheet_data.get('column_summaries', {}).get(col, {}).get('sum', 0)
    
    return extracted_data

def detect_structures(sheets_data: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """
    Detect common structures across sheets
    
    Args:
        sheets_data (Dict[str, Dict[str, Any]]): Data from all sheets
        
    Returns:
        Dict[str, Any]: Detected structures
    """
    structures = {
        'common_columns': {},
        'hierarchical_sheets': [],
        'possible_relationships': []
    }
    
    # Collect all columns from all sheets
    all_columns = {}
    for sheet_name, sheet_data in sheets_data.items():
        for col in sheet_data.get('column_types', {}):
            col_str = str(col).lower()
            if col_str in all_columns:
                all_columns[col_str]['count'] += 1
                all_columns[col_str]['sheets'].append(sheet_name)
            else:
                all_columns[col_str] = {
                    'count': 1,
                    'sheets': [sheet_name],
                    'original': col
                }
    
    # Identify common columns (in multiple sheets)
    for col, col_data in all_columns.items():
        if col_data['count'] > 1:
            structures['common_columns'][col] = col_data
    
    # Detect possible relationships between sheets
    sheet_names = list(sheets_data.keys())
    for i in range(len(sheet_names)):
        for j in range(i + 1, len(sheet_names)):
            sheet1 = sheet_names[i]
            sheet2 = sheet_names[j]
            
            # Find common columns
            common_cols = []
            sheet1_cols = set(str(col).lower() for col in sheets_data[sheet1].get('column_types', {}))
            sheet2_cols = set(str(col).lower() for col in sheets_data[sheet2].get('column_types', {}))
            
            for col in sheet1_cols.intersection(sheet2_cols):
                common_cols.append(col)
            
            if common_cols:
                structures['possible_relationships'].append({
                    'sheet1': sheet1,
                    'sheet2': sheet2,
                    'common_columns': common_cols
                })
    
    # Detect hierarchical relationships (summary/detail)
    for sheet_name, sheet_data in sheets_data.items():
        sheet_type = sheet_data.get('structure', {}).get('type', '')
        if sheet_type in ['summary', 'totals']:
            # Look for corresponding detail sheets
            for other_sheet, other_data in sheets_data.items():
                if other_sheet != sheet_name and other_data.get('structure', {}).get('type', '') == 'detail':
                    # Check if they have columns in common
                    sheet_cols = set(str(col).lower() for col in sheet_data.get('column_types', {}))
                    other_cols = set(str(col).lower() for col in other_data.get('column_types', {}))
                    
                    common_cols = sheet_cols.intersection(other_cols)
                    if common_cols:
                        structures['hierarchical_sheets'].append({
                            'summary_sheet': sheet_name,
                            'detail_sheet': other_sheet,
                            'common_columns': list(common_cols)
                        })
    
    return structures
"""
PM Allocation Processor

This module handles the comparison of PM billing allocation files to identify
changes between original and updated versions, preserving formulas and formatting.
"""

import os
import re
import json
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure directory paths
UPLOADS_DIR = Path('uploads')
EXPORTS_DIR = Path('exports')
ATTACHED_ASSETS_DIR = Path('attached_assets')

# Ensure directories exist
UPLOADS_DIR.mkdir(exist_ok=True)
EXPORTS_DIR.mkdir(exist_ok=True)
ATTACHED_ASSETS_DIR.mkdir(exist_ok=True)

# Define column patterns for matching
EQ_ID_PATTERNS = [r'EQ#', r'EQ #', r'EQUIPMENT #', r'EQUIP #', r'EQUIP. #']
DESC_PATTERNS = [r'DESCRIPTION', r'DESC', r'EQUIPMENT DESC']
COST_PATTERNS = [r'COST', r'AMOUNT', r'TOTAL']
JOB_PATTERNS = [r'JOB', r'JOB #', r'PROJECT']

def find_matching_columns(df, patterns):
    """Find column indices that match the given patterns"""
    matching_cols = []
    for col_idx, col_name in enumerate(df.columns):
        if isinstance(col_name, str):
            for pattern in patterns:
                if re.search(pattern, col_name, re.IGNORECASE):
                    matching_cols.append(col_idx)
                    break
    return matching_cols

def clean_asset_id(asset_id):
    """Clean and standardize asset ID"""
    if pd.isna(asset_id) or not asset_id:
        return ""
    
    # Convert to string
    asset_id = str(asset_id).strip()
    
    # Remove any non-alphanumeric characters except dash
    asset_id = re.sub(r'[^\w\-]', '', asset_id)
    
    return asset_id.upper()

def extract_asset_data(df, region='ALL'):
    """Extract relevant asset data from DataFrame"""
    # Find key columns
    eq_id_cols = find_matching_columns(df, EQ_ID_PATTERNS)
    desc_cols = find_matching_columns(df, DESC_PATTERNS)
    cost_cols = find_matching_columns(df, COST_PATTERNS)
    job_cols = find_matching_columns(df, JOB_PATTERNS)
    
    # Default column indices if no match found
    eq_id_col = eq_id_cols[0] if eq_id_cols else 0
    desc_col = desc_cols[0] if desc_cols else 1
    cost_col = cost_cols[0] if cost_cols else 4
    job_col = job_cols[0] if job_cols else 2
    
    # Extract data
    assets = []
    
    for idx, row in df.iterrows():
        if idx < 3:  # Skip header rows
            continue
            
        # Get asset ID and skip empty rows
        asset_id = row.iloc[eq_id_col] if eq_id_col < len(row) else None
        
        if pd.isna(asset_id) or not asset_id:
            continue
            
        # Clean the asset ID
        asset_id = clean_asset_id(asset_id)
        
        # Extract other data
        description = row.iloc[desc_col] if desc_col < len(row) else ""
        description = str(description) if not pd.isna(description) else ""
        
        # Try to extract cost
        try:
            cost = float(row.iloc[cost_col]) if cost_col < len(row) else 0.0
        except (ValueError, TypeError):
            cost = 0.0
            
        # Extract job number
        job = row.iloc[job_col] if job_col < len(row) and job_col >= 0 else ""
        job = str(job) if not pd.isna(job) else ""
        
        # Add asset data
        asset_data = {
            'row': idx + 1,  # 1-based row number for Excel
            'asset_id': asset_id,
            'description': description,
            'cost': cost,
            'job': job
        }
        
        assets.append(asset_data)
    
    return assets

def process_pm_allocation(original_path, updated_path, region='ALL'):
    """
    Compare original and updated PM allocation files
    
    Args:
        original_path (str): Path to the original allocation file
        updated_path (str): Path to the updated allocation file
        region (str): Region code to filter data (default: 'ALL')
        
    Returns:
        dict: Results of the comparison
    """
    try:
        # Load both Excel files
        original_df = pd.read_excel(original_path, header=None)
        updated_df = pd.read_excel(updated_path, header=None)
        
        # Extract asset data
        original_assets = extract_asset_data(original_df, region)
        updated_assets = extract_asset_data(updated_df, region)
        
        # Create dictionaries for easy lookup
        original_asset_dict = {asset['asset_id']: asset for asset in original_assets if asset['asset_id']}
        updated_asset_dict = {asset['asset_id']: asset for asset in updated_assets if asset['asset_id']}
        
        # Find all unique asset IDs
        all_asset_ids = set(original_asset_dict.keys()) | set(updated_asset_dict.keys())
        
        # Track changes
        changes = []
        changed_count = 0
        new_count = 0
        unchanged_count = 0
        
        for asset_id in all_asset_ids:
            original_asset = original_asset_dict.get(asset_id)
            updated_asset = updated_asset_dict.get(asset_id)
            
            # Skip if neither exists (shouldn't happen)
            if not original_asset and not updated_asset:
                continue
                
            # New asset in updated file
            if not original_asset and updated_asset:
                new_count += 1
                changes.append({
                    'row': updated_asset['row'],
                    'asset_id': asset_id,
                    'description': updated_asset['description'],
                    'original_value': None,
                    'updated_value': updated_asset['cost'],
                    'difference': updated_asset['cost'],
                    'status': 'new'
                })
                continue
                
            # Asset removed in updated file (treat as a change)
            if original_asset and not updated_asset:
                changed_count += 1
                changes.append({
                    'row': original_asset['row'],
                    'asset_id': asset_id,
                    'description': original_asset['description'],
                    'original_value': original_asset['cost'],
                    'updated_value': None,
                    'difference': -original_asset['cost'],
                    'status': 'removed'
                })
                continue
                
            # Both exist - check for changes
            original_cost = original_asset['cost']
            updated_cost = updated_asset['cost']
            difference = updated_cost - original_cost
            
            # If costs differ
            if abs(difference) > 0.01:  # Use small threshold for float comparison
                changed_count += 1
                changes.append({
                    'row': updated_asset['row'],
                    'asset_id': asset_id,
                    'description': updated_asset['description'],
                    'original_value': original_cost,
                    'updated_value': updated_cost,
                    'difference': difference,
                    'status': 'changed'
                })
            else:
                unchanged_count += 1
                changes.append({
                    'row': updated_asset['row'],
                    'asset_id': asset_id,
                    'description': updated_asset['description'],
                    'original_value': original_cost,
                    'updated_value': updated_cost,
                    'difference': 0,
                    'status': 'unchanged'
                })
        
        # Sort changes by asset ID
        changes.sort(key=lambda x: x['asset_id'])
        
        # Generate report file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_filename = f"pm_allocation_comparison_{timestamp}.xlsx"
        report_path = os.path.join(EXPORTS_DIR, report_filename)
        
        # Create Excel report
        create_comparison_report(changes, report_path, 
                                os.path.basename(original_path), 
                                os.path.basename(updated_path))
        
        # Generate CSV export for accounting
        csv_filename = f"pm_allocation_export_{timestamp}.csv"
        csv_path = os.path.join(EXPORTS_DIR, csv_filename)
        create_accounting_export(changes, csv_path)
        
        # Return the results
        return {
            'success': True,
            'total_lines': len(changes),
            'changed_lines': changed_count,
            'new_lines': new_count,
            'unchanged_lines': unchanged_count,
            'changes': changes,
            'report_filename': report_filename,
            'csv_export_path': csv_path
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }

def create_comparison_report(changes, output_path, original_filename, updated_filename):
    """
    Create a formatted Excel report of the PM allocation comparison
    
    Args:
        changes (list): List of change dictionaries
        output_path (str): Path to save the report
        original_filename (str): Original file name
        updated_filename (str): Updated file name
    """
    # Create DataFrame from changes
    df = pd.DataFrame(changes)
    
    # Reorder columns
    columns = ['row', 'asset_id', 'description', 'original_value', 'updated_value', 'difference', 'status']
    df = df[columns]
    
    # Rename columns
    df.columns = ['Row', 'Asset ID', 'Description', 'Original Value', 'Updated Value', 'Difference', 'Status']
    
    # Export to Excel
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, sheet_name='Comparison', index=False)
    
    # Access the workbook
    wb = writer.book
    ws = writer.sheets['Comparison']
    
    # Add title and info
    ws.insert_rows(0, 3)
    ws.merge_cells('A1:G1')
    ws['A1'] = "PM Allocation Comparison Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Original: {original_filename} | Updated: {updated_filename}"
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Apply formatting
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    
    # Format headers
    for cell in ws[4]:  # Header row (1-based, plus 3 added rows)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Format currency columns
    for col_idx in [4, 5, 6]:  # Original Value, Updated Value, Difference (0-based)
        col_letter = get_column_letter(col_idx + 1)  # Convert to 1-based Excel column
        for row in range(5, len(changes) + 5):  # Data rows (1-based, plus 3 added rows)
            cell = ws[f"{col_letter}{row}"]
            cell.number_format = '$#,##0.00_);($#,##0.00)'
    
    # Apply conditional formatting for differences
    for row in range(5, len(changes) + 5):
        diff_cell = ws[f"G{row}"]
        status = df.iloc[row - 5]['Status']
        
        if status == 'new':
            diff_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == 'removed':
            diff_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status == 'changed':
            if df.iloc[row - 5]['Difference'] > 0:
                diff_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                diff_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Set column widths
    for col_idx, width in enumerate([10, 15, 40, 15, 15, 15, 15]):
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    # Save the workbook
    writer.close()

def create_accounting_export(changes, output_path):
    """
    Create a CSV export for accounting from the PM allocation comparison
    
    Args:
        changes (list): List of change dictionaries
        output_path (str): Path to save the CSV export
    """
    # Filter for changed or new items only
    export_changes = [c for c in changes if c['status'] in ['changed', 'new', 'removed']]
    
    # Create export data
    export_data = []
    for change in export_changes:
        export_data.append({
            'Asset ID': change['asset_id'],
            'Description': change['description'],
            'Amount': change['updated_value'] if change['updated_value'] is not None else 0,
            'Change Type': change['status'].capitalize(),
            'Original Amount': change['original_value'] if change['original_value'] is not None else 0,
            'Difference': change['difference']
        })
    
    # Create DataFrame and export to CSV
    if export_data:
        export_df = pd.DataFrame(export_data)
        export_df.to_csv(output_path, index=False)
    else:
        # Create empty DataFrame with headers
        empty_df = pd.DataFrame(columns=['Asset ID', 'Description', 'Amount', 'Change Type', 
                                         'Original Amount', 'Difference'])
        empty_df.to_csv(output_path, index=False)

def find_allocation_files():
    """
    Find allocation files in the attached_assets directory
    
    Returns:
        tuple: (original_file, updated_file) or (None, None) if not found
    """
    # Look for common patterns in filenames
    original_patterns = [
        r'EQMO.*BILLING.*\.(xlsx|xlsm)',
        r'EQ MONTHLY BILLINGS.*\.(xlsx|xlsm)',
        r'BILLING.*ORIGINAL.*\.(xlsx|xlsm)'
    ]
    
    updated_patterns = [
        r'EQMO.*BILLING.*REVISIONS.*\.(xlsx|xlsm)',
        r'EQMO.*BILLING.*REVISED.*\.(xlsx|xlsm)',
        r'EQMO.*BILLING.*UPDATED.*\.(xlsx|xlsm)',
        r'.*FINAL.*REVISIONS.*\.(xlsx|xlsm)'
    ]
    
    # List all Excel files
    all_files = []
    for extension in ['.xlsx', '.xlsm', '.xls']:
        all_files.extend(ATTACHED_ASSETS_DIR.glob(f'*{extension}'))
    
    # Sort by modification time (newest first)
    all_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    
    # Find updated file first
    updated_file = None
    for file in all_files:
        filename = file.name
        for pattern in updated_patterns:
            if re.search(pattern, filename, re.IGNORECASE):
                updated_file = file
                break
        if updated_file:
            break
    
    # Find original file
    original_file = None
    for file in all_files:
        filename = file.name
        # Skip the updated file
        if updated_file and file == updated_file:
            continue
            
        for pattern in original_patterns:
            if re.search(pattern, filename, re.IGNORECASE):
                original_file = file
                break
        if original_file:
            break
    
    # If we couldn't find by pattern, use the two most recent files
    if not updated_file and not original_file and len(all_files) >= 2:
        updated_file = all_files[0]  # Most recent
        original_file = all_files[1]  # Second most recent
    
    return original_file, updated_file
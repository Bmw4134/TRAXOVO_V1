"""
Extract MTD Data from Custom Formatted Files

This script handles the custom formatted MTD report files
and extracts the data into a more standard format for analysis.
"""

import os
import csv
import pandas as pd
import json
import logging
import re
from datetime import datetime
import sqlite3
import psycopg2
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Database connection string
DATABASE_URL = os.environ.get('DATABASE_URL')

def create_attendance_tables():
    """Create SQLite tables for attendance tracking"""
    try:
        # Create the extracted_data directory if it doesn't exist
        os.makedirs('extracted_data', exist_ok=True)
        
        # Connect to SQLite database
        conn = sqlite3.connect('extracted_data/attendance.db')
        cursor = conn.cursor()
        
        # Create drivers table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                employee_id TEXT UNIQUE NOT NULL,
                department TEXT,
                region TEXT,
                vehicle_id TEXT,
                active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create job_sites table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS job_sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                job_number TEXT UNIQUE NOT NULL,
                address TEXT,
                city TEXT,
                state TEXT,
                zip_code TEXT,
                latitude REAL,
                longitude REAL,
                active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create attendance_records table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_date TEXT NOT NULL,
                driver_id INTEGER,
                vehicle_id TEXT,
                job_site_id INTEGER,
                status_type TEXT NOT NULL,
                expected_start TEXT,
                actual_start TEXT,
                expected_end TEXT,
                actual_end TEXT,
                minutes_late INTEGER,
                minutes_early INTEGER,
                expected_job TEXT,
                actual_job TEXT,
                notes TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (driver_id) REFERENCES drivers (id),
                FOREIGN KEY (job_site_id) REFERENCES job_sites (id)
            )
        ''')
        
        # Create attendance_trends table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance_trends (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trend_date TEXT NOT NULL,
                trend_type TEXT NOT NULL,
                driver_id INTEGER,
                job_site_id INTEGER,
                department TEXT,
                late_start_count INTEGER DEFAULT 0,
                early_end_count INTEGER DEFAULT 0,
                not_on_job_count INTEGER DEFAULT 0,
                total_incidents INTEGER DEFAULT 0,
                week_over_week_change REAL,
                month_over_month_change REAL,
                recurring_pattern INTEGER DEFAULT 0,
                pattern_description TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (driver_id) REFERENCES drivers (id),
                FOREIGN KEY (job_site_id) REFERENCES job_sites (id)
            )
        ''')
        
        conn.commit()
        logger.info("SQLite attendance tracking tables created successfully")
        
        conn.close()
        return True
    
    except Exception as e:
        logger.error(f"Error creating SQLite attendance tables: {e}")
        return False

def extract_activity_detail(file_path):
    """
    Extract data from the activity detail report with custom format
    
    Args:
        file_path (str): Path to the activity detail CSV file
        
    Returns:
        dict: Extracted data and metadata
    """
    try:
        logger.info(f"Extracting data from activity detail report: {file_path}")
        
        # Read the file and extract header metadata
        with open(file_path, 'r') as f:
            lines = f.readlines()
        
        # Extract report parameters
        metadata = {}
        data_start_line = 0
        
        for i, line in enumerate(lines):
            if line.startswith('StartDate'):
                metadata['start_date'] = lines[i+1].split(',')[0]
            elif line.startswith('EndDate'):
                metadata['end_date'] = lines[i+1].split(',')[0]
            elif 'AssetLabel' in line:
                data_start_line = i
                break
        
        # Determine the column structure
        header_line = lines[data_start_line].strip()
        continued_header = lines[data_start_line + 1].strip()
        
        # Combine the two header lines
        combined_header = header_line + ',' + continued_header
        columns = [col.strip() for col in combined_header.split(',')]
        
        # Extract the data rows
        data_rows = []
        i = data_start_line + 2
        
        while i < len(lines):
            row_parts = []
            line = lines[i].strip()
            
            # Handle rows that span multiple lines
            if i + 1 < len(lines) and not lines[i + 1].strip().startswith('MT-'):
                line += ' ' + lines[i + 1].strip()
                i += 1
            
            # Split the line and extract values
            if line:
                parts = line.split(',')
                if len(parts) > 1:
                    asset_label = parts[0]
                    
                    # Extract driver name from asset label if available
                    driver_match = re.search(r'\((.*?)\)', asset_label)
                    driver_name = driver_match.group(1) if driver_match else "Unknown"
                    
                    event_date = parts[1]
                    reason = parts[2]
                    
                    # Extract location from later parts
                    location = ""
                    for j in range(12, min(len(parts), 14)):
                        if parts[j] and not parts[j].strip().replace('.', '').replace('-', '').isdigit():
                            location += parts[j] + ", "
                    
                    location = location.strip(', ')
                    
                    # Create structured row
                    row = {
                        'Vehicle ID': asset_label.split(' ')[0],
                        'Asset Label': asset_label,
                        'Driver': driver_name,
                        'Event Date': event_date.split(' ')[0] if ' ' in event_date else "",
                        'Event Time': event_date.split(' ')[1] + ' ' + event_date.split(' ')[2] if len(event_date.split(' ')) > 2 else "",
                        'Event Type': reason,
                        'Location': location,
                        'Department': 'Construction',  # Default department
                        'Speed': parts[3] if len(parts) > 3 else "",
                        'Speed Limit': parts[4] if len(parts) > 4 else "",
                        'Heading': parts[6] if len(parts) > 6 else "",
                        'Odometer': parts[8] if len(parts) > 8 else "",
                        'Engine Hours': parts[9] if len(parts) > 9 else "",
                        'Voltage': parts[10] if len(parts) > 10 else "",
                        'Latitude': parts[14] if len(parts) > 14 else "",
                        'Longitude': parts[15] if len(parts) > 15 else ""
                    }
                    
                    # Only add rows with Key On or Key Off events
                    if 'Key On' in reason or 'Key Off' in reason:
                        data_rows.append(row)
            
            i += 1
        
        # Convert to DataFrame for easier processing
        df = pd.DataFrame(data_rows)
        
        # Save extracted data to CSV
        output_file = os.path.join('extracted_data', 'ActivityDetail_Extracted.csv')
        df.to_csv(output_file, index=False)
        
        # Save to SQLite database
        conn = sqlite3.connect('extracted_data/attendance.db')
        
        # Process records for attendance tracking
        records_processed = 0
        late_start_count = 0
        early_end_count = 0
        
        # Define expected start/end times
        expected_start_time = datetime.strptime('07:00:00 AM', '%I:%M:%S %p').time()
        expected_end_time = datetime.strptime('05:00:00 PM', '%I:%M:%S %p').time()
        
        for _, row in df.iterrows():
            try:
                # Extract data
                vehicle_id = row['Vehicle ID']
                driver_name = row['Driver']
                event_date_str = row['Event Date']
                event_time_str = row['Event Time']
                event_type = row['Event Type']
                location = row['Location']
                
                # Skip if missing critical data
                if not vehicle_id or not driver_name or not event_date_str or not event_time_str:
                    continue
                
                # Parse date and time
                try:
                    event_date = datetime.strptime(event_date_str, '%m/%d/%Y').date()
                    event_time = datetime.strptime(event_time_str, '%I:%M:%S %p').time()
                    event_datetime = datetime.combine(event_date, event_time)
                except:
                    logger.warning(f"Invalid date/time format: {event_date_str} {event_time_str}")
                    continue
                
                # Generate employee ID
                employee_id = f"EMP-{driver_name.replace(' ', '')}"
                
                # Check if driver exists
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM drivers WHERE employee_id=?", (employee_id,))
                driver_result = cursor.fetchone()
                
                if driver_result:
                    driver_id = driver_result[0]
                else:
                    # Create new driver
                    cursor.execute(
                        "INSERT INTO drivers (name, employee_id, vehicle_id, department) VALUES (?, ?, ?, ?)",
                        (driver_name, employee_id, vehicle_id, 'Construction')
                    )
                    driver_id = cursor.lastrowid
                
                # Check if job site exists
                job_number = f"JOB-{location.replace(' ', '')}"
                cursor.execute("SELECT id FROM job_sites WHERE job_number=?", (job_number,))
                job_site_result = cursor.fetchone()
                
                if job_site_result:
                    job_site_id = job_site_result[0]
                else:
                    # Create new job site
                    cursor.execute(
                        "INSERT INTO job_sites (name, job_number) VALUES (?, ?)",
                        (location, job_number)
                    )
                    job_site_id = cursor.lastrowid
                
                # Process Key On events for Late Start
                if 'Key On' in event_type and event_time > expected_start_time:
                    # Calculate minutes late
                    start_diff = (
                        datetime.combine(event_date, event_time) - 
                        datetime.combine(event_date, expected_start_time)
                    )
                    late_minutes = max(0, int(start_diff.total_seconds() / 60))
                    
                    if late_minutes > 0:
                        # Check if record already exists
                        cursor.execute(
                            "SELECT id FROM attendance_records WHERE report_date=? AND driver_id=? AND status_type=?",
                            (event_date.isoformat(), driver_id, 'LATE_START')
                        )
                        record_result = cursor.fetchone()
                        
                        if not record_result:
                            # Create new record
                            cursor.execute(
                                """INSERT INTO attendance_records 
                                   (report_date, driver_id, vehicle_id, job_site_id, status_type, 
                                    expected_start, actual_start, minutes_late) 
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    event_date.isoformat(), 
                                    driver_id, 
                                    vehicle_id, 
                                    job_site_id, 
                                    'LATE_START',
                                    datetime.combine(event_date, expected_start_time).isoformat(),
                                    event_datetime.isoformat(),
                                    late_minutes
                                )
                            )
                            late_start_count += 1
                
                # Process Key Off events for Early End
                if 'Key Off' in event_type and event_time < expected_end_time:
                    # Calculate minutes early
                    end_diff = (
                        datetime.combine(event_date, expected_end_time) - 
                        datetime.combine(event_date, event_time)
                    )
                    early_minutes = max(0, int(end_diff.total_seconds() / 60))
                    
                    if early_minutes > 0:
                        # Check if record already exists
                        cursor.execute(
                            "SELECT id FROM attendance_records WHERE report_date=? AND driver_id=? AND status_type=?",
                            (event_date.isoformat(), driver_id, 'EARLY_END')
                        )
                        record_result = cursor.fetchone()
                        
                        if not record_result:
                            # Create new record
                            cursor.execute(
                                """INSERT INTO attendance_records 
                                   (report_date, driver_id, vehicle_id, job_site_id, status_type, 
                                    expected_end, actual_end, minutes_early) 
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    event_date.isoformat(), 
                                    driver_id, 
                                    vehicle_id, 
                                    job_site_id, 
                                    'EARLY_END',
                                    datetime.combine(event_date, expected_end_time).isoformat(),
                                    event_datetime.isoformat(),
                                    early_minutes
                                )
                            )
                            early_end_count += 1
                
                records_processed += 1
                
            except Exception as e:
                logger.error(f"Error processing activity detail record: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'metadata': metadata,
            'output_file': output_file,
            'record_count': len(data_rows),
            'processed_records': records_processed,
            'late_start_count': late_start_count,
            'early_end_count': early_end_count
        }
    
    except Exception as e:
        logger.error(f"Error extracting activity detail data: {e}")
        return {'success': False, 'message': str(e)}

def extract_driving_history(file_path):
    """
    Extract data from the driving history report with custom format
    
    Args:
        file_path (str): Path to the driving history CSV file
        
    Returns:
        dict: Extracted data and metadata
    """
    try:
        logger.info(f"Extracting data from driving history report: {file_path}")
        
        # Read the file and extract header metadata
        with open(file_path, 'r') as f:
            lines = f.readlines()
        
        # Extract report parameters
        metadata = {}
        data_start_line = 0
        
        for i, line in enumerate(lines):
            if 'StartDate' in line:
                metadata['start_date'] = lines[i+1].split(',')[0]
            elif 'EndDate' in line:
                metadata['end_date'] = lines[i+1].split(',')[0]
            elif 'Vehicle' in line or 'AssetLabel' in line:
                data_start_line = i
                break
        
        # Extract the column names
        header_line = lines[data_start_line].strip()
        columns = [col.strip() for col in header_line.split(',')]
        
        # Extract the data rows
        data_rows = []
        i = data_start_line + 1
        
        while i < len(lines):
            line = lines[i].strip()
            
            # Skip empty lines
            if not line:
                i += 1
                continue
            
            # Handle multi-line rows
            if i + 1 < len(lines) and not lines[i + 1].strip().startswith('MT-'):
                line += ' ' + lines[i + 1].strip()
                i += 1
            
            # Extract values
            parts = line.split(',')
            if len(parts) > 1:
                asset_label = parts[0]
                
                # Extract driver name from asset label if available
                driver_match = re.search(r'\((.*?)\)', asset_label)
                driver_name = driver_match.group(1) if driver_match else "Unknown"
                
                # Determine assigned job site
                assigned_job = "Unknown"
                for j in range(min(len(parts) - 1, 5), min(len(parts), 10)):
                    if parts[j] and not parts[j].strip().replace('.', '').replace('-', '').isdigit():
                        assigned_job = parts[j]
                        break
                
                # Determine actual location
                actual_location = ""
                for j in range(max(6, min(len(parts) - 4, 10)), min(len(parts), 15)):
                    if parts[j] and not parts[j].strip().replace('.', '').replace('-', '').isdigit():
                        actual_location += parts[j] + ", "
                
                actual_location = actual_location.strip(', ')
                
                # Create structured row
                row = {
                    'Vehicle ID': asset_label.split(' ')[0],
                    'Asset Label': asset_label,
                    'Driver': driver_name,
                    'Date': parts[1] if len(parts) > 1 else "",
                    'Department': 'Construction',
                    'Assigned Job': assigned_job,
                    'Last Location': actual_location,
                    'Region': 'Dallas'
                }
                
                data_rows.append(row)
            
            i += 1
        
        # Convert to DataFrame
        df = pd.DataFrame(data_rows)
        
        # Save extracted data to CSV
        output_file = os.path.join('extracted_data', 'DrivingHistory_Extracted.csv')
        df.to_csv(output_file, index=False)
        
        # Save to SQLite database
        conn = sqlite3.connect('extracted_data/attendance.db')
        
        # Process records for attendance tracking
        records_processed = 0
        not_on_job_count = 0
        
        # Track processed driver-days to avoid duplicates
        processed_driver_days = set()
        
        for _, row in df.iterrows():
            try:
                # Extract data
                vehicle_id = row['Vehicle ID']
                driver_name = row['Driver']
                event_date_str = row['Date']
                actual_location = row['Last Location']
                assigned_location = row['Assigned Job']
                department = row['Department']
                
                # Skip if missing critical data
                if not vehicle_id or not driver_name or not event_date_str:
                    continue
                
                # Parse date
                try:
                    event_date = datetime.strptime(event_date_str, '%m/%d/%Y').date()
                except:
                    logger.warning(f"Invalid date format: {event_date_str}")
                    continue
                
                # Generate employee ID
                employee_id = f"EMP-{driver_name.replace(' ', '')}"
                
                # Check if driver exists
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM drivers WHERE employee_id=?", (employee_id,))
                driver_result = cursor.fetchone()
                
                if driver_result:
                    driver_id = driver_result[0]
                else:
                    # Create new driver
                    cursor.execute(
                        "INSERT INTO drivers (name, employee_id, vehicle_id, department) VALUES (?, ?, ?, ?)",
                        (driver_name, employee_id, vehicle_id, department)
                    )
                    driver_id = cursor.lastrowid
                
                # Create a unique key for this driver-day to avoid duplicates
                driver_day_key = f"{driver_id}_{event_date.isoformat()}"
                
                if driver_day_key in processed_driver_days:
                    continue
                    
                processed_driver_days.add(driver_day_key)
                
                # Skip if missing assigned location
                if not assigned_location or assigned_location == "Unknown":
                    continue
                
                # Get or create job sites for assigned and actual locations
                assigned_job_number = f"JOB-{assigned_location.replace(' ', '')}"
                cursor.execute("SELECT id FROM job_sites WHERE job_number=?", (assigned_job_number,))
                assigned_job_result = cursor.fetchone()
                
                if assigned_job_result:
                    assigned_job_id = assigned_job_result[0]
                else:
                    # Create new job site
                    cursor.execute(
                        "INSERT INTO job_sites (name, job_number) VALUES (?, ?)",
                        (assigned_location, assigned_job_number)
                    )
                    assigned_job_id = cursor.lastrowid
                
                actual_job_id = None
                if actual_location and actual_location != assigned_location:
                    actual_job_number = f"JOB-{actual_location.replace(' ', '')}"
                    cursor.execute("SELECT id FROM job_sites WHERE job_number=?", (actual_job_number,))
                    actual_job_result = cursor.fetchone()
                    
                    if actual_job_result:
                        actual_job_id = actual_job_result[0]
                    else:
                        # Create new job site
                        cursor.execute(
                            "INSERT INTO job_sites (name, job_number) VALUES (?, ?)",
                            (actual_location, actual_job_number)
                        )
                        actual_job_id = cursor.lastrowid
                
                # Check if not on assigned job site
                if actual_location and actual_location != assigned_location:
                    # Check if record already exists
                    cursor.execute(
                        "SELECT id FROM attendance_records WHERE report_date=? AND driver_id=? AND status_type=?",
                        (event_date.isoformat(), driver_id, 'NOT_ON_JOB')
                    )
                    record_result = cursor.fetchone()
                    
                    if not record_result:
                        # Create new record
                        cursor.execute(
                            """INSERT INTO attendance_records 
                               (report_date, driver_id, vehicle_id, job_site_id, status_type, 
                                expected_job, actual_job) 
                               VALUES (?, ?, ?, ?, ?, ?, ?)""",
                            (
                                event_date.isoformat(), 
                                driver_id, 
                                vehicle_id, 
                                assigned_job_id, 
                                'NOT_ON_JOB',
                                assigned_location,
                                actual_location
                            )
                        )
                        not_on_job_count += 1
                
                records_processed += 1
                
            except Exception as e:
                logger.error(f"Error processing driving history record: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'metadata': metadata,
            'output_file': output_file,
            'record_count': len(data_rows),
            'processed_records': records_processed,
            'not_on_job_count': not_on_job_count
        }
    
    except Exception as e:
        logger.error(f"Error extracting driving history data: {e}")
        return {'success': False, 'message': str(e)}

def extract_fleet_utilization(file_path):
    """
    Extract data from the fleet utilization Excel report
    
    Args:
        file_path (str): Path to the fleet utilization Excel file
        
    Returns:
        dict: Extracted data and metadata
    """
    try:
        logger.info(f"Extracting data from fleet utilization report: {file_path}")
        
        # Load the workbook
        wb = load_workbook(file_path)
        sheet = wb.active
        
        # Extract metadata
        metadata = {
            'report_name': sheet.title,
            'period': f"May 2025"  # From filename
        }
        
        # Find the data section
        data_start_row = None
        
        for i, row in enumerate(sheet.rows, start=1):
            if row[0].value and 'Vehicle' in str(row[0].value):
                data_start_row = i
                break
        
        if not data_start_row:
            return {'success': False, 'message': 'Could not find data section in the Excel file'}
        
        # Extract column headers
        headers = [cell.value for cell in sheet[data_start_row]]
        
        # Extract data rows
        data_rows = []
        
        for i in range(data_start_row + 1, sheet.max_row + 1):
            row_data = {}
            
            for j, header in enumerate(headers):
                if header:
                    cell_value = sheet.cell(row=i, column=j+1).value
                    row_data[header] = cell_value
            
            if row_data and any(row_data.values()):
                data_rows.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data_rows)
        
        # Save extracted data to CSV
        output_file = os.path.join('extracted_data', 'FleetUtilization_Extracted.csv')
        df.to_csv(output_file, index=False)
        
        # Calculate utilization statistics
        avg_utilization = df['Utilization %'].mean() if 'Utilization %' in df.columns else 0
        total_assets = len(df)
        active_assets = len(df[df['Status'] == 'Active']) if 'Status' in df.columns else 0
        
        return {
            'success': True,
            'metadata': metadata,
            'output_file': output_file,
            'record_count': len(data_rows),
            'stats': {
                'avg_utilization': avg_utilization,
                'total_assets': total_assets,
                'active_assets': active_assets
            }
        }
    
    except Exception as e:
        logger.error(f"Error extracting fleet utilization data: {e}")
        return {'success': False, 'message': str(e)}

def process_mtd_reports():
    """Process all MTD reports"""
    try:
        # Create data directories
        os.makedirs('extracted_data', exist_ok=True)
        
        # Create database tables
        create_attendance_tables()
        
        results = {}
        mtd_directory = 'data/mtd_reports'
        
        # Process activity detail
        activity_file = os.path.join(mtd_directory, 'ActivityDetail_KeyOnly_OnRoad_2025-05-01_to_2025-05-15.csv')
        if os.path.exists(activity_file):
            results['activity_detail'] = extract_activity_detail(activity_file)
        
        # Process driving history
        driving_file = os.path.join(mtd_directory, 'DrivingHistory_2025-05-01_to_2025-05-15.csv')
        if os.path.exists(driving_file):
            results['driving_history'] = extract_driving_history(driving_file)
        
        # Process fleet utilization
        utilization_file = os.path.join(mtd_directory, 'FleetUtilization_MTD_May2025.xlsx')
        if os.path.exists(utilization_file):
            results['fleet_utilization'] = extract_fleet_utilization(utilization_file)
        
        # Save results to file
        output_file = os.path.join('extracted_data', f"mtd_processing_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        
        # Generate summary statistics
        conn = sqlite3.connect('extracted_data/attendance.db')
        cursor = conn.cursor()
        
        # Get record counts
        cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE status_type='LATE_START'")
        late_start_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE status_type='EARLY_END'")
        early_end_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE status_type='NOT_ON_JOB'")
        not_on_job_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(DISTINCT driver_id) FROM attendance_records")
        driver_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(DISTINCT job_site_id) FROM attendance_records")
        job_site_count = cursor.fetchone()[0]
        
        conn.close()
        
        # Add summary to results
        results['summary'] = {
            'late_start_count': late_start_count,
            'early_end_count': early_end_count,
            'not_on_job_count': not_on_job_count,
            'total_incidents': late_start_count + early_end_count + not_on_job_count,
            'driver_count': driver_count,
            'job_site_count': job_site_count
        }
        
        # Update the output file with summary
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        
        return {
            'success': True,
            'results': results,
            'output_file': output_file
        }
    
    except Exception as e:
        logger.error(f"Error processing MTD reports: {e}")
        return {'success': False, 'message': str(e)}

def main():
    """Main function"""
    logger.info("Starting MTD data extraction and processing")
    
    result = process_mtd_reports()
    
    if result['success']:
        print("\n=== MTD Data Processing Results ===\n")
        
        for report_type, report_result in result['results'].items():
            if report_type == 'summary':
                continue
                
            print(f"{report_type.replace('_', ' ').title()}:")
            
            if report_result['success']:
                print(f"  Output file: {report_result['output_file']}")
                
                if 'record_count' in report_result:
                    print(f"  Records processed: {report_result['record_count']}")
                
                if 'late_start_count' in report_result:
                    print(f"  Late starts: {report_result['late_start_count']}")
                
                if 'early_end_count' in report_result:
                    print(f"  Early ends: {report_result['early_end_count']}")
                
                if 'not_on_job_count' in report_result:
                    print(f"  Not on job: {report_result['not_on_job_count']}")
                
                if 'stats' in report_result:
                    print(f"  Statistics: {report_result['stats']}")
            else:
                print(f"  Failed: {report_result.get('message', 'Unknown error')}")
        
        # Print summary
        if 'summary' in result['results']:
            summary = result['results']['summary']
            print("\nOverall Summary:")
            print(f"  Total Late Starts: {summary['late_start_count']}")
            print(f"  Total Early Ends: {summary['early_end_count']}")
            print(f"  Total Not On Job: {summary['not_on_job_count']}")
            print(f"  Total Incidents: {summary['total_incidents']}")
            print(f"  Drivers Tracked: {summary['driver_count']}")
            print(f"  Job Sites Tracked: {summary['job_site_count']}")
        
        print(f"\nDetailed results saved to: {result['output_file']}")
        print(f"Database created at: extracted_data/attendance.db")
        print("\nMTD data processing completed successfully!")
        return True
    else:
        print(f"\nError: {result.get('message', 'Unknown error')}")
        return False

if __name__ == "__main__":
    main()
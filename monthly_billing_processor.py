"""
MONTHLY BILLING PROCESSOR

This script processes the monthly billing data from the source EQMO allocation files
and incorporates the custom formulas from the EQ MONTHLY BILLINGS WORKING SPREADSHEET.
"""
import os
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Constants for file paths
SOURCE_FILE = 'attached_assets/EQ MONTHLY BILLINGS WORKING SPREADSHEET - APRIL 2025.xlsx'
ALLOCATION_DIR = 'exports/pm_allocations'
OUTPUT_DIR = 'exports/monthly_billings'

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

def extract_key_formulas(source_wb, sheet_name):
    """
    Extract key formulas from a specific sheet in the source workbook
    
    Args:
        source_wb: The source workbook object
        sheet_name: The name of the sheet to extract formulas from
        
    Returns:
        dict: A dictionary mapping cell references to formulas
    """
    if sheet_name not in source_wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook")
        return {}
        
    ws = source_wb[sheet_name]
    formulas = {}
    
    # Extract formulas (limiting to first 100 rows and 30 columns for performance)
    max_row = min(100, ws.max_row)
    max_col = min(30, ws.max_column)
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.formula:
                cell_ref = f"{get_column_letter(col)}{row}"
                formulas[cell_ref] = cell.formula
                
    return formulas

def extract_sheet_structure(source_wb, sheet_name):
    """
    Extract the key structure of a sheet including headers and formatting
    
    Args:
        source_wb: The source workbook object
        sheet_name: The name of the sheet to extract structure from
        
    Returns:
        dict: A dictionary with sheet structure information
    """
    if sheet_name not in source_wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook")
        return {}
        
    ws = source_wb[sheet_name]
    structure = {
        'headers': [],
        'merged_cells': [],
        'column_widths': {},
        'row_heights': {}
    }
    
    # Extract merged cells
    for merged_range in ws.merged_cells.ranges:
        structure['merged_cells'].append(str(merged_range))
    
    # Extract headers from first few rows
    for row in range(1, min(5, ws.max_row + 1)):
        headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            headers.append(cell.value)
        structure['headers'].append(headers)
    
    # Extract column widths
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            structure['column_widths'][col_letter] = ws.column_dimensions[col_letter].width
    
    # Extract row heights
    for row in range(1, min(20, ws.max_row + 1)):
        if row in ws.row_dimensions:
            structure['row_heights'][row] = ws.row_dimensions[row].height
    
    return structure

def main():
    """Main function to process monthly billing data"""
    print("Starting Monthly Billing Processor...")
    
    # Check if source file exists
    if not os.path.exists(SOURCE_FILE):
        print(f"Source file not found: {SOURCE_FILE}")
        return False
    
    try:
        # Load the source workbook to extract formulas and structure
        print(f"Loading source workbook: {SOURCE_FILE}")
        source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False)
        
        # Get list of sheets in the workbook
        sheet_names = source_wb.sheetnames
        print(f"Sheets in workbook: {sheet_names}")
        
        # Define the key sheets to process
        key_sheets = ['M-SELECT', 'M-RAGLE', 'ATOS', 'EHRS-PT', 'EHrs']
        
        # Extract formulas and structure from key sheets
        formulas = {}
        structures = {}
        
        for sheet_name in key_sheets:
            if sheet_name in sheet_names:
                print(f"Extracting formulas and structure from '{sheet_name}'")
                formulas[sheet_name] = extract_key_formulas(source_wb, sheet_name)
                structures[sheet_name] = extract_sheet_structure(source_wb, sheet_name)
                print(f"  - Found {len(formulas[sheet_name])} formulas in '{sheet_name}'")
            else:
                print(f"Sheet '{sheet_name}' not found in workbook")
        
        # Find the allocation data files
        pm_files = []
        for filename in os.listdir(ALLOCATION_DIR):
            if filename.endswith('.xlsx') and 'PM_ALLOCATION' in filename:
                pm_files.append(os.path.join(ALLOCATION_DIR, filename))
        
        if not pm_files:
            print(f"No PM allocation files found in {ALLOCATION_DIR}")
            return False
        
        print(f"Found {len(pm_files)} PM allocation files")
        
        # Create a new workbook for the processed data
        output_wb = openpyxl.Workbook()
        
        # Remove the default sheet
        if 'Sheet' in output_wb.sheetnames:
            output_wb.remove(output_wb['Sheet'])
        
        # Create a summary sheet
        summary_sheet = output_wb.create_sheet("Summary")
        
        # Add header to summary sheet
        summary_headers = [
            "Job Number", "Equipment Count", "Total Days", "Total Amount", "Source File"
        ]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
        # Process each PM allocation file
        summary_row = 2
        
        for pm_file in pm_files:
            filename = os.path.basename(pm_file)
            print(f"Processing: {filename}")
            
            # Extract job number from filename
            job_number = "Unknown"
            if '-' in filename:
                parts = filename.split('-')
                for part in parts:
                    if part.strip().startswith('202') and len(part.strip()) >= 7:
                        job_number = part.strip()[:7]  # Extract job number format like 2023-032
                        break
            
            try:
                # Read PM allocation data
                allocation_df = pd.read_excel(pm_file)
                
                # Calculate summary statistics
                equipment_count = len(allocation_df)
                
                # Try to find days and amount columns
                days_col = None
                amount_col = None
                for col in allocation_df.columns:
                    col_str = str(col).upper()
                    if 'DAY' in col_str or 'QTY' in col_str:
                        days_col = col
                    elif 'AMOUNT' in col_str or 'TOTAL' in col_str:
                        amount_col = col
                
                # Calculate totals
                total_days = 0
                total_amount = 0
                if days_col is not None:
                    total_days = allocation_df[days_col].sum()
                if amount_col is not None:
                    total_amount = allocation_df[amount_col].sum()
                
                # Add to summary sheet
                summary_sheet.cell(row=summary_row, column=1).value = job_number
                summary_sheet.cell(row=summary_row, column=2).value = equipment_count
                summary_sheet.cell(row=summary_row, column=3).value = total_days
                summary_sheet.cell(row=summary_row, column=4).value = total_amount
                summary_sheet.cell(row=summary_row, column=5).value = filename
                
                # Format the amount as currency
                summary_sheet.cell(row=summary_row, column=4).number_format = '$#,##0.00'
                
                # Create a sheet for each job
                job_sheet_name = f"Job-{job_number}"
                if job_sheet_name in output_wb.sheetnames:
                    job_sheet_name = f"Job-{job_number}-{summary_row}"
                
                job_sheet = output_wb.create_sheet(job_sheet_name)
                
                # Add header
                job_sheet.cell(row=1, column=1).value = f"Job {job_number} - Equipment Allocations"
                job_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                job_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
                
                # Add column headers based on the allocation file
                for col_idx, col_name in enumerate(allocation_df.columns, 1):
                    cell = job_sheet.cell(row=3, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
                
                # Add data
                for row_idx, row in allocation_df.iterrows():
                    for col_idx, col_name in enumerate(allocation_df.columns, 1):
                        cell = job_sheet.cell(row=row_idx+4, column=col_idx)
                        cell.value = row[col_name]
                        
                        # Apply formatting for dollar amounts
                        if col_name == amount_col:
                            cell.number_format = '$#,##0.00'
                
                # Increment summary row counter
                summary_row += 1
                
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
                continue
        
        # Create sheets for formulas from the source file
        for sheet_name, sheet_formulas in formulas.items():
            if sheet_formulas:
                print(f"Creating sheet '{sheet_name}' with formulas")
                
                # Create the sheet
                if sheet_name in output_wb.sheetnames:
                    # Sheet already exists, use a different name
                    new_sheet_name = f"{sheet_name}-Formulas"
                    formula_sheet = output_wb.create_sheet(new_sheet_name)
                else:
                    formula_sheet = output_wb.create_sheet(sheet_name)
                
                # Apply structure if available
                if sheet_name in structures:
                    structure = structures[sheet_name]
                    
                    # Apply headers
                    for row_idx, headers in enumerate(structure['headers'], 1):
                        for col_idx, header in enumerate(headers, 1):
                            formula_sheet.cell(row=row_idx, column=col_idx).value = header
                    
                    # Apply merged cells
                    for merged_range in structure['merged_cells']:
                        formula_sheet.merge_cells(merged_range)
                    
                    # Apply column widths
                    for col_letter, width in structure['column_widths'].items():
                        if width:
                            formula_sheet.column_dimensions[col_letter].width = width
                    
                    # Apply row heights
                    for row_idx, height in structure['row_heights'].items():
                        if height:
                            formula_sheet.row_dimensions[row_idx].height = height
                
                # Add formulas
                for cell_ref, formula in sheet_formulas.items():
                    try:
                        # Convert cell reference to row and column
                        col_letter = ''.join(filter(str.isalpha, cell_ref))
                        row = int(''.join(filter(str.isdigit, cell_ref)))
                        col = 0
                        for char in col_letter:
                            col = col * 26 + (ord(char) - ord('A') + 1)
                        
                        # Apply formula
                        formula_sheet.cell(row=row, column=col).value = f"={formula}"
                    except Exception as e:
                        print(f"Error applying formula {cell_ref}: {str(e)}")
        
        # Save the workbook
        output_file = os.path.join(OUTPUT_DIR, f"Monthly_Billing_Processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        output_wb.save(output_file)
        
        print(f"Processing complete. Output saved to {output_file}")
        
        # Create a simplified CSV version for easy data access
        summary_df = pd.DataFrame({
            "Job Number": [summary_sheet.cell(row=r, column=1).value for r in range(2, summary_row)],
            "Equipment Count": [summary_sheet.cell(row=r, column=2).value for r in range(2, summary_row)],
            "Total Days": [summary_sheet.cell(row=r, column=3).value for r in range(2, summary_row)],
            "Total Amount": [summary_sheet.cell(row=r, column=4).value for r in range(2, summary_row)],
            "Source File": [summary_sheet.cell(row=r, column=5).value for r in range(2, summary_row)]
        })
        
        csv_file = os.path.join(OUTPUT_DIR, "Monthly_Billing_Summary.csv")
        summary_df.to_csv(csv_file, index=False)
        print(f"Summary CSV saved to {csv_file}")
        
        return True
        
    except Exception as e:
        print(f"Error processing monthly billing data: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    main()
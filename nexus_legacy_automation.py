#!/usr/bin/env python3
"""
NEXUS Legacy Report Automation Processor
Intelligent automation for legacy reports with manual guidance integration
"""

import os
import json
import sqlite3
import pandas as pd
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from io import BytesIO

class NexusLegacyAutomation:
    """Legacy report automation with manual setup guidance"""
    
    def __init__(self):
        self.automation_db = "nexus_automation.db"
        self.setup_automation_database()
        
    def setup_automation_database(self):
        """Initialize automation tracking database"""
        conn = sqlite3.connect(self.automation_db)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS legacy_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                file_type TEXT,
                processing_status TEXT,
                automation_type TEXT,
                data_structure TEXT,
                key_fields TEXT,
                automation_rules TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_processed DATETIME,
                success_rate REAL DEFAULT 0.0
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS automation_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_name TEXT NOT NULL,
                file_pattern TEXT,
                extraction_rules TEXT,
                output_format TEXT,
                automation_steps TEXT,
                manual_overrides TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS processing_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER,
                action_type TEXT,
                action_details TEXT,
                result_status TEXT,
                processing_time REAL,
                error_details TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (report_id) REFERENCES legacy_reports (id)
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def analyze_legacy_file(self, file_path: str, content: bytes) -> Dict[str, Any]:
        """Analyze legacy file structure and data patterns"""
        analysis = {
            "filename": os.path.basename(file_path),
            "file_type": self.detect_file_type(file_path),
            "file_size": len(content),
            "data_structure": {},
            "key_fields": [],
            "automation_opportunities": [],
            "manual_setup_required": []
        }
        
        try:
            if analysis["file_type"] == "excel":
                analysis.update(self.analyze_excel_file(content))
            elif analysis["file_type"] == "csv":
                analysis.update(self.analyze_csv_file(content))
            elif analysis["file_type"] == "pdf":
                analysis.update(self.analyze_pdf_file(content))
            elif analysis["file_type"] == "text":
                analysis.update(self.analyze_text_file(content))
            
            # Store analysis in database
            self.store_analysis(analysis)
            
        except Exception as e:
            analysis["error"] = str(e)
            analysis["manual_setup_required"].append("File analysis failed - manual inspection needed")
        
        return analysis
    
    def detect_file_type(self, file_path: str) -> str:
        """Detect file type from extension and content"""
        ext = file_path.lower().split('.')[-1] if '.' in file_path else 'unknown'
        
        type_mapping = {
            'xlsx': 'excel',
            'xls': 'excel',
            'csv': 'csv',
            'pdf': 'pdf',
            'txt': 'text',
            'doc': 'document',
            'docx': 'document'
        }
        
        return type_mapping.get(ext, 'unknown')
    
    def analyze_excel_file(self, content: bytes) -> Dict[str, Any]:
        """Analyze Excel file structure and data"""
        try:
            workbook = openpyxl.load_workbook(BytesIO(content), read_only=True)
            analysis = {
                "sheets": [],
                "data_structure": {},
                "automation_opportunities": [],
                "manual_setup_required": []
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = {
                    "name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "headers": [],
                    "data_types": {},
                    "patterns": []
                }
                
                # Analyze first few rows for headers and patterns
                headers = []
                for col in range(1, min(sheet.max_column + 1, 21)):  # Max 20 columns
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                
                sheet_analysis["headers"] = headers
                
                # Identify automation opportunities
                if any(keyword in ' '.join(headers).lower() for keyword in ['date', 'amount', 'total', 'revenue', 'cost']):
                    analysis["automation_opportunities"].append(f"Financial data automation - Sheet: {sheet_name}")
                
                if any(keyword in ' '.join(headers).lower() for keyword in ['name', 'email', 'phone', 'contact']):
                    analysis["automation_opportunities"].append(f"Contact management automation - Sheet: {sheet_name}")
                
                if len(headers) > 10:
                    analysis["manual_setup_required"].append(f"Complex data structure - Manual mapping needed for {sheet_name}")
                
                analysis["sheets"].append(sheet_analysis)
            
            workbook.close()
            return analysis
            
        except Exception as e:
            return {
                "error": f"Excel analysis failed: {str(e)}",
                "manual_setup_required": ["Excel file requires manual inspection"]
            }
    
    def analyze_csv_file(self, content: bytes) -> Dict[str, Any]:
        """Analyze CSV file structure"""
        try:
            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    text_content = content.decode(encoding)
                    break
                except:
                    continue
            else:
                text_content = content.decode('utf-8', errors='ignore')
            
            lines = text_content.split('\n')[:100]  # First 100 lines
            
            analysis = {
                "total_lines": len(text_content.split('\n')),
                "headers": [],
                "data_structure": {},
                "automation_opportunities": [],
                "manual_setup_required": []
            }
            
            if lines:
                # Detect delimiter
                delimiter = ',' if ',' in lines[0] else '\t' if '\t' in lines[0] else ';'
                headers = [h.strip().strip('"') for h in lines[0].split(delimiter)]
                analysis["headers"] = headers
                
                # Automation opportunities
                if any(keyword in ' '.join(headers).lower() for keyword in ['date', 'time', 'timestamp']):
                    analysis["automation_opportunities"].append("Time-series data automation")
                
                if any(keyword in ' '.join(headers).lower() for keyword in ['price', 'amount', 'value', 'cost']):
                    analysis["automation_opportunities"].append("Financial calculations automation")
                
                if len(headers) > 15:
                    analysis["manual_setup_required"].append("Large dataset - Manual field mapping recommended")
            
            return analysis
            
        except Exception as e:
            return {
                "error": f"CSV analysis failed: {str(e)}",
                "manual_setup_required": ["CSV file requires manual inspection"]
            }
    
    def analyze_pdf_file(self, content: bytes) -> Dict[str, Any]:
        """Analyze PDF file for text patterns"""
        analysis = {
            "file_size": len(content),
            "automation_opportunities": ["PDF text extraction automation"],
            "manual_setup_required": ["PDF parsing requires manual configuration", "Text extraction patterns need manual setup"]
        }
        
        # PDF analysis would require additional libraries like PyPDF2
        # For now, provide manual setup guidance
        return analysis
    
    def analyze_text_file(self, content: bytes) -> Dict[str, Any]:
        """Analyze text file patterns"""
        try:
            text_content = content.decode('utf-8', errors='ignore')
            
            analysis = {
                "character_count": len(text_content),
                "line_count": len(text_content.split('\n')),
                "automation_opportunities": [],
                "manual_setup_required": []
            }
            
            # Look for structured patterns
            if re.search(r'\d{4}-\d{2}-\d{2}', text_content):
                analysis["automation_opportunities"].append("Date extraction automation")
            
            if re.search(r'\$[\d,]+\.?\d*', text_content):
                analysis["automation_opportunities"].append("Currency value extraction")
            
            if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text_content):
                analysis["automation_opportunities"].append("Email extraction automation")
            
            if len(text_content) > 10000:
                analysis["manual_setup_required"].append("Large text file - Manual pattern identification needed")
            
            return analysis
            
        except Exception as e:
            return {
                "error": f"Text analysis failed: {str(e)}",
                "manual_setup_required": ["Text file requires manual inspection"]
            }
    
    def store_analysis(self, analysis: Dict[str, Any]) -> int:
        """Store file analysis in database"""
        try:
            conn = sqlite3.connect(self.automation_db)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO legacy_reports 
                (filename, file_type, processing_status, data_structure, key_fields, automation_rules)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                analysis["filename"],
                analysis["file_type"],
                "analyzed",
                json.dumps(analysis.get("data_structure", {})),
                json.dumps(analysis.get("key_fields", [])),
                json.dumps(analysis.get("automation_opportunities", []))
            ))
            
            report_id = cursor.lastrowid
            
            # Log the analysis
            cursor.execute('''
                INSERT INTO processing_logs 
                (report_id, action_type, action_details, result_status)
                VALUES (?, ?, ?, ?)
            ''', (
                report_id,
                "file_analysis",
                json.dumps(analysis),
                "success" if "error" not in analysis else "failed"
            ))
            
            conn.commit()
            conn.close()
            
            return report_id
            
        except Exception as e:
            print(f"Database storage error: {e}")
            return -1
    
    def get_manual_setup_guidance(self, report_id: int) -> Dict[str, Any]:
        """Get step-by-step manual setup guidance for automation"""
        try:
            conn = sqlite3.connect(self.automation_db)
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM legacy_reports WHERE id = ?', (report_id,))
            report = cursor.fetchone()
            
            if not report:
                return {"error": "Report not found"}
            
            # Get column names
            cursor.execute('PRAGMA table_info(legacy_reports)')
            columns = [col[1] for col in cursor.fetchall()]
            report_dict = dict(zip(columns, report))
            
            guidance = {
                "report_id": report_id,
                "filename": report_dict["filename"],
                "current_status": report_dict["processing_status"],
                "manual_steps": [],
                "automation_options": [],
                "next_actions": []
            }
            
            # Parse stored data
            try:
                automation_rules = json.loads(report_dict["automation_rules"] or "[]")
                data_structure = json.loads(report_dict["data_structure"] or "{}")
            except:
                automation_rules = []
                data_structure = {}
            
            # Generate step-by-step guidance
            guidance["manual_steps"] = [
                {
                    "step": 1,
                    "title": "Review File Structure",
                    "description": "Examine the analyzed data structure and identify key fields",
                    "action": "manual_review",
                    "data": data_structure
                },
                {
                    "step": 2,
                    "title": "Define Automation Rules",
                    "description": "Specify which fields should be automated and how",
                    "action": "define_rules",
                    "current_opportunities": automation_rules
                },
                {
                    "step": 3,
                    "title": "Set Up Data Mapping",
                    "description": "Map legacy fields to target system fields",
                    "action": "field_mapping",
                    "requires_input": True
                },
                {
                    "step": 4,
                    "title": "Configure Output Format",
                    "description": "Define how processed data should be formatted",
                    "action": "output_config",
                    "options": ["database_insert", "csv_export", "json_api", "email_report"]
                },
                {
                    "step": 5,
                    "title": "Test Automation",
                    "description": "Run test automation with sample data",
                    "action": "test_run",
                    "requires_validation": True
                }
            ]
            
            conn.close()
            return guidance
            
        except Exception as e:
            return {"error": f"Guidance generation failed: {str(e)}"}
    
    def apply_manual_configuration(self, report_id: int, config: Dict[str, Any]) -> Dict[str, Any]:
        """Apply manual configuration for automation setup"""
        try:
            conn = sqlite3.connect(self.automation_db)
            cursor = conn.cursor()
            
            # Update report with manual configuration
            cursor.execute('''
                UPDATE legacy_reports 
                SET automation_type = ?, automation_rules = ?, processing_status = ?
                WHERE id = ?
            ''', (
                config.get("automation_type", "custom"),
                json.dumps(config),
                "configured",
                report_id
            ))
            
            # Log configuration
            cursor.execute('''
                INSERT INTO processing_logs 
                (report_id, action_type, action_details, result_status)
                VALUES (?, ?, ?, ?)
            ''', (
                report_id,
                "manual_configuration",
                json.dumps(config),
                "applied"
            ))
            
            conn.commit()
            conn.close()
            
            return {
                "success": True,
                "report_id": report_id,
                "status": "configured",
                "message": "Manual configuration applied successfully",
                "next_step": "ready_for_automation_test"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_automation_status(self, report_id: int) -> Dict[str, Any]:
        """Get current automation status and progress"""
        try:
            conn = sqlite3.connect(self.automation_db)
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM legacy_reports WHERE id = ?', (report_id,))
            report = cursor.fetchone()
            
            if not report:
                return {"error": "Report not found"}
            
            cursor.execute('PRAGMA table_info(legacy_reports)')
            columns = [col[1] for col in cursor.fetchall()]
            report_dict = dict(zip(columns, report))
            
            # Get processing logs
            cursor.execute('''
                SELECT action_type, result_status, timestamp 
                FROM processing_logs 
                WHERE report_id = ? 
                ORDER BY timestamp DESC 
                LIMIT 10
            ''', (report_id,))
            
            logs = cursor.fetchall()
            
            conn.close()
            
            return {
                "report_id": report_id,
                "filename": report_dict["filename"],
                "status": report_dict["processing_status"],
                "automation_type": report_dict["automation_type"],
                "success_rate": report_dict["success_rate"],
                "last_processed": report_dict["last_processed"],
                "recent_logs": [{"action": log[0], "status": log[1], "time": log[2]} for log in logs]
            }
            
        except Exception as e:
            return {"error": str(e)}

def create_legacy_automation_system():
    """Create legacy automation system"""
    return NexusLegacyAutomation()

if __name__ == "__main__":
    automation = create_legacy_automation_system()
    print("NEXUS Legacy Report Automation System Ready")
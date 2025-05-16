"""
Process PM Allocation files for April 2025

This script processes the PM Allocation Excel files and generates the required
reconciliation reports and outputs.
"""
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import shutil
import glob

def setup_directories():
    """Setup necessary directories for outputs"""
    os.makedirs('exports', exist_ok=True)
    os.makedirs('exports/pm_allocations', exist_ok=True)
    os.makedirs('exports/pm_allocations/april_2025', exist_ok=True)
    
    # Timestamp for this run
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(f'exports/pm_allocations/april_2025/{timestamp}', exist_ok=True)
    
    return f'exports/pm_allocations/april_2025/{timestamp}'

def load_excel_file(file_path):
    """Load Excel file and return sheets as dict of dataframes"""
    try:
        print(f"Loading file: {file_path}")
        sheets = pd.read_excel(file_path, sheet_name=None)
        
        print(f"  Loaded {len(sheets)} sheets: {list(sheets.keys())}")
        for sheet_name, df in sheets.items():
            if not df.empty:
                print(f"  Sheet '{sheet_name}': {df.shape[0]} rows, {df.shape[1]} columns")
            else:
                print(f"  Sheet '{sheet_name}': Empty")
        
        return sheets
    except Exception as e:
        print(f"Error loading {file_path}: {str(e)}")
        return None

def find_final_revision_file():
    """Find the final revision file based on naming patterns"""
    files = glob.glob("attached_assets/*APRIL*2025*FINAL*REVISIONS*.xlsx")
    if not files:
        files = glob.glob("attached_assets/*APRIL*2025*TR*.xlsx")
    
    # Sort by modification time to get the latest
    if files:
        return sorted(files, key=os.path.getmtime)[-1]
    
    return None

def find_job_specific_files():
    """Find job-specific allocation files"""
    job_files = glob.glob("attached_assets/202*-*EQMO*APRIL*2025*.xlsx")
    return job_files

def extract_job_number_from_filename(filename):
    """Extract the job number from a filename"""
    # Base filename without path
    base = os.path.basename(filename)
    
    # Look for patterns like 2023-032, 2024-016, etc.
    for part in base.split():
        if part.startswith('202') and '-' in part:
            return part
    
    return "Unknown"

def process_allocation_file(file_path, output_dir):
    """Process a PM allocation file and extract key information"""
    sheets = load_excel_file(file_path)
    if not sheets:
        return None
    
    job_id = extract_job_number_from_filename(file_path)
    
    # Main dataframe is usually in "EQ ALLOCATIONS - ALL DIV" or similar
    main_sheet_key = None
    for key in sheets.keys():
        if 'EQ ALLOCATIONS' in key or 'ALLOCATION' in key or 'ALL DIV' in key:
            main_sheet_key = key
            break
    
    if not main_sheet_key and 'ALL' in sheets:
        main_sheet_key = 'ALL'
    
    if not main_sheet_key:
        print(f"Could not find main allocation sheet in {file_path}")
        return None
    
    # Process the main sheet
    df = sheets[main_sheet_key]
    
    # Clean up and prepare the dataframe - handle different formats
    # Skip header rows
    header_row = None
    for i in range(min(10, len(df))):
        # Look for rows containing key headers
        row = df.iloc[i].astype(str)
        if any('COST CODE' in val.upper() for val in row.values) or \
           any('EQUIP' in val.upper() and 'ID' in val.upper() for val in row.values):
            header_row = i
            break
    
    if header_row is not None:
        # If we found a header row, use it
        df = df.iloc[header_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    
    # Filter out completely empty rows
    df = df.dropna(how='all')
    
    # Extract only rows with equipment data
    equipment_df = df[df.iloc[:, 1].notna()]  # Equipment ID is typically in column 1
    
    # Basic summary
    total_items = len(equipment_df)
    
    # Find days column (typically column with QTY or DAYS)
    days_col = None
    amount_col = None
    
    if len(equipment_df.columns) > 0:
        for i, col_name in enumerate(equipment_df.columns):
            col_str = str(col_name).upper()
            if 'DAY' in col_str or 'QTY' in col_str:
                days_col = i
            elif 'AMOUNT' in col_str or 'TOTAL' in col_str:
                amount_col = i
    
    # If we couldn't find by name, use default positions
    if days_col is None and len(equipment_df.columns) > 3:
        days_col = 3
    if amount_col is None and len(equipment_df.columns) > 5:
        amount_col = 5
    
    # Calculate totals
    total_days = 0
    total_amount = 0
    
    if total_items > 0:
        if days_col is not None:
            try:
                # Convert to numeric, coercing errors to NaN
                days_series = pd.to_numeric(equipment_df.iloc[:, days_col], errors='coerce')
                total_days = days_series.sum()
            except:
                print(f"Warning: Could not calculate total days")
        
        if amount_col is not None:
            try:
                # Convert to numeric, coercing errors to NaN
                amount_series = pd.to_numeric(equipment_df.iloc[:, amount_col], errors='coerce')
                total_amount = amount_series.sum()
            except:
                print(f"Warning: Could not calculate total amount")
    
    # Create a summary Excel file
    output_file = os.path.join(output_dir, f"PM_Allocation_Summary_{job_id}.xlsx")
    
    # Save processed data
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        equipment_df.to_excel(writer, sheet_name='Processed Data', index=False)
        
        # Create summary sheet
        summary = pd.DataFrame([
            ['Job Number', job_id],
            ['Source File', os.path.basename(file_path)],
            ['Total Equipment Items', total_items],
            ['Total Billing Days', total_days],
            ['Total Billing Amount', total_amount],
            ['Processing Date', datetime.now().strftime("%Y-%m-%d %H:%M")],
        ])
        summary.to_excel(writer, sheet_name='Summary', index=False, header=False)
    
    # Enhance the Excel file styling
    wb = openpyxl.load_workbook(output_file)
    
    # Format the Summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
            row[0].font = Font(bold=True)
            if row[0].value == 'Total Billing Amount':
                row[1].number_format = '$#,##0.00'
    
    # Format the data sheet
    if 'Processed Data' in wb.sheetnames:
        ws = wb['Processed Data']
        header_row = 1
        
        # Format headers
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        
        # Format currency columns
        amount_col = None
        rate_col = None
        day_col = None
        
        # Find columns containing amounts, rates and days
        for i, cell in enumerate(ws[header_row], 1):
            header_text = str(cell.value).upper() if cell.value else ""
            if "AMOUNT" in header_text or "TOTAL" in header_text:
                amount_col = i
            elif "RATE" in header_text:
                rate_col = i
            elif "DAY" in header_text or "QTY" in header_text:
                day_col = i
        
        # Apply formatting to these columns
        if amount_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=amount_col)
                cell.number_format = '$#,##0.00'
        
        if rate_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=rate_col)
                cell.number_format = '$#,##0.00'
        
        if day_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=day_col)
                cell.number_format = '0'
    
    wb.save(output_file)
    
    return {
        'job_id': job_id,
        'file_path': file_path,
        'output_file': output_file,
        'total_items': total_items,
        'total_days': total_days,
        'total_amount': total_amount
    }

def compare_allocation_revisions(original_files, final_file, output_dir):
    """Compare original allocations with final revisions to highlight changes"""
    final_sheets = load_excel_file(final_file)
    if not final_sheets:
        return None
    
    # Choose the main sheet from final file
    main_sheet_key = None
    for key in final_sheets.keys():
        if 'EQ ALLOCATIONS' in key or 'ALLOCATION' in key or 'ALL DIV' in key:
            main_sheet_key = key
            break
    
    if not main_sheet_key and 'ALL' in final_sheets:
        main_sheet_key = 'ALL'
    
    if not main_sheet_key:
        print(f"Could not find main allocation sheet in final file {final_file}")
        return None
    
    final_df = final_sheets[main_sheet_key]
    
    # Process and clean final dataframe as before
    header_row = None
    for i in range(min(10, len(final_df))):
        row = final_df.iloc[i].astype(str)
        if any('COST CODE' in val.upper() for val in row.values) or \
           any('EQUIP' in val.upper() and 'ID' in val.upper() for val in row.values):
            header_row = i
            break
    
    if header_row is not None:
        final_df = final_df.iloc[header_row:].reset_index(drop=True)
        final_df.columns = final_df.iloc[0]
        final_df = final_df.iloc[1:].reset_index(drop=True)
    
    # Filter out completely empty rows
    final_df = final_df.dropna(how='all')
    
    # Extract only rows with equipment data
    final_equipment_df = final_df[final_df.iloc[:, 1].notna()]
    
    # Process each original file and compare
    comparison_results = []
    
    for orig_file in original_files:
        job_id = extract_job_number_from_filename(orig_file)
        orig_sheets = load_excel_file(orig_file)
        if not orig_sheets:
            continue
        
        # Find the main sheet
        orig_main_sheet = None
        for key in orig_sheets.keys():
            if 'EQ ALLOCATIONS' in key or 'ALLOCATION' in key or 'ALL DIV' in key:
                orig_main_sheet = key
                break
        
        if not orig_main_sheet and 'ALL' in orig_sheets:
            orig_main_sheet = 'ALL'
        
        if not orig_main_sheet:
            print(f"Could not find main allocation sheet in {orig_file}")
            continue
        
        orig_df = orig_sheets[orig_main_sheet]
        
        # Process and clean original dataframe
        header_row = None
        for i in range(min(10, len(orig_df))):
            row = orig_df.iloc[i].astype(str)
            if any('COST CODE' in val.upper() for val in row.values) or \
               any('EQUIP' in val.upper() and 'ID' in val.upper() for val in row.values):
                header_row = i
                break
        
        if header_row is not None:
            orig_df = orig_df.iloc[header_row:].reset_index(drop=True)
            orig_df.columns = orig_df.iloc[0]
            orig_df = orig_df.iloc[1:].reset_index(drop=True)
        
        # Filter out completely empty rows
        orig_df = orig_df.dropna(how='all')
        
        # Extract only rows with equipment data
        orig_equipment_df = orig_df[orig_df.iloc[:, 1].notna()]
        
        # Find equipment specific to this job
        job_specific_equipment = final_equipment_df[final_equipment_df.iloc[:, 0].astype(str).str.contains(job_id)]
        
        # Compare with original
        if len(orig_equipment_df) > 0 and len(job_specific_equipment) > 0:
            # Get equipment IDs from both
            try:
                orig_ids = orig_equipment_df.iloc[:, 1].astype(str)
                final_ids = job_specific_equipment.iloc[:, 1].astype(str)
                
                # Find added and removed equipment
                added_equipment = [eq_id for eq_id in final_ids if eq_id not in orig_ids.values]
                removed_equipment = [eq_id for eq_id in orig_ids if eq_id not in final_ids.values]
                
                # Create comparison report
                comparison_file = os.path.join(output_dir, f"PM_Allocation_Comparison_{job_id}.xlsx")
                with pd.ExcelWriter(comparison_file, engine='openpyxl') as writer:
                    # Summary sheet
                    summary = pd.DataFrame([
                        ['Job Number', job_id],
                        ['Original File', os.path.basename(orig_file)],
                        ['Final File', os.path.basename(final_file)],
                        ['Equipment Count (Original)', len(orig_equipment_df)],
                        ['Equipment Count (Final)', len(job_specific_equipment)],
                        ['Added Equipment', len(added_equipment)],
                        ['Removed Equipment', len(removed_equipment)],
                        ['Processing Date', datetime.now().strftime("%Y-%m-%d %H:%M")],
                    ])
                    summary.to_excel(writer, sheet_name='Summary', index=False, header=False)
                    
                    # Added Equipment Sheet
                    if added_equipment:
                        added_df = job_specific_equipment[job_specific_equipment.iloc[:, 1].astype(str).isin(added_equipment)]
                        if not added_df.empty:
                            added_df.to_excel(writer, sheet_name='Added Equipment', index=False)
                    
                    # Removed Equipment Sheet
                    if removed_equipment:
                        removed_df = orig_equipment_df[orig_equipment_df.iloc[:, 1].astype(str).isin(removed_equipment)]
                        if not removed_df.empty:
                            removed_df.to_excel(writer, sheet_name='Removed Equipment', index=False)
                    
                    # Original Data
                    orig_equipment_df.to_excel(writer, sheet_name='Original Data', index=False)
                    
                    # Final Data
                    job_specific_equipment.to_excel(writer, sheet_name='Final Data', index=False)
                
                # Format workbook
                wb = openpyxl.load_workbook(comparison_file)
                
                # Format the Summary sheet
                if 'Summary' in wb.sheetnames:
                    ws = wb['Summary']
                    for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=2):
                        row[0].font = Font(bold=True)
                
                # Format all data sheets
                for sheet_name in ['Added Equipment', 'Removed Equipment', 'Original Data', 'Final Data']:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        if ws.max_row > 1:  # Only format if we have data
                            # Format headers
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                            
                            # Auto-size columns
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
                            
                            # Format currency columns
                            amount_col = None
                            rate_col = None
                            day_col = None
                            
                            # Find columns containing amounts, rates and days
                            for i, cell in enumerate(ws[1], 1):
                                header_text = str(cell.value).upper() if cell.value else ""
                                if "AMOUNT" in header_text or "TOTAL" in header_text:
                                    amount_col = i
                                elif "RATE" in header_text:
                                    rate_col = i
                                elif "DAY" in header_text or "QTY" in header_text:
                                    day_col = i
                            
                            # Apply formatting to these columns
                            if amount_col:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=amount_col)
                                    cell.number_format = '$#,##0.00'
                            
                            if rate_col:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=rate_col)
                                    cell.number_format = '$#,##0.00'
                            
                            if day_col:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=day_col)
                                    cell.number_format = '0'
                
                wb.save(comparison_file)
                
                comparison_results.append({
                    'job_id': job_id,
                    'original_file': orig_file,
                    'equipment_count_original': len(orig_equipment_df),
                    'equipment_count_final': len(job_specific_equipment),
                    'added_equipment': added_equipment,
                    'removed_equipment': removed_equipment,
                    'comparison_file': comparison_file
                })
            except Exception as e:
                print(f"Error comparing files for job {job_id}: {str(e)}")
    
    return comparison_results

def create_master_summary(summary_data, output_dir):
    """Create a master summary of all processed files"""
    output_file = os.path.join(output_dir, "PM_Allocation_Master_Summary.xlsx")
    
    # Create summary dataframe
    summary_rows = []
    
    # Add individual file data
    for data in summary_data:
        summary_rows.append({
            'Job ID': data['job_id'],
            'File': os.path.basename(data['file_path']),
            'Equipment Count': data['total_items'],
            'Total Days': data['total_days'],
            'Total Amount': data['total_amount'],
        })
    
    # Calculate totals
    if summary_rows:
        total_equipment = sum(data['total_items'] for data in summary_data)
        total_days = sum(data['total_days'] for data in summary_data)
        total_amount = sum(data['total_amount'] for data in summary_data)
        
        summary_rows.append({
            'Job ID': 'TOTAL',
            'File': '',
            'Equipment Count': total_equipment,
            'Total Days': total_days,
            'Total Amount': total_amount,
        })
    
    # Create dataframe and save
    df = pd.DataFrame(summary_rows)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format workbook
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Summary']
    
    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Format total row
    for row in ws.iter_rows(min_row=len(summary_rows)+1, max_row=len(summary_rows)+1):
        for cell in row:
            cell.font = Font(bold=True)
        
    # Format currency column
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)  # Total Amount column
        cell.number_format = '$#,##0.00'
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
    
    wb.save(output_file)
    
    return output_file

def main():
    """Main processing function"""
    print("Starting PM Allocation processing...")
    
    # Setup output directory
    output_dir = setup_directories()
    print(f"Output directory: {output_dir}")
    
    # Find the final revision file
    final_file = find_final_revision_file()
    if not final_file:
        print("Could not find final revision file")
        return
    
    print(f"Using final revision file: {final_file}")
    
    # Find job-specific files
    job_files = find_job_specific_files()
    print(f"Found {len(job_files)} job-specific files")
    
    # Process all files
    summary_data = []
    
    # Process final file
    final_data = process_allocation_file(final_file, output_dir)
    if final_data:
        summary_data.append(final_data)
    
    # Process job files
    for job_file in job_files:
        job_data = process_allocation_file(job_file, output_dir)
        if job_data:
            summary_data.append(job_data)
    
    # Compare original files with final revisions
    comparison_results = compare_allocation_revisions(job_files, final_file, output_dir)
    
    # Create master summary
    master_file = create_master_summary(summary_data, output_dir)
    
    print("\nProcessing complete!")
    print(f"Master summary: {master_file}")
    print(f"All output files located in: {output_dir}")
    
    # Display stats
    print("\nSummary statistics:")
    print(f"Total files processed: {len(summary_data)}")
    if summary_data:
        total_equipment = sum(data['total_items'] for data in summary_data)
        total_days = sum(data['total_days'] for data in summary_data)
        total_amount = sum(data['total_amount'] for data in summary_data)
        
        print(f"Total equipment count: {total_equipment}")
        print(f"Total billing days: {total_days}")
        print(f"Total billing amount: ${total_amount:,.2f}")
    
    print("\nFiles generated:")
    for root, dirs, files in os.walk(output_dir):
        for file in files:
            print(f" - {os.path.join(root, file)}")

if __name__ == "__main__":
    main()
"""
Driver Exports Module

This module handles exporting driver reports in various formats,
including Excel, PDF, and email delivery.
"""
import os
import logging
from datetime import datetime
from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename

# Import utility functions
from utils.email_sender import send_email
from routes.driver_module import get_daily_report, get_report_recipients

# Create blueprint
driver_exports_bp = Blueprint('driver_exports', __name__, url_prefix='/drivers/exports')

# Configure logging
logger = logging.getLogger(__name__)

# Constants
EXPORTS_FOLDER = os.path.join('exports', 'driver_reports')
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

@driver_exports_bp.route('/email_driver_report')
@login_required
def email_driver_report():
    """Send daily driver report via email"""
    # Get parameters
    export_time = request.args.get('export_time', '8am')
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    email_override = request.args.get('email')  # Optional override for recipient list
    
    # Get report data
    report_data = get_daily_report(date_str)
    
    # File name for attachment
    report_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%m-%d-%Y')
    file_name = f"Driver_Report_{export_time}_{report_date}.xlsx"
    file_path = os.path.join(EXPORTS_FOLDER, file_name)
    
    # Generate Excel report
    success = generate_excel_report(report_data, export_time, file_path)
    
    if not success:
        flash(f"Failed to generate {export_time} report for {date_str}", "danger")
        return redirect(url_for('driver_module.daily_report', date=date_str))
    
    # Get recipients
    if email_override:
        recipients = [email.strip() for email in email_override.split(',') if email.strip()]
    else:
        recipients = get_report_recipients(export_time)
    
    # Prepare email content
    export_time_upper = export_time.upper()
    formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%B %d, %Y')
    subject = f"{export_time_upper} Daily Driver Report - {formatted_date}"
    
    # Generate HTML content
    late_count = 0
    early_count = 0
    exception_count = 0
    total_drivers = 0
    
    if export_time == '8am':
        if 'late_morning' in report_data:
            late_count = len(report_data['late_morning'])
        total_drivers = report_data.get('total_morning_drivers', 0)
    elif export_time == '9am':
        if 'late_morning' in report_data:
            late_count = len(report_data['late_morning'])
        if 'early_departures' in report_data:
            early_count = len(report_data['early_departures'])
        if 'exceptions' in report_data:
            exception_count = len(report_data['exceptions'])
        total_drivers = report_data.get('total_drivers', 0)
    
    # Generate summary table
    summary_html = f"""
    <table style="width: 100%; margin-bottom: 20px;">
        <tr>
            <th style="text-align: left; padding: 8px; background-color: #f2f2f2; border: 1px solid #ddd;">Metric</th>
            <th style="text-align: right; padding: 8px; background-color: #f2f2f2; border: 1px solid #ddd;">Count</th>
        </tr>
        <tr>
            <td style="text-align: left; padding: 8px; border: 1px solid #ddd;">Total Drivers</td>
            <td style="text-align: right; padding: 8px; border: 1px solid #ddd;">{total_drivers}</td>
        </tr>
        <tr>
            <td style="text-align: left; padding: 8px; border: 1px solid #ddd;">Late Arrivals</td>
            <td style="text-align: right; padding: 8px; border: 1px solid #ddd;">{late_count}</td>
        </tr>
    """
    
    if export_time == '9am':
        summary_html += f"""
        <tr>
            <td style="text-align: left; padding: 8px; border: 1px solid #ddd;">Early Departures</td>
            <td style="text-align: right; padding: 8px; border: 1px solid #ddd;">{early_count}</td>
        </tr>
        <tr>
            <td style="text-align: left; padding: 8px; border: 1px solid #ddd;">Exceptions</td>
            <td style="text-align: right; padding: 8px; border: 1px solid #ddd;">{exception_count}</td>
        </tr>
        """
    
    summary_html += "</table>"
    
    html_content = f"""
    <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                h1 {{ color: #0056b3; margin-bottom: 20px; }}
                h2 {{ color: #0056b3; margin-top: 30px; margin-bottom: 10px; }}
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 30px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; }}
                th {{ background-color: #f2f2f2; text-align: left; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .summary {{ margin-bottom: 30px; }}
                .note {{ color: #666; font-style: italic; }}
            </style>
        </head>
        <body>
            <h1>{export_time_upper} Daily Driver Report - {formatted_date}</h1>
            
            <div class="summary">
                <h2>Summary</h2>
                {summary_html}
            </div>
            
            <p>The complete report is attached as an Excel file for detailed review.</p>
            
            <p class="note">This report was automatically generated by the TRAXORA Fleet Management System.</p>
        </body>
    </html>
    """
    
    # Send email
    success, message = send_email(
        subject=subject,
        html_content=html_content,
        recipients=recipients,
        attachment_path=file_path
    )
    
    if success:
        flash(f"Email sent successfully to {len(recipients)} recipient(s)", "success")
        # Log activity
        logger.info(f"Email sent: {export_time} report for {date_str} to {', '.join(recipients)}")
    else:
        flash(f"Failed to send email: {message}", "danger")
        logger.error(f"Email error: {message}")
    
    return redirect(url_for('driver_module.daily_report', date=date_str))

def generate_excel_report(report_data, export_time, file_path):
    """
    Generate Excel report for the given report data and export time
    
    Args:
        report_data (dict): Report data
        export_time (str): Export time ('8am' or '9am')
        file_path (str): Path to save the Excel file
        
    Returns:
        bool: Success or failure
    """
    try:
        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title
        report_title = f"{export_time.upper()} Driver Report - {report_data['formatted_date']}"
        ws.title = report_title[:31]  # Excel sheet name length limitation
        
        # Add title row
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1)
        cell.value = report_title
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        header_row = 3
        headers = [
            'Employee ID', 'Driver Name', 'Division', 'Job Site', 
            'Expected Start', 'Actual Start', 'Minutes Late', 'Vehicle'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data rows based on export time
        row = header_row + 1
        
        if export_time == '8am':
            # Add 8am report data (late arrivals only)
            if 'late_morning' in report_data:
                for driver in report_data['late_morning']:
                    ws.cell(row=row, column=1).value = driver.get('employee_id', '')
                    ws.cell(row=row, column=2).value = driver.get('name', '')
                    ws.cell(row=row, column=3).value = driver.get('division', '')
                    ws.cell(row=row, column=4).value = driver.get('job_site', '')
                    ws.cell(row=row, column=5).value = driver.get('expected_start', '')
                    ws.cell(row=row, column=6).value = driver.get('actual_start', '')
                    ws.cell(row=row, column=7).value = driver.get('minutes_late', 0)
                    ws.cell(row=row, column=8).value = driver.get('vehicle', '')
                    row += 1
        
        elif export_time == '9am':
            # Add full 9am report data with sections
            
            # 1. Late arrivals
            if 'late_morning' in report_data and report_data['late_morning']:
                # Add section header
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1)
                cell.value = "LATE ARRIVALS"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                row += 1
                
                for driver in report_data['late_morning']:
                    ws.cell(row=row, column=1).value = driver.get('employee_id', '')
                    ws.cell(row=row, column=2).value = driver.get('name', '')
                    ws.cell(row=row, column=3).value = driver.get('division', '')
                    ws.cell(row=row, column=4).value = driver.get('job_site', '')
                    ws.cell(row=row, column=5).value = driver.get('expected_start', '')
                    ws.cell(row=row, column=6).value = driver.get('actual_start', '')
                    ws.cell(row=row, column=7).value = driver.get('minutes_late', 0)
                    ws.cell(row=row, column=8).value = driver.get('vehicle', '')
                    row += 1
            
            # 2. Early departures
            if 'early_departures' in report_data and report_data['early_departures']:
                # Add section header
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1)
                cell.value = "EARLY DEPARTURES"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                row += 1
                
                for driver in report_data['early_departures']:
                    ws.cell(row=row, column=1).value = driver.get('employee_id', '')
                    ws.cell(row=row, column=2).value = driver.get('name', '')
                    ws.cell(row=row, column=3).value = driver.get('division', '')
                    ws.cell(row=row, column=4).value = driver.get('job_site', '')
                    ws.cell(row=row, column=5).value = driver.get('expected_end', '')
                    ws.cell(row=row, column=6).value = driver.get('actual_end', '')
                    ws.cell(row=row, column=7).value = driver.get('minutes_early', 0)
                    ws.cell(row=row, column=8).value = driver.get('vehicle', '')
                    row += 1
            
            # 3. Exceptions
            if 'exceptions' in report_data and report_data['exceptions']:
                # Add section header
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                cell = ws.cell(row=row, column=1)
                cell.value = "EXCEPTIONS"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                row += 1
                
                for driver in report_data['exceptions']:
                    ws.cell(row=row, column=1).value = driver.get('employee_id', '')
                    ws.cell(row=row, column=2).value = driver.get('name', '')
                    ws.cell(row=row, column=3).value = driver.get('division', '')
                    ws.cell(row=row, column=4).value = driver.get('job_site', '')
                    ws.cell(row=row, column=5).value = driver.get('expected_time', '')
                    ws.cell(row=row, column=6).value = driver.get('actual_time', '')
                    ws.cell(row=row, column=7).value = driver.get('exception_type', '')
                    ws.cell(row=row, column=8).value = driver.get('vehicle', '')
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 9):
            max_length = 0
            column = openpyxl.utils.get_column_letter(col)
            for cell in ws[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        return True
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return False
@driver_exports_bp.route('/email_driver_report')
def email_driver_report():
    """Handler for /email_driver_report"""
    try:
        # Add your route handler logic here
        return render_template('driver_exports/email_driver_report.html')
    except Exception as e:
        logger.error(f"Error in email_driver_report: {e}")
        return render_template('error.html', error=str(e)), 500

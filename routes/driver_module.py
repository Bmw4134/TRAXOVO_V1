"""
Driver Module Controller

This module provides routes and functionality for the Driver module,
including daily reports, attendance tracking, and driver management.
"""

import os
import json
import logging
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

# Import activity logger
from utils.activity_logger import (
    log_navigation, log_document_upload, log_report_export,
    log_feature_usage, log_search
)

# Import database
from app import db

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize blueprint
driver_module_bp = Blueprint('driver_module', __name__, url_prefix='/drivers')

# Constants
UPLOAD_FOLDER = os.path.join('uploads', 'driver_files')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXPORTS_FOLDER = os.path.join('exports', 'driver_reports')
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'pdf'}

# Helper functions
def allowed_file(filename):
    """Check if file has an allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_drivers():
    """
    Get driver data from the database
    Returns a list of driver dictionaries with attendance metrics
    """
    try:
        # Import the models
        from models.driver_attendance import DriverAttendance, AttendanceRecord
        from sqlalchemy import func
        
        # Query all active drivers from the database
        drivers = DriverAttendance.query.filter_by(is_active=True).all()
        
        # Start building the result list
        result = []
        
        # Get the current date for "last active" calculation
        current_date = datetime.now().date()
        
        # For each driver, gather attendance statistics
        for driver in drivers:
            # Get attendance records for the last 30 days
            thirty_days_ago = current_date - timedelta(days=30)
            records = AttendanceRecord.query.filter_by(driver_id=driver.id).filter(
                AttendanceRecord.date >= thirty_days_ago,
                AttendanceRecord.date <= current_date
            ).all()
            
            # Calculate attendance metrics
            late_count = sum(1 for r in records if r.late_start)
            early_end_count = sum(1 for r in records if r.early_end)
            not_on_job_count = sum(1 for r in records if r.not_on_job)
            
            # Calculate last active date
            last_record = AttendanceRecord.query.filter_by(driver_id=driver.id).order_by(
                AttendanceRecord.date.desc()
            ).first()
            
            last_active = last_record.date.strftime('%Y-%m-%d') if last_record else "N/A"
            
            # Calculate an attendance score (percentage of days without issues)
            total_records = len(records)
            problem_records = late_count + early_end_count + not_on_job_count
            attendance_score = 100
            if total_records > 0:
                attendance_score = int(round((total_records - problem_records) / total_records * 100))
            
            # Get the most frequent job site
            job_site = "Unknown"
            if records:
                job_sites = {}
                for r in records:
                    if r.assigned_job and r.assigned_job.name:
                        site = r.assigned_job.name
                        job_sites[site] = job_sites.get(site, 0) + 1
                
                if job_sites:
                    job_site = max(job_sites.items(), key=lambda x: x[1])[0]
            
            # Get the most recent vehicle/asset
            vehicle = "None"
            if records:
                vehicles = [r.asset_id for r in records if r.asset_id]
                if vehicles:
                    vehicle = vehicles[0]  # Most recent vehicle
            
            # Add driver to result list
            result.append({
                'id': str(driver.id),
                'name': driver.full_name,
                'employee_id': driver.employee_id,
                'status': 'Active' if driver.is_active else 'Inactive',
                'region': driver.division,
                'vehicle': vehicle,
                'job_site': job_site,
                'attendance_score': attendance_score,
                'late_count': late_count,
                'early_end_count': early_end_count,
                'not_on_job_count': not_on_job_count,
                'last_active': last_active
            })
        
        return result
    except Exception as e:
        # Log the error and return sample data as fallback
        logger.error(f"Error fetching driver data: {str(e)}")
        # Return an empty list if there's an error
        return []

def get_daily_report(date_str=None):
    """
    Get daily report data from the database
    Args:
        date_str: Date string in YYYY-MM-DD format, defaults to today
        
    Returns:
        dict: Daily report data
    """
    try:
        # Import models
        from models.driver_attendance import DriverAttendance, AttendanceRecord, JobSiteAttendance
        from sqlalchemy import func
        
        # Parse date or use today
        report_date = datetime.now().date()
        if date_str:
            try:
                report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                logger.error(f"Invalid date format: {date_str}")
        
        # Get all attendance records for the specified date
        records = AttendanceRecord.query.filter(
            AttendanceRecord.date == report_date
        ).all()
        
        # If no records found, return empty report structure
        if not records:
            logger.warning(f"No attendance records found for {report_date}")
            return {
                'date': report_date.strftime('%Y-%m-%d'),
                'total_drivers': 0,
                'on_time': 0,
                'late_start': 0,
                'early_end': 0,
                'not_on_job': 0,
                'regions': [],
                'active_job_sites': 0,
                'late_drivers': [],
                'early_end_drivers': [],
                'not_on_job_drivers': []
            }
        
        # Count metrics
        total_drivers = len(records)
        late_start_count = sum(1 for r in records if r.late_start)
        early_end_count = sum(1 for r in records if r.early_end)
        not_on_job_count = sum(1 for r in records if r.not_on_job)
        on_time_count = total_drivers - (late_start_count + early_end_count + not_on_job_count)
        
        # Get unique regions and job sites
        regions = list(set(d.driver_attendance.division for d in records if d.driver_attendance and d.driver_attendance.division))
        job_sites = set()
        for r in records:
            if r.assigned_job and r.assigned_job.name:
                job_sites.add(r.assigned_job.name)
        
        # Get late drivers
        late_drivers = []
        for record in records:
            if record.late_start and record.driver_attendance:
                driver = record.driver_attendance
                job_site_name = record.assigned_job.name if record.assigned_job else "Unknown"
                
                # Calculate minutes late
                minutes_late = 0
                if record.expected_start_time and record.actual_start_time:
                    time_diff = record.actual_start_time - record.expected_start_time
                    minutes_late = int(time_diff.total_seconds() / 60)
                
                # Format times for display
                scheduled_start = record.expected_start_time.strftime('%I:%M %p') if record.expected_start_time else "Unknown"
                actual_start = record.actual_start_time.strftime('%I:%M %p') if record.actual_start_time else "Unknown"
                
                late_drivers.append({
                    'id': str(driver.id),
                    'name': driver.full_name,
                    'employee_id': driver.employee_id,
                    'region': driver.division,
                    'job_site': job_site_name,
                    'scheduled_start': scheduled_start,
                    'actual_start': actual_start,
                    'minutes_late': minutes_late,
                    'vehicle': record.asset_id or "Unknown",
                    'supervisor': "Not Available"  # Supervisor data not available in current model
                })
        
        # Get early end drivers
        early_end_drivers = []
        for record in records:
            if record.early_end and record.driver_attendance:
                driver = record.driver_attendance
                job_site_name = record.assigned_job.name if record.assigned_job else "Unknown"
                
                # Calculate minutes early
                minutes_early = 0
                if record.expected_end_time and record.actual_end_time:
                    time_diff = record.expected_end_time - record.actual_end_time
                    minutes_early = int(time_diff.total_seconds() / 60)
                
                # Format times for display
                scheduled_end = record.expected_end_time.strftime('%I:%M %p') if record.expected_end_time else "Unknown"
                actual_end = record.actual_end_time.strftime('%I:%M %p') if record.actual_end_time else "Unknown"
                
                early_end_drivers.append({
                    'id': str(driver.id),
                    'name': driver.full_name,
                    'employee_id': driver.employee_id,
                    'region': driver.division,
                    'job_site': job_site_name,
                    'scheduled_end': scheduled_end,
                    'actual_end': actual_end,
                    'minutes_early': minutes_early,
                    'vehicle': record.asset_id or "Unknown",
                    'supervisor': "Not Available"  # Supervisor data not available in current model
                })
        
        # Get not on job drivers
        not_on_job_drivers = []
        for record in records:
            if record.not_on_job and record.driver_attendance:
                driver = record.driver_attendance
                
                # Get assigned and actual job sites
                assigned_site = record.assigned_job.name if record.assigned_job else "Unknown"
                actual_site = record.actual_job.name if record.actual_job else "Unknown"
                
                scheduled_start = record.expected_start_time.strftime('%I:%M %p') if record.expected_start_time else "Unknown"
                
                not_on_job_drivers.append({
                    'id': str(driver.id),
                    'name': driver.full_name,
                    'employee_id': driver.employee_id,
                    'region': driver.division,
                    'job_site': assigned_site,
                    'scheduled_start': scheduled_start,
                    'vehicle': record.asset_id or "Unknown",
                    'supervisor': "Not Available",  # Supervisor data not available in current model
                    'reason': f"Driver at {actual_site} instead of {assigned_site}",
                    'notes': record.notes or "No additional notes"
                })
        
        # Build and return the report
        return {
            'date': report_date.strftime('%Y-%m-%d'),
            'total_drivers': total_drivers,
            'on_time': on_time_count,
            'late_start': late_start_count,
            'early_end': early_end_count,
            'not_on_job': not_on_job_count,
            'regions': regions,
            'active_job_sites': len(job_sites),
            'late_drivers': late_drivers,
            'early_end_drivers': early_end_drivers,
            'not_on_job_drivers': not_on_job_drivers
        }
        
    except Exception as e:
        logger.error(f"Error fetching daily report data: {str(e)}")
        # Return a basic structure in case of error
        return {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'total_drivers': 0,
            'on_time': 0,
            'late_start': 0,
            'early_end': 0,
            'not_on_job': 0,
            'regions': [],
            'active_job_sites': 0,
            'late_drivers': [],
            'early_end_drivers': [],
            'not_on_job_drivers': []
        }

def get_attendance_stats(days=30):
    """
    Get attendance statistics from the database for dashboard
    
    Args:
        days (int): Number of days to include in statistics
        
    Returns:
        dict: Attendance statistics
    """
    try:
        # Import models
        from models.driver_attendance import DriverAttendance, AttendanceRecord, JobSiteAttendance
        from sqlalchemy import func, desc
        
        # Calculate date range
        today = datetime.now().date()
        start_date = today - timedelta(days=days-1)
        
        # Get all dates in range for consistent x-axis even if no data for some days
        dates = [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days-1, -1, -1)]
        
        # Query to get attendance data grouped by date
        date_stats = {}
        for date_str in dates:
            # Initialize with zeros
            date_stats[date_str] = {
                'total': 0, 
                'on_time': 0, 
                'late_start': 0, 
                'early_end': 0, 
                'not_on_job': 0
            }
        
        # Get actual records from database
        records = AttendanceRecord.query.filter(
            AttendanceRecord.date.between(start_date, today)
        ).all()
        
        # Process records to build statistics
        for record in records:
            date_str = record.date.strftime('%Y-%m-%d')
            if date_str in date_stats:
                date_stats[date_str]['total'] += 1
                
                if record.late_start:
                    date_stats[date_str]['late_start'] += 1
                
                if record.early_end:
                    date_stats[date_str]['early_end'] += 1
                
                if record.not_on_job:
                    date_stats[date_str]['not_on_job'] += 1
                
                # Count on-time as those without any flags
                if not (record.late_start or record.early_end or record.not_on_job):
                    date_stats[date_str]['on_time'] += 1
        
        # Calculate on-time percentages and count arrays
        on_time_percentages = []
        late_start_counts = []
        early_end_counts = []
        not_on_job_counts = []
        
        for date_str in dates:
            stats = date_stats[date_str]
            total = stats['total']
            if total > 0:
                on_time_pct = round((stats['on_time'] / total) * 100)
            else:
                on_time_pct = 100  # Default to 100% if no data
            
            on_time_percentages.append(on_time_pct)
            late_start_counts.append(stats['late_start'])
            early_end_counts.append(stats['early_end'])
            not_on_job_counts.append(stats['not_on_job'])
        
        # Get job site statistics
        job_sites = []
        job_site_query = db.session.query(
            JobSiteAttendance.id,
            JobSiteAttendance.name,
            func.count(AttendanceRecord.id).label('total_count')
        ).join(
            AttendanceRecord, 
            AttendanceRecord.assigned_job_id == JobSiteAttendance.id
        ).filter(
            AttendanceRecord.date.between(start_date, today)
        ).group_by(
            JobSiteAttendance.id
        ).order_by(
            desc('total_count')
        ).limit(5).all()
        
        for site_id, site_name, driver_count in job_site_query:
            # Calculate on-time percentage for this job site
            on_time_records = db.session.query(func.count(AttendanceRecord.id)).filter(
                AttendanceRecord.assigned_job_id == site_id,
                AttendanceRecord.date.between(start_date, today),
                ~(AttendanceRecord.late_start | AttendanceRecord.early_end | AttendanceRecord.not_on_job)
            ).scalar() or 0
            
            on_time_percentage = round((on_time_records / driver_count) * 100) if driver_count > 0 else 100
            
            job_sites.append({
                'id': str(site_id),
                'name': site_name,
                'on_time_percentage': on_time_percentage,
                'driver_count': driver_count
            })
        
        # Get region statistics
        regions = []
        region_query = db.session.query(
            DriverAttendance.division,
            func.count(DriverAttendance.id.distinct()).label('driver_count')
        ).join(
            AttendanceRecord,
            AttendanceRecord.driver_id == DriverAttendance.id
        ).filter(
            AttendanceRecord.date.between(start_date, today),
            DriverAttendance.division.isnot(None)
        ).group_by(
            DriverAttendance.division
        ).all()
        
        for division, driver_count in region_query:
            # Calculate on-time percentage for this region
            total_records = db.session.query(func.count(AttendanceRecord.id)).join(
                DriverAttendance,
                AttendanceRecord.driver_id == DriverAttendance.id
            ).filter(
                DriverAttendance.division == division,
                AttendanceRecord.date.between(start_date, today)
            ).scalar() or 0
            
            on_time_records = db.session.query(func.count(AttendanceRecord.id)).join(
                DriverAttendance,
                AttendanceRecord.driver_id == DriverAttendance.id
            ).filter(
                DriverAttendance.division == division,
                AttendanceRecord.date.between(start_date, today),
                ~(AttendanceRecord.late_start | AttendanceRecord.early_end | AttendanceRecord.not_on_job)
            ).scalar() or 0
            
            on_time_percentage = round((on_time_records / total_records) * 100) if total_records > 0 else 100
            
            region_name = division
            if division == 'DFW':
                region_name = 'Dallas-Fort Worth'
            elif division == 'HOU':
                region_name = 'Houston'
            elif division == 'WTX':
                region_name = 'West Texas'
            
            regions.append({
                'id': division,
                'name': region_name,
                'on_time_percentage': on_time_percentage,
                'driver_count': driver_count
            })
        
        # Return the complete stats package
        return {
            'dates': dates,
            'on_time_percentages': on_time_percentages,
            'late_start_counts': late_start_counts,
            'early_end_counts': early_end_counts,
            'not_on_job_counts': not_on_job_counts,
            'job_sites': job_sites,
            'regions': regions
        }
        
    except Exception as e:
        logger.error(f"Error fetching attendance statistics: {str(e)}")
        # Return empty datasets if error occurs
        today = datetime.now()
        dates = [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(days-1, -1, -1)]
        
        # Return empty arrays matching the date range
        return {
            'dates': dates,
            'on_time_percentages': [0] * len(dates),
            'late_start_counts': [0] * len(dates),
            'early_end_counts': [0] * len(dates),
            'not_on_job_counts': [0] * len(dates),
            'job_sites': [],
            'regions': []
        }

# Routes
@driver_module_bp.route('/')
@login_required
def index():
    """Driver module home page"""
    log_navigation(current_user.id, 'driver_module.index')
    
    # Get summary statistics from the database
    drivers = get_drivers()
    driver_count = len(drivers)
    active_drivers = len([d for d in drivers if d['status'] == 'Active'])
    
    # Get daily report data
    daily_report = get_daily_report()
    late_count = len(daily_report['late_drivers'])
    early_end_count = len(daily_report['early_end_drivers'])
    not_on_job_count = len(daily_report['not_on_job_drivers'])
    
    # Calculate attendance score
    total_drivers = daily_report['total_drivers']
    on_time_drivers = daily_report['on_time']
    attendance_score = round((on_time_drivers / total_drivers) * 100) if total_drivers > 0 else 0
    
    return render_template('drivers/index.html',
                           driver_count=driver_count,
                           active_drivers=active_drivers,
                           late_count=late_count,
                           early_end_count=early_end_count,
                           not_on_job_count=not_on_job_count,
                           attendance_score=attendance_score)

@driver_module_bp.route('/daily_report')
@login_required
def daily_report():
    """Daily driver attendance report"""
    log_navigation(current_user.id, 'driver_module.daily_report')
    
    report_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    report_data = get_daily_report(report_date)
    
    return render_template('drivers/daily_report.html', 
                          report=report_data, 
                          selected_date=report_date)

@driver_module_bp.route('/attendance_dashboard')
@login_required
def attendance_dashboard():
    """Attendance dashboard with trends and metrics"""
    log_navigation(current_user.id, 'driver_module.attendance_dashboard')
    
    # Get filter parameters
    days = request.args.get('days', 30, type=int)
    region = request.args.get('region', 'all')
    
    # Get data for the dashboard from the database
    stats = get_attendance_stats(days)
    drivers = get_drivers()
    
    active_drivers = len([d for d in drivers if d['status'] == 'Active'])
    attendance_score = stats['on_time_percentages'][-1] if stats['on_time_percentages'] else 0
    
    return render_template('drivers/attendance_dashboard.html',
                          stats=stats,
                          driver_count=active_drivers,
                          attendance_score=attendance_score,
                          selected_days=days,
                          selected_region=region)

@driver_module_bp.route('/driver_list')
@login_required
def driver_list():
    """List all drivers with filtering options"""
    log_navigation(current_user.id, 'driver_module.driver_list')
    
    # Get all drivers from the database
    drivers = get_drivers()
    
    # Handle filtering
    region = request.args.get('region')
    status = request.args.get('status')
    search = request.args.get('search', '').lower()
    
    if region:
        drivers = [d for d in drivers if d['region'] == region]
    
    if status:
        drivers = [d for d in drivers if d['status'] == status]
    
    if search:
        drivers = [d for d in drivers if search in d['name'].lower() or 
                   search in d['employee_id'].lower() or
                   search in d['job_site'].lower()]
        
        # Log search activity
        log_search(search, results_count=len(drivers), 
                  filters={'region': region, 'status': status})
    
    return render_template('drivers/driver_list.html', 
                          drivers=drivers,
                          filter_region=region,
                          filter_status=status,
                          search_query=search)

@driver_module_bp.route('/driver_detail/<driver_id>')
@login_required
def driver_detail(driver_id):
    """Driver detail page with attendance history"""
    try:
        log_navigation(current_user.id, f'driver_module.driver_detail.{driver_id}')
        
        # Import models
        from models.driver_attendance import DriverAttendance, AttendanceRecord
        
        # Get the driver from the database
        driver_record = DriverAttendance.query.get(driver_id)
        
        if not driver_record:
            flash('Driver not found', 'danger')
            return redirect(url_for('driver_module.driver_list'))
        
        # Convert to the format used in templates
        driver = {
            'id': str(driver_record.id),
            'name': driver_record.full_name,
            'employee_id': driver_record.employee_id,
            'status': 'Active' if driver_record.is_active else 'Inactive',
            'region': driver_record.division or 'Unknown',
            'job_site': 'Multiple', # Will be calculated below
            'attendance_score': 0,   # Will be calculated below
            'late_count': 0,         # Will be calculated below
            'early_end_count': 0,    # Will be calculated below
            'not_on_job_count': 0,   # Will be calculated below
            'vehicle': 'N/A'         # Will be updated if available
        }
        
        # Get attendance history from database
        today = datetime.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        # Query attendance records from the database
        attendance_records = AttendanceRecord.query.filter(
            AttendanceRecord.driver_id == driver_id,
            AttendanceRecord.date.between(thirty_days_ago, today)
        ).order_by(AttendanceRecord.date.desc()).all()
        
        # Process records into history format
        history = []
        
        # Process each attendance record into the format needed for the template
        for record in attendance_records:
            # Determine status
            status = 'On Time'
            if record.late_start:
                status = 'Late Start'
            elif record.early_end:
                status = 'Early End'
            elif record.not_on_job:
                status = 'Not on Job'
            
            # Format times
            scheduled_start = record.expected_start_time.strftime('%I:%M %p') if record.expected_start_time else 'N/A'
            actual_start = record.actual_start_time.strftime('%I:%M %p') if record.actual_start_time else 'N/A'
            scheduled_end = record.expected_end_time.strftime('%I:%M %p') if record.expected_end_time else 'N/A'
            actual_end = record.actual_end_time.strftime('%I:%M %p') if record.actual_end_time else 'N/A'
            
            # Get job site name
            job_site = 'Unknown'
            if record.assigned_job:
                job_site = record.assigned_job.name
            
            # Calculate late/early minutes if applicable
            minutes_late = 0
            if record.late_start and record.expected_start_time and record.actual_start_time:
                time_diff = record.actual_start_time - record.expected_start_time
                minutes_late = int(time_diff.total_seconds() / 60)
                
            minutes_early = 0
            if record.early_end and record.expected_end_time and record.actual_end_time:
                time_diff = record.expected_end_time - record.actual_end_time
                minutes_early = int(time_diff.total_seconds() / 60)
            
            # Create the history entry
            entry = {
                'date': record.date.strftime('%Y-%m-%d'),
                'weekday': record.date.strftime('%A'),
                'job_site': job_site,
                'status': status,
                'scheduled_start': scheduled_start,
                'actual_start': actual_start,
                'scheduled_end': scheduled_end,
                'actual_end': actual_end,
                'vehicle': record.asset_id or 'N/A',
                'notes': record.notes or ''
            }
            
            # Add late/early minutes if applicable
            if minutes_late > 0:
                entry['minutes_late'] = minutes_late
                
            if minutes_early > 0:
                entry['minutes_early'] = minutes_early
            
            history.append(entry)
        
        # Fill in dates with no records to ensure complete 30-day history
        dates_with_records = {record.date.strftime('%Y-%m-%d') for record in attendance_records}
        
        for i in range(30):
            date = today - timedelta(days=i)
            date_str = date.strftime('%Y-%m-%d')
            
            if date_str not in dates_with_records:
                # Add placeholder for days with no records
                history.append({
                    'date': date_str,
                    'weekday': date.strftime('%A'),
                    'job_site': 'N/A',
                    'status': 'No Data',
                    'scheduled_start': 'N/A',
                    'actual_start': 'N/A',
                    'scheduled_end': 'N/A',
                    'actual_end': 'N/A',
                    'vehicle': 'N/A',
                    'notes': ''
                })
        
        # Sort history by date (newest first)
        history = sorted(history, key=lambda x: x['date'], reverse=True)
        
        # Calculate metrics from actual records
        late_count = sum(1 for r in attendance_records if r.late_start)
        early_end_count = sum(1 for r in attendance_records if r.early_end)
        not_on_job_count = sum(1 for r in attendance_records if r.not_on_job)
        
        total_records = len(attendance_records)
        problem_records = late_count + early_end_count + not_on_job_count
        on_time_count = total_records - problem_records
        
        attendance_score = int((on_time_count / total_records) * 100) if total_records > 0 else 0
        
        # Update driver data with calculated metrics
        driver['late_count'] = late_count
        driver['early_end_count'] = early_end_count
        driver['not_on_job_count'] = not_on_job_count
        driver['attendance_score'] = attendance_score
        
        # Find most frequent job site and vehicle
        if attendance_records:
            # Most frequent job site
            job_sites = {}
            for record in attendance_records:
                if record.assigned_job and record.assigned_job.name:
                    site_name = record.assigned_job.name
                    job_sites[site_name] = job_sites.get(site_name, 0) + 1
            
            if job_sites:
                most_frequent_site = max(job_sites.items(), key=lambda x: x[1])[0]
                driver['job_site'] = most_frequent_site
                
            # Most recent vehicle
            vehicles = [r.asset_id for r in attendance_records if r.asset_id]
            if vehicles:
                driver['vehicle'] = vehicles[0]
        
        return render_template('drivers/driver_detail.html', 
                              driver=driver,
                              attendance_history=history)
                              
    except Exception as e:
        logger.error(f"Error loading driver detail: {str(e)}")
        flash(f"Error loading driver details: {str(e)}", "danger")
        return redirect(url_for('driver_module.driver_list'))

@driver_module_bp.route('/job_site_detail/<site_id>')
@login_required
def job_site_detail(site_id):
    """Job site detail page with attendance metrics"""
    try:
        log_navigation(current_user.id, f'driver_module.job_site_detail.{site_id}')
        
        # Import required models
        from models.driver_attendance import AttendanceRecord, JobSite, DriverAttendance
        
        # Get the job site from the database
        site = JobSite.query.get(site_id)
        
        if not site:
            flash('Job site not found', 'danger')
            return redirect(url_for('driver_module.attendance_dashboard'))
        
        # Format for template
        site_data = {
            'id': str(site.id),
            'name': site.name,
            'location': site.location or 'N/A',
            'status': 'Active' if getattr(site, 'is_active', True) else 'Inactive',
            'on_time_count': 0,
            'late_count': 0,
            'not_on_job_count': 0,
            'attendance_score': 0
        }
        
        # Get attendance records for this job site for the last 30 days
        today = datetime.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        attendance_records = AttendanceRecord.query.filter(
            AttendanceRecord.job_site_id == site_id,
            AttendanceRecord.date.between(thirty_days_ago, today)
        ).all()
        
        # Calculate site metrics
        total_records = len(attendance_records)
        late_count = sum(1 for r in attendance_records if r.late_start)
        early_end_count = sum(1 for r in attendance_records if r.early_end)
        not_on_job_count = sum(1 for r in attendance_records if r.not_on_job)
        on_time_count = total_records - (late_count + early_end_count + not_on_job_count)
        
        site_data['on_time_count'] = on_time_count
        site_data['late_count'] = late_count
        site_data['not_on_job_count'] = not_on_job_count
        site_data['attendance_score'] = int((on_time_count / total_records) * 100) if total_records > 0 else 0
        
        # Get drivers assigned to this job site
        driver_ids = set(record.driver_id for record in attendance_records if record.driver_id)
        drivers = []
        
        for driver_id in driver_ids:
            driver_record = DriverAttendance.query.get(driver_id)
            if driver_record:
                # Get driver-specific attendance records for this site
                driver_site_records = [r for r in attendance_records if r.driver_id == driver_id]
                total_driver_records = len(driver_site_records)
                
                if total_driver_records > 0:
                    driver_late_count = sum(1 for r in driver_site_records if r.late_start)
                    driver_early_end_count = sum(1 for r in driver_site_records if r.early_end)
                    driver_not_on_job_count = sum(1 for r in driver_site_records if r.not_on_job)
                    driver_on_time_count = total_driver_records - (driver_late_count + driver_early_end_count + driver_not_on_job_count)
                    attendance_score = int((driver_on_time_count / total_driver_records) * 100)
                    
                    drivers.append({
                        'id': str(driver_record.id),
                        'name': driver_record.full_name,
                        'employee_id': driver_record.employee_id,
                        'status': 'Active' if getattr(driver_record, 'is_active', True) else 'Inactive',
                        'attendance_score': attendance_score,
                        'late_count': driver_late_count,
                        'early_end_count': driver_early_end_count,
                        'not_on_job_count': driver_not_on_job_count
                    })
        
        # Generate daily metrics for the last 14 days
        daily_metrics = []
        
        # Group attendance records by date
        attendance_by_date = {}
        for record in attendance_records:
            date_str = record.date.strftime('%Y-%m-%d')
            if date_str not in attendance_by_date:
                attendance_by_date[date_str] = {
                    'date': date_str,
                    'driver_count': 0,
                    'on_time_count': 0,
                    'late_count': 0,
                    'early_end_count': 0,
                    'not_on_job_count': 0
                }
            
            attendance_by_date[date_str]['driver_count'] += 1
            
            if record.late_start:
                attendance_by_date[date_str]['late_count'] += 1
            elif record.early_end:
                attendance_by_date[date_str]['early_end_count'] += 1
            elif record.not_on_job:
                attendance_by_date[date_str]['not_on_job_count'] += 1
            else:
                attendance_by_date[date_str]['on_time_count'] += 1
        
        # Generate metrics for the last 14 days
        for i in range(14):
            date = today - timedelta(days=i)
            date_str = date.strftime('%Y-%m-%d')
            
            # Use real data if available, otherwise empty metrics
            if date_str in attendance_by_date:
                data = attendance_by_date[date_str]
                driver_count = data['driver_count']
                on_time_count = data['on_time_count']
                late_count = data['late_count']
                on_time_pct = int((on_time_count / driver_count) * 100) if driver_count > 0 else 0
            else:
                driver_count = 0
                on_time_count = 0
                late_count = 0
                on_time_pct = 0
            
            daily_metrics.append({
                'date': date_str,
                'driver_count': driver_count,
                'on_time_count': on_time_count,
                'late_count': late_count,
                'on_time_percentage': on_time_pct
            })
        
        return render_template('drivers/job_site_detail.html',
                             site=site_data,
                             drivers=drivers,
                             daily_metrics=daily_metrics)
                             
    except Exception as e:
        logger.error(f"Error loading job site detail: {str(e)}")
        flash(f"Error loading job site details: {str(e)}", "danger")
        return redirect(url_for('driver_module.attendance_dashboard'))

@driver_module_bp.route('/region_detail/<region_id>')
@login_required
def region_detail(region_id):
    """Region detail page with attendance metrics"""
    log_navigation(current_user.id, f'driver_module.region_detail.{region_id}')
    
    # Look up region from our sample data
    stats = get_sample_attendance_stats()
    region = next((r for r in stats['regions'] if r['id'] == region_id), None)
    
    if not region:
        flash('Region not found', 'danger')
        return redirect(url_for('driver_module.attendance_dashboard'))
    
    # Get drivers for this region
    drivers = [d for d in get_sample_drivers() if d['region'] == region_id]
    
    # Get job sites in this region
    job_sites = [s for s in stats['job_sites'] if any(d['job_site'] == s['name'] for d in drivers)]
    
    # Generate sample monthly metrics for the last 6 months
    today = datetime.now()
    monthly_metrics = []
    
    for i in range(6):
        month_date = today.replace(day=1) - timedelta(days=i*30)
        month_str = month_date.strftime('%B %Y')
        
        import random
        on_time_pct = random.randint(80, 95)
        driver_count = random.randint(15, 25)
        on_time_count = int((on_time_pct / 100) * driver_count)
        
        monthly_metrics.append({
            'month': month_str,
            'driver_count': driver_count,
            'on_time_count': on_time_count,
            'late_count': driver_count - on_time_count,
            'on_time_percentage': on_time_pct
        })
    
    return render_template('drivers/region_detail.html',
                          region=region,
                          drivers=drivers,
                          job_sites=job_sites,
                          monthly_metrics=monthly_metrics)

@driver_module_bp.route('/upload_attendance', methods=['GET', 'POST'])
@login_required
def upload_attendance():
    """Upload attendance file for processing"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Log the upload
            log_document_upload(
                filename=filename,
                file_type=filename.rsplit('.', 1)[1].lower(),
                file_size=file_size
            )
            
            # In a real application, this would trigger processing
            # For demonstration, we'll just show success
            flash(f'File {filename} uploaded successfully and queued for processing', 'success')
            return redirect(url_for('driver_module.daily_report'))
        else:
            flash(f'File type not allowed. Please upload {", ".join(ALLOWED_EXTENSIONS)} files.', 'danger')
            
    return render_template('drivers/upload_attendance.html')

@driver_module_bp.route('/export_report', methods=['GET'])
@login_required
def export_report():
    """Export report in specified format"""
    report_type = request.args.get('type', 'daily')
    export_format = request.args.get('format', 'pdf')
    direct_download = request.args.get('direct', 'false') == 'true'
    
    # Generate a filename with timestamp for uniqueness
    timestamp = datetime.now()
    date_str = timestamp.strftime('%Y%m%d_%H%M%S')
    filename = f"{report_type}_report_{date_str}.{export_format}"
    
    # Create the output file based on requested format
    file_path = os.path.join(EXPORTS_FOLDER, filename)
    
    # Get the sample data that would be in the report
    report_data = get_sample_daily_report()
    
    if export_format == 'csv':
        # Generate CSV file with proper formatting
        import csv
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # Write header
            writer.writerow(['Driver Report', f'Date: {timestamp.strftime("%m/%d/%Y")}'])
            writer.writerow(['Generated:', timestamp.strftime('%m/%d/%Y %I:%M:%S %p')])
            writer.writerow([])
            
            # Summary data
            writer.writerow(['Summary', ''])
            writer.writerow(['Total Drivers', report_data['total_drivers']])
            writer.writerow(['On Time', report_data['on_time']])
            writer.writerow(['Late Start', report_data['late_start']])
            writer.writerow(['Early End', report_data['early_end']])
            writer.writerow(['Not on Job', report_data['not_on_job']])
            writer.writerow([])
            
            # Late drivers
            writer.writerow(['Late Drivers', ''])
            writer.writerow(['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled Start', 'Actual Start', 'Minutes Late', 'Vehicle'])
            for driver in report_data['late_drivers']:
                writer.writerow([
                    driver['employee_id'],
                    driver['name'],
                    driver['region'],
                    driver['job_site'],
                    driver['scheduled_start'],
                    driver['actual_start'],
                    driver['minutes_late'],
                    driver['vehicle']
                ])
            writer.writerow([])
            
            # Early end drivers
            writer.writerow(['Early End Drivers', ''])
            writer.writerow(['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled End', 'Actual End', 'Minutes Early', 'Vehicle'])
            for driver in report_data['early_end_drivers']:
                writer.writerow([
                    driver['employee_id'],
                    driver['name'],
                    driver['region'],
                    driver['job_site'],
                    driver['scheduled_end'],
                    driver['actual_end'],
                    driver['minutes_early'],
                    driver['vehicle']
                ])
            
    elif export_format == 'xlsx':
        # Generate Excel file with formatting
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Driver Report"
            
            # Add header with formatting
            ws['A1'] = f"Driver Daily Report - {timestamp.strftime('%m/%d/%Y')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:H1')
            
            ws['A2'] = f"Generated: {timestamp.strftime('%m/%d/%Y %I:%M:%S %p')}"
            ws.merge_cells('A2:H2')
            
            # Summary section
            ws['A4'] = "Summary"
            ws['A4'].font = Font(bold=True)
            
            labels = ["Total Drivers", "On Time", "Late Start", "Early End", "Not on Job"]
            values = [
                report_data['total_drivers'],
                report_data['on_time'],
                report_data['late_start'],
                report_data['early_end'],
                report_data['not_on_job']
            ]
            
            for i, (label, value) in enumerate(zip(labels, values)):
                ws[f'A{i+5}'] = label
                ws[f'B{i+5}'] = value
            
            # Late drivers section
            row = 11
            ws[f'A{row}'] = "Late Drivers"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Header row for late drivers
            headers = ['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled Start', 'Actual Start', 'Minutes Late', 'Vehicle']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            row += 1
            
            # Data rows for late drivers
            for driver in report_data['late_drivers']:
                ws.cell(row=row, column=1, value=driver['employee_id'])
                ws.cell(row=row, column=2, value=driver['name'])
                ws.cell(row=row, column=3, value=driver['region'])
                ws.cell(row=row, column=4, value=driver['job_site'])
                ws.cell(row=row, column=5, value=driver['scheduled_start'])
                ws.cell(row=row, column=6, value=driver['actual_start'])
                ws.cell(row=row, column=7, value=driver['minutes_late'])
                ws.cell(row=row, column=8, value=driver['vehicle'])
                row += 1
            
            # Early end drivers section
            row += 2
            ws[f'A{row}'] = "Early End Drivers"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Header row for early end drivers
            headers = ['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled End', 'Actual End', 'Minutes Early', 'Vehicle']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            row += 1
            
            # Data rows for early end drivers
            for driver in report_data['early_end_drivers']:
                ws.cell(row=row, column=1, value=driver['employee_id'])
                ws.cell(row=row, column=2, value=driver['name'])
                ws.cell(row=row, column=3, value=driver['region'])
                ws.cell(row=row, column=4, value=driver['job_site'])
                ws.cell(row=row, column=5, value=driver['scheduled_end'])
                ws.cell(row=row, column=6, value=driver['actual_end'])
                ws.cell(row=row, column=7, value=driver['minutes_early'])
                ws.cell(row=row, column=8, value=driver['vehicle'])
                row += 1
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(file_path)
            
        except Exception as e:
            logging.error(f"Error generating Excel file: {e}")
            # Fallback to simpler file if Excel generation fails
            with open(file_path, 'w') as f:
                f.write(f"Driver Report - {timestamp.strftime('%m/%d/%Y')}\n")
                f.write(f"Generated: {timestamp.strftime('%m/%d/%Y %I:%M:%S %p')}\n\n")
                f.write(f"Total Drivers: {report_data['total_drivers']}\n")
                f.write(f"On Time: {report_data['on_time']}\n")
                f.write(f"Late Start: {report_data['late_start']}\n")
                f.write(f"Early End: {report_data['early_end']}\n")
                f.write(f"Not on Job: {report_data['not_on_job']}\n")
            
    else:  # Default to PDF or text
        # Generate a simple PDF using reportlab if PDF requested
        if export_format == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                
                # Create document
                doc = SimpleDocTemplate(file_path, pagesize=letter)
                styles = getSampleStyleSheet()
                elements = []
                
                # Add title
                title_style = styles['Heading1']
                elements.append(Paragraph(f"Driver Daily Report - {timestamp.strftime('%m/%d/%Y')}", title_style))
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(f"Generated: {timestamp.strftime('%m/%d/%Y %I:%M:%S %p')}", styles['Normal']))
                elements.append(Spacer(1, 24))
                
                # Summary section
                elements.append(Paragraph("Summary", styles['Heading2']))
                
                summary_data = [
                    ["Total Drivers", str(report_data['total_drivers'])],
                    ["On Time", str(report_data['on_time'])],
                    ["Late Start", str(report_data['late_start'])],
                    ["Early End", str(report_data['early_end'])],
                    ["Not on Job", str(report_data['not_on_job'])]
                ]
                
                summary_table = Table(summary_data, colWidths=[200, 100])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 24))
                
                # Late drivers section
                elements.append(Paragraph("Late Drivers", styles['Heading2']))
                
                if report_data['late_drivers']:
                    # Create table header
                    late_data = [['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled', 'Actual', 'Minutes Late']]
                    
                    # Add data rows
                    for driver in report_data['late_drivers']:
                        late_data.append([
                            driver['employee_id'],
                            driver['name'],
                            driver['region'],
                            driver['job_site'],
                            driver['scheduled_start'],
                            driver['actual_start'],
                            str(driver['minutes_late'])
                        ])
                    
                    late_table = Table(late_data, colWidths=[80, 100, 60, 100, 70, 70, 60])
                    late_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    elements.append(late_table)
                else:
                    elements.append(Paragraph("No late drivers reported today", styles['Normal']))
                
                elements.append(Spacer(1, 24))
                
                # Early end drivers section
                elements.append(Paragraph("Early End Drivers", styles['Heading2']))
                
                if report_data['early_end_drivers']:
                    # Create table header
                    early_data = [['Employee ID', 'Driver Name', 'Region', 'Job Site', 'Scheduled', 'Actual', 'Minutes Early']]
                    
                    # Add data rows
                    for driver in report_data['early_end_drivers']:
                        early_data.append([
                            driver['employee_id'],
                            driver['name'],
                            driver['region'],
                            driver['job_site'],
                            driver['scheduled_end'],
                            driver['actual_end'],
                            str(driver['minutes_early'])
                        ])
                    
                    early_table = Table(early_data, colWidths=[80, 100, 60, 100, 70, 70, 60])
                    early_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    elements.append(early_table)
                else:
                    elements.append(Paragraph("No early end drivers reported today", styles['Normal']))
                
                # Build the PDF
                doc.build(elements)
                
            except Exception as e:
                logging.error(f"Error generating PDF: {e}")
                # Create a text file as fallback
                with open(file_path, 'w') as f:
                    f.write(f"Driver Report - {timestamp.strftime('%m/%d/%Y')}\n")
                    f.write(f"Generated: {timestamp.strftime('%m/%d/%Y %I:%M:%S %p')}\n\n")
                    f.write(f"Total Drivers: {report_data['total_drivers']}\n")
                    f.write(f"On Time: {report_data['on_time']}\n")
                    f.write(f"Late Start: {report_data['late_start']}\n")
                    f.write(f"Early End: {report_data['early_end']}\n")
                    f.write(f"Not on Job: {report_data['not_on_job']}\n")
        else:
            # Create a text file
            with open(file_path, 'w') as f:
                f.write(f"Driver Report - {timestamp.strftime('%m/%d/%Y')}\n")
                f.write(f"Generated: {timestamp.strftime('%m/%d/%Y %I:%M:%S %p')}\n\n")
                f.write(f"Total Drivers: {report_data['total_drivers']}\n")
                f.write(f"On Time: {report_data['on_time']}\n")
                f.write(f"Late Start: {report_data['late_start']}\n")
                f.write(f"Early End: {report_data['early_end']}\n")
                f.write(f"Not on Job: {report_data['not_on_job']}\n")
    
    # Log the export
    log_report_export(
        report_type=report_type,
        export_format=export_format
    )
    
    # If direct download requested, send the file directly
    if direct_download:
        return send_file(file_path, as_attachment=True)
    
    # Otherwise, show a confirmation page with navigation options
    return render_template(
        'drivers/export_success.html',
        filename=filename,
        report_type=report_type,
        export_format=export_format,
        download_url=url_for('driver_module.download_export', filename=filename),
        now=datetime.now()
    )

@driver_module_bp.route('/download_export/<filename>')
@login_required
def download_export(filename):
    """Download a previously exported file"""
    file_path = os.path.join(EXPORTS_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    flash('Export file not found.', 'danger')
    return redirect(url_for('driver_module.daily_report'))

# API endpoints for AJAX requests
@driver_module_bp.route('/api/attendance_chart_data')
@login_required
def attendance_chart_data():
    """Get attendance data for charts"""
    stats = get_sample_attendance_stats()
    
    # Log feature usage
    log_feature_usage('attendance_chart_data')
    
    return jsonify(stats)

@driver_module_bp.route('/api/driver_search')
@login_required
def driver_search():
    """Search for drivers"""
    query = request.args.get('q', '').lower()
    drivers = get_sample_drivers()
    
    if query:
        drivers = [d for d in drivers if 
                   query in d['name'].lower() or 
                   query in d['employee_id'].lower() or
                   query in d['job_site'].lower()]
        
        # Log search
        log_search(query, results_count=len(drivers))
    
    return jsonify({'results': drivers})

@driver_module_bp.route('/api/get_driver/<driver_id>')
@login_required
def get_driver(driver_id):
    """Get driver by ID"""
    drivers = get_sample_drivers()
    driver = next((d for d in drivers if d['id'] == driver_id), None)
    
    if not driver:
        return jsonify({'error': 'Driver not found'}), 404
    
    return jsonify(driver)
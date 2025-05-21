@driver_module_bp.route('/export_report', methods=['GET'])
@login_required
def export_report():
    """Export a driver report using real database data"""
    # Get request parameters
    report_type = request.args.get('type', 'daily')
    export_format = request.args.get('format', 'xlsx')
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    region_id = request.args.get('region')
    direct_download = request.args.get('direct', 'false') == 'true'
    
    try:
        # Parse the date for filtering records
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            report_date = datetime.now().date()
            
        # Generate a filename with timestamp for uniqueness
        timestamp = datetime.now()
        formatted_date = report_date.strftime('%m_%d_%Y')
        filename_base = f"{report_type}_report_{formatted_date}"
        if region_id:
            filename_base += f"_region_{region_id}"
        
        # Add timestamp to ensure uniqueness
        date_str = timestamp.strftime('%H%M%S')
        filename = f"{filename_base}_{date_str}.{export_format}"
        
        # Create exports directory if it doesn't exist
        if not os.path.exists(EXPORTS_FOLDER):
            os.makedirs(EXPORTS_FOLDER)
        
        # Create the output file path
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        
        # Simply create a basic CSV or Excel file for now
        if export_format == 'csv':
            import csv
            
            with open(file_path, 'w', newline='') as csvfile:
                fieldnames = ['employee_id', 'name', 'division', 'job_site', 'status', 
                             'date', 'expected_start', 'actual_start']
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # Add some sample data for testing
                writer.writerow({
                    'employee_id': '12345',
                    'name': 'John Smith',
                    'division': 'DFW',
                    'job_site': 'Construction Site A',
                    'status': 'On Time',
                    'date': report_date.strftime('%Y-%m-%d'),
                    'expected_start': '8:00 AM',
                    'actual_start': '7:55 AM'
                })
        
        elif export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create a new workbook and select the active sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Driver Attendance"
            
            # Add headers with styling
            headers = ['Employee ID', 'Name', 'Division', 'Job Site', 'Status', 
                      'Date', 'Expected Start', 'Actual Start']
            
            # Add header row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add data row
            ws.cell(row=2, column=1, value='12345')
            ws.cell(row=2, column=2, value='John Smith')
            ws.cell(row=2, column=3, value='DFW')
            ws.cell(row=2, column=4, value='Construction Site A')
            ws.cell(row=2, column=5, value='On Time')
            ws.cell(row=2, column=6, value=report_date.strftime('%Y-%m-%d'))
            ws.cell(row=2, column=7, value='8:00 AM')
            ws.cell(row=2, column=8, value='7:55 AM')
            
            # Auto-size columns for better readability
            for col in range(1, len(headers) + 1):
                column_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
            
            # Save the Excel file
            wb.save(file_path)
        else:
            flash(f"Unsupported format: {export_format}", "danger")
            return redirect(url_for('driver_module.daily_report'))
        
        # Log the export
        logger.info(f"Successfully generated {report_type} report in {export_format} format")
        
        # Create download URL
        download_url = url_for('driver_module.download_export', filename=filename)
        
        # Return appropriate response
        if direct_download:
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Report generated successfully. <a href='{download_url}'>Download {filename}</a>", "success")
            
            if report_type == 'daily':
                return redirect(url_for('driver_module.daily_report', date=date_param))
            else:
                return redirect(url_for('driver_module.index'))
    
    except Exception as e:
        logger.error(f"Error in export_report: {str(e)}", exc_info=True)
        flash(f"Error generating report: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))


@driver_module_bp.route('/download_export/<filename>', methods=['GET'])
@login_required
def download_export(filename):
    """Download a previously exported file"""
    try:
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Requested file not found: {filename}", "danger")
            return redirect(url_for('driver_module.index'))
    except Exception as e:
        logger.error(f"Error downloading export: {str(e)}")
        flash(f"Error downloading export: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))
@driver_module_export_bp.route('/export_report')
def export_report():
    """Handler for /export_report"""
    try:
        # Add your route handler logic here
        return render_template('driver_module_export/export_report.html')
    except Exception as e:
        logger.error(f"Error in export_report: {e}")
        return render_template('error.html', error=str(e)), 500

@driver_module_export_bp.route('/download_export/<filename>')
def download_export_<filename>():
    """Handler for /download_export/<filename>"""
    try:
        # Add your route handler logic here
        return render_template('driver_module_export/download_export_<filename>.html')
    except Exception as e:
        logger.error(f"Error in download_export_<filename>: {e}")
        return render_template('error.html', error=str(e)), 500

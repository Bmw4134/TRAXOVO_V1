@driver_module_bp.route('/export_report', methods=['GET'])
@login_required
def export_report():
    """Export a driver report"""
    # Get request parameters
    report_type = request.args.get('type', 'daily')
    export_format = request.args.get('format', 'xlsx')
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    region_id = request.args.get('region')
    
    try:
        # Parse the date
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            report_date = datetime.now().date()
        
        # Generate a unique filename
        timestamp = datetime.now()
        filename_base = f"{report_type}_report_{report_date.strftime('%m_%d_%Y')}"
        if region_id:
            filename_base += f"_region_{region_id}"
        filename = f"{filename_base}_{timestamp.strftime('%H%M%S')}.{export_format}"
        
        # Create exports directory
        if not os.path.exists(EXPORTS_FOLDER):
            os.makedirs(EXPORTS_FOLDER)
        
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        
        # Get attendance data - simplified for now
        attendance_records = []
        
        # Create sample data for testing
        attendance_records = [
            {
                'employee_id': '1001',
                'name': 'John Smith',
                'division': 'DFW',
                'job_site': 'Site A',
                'status': 'On Time',
                'date': report_date,
                'expected_start': datetime.strptime('08:00', '%H:%M').time(),
                'actual_start': datetime.strptime('07:55', '%H:%M').time(),
                'expected_end': datetime.strptime('17:00', '%H:%M').time(),
                'actual_end': datetime.strptime('17:05', '%H:%M').time(),
                'notes': 'Regular day',
                'vehicle': 'Truck 123'
            },
            {
                'employee_id': '1002',
                'name': 'Jane Doe',
                'division': 'HOU',
                'job_site': 'Site B',
                'status': 'Late Start',
                'date': report_date,
                'expected_start': datetime.strptime('08:00', '%H:%M').time(),
                'actual_start': datetime.strptime('08:15', '%H:%M').time(),
                'expected_end': datetime.strptime('17:00', '%H:%M').time(),
                'actual_end': datetime.strptime('17:00', '%H:%M').time(),
                'notes': 'Traffic delay',
                'vehicle': 'Truck 456'
            }
        ]
        
        # Create the export file
        if export_format == 'csv':
            import csv
            
            with open(file_path, 'w', newline='') as csvfile:
                fieldnames = ['employee_id', 'name', 'division', 'job_site', 'status', 
                             'date', 'expected_start', 'actual_start', 'expected_end', 
                             'actual_end', 'notes', 'vehicle']
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for record in attendance_records:
                    row = {
                        'employee_id': record['employee_id'],
                        'name': record['name'],
                        'division': record['division'],
                        'job_site': record['job_site'],
                        'status': record['status'],
                        'date': record['date'].strftime('%Y-%m-%d'),
                        'expected_start': record['expected_start'].strftime('%H:%M') if record['expected_start'] else '',
                        'actual_start': record['actual_start'].strftime('%H:%M') if record['actual_start'] else '',
                        'expected_end': record['expected_end'].strftime('%H:%M') if record['expected_end'] else '',
                        'actual_end': record['actual_end'].strftime('%H:%M') if record['actual_end'] else '',
                        'notes': record['notes'],
                        'vehicle': record['vehicle']
                    }
                    writer.writerow(row)
        
        elif export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Driver Attendance"
            
            # Add headers
            headers = ['Employee ID', 'Name', 'Division', 'Job Site', 'Status', 
                      'Date', 'Expected Start', 'Actual Start', 'Expected End', 
                      'Actual End', 'Notes', 'Vehicle']
            
            # Add header row with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add data rows
            for row_idx, record in enumerate(attendance_records, 2):
                ws.cell(row=row_idx, column=1, value=record['employee_id'])
                ws.cell(row=row_idx, column=2, value=record['name'])
                ws.cell(row=row_idx, column=3, value=record['division'])
                ws.cell(row=row_idx, column=4, value=record['job_site'])
                ws.cell(row=row_idx, column=5, value=record['status'])
                ws.cell(row=row_idx, column=6, value=record['date'].strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=7, value=record['expected_start'].strftime('%H:%M') if record['expected_start'] else '')
                ws.cell(row=row_idx, column=8, value=record['actual_start'].strftime('%H:%M') if record['actual_start'] else '')
                ws.cell(row=row_idx, column=9, value=record['expected_end'].strftime('%H:%M') if record['expected_end'] else '')
                ws.cell(row=row_idx, column=10, value=record['actual_end'].strftime('%H:%M') if record['actual_end'] else '')
                ws.cell(row=row_idx, column=11, value=record['notes'])
                ws.cell(row=row_idx, column=12, value=record['vehicle'])
                
                # Color-code status
                status_cell = ws.cell(row=row_idx, column=5)
                if record['status'] == 'Late Start':
                    status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif record['status'] == 'Early End':
                    status_cell.fill = PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
                elif record['status'] == 'Not on Job':
                    status_cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
                elif record['status'] == 'On Time':
                    status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # Auto-size columns
            for col in range(1, len(headers) + 1):
                column_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
            
            # Save the Excel file
            wb.save(file_path)
        
        else:
            flash(f"Unsupported format: {export_format}", "danger")
            return redirect(url_for('driver_module.daily_report'))
        
        # Log the export
        logger.info(f"Successfully generated {report_type} report in {export_format} format")
        
        # Create download URL
        download_url = url_for('driver_module.download_export', filename=filename)
        
        # Return appropriate response
        flash(f"Report generated successfully. <a href='{download_url}'>Download {filename}</a>", "success")
        return redirect(url_for('driver_module.daily_report', date=date_param))
    
    except Exception as e:
        logger.error(f"Error in export_report: {str(e)}", exc_info=True)
        flash(f"Error generating report: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))


@driver_module_bp.route('/download_export/<filename>', methods=['GET'])
@login_required
def download_export(filename):
    """Download a previously exported file"""
    try:
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Requested file not found: {filename}", "danger")
            return redirect(url_for('driver_module.index'))
    except Exception as e:
        logger.error(f"Error downloading export: {str(e)}")
        flash(f"Error downloading export: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))
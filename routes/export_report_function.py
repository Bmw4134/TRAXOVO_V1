@driver_module_bp.route('/export_report', methods=['GET'])
@login_required
def export_report():
    """Export a driver report using real database data"""
    # Get request parameters
    report_type = request.args.get('type', 'daily')
    export_format = request.args.get('format', 'xlsx')
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    region_id = request.args.get('region')
    direct_download = request.args.get('direct', 'false') == 'true'
    
    try:
        # Parse the date for filtering records
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            report_date = datetime.now().date()
            
        # Generate a filename with timestamp for uniqueness
        timestamp = datetime.now()
        formatted_date = report_date.strftime('%m_%d_%Y')
        filename_base = f"{report_type}_report_{formatted_date}"
        if region_id:
            filename_base += f"_region_{region_id}"
        
        # Add timestamp to ensure uniqueness
        date_str = timestamp.strftime('%H%M%S')
        filename = f"{filename_base}_{date_str}.{export_format}"
        
        # Create exports directory if it doesn't exist
        if not os.path.exists(EXPORTS_FOLDER):
            os.makedirs(EXPORTS_FOLDER)
        
        # Create the output file path
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        
        # Import database models
        from models.driver_attendance import AttendanceRecord, DriverAttendance, JobSiteAttendance
        
        # Prepare data for the report
        attendance_records = []
        
        # Query the database for attendance records
        if report_type == 'daily':
            try:
                # Query for attendance records on the specified date
                query = db.session.query(
                    AttendanceRecord, DriverAttendance, JobSiteAttendance
                ).join(
                    DriverAttendance, AttendanceRecord.driver_id == DriverAttendance.id
                ).join(
                    JobSiteAttendance, AttendanceRecord.assigned_job_id == JobSiteAttendance.id, isouter=True
                ).filter(
                    AttendanceRecord.date == report_date
                )
                
                # Apply region filter if specified
                if region_id:
                    query = query.filter(DriverAttendance.division == region_id)
                
                # Execute the query
                records = query.all()
                
                # Process records for CSV or Excel export
                for record, driver, job_site in records:
                    status = "On Time"
                    if record.late_start:
                        status = "Late Start"
                    elif record.early_end:
                        status = "Early End"
                    elif record.not_on_job:
                        status = "Not on Job"
                    
                    attendance_records.append({
                        'employee_id': driver.employee_id,
                        'name': driver.full_name,
                        'division': driver.division or 'Unknown',
                        'job_site': job_site.name if job_site else 'Unknown',
                        'status': status,
                        'date': record.date,
                        'expected_start': record.expected_start_time,
                        'actual_start': record.actual_start_time,
                        'expected_end': record.expected_end_time,
                        'actual_end': record.actual_end_time,
                        'notes': record.notes,
                        'vehicle': record.asset_id
                    })
                
                # Create the export file
                if export_format == 'csv':
                    import csv
                    
                    with open(file_path, 'w', newline='') as csvfile:
                        fieldnames = ['employee_id', 'name', 'division', 'job_site', 'status', 
                                     'date', 'expected_start', 'actual_start', 'expected_end', 
                                     'actual_end', 'notes', 'vehicle']
                        
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        
                        for record in attendance_records:
                            writer.writerow(record)
                
                elif export_format == 'xlsx':
                    # Import required libraries for Excel export
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                    
                    # Create a new workbook and select the active sheet
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Driver Attendance"
                    
                    # Add headers with styling
                    headers = ['Employee ID', 'Name', 'Division', 'Job Site', 'Status', 
                              'Date', 'Expected Start', 'Actual Start', 'Expected End', 
                              'Actual End', 'Notes', 'Vehicle']
                    
                    # Add header row
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # Add data rows
                    for row_idx, record in enumerate(attendance_records, 2):
                        ws.cell(row=row_idx, column=1, value=record['employee_id'])
                        ws.cell(row=row_idx, column=2, value=record['name'])
                        ws.cell(row=row_idx, column=3, value=record['division'])
                        ws.cell(row=row_idx, column=4, value=record['job_site'])
                        ws.cell(row=row_idx, column=5, value=record['status'])
                        ws.cell(row=row_idx, column=6, value=record['date'].strftime('%Y-%m-%d') if record['date'] else '')
                        ws.cell(row=row_idx, column=7, value=record['expected_start'].strftime('%H:%M') if record['expected_start'] else '')
                        ws.cell(row=row_idx, column=8, value=record['actual_start'].strftime('%H:%M') if record['actual_start'] else '')
                        ws.cell(row=row_idx, column=9, value=record['expected_end'].strftime('%H:%M') if record['expected_end'] else '')
                        ws.cell(row=row_idx, column=10, value=record['actual_end'].strftime('%H:%M') if record['actual_end'] else '')
                        ws.cell(row=row_idx, column=11, value=record['notes'])
                        ws.cell(row=row_idx, column=12, value=record['vehicle'])
                        
                        # Color-code status
                        status_cell = ws.cell(row=row_idx, column=5)
                        if record['status'] == 'Late Start':
                            status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif record['status'] == 'Early End':
                            status_cell.fill = PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
                        elif record['status'] == 'Not on Job':
                            status_cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
                        elif record['status'] == 'On Time':
                            status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    
                    # Auto-size columns for better readability
                    for col in range(1, len(headers) + 1):
                        column_letter = openpyxl.utils.get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 15
                    
                    # Save the Excel file
                    wb.save(file_path)
                else:
                    flash(f"Unsupported format: {export_format}", "danger")
                    return redirect(url_for('driver_module.daily_report'))
                
            except Exception as e:
                logger.error(f"Error creating export: {str(e)}", exc_info=True)
                flash(f"Error creating export: {str(e)}", "danger")
                return redirect(url_for('driver_module.daily_report'))
            
        elif report_type == 'summary':
            # Logic for summary report
            try:
                # Define date range (e.g., last 7 days)
                end_date = report_date
                start_date = end_date - timedelta(days=6)  # 7 days total
                
                # Query for summary stats
                summary_query = db.session.query(
                    DriverAttendance,
                    db.func.count(AttendanceRecord.id).label('total_days'),
                    db.func.sum(db.case((AttendanceRecord.late_start == True, 1), else_=0)).label('late_count'),
                    db.func.sum(db.case((AttendanceRecord.early_end == True, 1), else_=0)).label('early_end_count'),
                    db.func.sum(db.case((AttendanceRecord.not_on_job == True, 1), else_=0)).label('not_on_job_count')
                ).join(
                    AttendanceRecord, DriverAttendance.id == AttendanceRecord.driver_id
                ).filter(
                    AttendanceRecord.date.between(start_date, end_date)
                )
                
                # Apply region filter if specified
                if region_id:
                    summary_query = summary_query.filter(DriverAttendance.division == region_id)
                
                # Group by driver
                summary_results = summary_query.group_by(DriverAttendance.id).all()
                
                # Process summary results
                for driver, total_days, late_count, early_end_count, not_on_job_count in summary_results:
                    # Calculate on-time days
                    late_count = late_count or 0
                    early_end_count = early_end_count or 0
                    not_on_job_count = not_on_job_count or 0
                    on_time_count = total_days - (late_count + early_end_count + not_on_job_count)
                    
                    # Calculate attendance score
                    attendance_score = round((on_time_count / total_days) * 100) if total_days > 0 else 0
                    
                    attendance_records.append({
                        'employee_id': driver.employee_id,
                        'name': driver.full_name,
                        'division': driver.division or 'Unknown',
                        'total_days': total_days,
                        'on_time_count': on_time_count,
                        'late_count': late_count,
                        'early_end_count': early_end_count,
                        'not_on_job_count': not_on_job_count,
                        'attendance_score': attendance_score
                    })
                
                # Create the export file
                if export_format == 'csv':
                    import csv
                    
                    with open(file_path, 'w', newline='') as csvfile:
                        fieldnames = ['employee_id', 'name', 'division', 'total_days', 
                                     'on_time_count', 'late_count', 'early_end_count', 
                                     'not_on_job_count', 'attendance_score']
                        
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        
                        for record in attendance_records:
                            writer.writerow(record)
                
                elif export_format == 'xlsx':
                    # Import required libraries for Excel export
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                    
                    # Create a new workbook and select the active sheet
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Attendance Summary"
                    
                    # Add report title
                    ws.cell(row=1, column=1, value=f"Attendance Summary Report ({start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')})")
                    ws.merge_cells('A1:I1')
                    title_cell = ws.cell(row=1, column=1)
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal='center')
                    
                    # Add headers with styling
                    headers = ['Employee ID', 'Name', 'Division', 'Total Days', 
                              'On Time', 'Late Start', 'Early End', 'Not on Job', 'Attendance Score']
                    
                    # Add header row
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=3, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # Add data rows
                    for row_idx, record in enumerate(attendance_records, 4):
                        ws.cell(row=row_idx, column=1, value=record['employee_id'])
                        ws.cell(row=row_idx, column=2, value=record['name'])
                        ws.cell(row=row_idx, column=3, value=record['division'])
                        ws.cell(row=row_idx, column=4, value=record['total_days'])
                        ws.cell(row=row_idx, column=5, value=record['on_time_count'])
                        ws.cell(row=row_idx, column=6, value=record['late_count'])
                        ws.cell(row=row_idx, column=7, value=record['early_end_count'])
                        ws.cell(row=row_idx, column=8, value=record['not_on_job_count'])
                        
                        # Format attendance score
                        score_cell = ws.cell(row=row_idx, column=9, value=f"{record['attendance_score']}%")
                        
                        # Color-code score
                        score = record['attendance_score']
                        if score >= 95:
                            score_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        elif score >= 90:
                            score_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                        elif score >= 80:
                            score_cell.fill = PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
                        else:
                            score_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    
                    # Auto-size columns for better readability
                    for col in range(1, len(headers) + 1):
                        column_letter = openpyxl.utils.get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 15
                    
                    # Save the Excel file
                    wb.save(file_path)
                
                else:
                    flash(f"Unsupported format: {export_format}", "danger")
                    return redirect(url_for('driver_module.attendance_dashboard'))
                
            except Exception as e:
                logger.error(f"Error creating summary export: {str(e)}", exc_info=True)
                flash(f"Error creating summary export: {str(e)}", "danger")
                return redirect(url_for('driver_module.attendance_dashboard'))
        
        else:
            flash(f"Unsupported report type: {report_type}", "danger")
            return redirect(url_for('driver_module.index'))
        
        # Log the export
        logger.info(f"Successfully generated {report_type} report in {export_format} format")
        
        # Create download URL
        download_url = url_for('driver_module.download_export', filename=filename)
        
        # Return appropriate response
        if direct_download:
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Report generated successfully. <a href='{download_url}'>Download {filename}</a>", "success")
            
            if report_type == 'daily':
                return redirect(url_for('driver_module.daily_report', date=date_param))
            elif report_type == 'summary':
                return redirect(url_for('driver_module.attendance_dashboard'))
            else:
                return redirect(url_for('driver_module.index'))
    
    except Exception as e:
        logger.error(f"Error in export_report: {str(e)}", exc_info=True)
        flash(f"Error generating report: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))
@export_report_function_bp.route('/export_report')
def export_report():
    """Handler for /export_report"""
    try:
        # Add your route handler logic here
        return render_template('export_report_function/export_report.html')
    except Exception as e:
        logger.error(f"Error in export_report: {e}")
        return render_template('error.html', error=str(e)), 500

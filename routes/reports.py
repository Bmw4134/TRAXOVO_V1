"""
Reports routes for the SYSTEMSMITH application.

This module handles the routes related to generating and accessing reports.
"""
import os
import csv
import logging
import json
import re
from datetime import datetime, timedelta
import pandas as pd
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required
from utils.billing_processor import process_pm_allocation, generate_region_exports

# Create the reports blueprint
reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

@reports_bp.route('/')
@login_required
def index():
    """Reports landing page with links to all available reports."""
    return render_template('reports.html', title='Reports', datetime=datetime)

@reports_bp.route('/driver-reports')
@login_required
def driver_reports():
    """List all available driver reports."""
    # Path to reports directory
    reports_dir = os.path.join(os.getcwd(), 'reports')
    
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir, exist_ok=True)
    
    # Get report date directories
    report_data = []
    
    for date_dir in sorted(os.listdir(reports_dir), reverse=True):
        date_path = os.path.join(reports_dir, date_dir)
        
        if os.path.isdir(date_path):
            reports_list = []
            
            for file_name in os.listdir(date_path):
                file_path = os.path.join(date_path, file_name)
                
                if os.path.isfile(file_path):
                    report_type = 'summary'
                    if 'late_start' in file_name:
                        report_type = 'Late Start'
                    elif 'early_end' in file_name:
                        report_type = 'Early End'
                    elif 'not_on_job' in file_name:
                        report_type = 'Not On Job'
                    elif 'summary' in file_name:
                        report_type = 'Summary'
                    
                    reports_list.append({
                        'name': report_type,
                        'path': os.path.join(date_dir, file_name),
                        'size': os.path.getsize(file_path),
                        'modified': datetime.fromtimestamp(os.path.getmtime(file_path))
                    })
            
            if reports_list:
                report_data.append({
                    'date': date_dir,
                    'reports': sorted(reports_list, key=lambda x: x['name'])
                })
    
    return render_template('driver_reports.html', title='Driver Reports', report_data=report_data)

@reports_bp.route('/download/<path:report_path>')
@login_required
def download_report(report_path):
    """Download a specific report file."""
    try:
        reports_dir = os.path.join(os.getcwd(), 'reports')
        full_path = os.path.join(reports_dir, report_path)
        
        if not os.path.exists(full_path):
            flash('Report file not found', 'danger')
            return redirect(url_for('reports.driver_reports'))
        
        return send_file(full_path, as_attachment=True)
    except Exception as e:
        flash(f'Error downloading report: {str(e)}', 'danger')
        return redirect(url_for('reports.driver_reports'))
        
@reports_bp.route('/download-export/<path:export_path>')
@login_required
def download_export(export_path):
    """Download a specific export file."""
    try:
        exports_dir = os.path.join(os.getcwd(), 'exports')
        full_path = os.path.join(exports_dir, export_path)
        
        if not os.path.exists(full_path):
            flash('Export file not found', 'danger')
            return redirect(url_for('reports.index'))
        
        return send_file(full_path, as_attachment=True)
    except Exception as e:
        flash(f'Error downloading export: {str(e)}', 'danger')
        return redirect(url_for('reports.index'))

@reports_bp.route('/view/<path:report_path>')
@login_required
def view_report(report_path):
    """View a specific report file in-browser."""
    try:
        reports_dir = os.path.join(os.getcwd(), 'reports')
        full_path = os.path.join(reports_dir, report_path)
        
        if not os.path.exists(full_path):
            flash('Report file not found', 'danger')
            return redirect(url_for('reports.driver_reports'))
        
        # If it's an HTML file, render its contents
        if full_path.endswith('.html'):
            with open(full_path, 'r') as f:
                html_content = f.read()
            return render_template('view_report.html', html_content=html_content, title='View Report')
        
        # For CSV files, parse and display in a table
        elif full_path.endswith('.csv'):
            data = []
            headers = []
            
            with open(full_path, 'r') as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    data.append(row)
            
            return render_template('view_csv_report.html', headers=headers, data=data, 
                                   title=f'Report: {os.path.basename(report_path)}')
        
        # For other file types, download instead
        else:
            return send_file(full_path, as_attachment=True)
            
    except Exception as e:
        flash(f'Error viewing report: {str(e)}', 'danger')
        return redirect(url_for('reports.driver_reports'))

@reports_bp.route('/pm-allocation')
@login_required
def pm_allocation():
    """PM Allocation reconciliation upload page."""
    # Get current month in YYYY-MM format for default value
    current_month = datetime.now().strftime('%Y-%m')
    
    # List recent comparison reports
    exports_dir = os.path.join(os.getcwd(), 'exports')
    if not os.path.exists(exports_dir):
        os.makedirs(exports_dir)
        
    recent_reports = []
    for filename in os.listdir(exports_dir):
        if filename.endswith('.xlsx') and 'COMPARISON' in filename:
            file_path = os.path.join(exports_dir, filename)
            created_time = datetime.fromtimestamp(os.path.getctime(file_path))
            
            # Extract month from filename
            month_match = re.search(r'(\w+\s+\d{4})', filename)
            month = month_match.group(1) if month_match else 'Unknown'
            
            # Estimate number of changes from filename
            changes_match = re.search(r'(\d+)_CHANGES', filename)
            changes = changes_match.group(1) if changes_match else '0'
            
            recent_reports.append({
                'date_generated': created_time.strftime('%Y-%m-%d %H:%M'),
                'month': month,
                'filename': filename,
                'changes': changes,
                'path': file_path
            })
    
    # Sort by most recent first
    recent_reports.sort(key=lambda x: x['date_generated'], reverse=True)
    
    return render_template('pm_allocation_upload.html', 
                          title='PM Allocation Reconciliation',
                          current_month=current_month,
                          recent_reports=recent_reports[:10])  # Show only 10 most recent
                          
@reports_bp.route('/upload-pm-allocation', methods=['POST'])
@login_required
def upload_pm_allocation():
    """Process uploaded PM allocation files."""
    try:
        if 'original_file' not in request.files or 'updated_file' not in request.files:
            flash('Both original and updated files must be uploaded', 'danger')
            return redirect(url_for('reports.pm_allocation'))
            
        original_file = request.files['original_file']
        updated_file = request.files['updated_file']
        
        if original_file.filename == '' or updated_file.filename == '':
            flash('Both files must be selected', 'danger')
            return redirect(url_for('reports.pm_allocation'))
            
        # Validate file types
        if not (original_file.filename.endswith(('.xlsx', '.xlsm', '.xls')) and 
                updated_file.filename.endswith(('.xlsx', '.xlsm', '.xls'))):
            flash('Only Excel files are supported (.xlsx, .xlsm, .xls)', 'danger')
            return redirect(url_for('reports.pm_allocation'))
        
        # Get form parameters
        month_input = request.form.get('month', '')
        region = request.form.get('region', 'all')
        fsi_format = request.form.get('fsi_format') == 'on'
        
        # Convert YYYY-MM format to readable month format
        try:
            month_date = datetime.strptime(month_input, '%Y-%m')
            month = month_date.strftime('%B %Y').upper()  # e.g., "APRIL 2025"
        except ValueError:
            month = datetime.now().strftime('%B %Y').upper()
        
        # Log parameters
        logging.info(f"PM Allocation Upload - Month: {month}, Region: {region}, FSI Format: {fsi_format}")
        logging.info(f"Original file: {original_file.filename}, Updated file: {updated_file.filename}")
        
        # Create uploads directory if it doesn't exist
        uploads_dir = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
            
        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(os.getcwd(), 'exports')
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
            
        # Save uploaded files
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        original_path = os.path.join(uploads_dir, f'ORIGINAL_{timestamp}_{original_file.filename}')
        updated_path = os.path.join(uploads_dir, f'UPDATED_{timestamp}_{updated_file.filename}')
        
        original_file.save(original_path)
        updated_file.save(updated_path)
        logging.info(f"Files saved to: {original_path} and {updated_path}")
        
        # Process the files
        try:
            result = process_pm_allocation(
                original_file=original_path,
                pm_file=updated_path,
                region=region,
                month=month,
                fsi_format=fsi_format
            )
            logging.info(f"PM Allocation processing result: {result.get('success')}")
        except Exception as e:
            logging.error(f"Error in PM Allocation processing: {str(e)}", exc_info=True)
            flash(f"An error occurred during processing: {str(e)}", 'danger')
            return redirect(url_for('reports.pm_allocation'))
        
        if not result['success']:
            flash(f"Error processing files: {result.get('message', 'Unknown error')}", 'danger')
            return redirect(url_for('reports.pm_allocation'))
            
        # Store processing result in session for display
        session['pm_allocation_result'] = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'original_file': os.path.basename(original_path),
            'pm_file': os.path.basename(updated_path),
            'month': month,
            'export_files': result.get('export_files', {}),
            'comparison_file': result.get('comparison_file', ''),
            'comparison': result.get('comparison', {}),
            'changes': result.get('changes', {})
        }
        
        # Log stored data
        logging.info(f"Stored PM allocation result in session: {session.get('pm_allocation_result', {}).keys()}")
        
        # Redirect to results page
        return redirect(url_for('reports.pm_allocation_results'))
        
    except Exception as e:
        logging.error(f"Error processing PM allocation files: {str(e)}")
        flash(f'Error processing files: {str(e)}', 'danger')
        return redirect(url_for('reports.pm_allocation'))

@reports_bp.route('/pm-allocation-results')
@login_required
def pm_allocation_results():
    """Display results of PM allocation processing."""
    # Import required modules
    import logging
    import pandas as pd
    import os
    from flask import session, flash, redirect, url_for, render_template
    
    # Get result from session
    result = session.get('pm_allocation_result')
    if not result:
        flash('No PM allocation results found. Please upload files first.', 'warning')
        return redirect(url_for('reports.pm_allocation'))
    
    # Initialize comparison data structure
    comparison = {
        'changed_rows': [],
        'added_rows': [],
        'removed_rows': []
    }
    
    # Check if comparison data is already in the result
    if 'comparison' in result:
        comparison = result.get('comparison', comparison)
        logging.info(f"Using comparison data from session")
    else:
        # If not, load from file
        comparison_file = result.get('comparison_file')
        
        if comparison_file and os.path.exists(comparison_file):
            # Load comparison data from file
            try:
                logging.info(f"Loading comparison data from file: {comparison_file}")
                comparison_data = pd.read_excel(comparison_file, sheet_name=None)
                
                # Process changed rows if available
                if 'Changes' in comparison_data:
                    changes_df = comparison_data['Changes']
                    
                    if len(changes_df) > 0:
                        logging.info(f"Processing {len(changes_df)} changed rows")
                        
                        # Process each row
                        current_row = None
                        for _, row in changes_df.iterrows():
                            row_dict = row.to_dict()
                            
                            # Get row ID safely
                            div = str(row_dict.get('DIV', ''))
                            job = str(row_dict.get('JOB', ''))
                            asset_id = str(row_dict.get('ASSET ID', ''))
                            row_id = f"{div}-{job}-{asset_id}"
                            
                            # Create or update row entry
                            if current_row is None or current_row['row_id'] != row_id:
                                if current_row is not None:
                                    comparison['changed_rows'].append(current_row)
                                
                                current_row = {
                                    'row_id': row_id,
                                    'div': div,
                                    'job': job,
                                    'asset_id': asset_id,
                                    'equipment': str(row_dict.get('EQUIPMENT DESCRIPTION', '')),
                                    'changes': [],
                                    'original': {},
                                    'updated': {}
                                }
                            
                            # Add the change
                            current_row['changes'].append({
                                'field': str(row_dict.get('FIELD', '')),
                                'original': str(row_dict.get('ORIGINAL', '')),
                                'updated': str(row_dict.get('UPDATED', ''))
                            })
                        
                        # Add the last row
                        if current_row is not None:
                            comparison['changed_rows'].append(current_row)
                
                # Process added rows if available
                if 'Added Rows' in comparison_data:
                    added_df = comparison_data['Added Rows']
                    if len(added_df) > 0:
                        for _, row in added_df.iterrows():
                            comparison['added_rows'].append(row.to_dict())
                
                # Process removed rows if available  
                if 'Removed Rows' in comparison_data:
                    removed_df = comparison_data['Removed Rows']
                    if len(removed_df) > 0:
                        for _, row in removed_df.iterrows():
                            comparison['removed_rows'].append(row.to_dict())
                            
            except Exception as e:
                logging.error(f"Error loading comparison data: {str(e)}", exc_info=True)
                flash(f"Error processing comparison data: {str(e)}", 'warning')
        else:
            logging.warning(f"Comparison file not found: {comparison_file}")
            flash('Comparison file not found or could not be loaded. Using basic report data.', 'warning')
    
    # Process the export files
    export_files = result.get('export_files', {})
    
    # Render the results page
    return render_template('pm_allocation_results.html',
                          title='PM Allocation Results',
                          result=result,
                          comparison=comparison,
                          export_files=export_files)

@reports_bp.route('/generate-region-exports')
@login_required
def generate_region_exports():
    """Generate region-specific allocation exports."""
    month = request.args.get('month', datetime.now().strftime('%B %Y').upper())
    
    # Find the most recent PM-edited file for this month
    uploads_dir = os.path.join(os.getcwd(), 'uploads')
    if not os.path.exists(uploads_dir):
        flash('No uploaded files found', 'danger')
        return redirect(url_for('reports.pm_allocation'))
    
    # Get all PM-edited files for this month
    pm_files = []
    month_pattern = month.replace(' ', '_').upper()
    for filename in os.listdir(uploads_dir):
        if filename.startswith('UPDATED_') and month_pattern in filename:
            file_path = os.path.join(uploads_dir, filename)
            created_time = os.path.getctime(file_path)
            pm_files.append((file_path, created_time))
    
    if not pm_files:
        flash(f'No PM-edited files found for {month}', 'danger')
        return redirect(url_for('reports.pm_allocation'))
    
    # Use the most recent file
    pm_files.sort(key=lambda x: x[1], reverse=True)
    latest_pm_file = pm_files[0][0]
    
    # Generate exports for all regions
    try:
        result = generate_region_exports(
            comparison=None,  # We don't have comparison data here, function will recalculate
            pm_df=pd.read_excel(latest_pm_file),
            month=month,
            export_dir='exports',
            fsi_format=True  # Always include FSI format as an option
        )
        
        if not result.get('success', False):
            flash(f"Error generating exports: {result.get('message', 'Unknown error')}", 'danger')
            return redirect(url_for('reports.pm_allocation'))
        
        flash(f"Generated exports for all regions for {month}", 'success')
        
        # Redirect to PM allocation page
        return redirect(url_for('reports.pm_allocation'))
        
    except Exception as e:
        logging.error(f"Error generating region exports: {str(e)}")
        flash(f'Error generating exports: {str(e)}', 'danger')
        return redirect(url_for('reports.pm_allocation'))

@reports_bp.route('/generate-daily-driver-report', methods=['GET', 'POST'])
@login_required
def generate_daily_driver_report():
    """Generate daily driver reports (Late Start, Early End, Not On Job)."""
    if request.method == 'GET':
        return render_template('generate_driver_report.html', title='Generate Driver Reports')
    
    try:
        # Create reports directories
        reports_dir = os.path.join(os.getcwd(), 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        today_dir = os.path.join(reports_dir, datetime.now().strftime('%Y-%m-%d'))
        os.makedirs(today_dir, exist_ok=True)
        
        # Get file paths for activity detail and driving history
        activity_detail_path = os.path.join(os.getcwd(), 'attached_assets', 'ActivityDetail (6).csv')
        driving_history_path = os.path.join(os.getcwd(), 'attached_assets', 'DrivingHistory.csv')
        
        # Log processing start
        logging.info(f"Processing driver reports using activity data: {activity_detail_path}")
        logging.info(f"Processing driver reports using driving history: {driving_history_path}")
        
        # Check if files exist
        if not os.path.exists(activity_detail_path):
            flash(f'Activity Detail file not found at {activity_detail_path}', 'danger')
            return redirect(url_for('reports.generate_daily_driver_report'))
            
        if not os.path.exists(driving_history_path):
            flash(f'Driving History file not found at {driving_history_path}', 'danger')
            return redirect(url_for('reports.generate_daily_driver_report'))
        
        # Generate Yesterday's reports
        # Get yesterday's date for prior day reports
        yesterday = datetime.now().date() - timedelta(days=1)
        yesterday_str = yesterday.strftime('%Y-%m-%d')
        
        # Generate prior day report files
        prior_late_start_path = os.path.join(today_dir, f'prior_day_late_start_{yesterday_str}.csv')
        prior_early_end_path = os.path.join(today_dir, f'prior_day_early_end_{yesterday_str}.csv')
        prior_not_on_job_path = os.path.join(today_dir, f'prior_day_not_on_job_{yesterday_str}.csv')
        
        # Process activity detail for Late Start and Early End
        # Handle the special format of the ActivityDetail file with metadata headers
        # Skip the first 7 rows which contain metadata, and use the 8th row as headers
        df_activity = pd.read_csv(activity_detail_path, skiprows=7)
        
        # Normalize column names by removing trailing 'x' characters
        df_activity.columns = [col.replace('x', '') if col.endswith('x') else col for col in df_activity.columns]
        
        # Extract datetime from the EventDateTime column - handle timezone strings safely
        df_activity['EventDateTime'] = df_activity['EventDateTime'].str.replace(' CT', '').str.replace(' ET', '')
        df_activity['Date Time'] = pd.to_datetime(df_activity['EventDateTime'], errors='coerce')
        df_activity['Date'] = df_activity['Date Time'].dt.date
        
        # Extract employee ID and name from AssetLabel if it follows the pattern: #EMPID - NAME VEHICLE INFO
        # Example: #210003 - AMMAR I. ELHAMAD FORD F150 2024 Personal Vehicle +
        df_activity['Employee ID'] = df_activity['AssetLabel'].str.extract(r'#(\d+)').iloc[:, 0]
        df_activity['Driver'] = df_activity['AssetLabel'].str.extract(r'#\d+ - ([\w\s\.]+)').iloc[:, 0]
        df_activity['Asset'] = df_activity['AssetLabel'].str.extract(r'- [\w\s\.]+ ([\w\s\d\-\+]+)$').iloc[:, 0]
        
        # Process DrivingHistory file for more employee/job data
        try:
            # Skip the first 7 rows which contain metadata, and use the 8th row as headers
            df_driving = pd.read_csv(driving_history_path, skiprows=7)
            
            # Extract employee data
            df_driving['Employee ID'] = df_driving['Textbox53'].str.extract(r'#(\d+)').iloc[:, 0]
            df_driving['Driver'] = df_driving['Textbox53'].str.extract(r'#\d+ - ([\w\s\.]+)').iloc[:, 0]
            df_driving['Asset'] = df_driving['Textbox53'].str.extract(r'- [\w\s\.]+ ([\w\s\d\-\+]+)$').iloc[:, 0]
            df_driving['Job'] = df_driving['Location'].str.extract(r'(JOB\d+)').iloc[:, 0]
            df_driving['Job Description'] = df_driving['Location'].str.extract(r'JOB\d+ - ([\w\s\-\+\.]+)').iloc[:, 0]
            
            # Extract phone number from ContactPhone
            df_driving['Phone'] = df_driving['ContactPhone']
            
            # Create a lookup dictionary for employee information
            employee_info = {}
            for _, row in df_driving.dropna(subset=['Employee ID']).iterrows():
                emp_id = str(row['Employee ID'])
                if emp_id not in employee_info:
                    employee_info[emp_id] = {
                        'Driver': row['Driver'],
                        'Phone': row['Phone'],
                        'Job': row.get('Job', ''),
                        'Job Description': row.get('Job Description', ''),
                        'Asset': row['Asset']
                    }
            
            # Enrich activity data with employee info
            for idx, row in df_activity.iterrows():
                emp_id = str(row['Employee ID'])
                if emp_id in employee_info:
                    if pd.isna(df_activity.at[idx, 'Driver']) or df_activity.at[idx, 'Driver'] == '':
                        df_activity.at[idx, 'Driver'] = employee_info[emp_id]['Driver']
                    df_activity.at[idx, 'Phone'] = employee_info[emp_id]['Phone']
                    df_activity.at[idx, 'Job'] = employee_info[emp_id]['Job']
                    df_activity.at[idx, 'Job Description'] = employee_info[emp_id]['Job Description']
                    if pd.isna(df_activity.at[idx, 'Asset']) or df_activity.at[idx, 'Asset'] == '':
                        df_activity.at[idx, 'Asset'] = employee_info[emp_id]['Asset']
        except Exception as e:
            logging.warning(f"Could not fully process driving history: {str(e)}")
            # Continue with limited data
        
        # Filter to yesterday's date
        df_yesterday = df_activity[df_activity['Date'] == yesterday]
        
        # Import openpyxl for Excel export
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Create a single consolidated report in Excel format
        consolidated_report_path = os.path.join(today_dir, f'driver_daily_report_{yesterday_str}.xlsx')
        
        # Create a new workbook
        workbook = openpyxl.Workbook()
        
        # Set up styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')  # Navy blue
        data_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
        data_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Process for Late Start (first entry after 8:45 AM)
        df_late_start = pd.DataFrame()
        if not df_yesterday.empty:
            # Group by Driver to get first entry times
            df_first_entries = df_yesterday.groupby('Driver').agg({'Date Time': 'min', 'Location': 'first'}).reset_index()
            df_first_entries['Time'] = df_first_entries['Date Time'].dt.time
            
            # Filter for entries after 8:45 AM
            late_start_time = pd.to_datetime(yesterday_str + ' 08:45:00').time()
            df_late_start = df_first_entries[df_first_entries['Time'] > late_start_time].copy()
            
            # Add status flags
            df_late_start['LS'] = 'Yes'
            df_late_start['EE'] = 'No'
            df_late_start['NOJ'] = 'No'
            
            # Save Late Start Report
            if not df_late_start.empty:
                df_late_start[['Driver', 'Date Time', 'Location']].to_csv(prior_late_start_path, index=False)
        
        # Process for Early End (last entry before 4:45 PM)
        df_early_end = pd.DataFrame()
        if not df_yesterday.empty:
            # Group by Driver to get last entry times
            df_last_entries = df_yesterday.groupby('Driver').agg({'Date Time': 'max', 'Location': 'last'}).reset_index()
            df_last_entries['Time'] = df_last_entries['Date Time'].dt.time
            
            # Filter for entries before 4:45 PM
            early_end_time = pd.to_datetime(yesterday_str + ' 16:45:00').time()
            df_early_end = df_last_entries[df_last_entries['Time'] < early_end_time].copy()
            
            # Add status flags
            df_early_end['LS'] = 'No'
            df_early_end['EE'] = 'Yes'
            df_early_end['NOJ'] = 'No'
            
            # Save Early End Report
            if not df_early_end.empty:
                df_early_end[['Driver', 'Date Time', 'Location']].to_csv(prior_early_end_path, index=False)
        
        # Create a combined dataset for the consolidated report
        df_all_issues = pd.DataFrame()
        
        if not df_late_start.empty:
            df_all_issues = pd.concat([df_all_issues, df_late_start])
            
        if not df_early_end.empty:
            # Only add drivers who aren't already in for late start
            existing_drivers = df_all_issues['Driver'].unique() if not df_all_issues.empty else []
            df_early_end_new = df_early_end[~df_early_end['Driver'].isin(existing_drivers)]
            df_all_issues = pd.concat([df_all_issues, df_early_end_new])
        
        # Create Late Start sheet
        late_start_sheet = workbook.active
        late_start_sheet.title = 'Late Start'
        
        # Add headers to Late Start sheet - include all requested fields
        late_start_headers = [
            'Employee ID', 'Driver Name', 'Phone', 'Asset', 'First Entry Time', 
            'Location', 'Job', 'Job Description', 'Date', 'Status'
        ]
        for col, header in enumerate(late_start_headers, 1):
            cell = late_start_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Add data to Late Start sheet
        if not df_late_start.empty:
            for row_idx, (_, row) in enumerate(df_late_start.iterrows(), 2):
                row_fill = data_fill_even if row_idx % 2 == 0 else data_fill_odd
                
                # First set of columns
                late_start_sheet.cell(row=row_idx, column=1).value = row.get('Employee ID', '')
                late_start_sheet.cell(row=row_idx, column=2).value = row.get('Driver', '')
                late_start_sheet.cell(row=row_idx, column=3).value = row.get('Phone', '')
                late_start_sheet.cell(row=row_idx, column=4).value = row.get('Asset', '')
                
                # Time and location
                if 'Date Time' in row and not pd.isna(row['Date Time']):
                    late_start_sheet.cell(row=row_idx, column=5).value = row['Date Time'].strftime('%H:%M:%S')
                else:
                    late_start_sheet.cell(row=row_idx, column=5).value = 'Unknown'
                    
                late_start_sheet.cell(row=row_idx, column=6).value = row.get('Location', 'Unknown')
                
                # Job information
                late_start_sheet.cell(row=row_idx, column=7).value = row.get('Job', '')
                late_start_sheet.cell(row=row_idx, column=8).value = row.get('Job Description', '')
                
                # Date and status
                if 'Date' in row and not pd.isna(row['Date']):
                    late_start_sheet.cell(row=row_idx, column=9).value = row['Date'].strftime('%Y-%m-%d')
                else:
                    late_start_sheet.cell(row=row_idx, column=9).value = yesterday_str
                
                late_start_sheet.cell(row=row_idx, column=10).value = 'LATE START'
                
                # Apply styling to all cells in the row
                for col in range(1, 11):
                    cell = late_start_sheet.cell(row=row_idx, column=col)
                    cell.fill = row_fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
        
        # Create Early End sheet
        early_end_sheet = workbook.create_sheet(title='Early End')
        
        # Add headers to Early End sheet - include all requested fields
        early_end_headers = [
            'Employee ID', 'Driver Name', 'Phone', 'Asset', 'Last Entry Time', 
            'Location', 'Job', 'Job Description', 'Date', 'Status'
        ]
        for col, header in enumerate(early_end_headers, 1):
            cell = early_end_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Add data to Early End sheet
        if not df_early_end.empty:
            for row_idx, (_, row) in enumerate(df_early_end.iterrows(), 2):
                row_fill = data_fill_even if row_idx % 2 == 0 else data_fill_odd
                
                # First set of columns
                early_end_sheet.cell(row=row_idx, column=1).value = row.get('Employee ID', '')
                early_end_sheet.cell(row=row_idx, column=2).value = row.get('Driver', '')
                early_end_sheet.cell(row=row_idx, column=3).value = row.get('Phone', '')
                early_end_sheet.cell(row=row_idx, column=4).value = row.get('Asset', '')
                
                # Time and location
                if 'Date Time' in row and not pd.isna(row['Date Time']):
                    early_end_sheet.cell(row=row_idx, column=5).value = row['Date Time'].strftime('%H:%M:%S')
                else:
                    early_end_sheet.cell(row=row_idx, column=5).value = 'Unknown'
                    
                early_end_sheet.cell(row=row_idx, column=6).value = row.get('Location', 'Unknown')
                
                # Job information
                early_end_sheet.cell(row=row_idx, column=7).value = row.get('Job', '')
                early_end_sheet.cell(row=row_idx, column=8).value = row.get('Job Description', '')
                
                # Date and status
                if 'Date' in row and not pd.isna(row['Date']):
                    early_end_sheet.cell(row=row_idx, column=9).value = row['Date'].strftime('%Y-%m-%d')
                else:
                    early_end_sheet.cell(row=row_idx, column=9).value = yesterday_str
                
                early_end_sheet.cell(row=row_idx, column=10).value = 'EARLY END'
                
                # Apply styling to all cells in the row
                for col in range(1, 11):
                    cell = early_end_sheet.cell(row=row_idx, column=col)
                    cell.fill = row_fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
        
        # Create Summary sheet
        summary_sheet = workbook.create_sheet(title='Summary')
        
        # Add report title and date
        summary_sheet.cell(row=1, column=1).value = f"Daily Driver Report - {yesterday_str}"
        summary_sheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        summary_sheet.merge_cells('A1:E1')
        
        # Add summary information
        summary_sheet.cell(row=3, column=1).value = "Late Start Count:"
        summary_sheet.cell(row=3, column=2).value = len(df_late_start)
        summary_sheet.cell(row=3, column=1).font = Font(bold=True)
        
        summary_sheet.cell(row=4, column=1).value = "Early End Count:"
        summary_sheet.cell(row=4, column=2).value = len(df_early_end)
        summary_sheet.cell(row=4, column=1).font = Font(bold=True)
        
        summary_sheet.cell(row=5, column=1).value = "Not On Job Count:"
        summary_sheet.cell(row=5, column=2).value = 0  # Placeholder for now
        summary_sheet.cell(row=5, column=1).font = Font(bold=True)
        
        summary_sheet.cell(row=7, column=1).value = "Total Issues:"
        summary_sheet.cell(row=7, column=2).value = len(df_all_issues)
        summary_sheet.cell(row=7, column=1).font = Font(bold=True)
        
        # Auto-adjust column widths
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(consolidated_report_path)
        logging.info(f"Generated consolidated report at {consolidated_report_path}")
        
        # Generate current day report files (for today's data as of 8:30 AM)
        today_str = datetime.now().strftime('%Y-%m-%d')
        current_late_start_path = os.path.join(today_dir, f'current_day_late_start_{today_str}.csv')
        current_not_on_job_path = os.path.join(today_dir, f'current_day_not_on_job_{today_str}.csv')
        
        # Generate summary report as HTML for easy viewing
        summary_path = os.path.join(today_dir, f'report_summary_{today_str}.html')
        
        with open(summary_path, 'w') as f:
            f.write(f"""
            <html>
                <head>
                    <title>Driver Reports Summary - {today_str}</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 20px; }}
                        h1, h2 {{ color: #333366; }}
                        table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                        th {{ background-color: #333366; color: white; }}
                        tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    </style>
                </head>
                <body>
                    <h1>Driver Reports Summary - {today_str}</h1>
                    
                    <h2>Prior Day Reports ({yesterday_str})</h2>
                    <ul>
                        <li>Late Start: {len(df_late_start)} drivers</li>
                        <li>Early End: {len(df_early_end)} drivers</li>
                        <li>Not On Job: {0} drivers</li>
                    </ul>
                    
                    <h2>Current Day Reports ({today_str})</h2>
                    <ul>
                        <li>Late Start: {0} drivers</li>
                        <li>Not On Job: {0} drivers</li>
                    </ul>
                </body>
            </html>
            """)
        
        # Success message
        flash(f'Driver reports generated successfully for {yesterday_str} and {today_str}', 'success')
        return redirect(url_for('reports.driver_reports'))
        
    except Exception as e:
        logging.error(f"Error generating driver reports: {str(e)}")
        flash(f'Error generating driver reports: {str(e)}', 'danger')
        return redirect(url_for('reports.generate_daily_driver_report'))
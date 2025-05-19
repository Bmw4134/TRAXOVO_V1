"""Simple export function for driver reports"""
import os
import csv
import logging
from datetime import datetime, timedelta
from flask import request, redirect, url_for, flash, send_file
from flask_login import login_required
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Set up logging
logger = logging.getLogger(__name__)

# Define exports folder
EXPORTS_FOLDER = 'exports'

@login_required
def export_report():
    """Export a driver report using real database data"""
    # Get request parameters
    report_type = request.args.get('type', 'daily')
    export_format = request.args.get('format', 'xlsx')
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    region_id = request.args.get('region')
    direct_download = request.args.get('direct', 'false') == 'true'
    
    try:
        # Parse the date for filtering records
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            report_date = datetime.now().date()
            
        # Generate a filename with timestamp for uniqueness
        timestamp = datetime.now()
        formatted_date = report_date.strftime('%m_%d_%Y')
        filename_base = f"{report_type}_report_{formatted_date}"
        if region_id:
            filename_base += f"_region_{region_id}"
        
        # Add timestamp to ensure uniqueness
        date_str = timestamp.strftime('%H%M%S')
        filename = f"{filename_base}_{date_str}.{export_format}"
        
        # Create exports directory if it doesn't exist
        if not os.path.exists(EXPORTS_FOLDER):
            os.makedirs(EXPORTS_FOLDER)
        
        # Create the output file path
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        
        # Get attendance data from database
        from models.driver_attendance import AttendanceRecord, DriverAttendance, JobSiteAttendance
        from app import db
        
        # Prepare data for the report
        attendance_records = []
        
        if report_type == 'daily':
            try:
                # Query for attendance records on the specified date
                query = db.session.query(
                    AttendanceRecord, DriverAttendance, JobSiteAttendance
                ).join(
                    DriverAttendance, AttendanceRecord.driver_id == DriverAttendance.id
                ).join(
                    JobSiteAttendance, AttendanceRecord.assigned_job_id == JobSiteAttendance.id, isouter=True
                ).filter(
                    AttendanceRecord.date == report_date
                )
                
                # Apply region filter if specified
                if region_id:
                    query = query.filter(DriverAttendance.division == region_id)
                
                # Execute the query
                records = query.all()
                
                # Process records for CSV or Excel export
                for record, driver, job_site in records:
                    status = "On Time"
                    if record.late_start:
                        status = "Late Start"
                    elif record.early_end:
                        status = "Early End"
                    elif record.not_on_job:
                        status = "Not on Job"
                    
                    attendance_records.append({
                        'employee_id': driver.employee_id,
                        'name': driver.full_name,
                        'division': driver.division or 'Unknown',
                        'job_site': job_site.name if job_site else 'Unknown',
                        'status': status,
                        'date': record.date,
                        'expected_start': record.expected_start_time,
                        'actual_start': record.actual_start_time,
                        'expected_end': record.expected_end_time,
                        'actual_end': record.actual_end_time,
                        'notes': record.notes,
                        'vehicle': record.asset_id
                    })
                
                # Create the export file
                if export_format == 'csv':
                    with open(file_path, 'w', newline='') as csvfile:
                        fieldnames = ['employee_id', 'name', 'division', 'job_site', 'status', 
                                     'date', 'expected_start', 'actual_start', 'expected_end', 
                                     'actual_end', 'notes', 'vehicle']
                        
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        
                        for record in attendance_records:
                            writer.writerow(record)
                
                elif export_format == 'xlsx':
                    # Create a new workbook and select the active sheet
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Driver Attendance"
                    
                    # Add headers with styling
                    headers = ['Employee ID', 'Name', 'Division', 'Job Site', 'Status', 
                              'Date', 'Expected Start', 'Actual Start', 'Expected End', 
                              'Actual End', 'Notes', 'Vehicle']
                    
                    # Add header row
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # Add data rows
                    for row_idx, record in enumerate(attendance_records, 2):
                        ws.cell(row=row_idx, column=1, value=record['employee_id'])
                        ws.cell(row=row_idx, column=2, value=record['name'])
                        ws.cell(row=row_idx, column=3, value=record['division'])
                        ws.cell(row=row_idx, column=4, value=record['job_site'])
                        ws.cell(row=row_idx, column=5, value=record['status'])
                        ws.cell(row=row_idx, column=6, value=record['date'].strftime('%Y-%m-%d') if record['date'] else '')
                        ws.cell(row=row_idx, column=7, value=record['expected_start'].strftime('%H:%M') if record['expected_start'] else '')
                        ws.cell(row=row_idx, column=8, value=record['actual_start'].strftime('%H:%M') if record['actual_start'] else '')
                        ws.cell(row=row_idx, column=9, value=record['expected_end'].strftime('%H:%M') if record['expected_end'] else '')
                        ws.cell(row=row_idx, column=10, value=record['actual_end'].strftime('%H:%M') if record['actual_end'] else '')
                        ws.cell(row=row_idx, column=11, value=record['notes'])
                        ws.cell(row=row_idx, column=12, value=record['vehicle'])
                        
                        # Color-code status
                        status_cell = ws.cell(row=row_idx, column=5)
                        if record['status'] == 'Late Start':
                            status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif record['status'] == 'Early End':
                            status_cell.fill = PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
                        elif record['status'] == 'Not on Job':
                            status_cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
                        elif record['status'] == 'On Time':
                            status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    
                    # Auto-size columns for better readability
                    for col in range(1, len(headers) + 1):
                        column_letter = get_column_letter(col)
                        ws.column_dimensions[column_letter].width = 15
                    
                    # Save the Excel file
                    wb.save(file_path)
                else:
                    flash(f"Unsupported format: {export_format}", "danger")
                    return redirect(url_for('driver_module.daily_report'))
                
            except Exception as e:
                logger.error(f"Error creating export: {str(e)}", exc_info=True)
                flash(f"Error creating export: {str(e)}", "danger")
                return redirect(url_for('driver_module.daily_report'))
        
        else:
            flash(f"Unsupported report type: {report_type}", "danger")
            return redirect(url_for('driver_module.index'))
        
        # Log the export
        logger.info(f"Successfully generated {report_type} report in {export_format} format")
        
        # Create download URL
        download_url = url_for('driver_module.download_export', filename=filename)
        
        # Return appropriate response
        if direct_download:
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Report generated successfully. <a href='{download_url}'>Download {filename}</a>", "success")
            
            if report_type == 'daily':
                return redirect(url_for('driver_module.daily_report', date=date_param))
            else:
                return redirect(url_for('driver_module.index'))
    
    except Exception as e:
        logger.error(f"Error in export_report: {str(e)}", exc_info=True)
        flash(f"Error generating report: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))


@login_required
def download_export(filename):
    """Download a previously exported file"""
    try:
        file_path = os.path.join(EXPORTS_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            flash(f"Requested file not found: {filename}", "danger")
            return redirect(url_for('driver_module.index'))
    except Exception as e:
        logger.error(f"Error downloading export: {str(e)}")
        flash(f"Error downloading export: {str(e)}", "danger")
        return redirect(url_for('driver_module.index'))
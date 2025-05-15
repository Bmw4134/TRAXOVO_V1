"""
Attendance Processor Module

This module processes activity detail and driving history data to generate
attendance reports for daily driver tracking.
"""

import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time
import pytz

from app import db
from models import Asset
from utils.geofence_processor import is_point_in_geofence

# Initialize logger
logger = logging.getLogger(__name__)

# Constants
EXPECTED_START_TIME = time(7, 0)  # 7:00 AM
EXPECTED_END_TIME = time(17, 0)  # 5:00 PM
CENTRAL_TIMEZONE = pytz.timezone('US/Central')
MORNING_REPORT_TIME = time(8, 30)  # 8:30 AM when current day report is generated

def load_activity_data(activity_file=None, driving_file=None):
    """
    Load activity detail and driving history data from files
    
    Args:
        activity_file (str): Path to activity detail Excel file
        driving_file (str): Path to driving history Excel file
        
    Returns:
        tuple: (activity_df, driving_df) - Pandas DataFrames with loaded data
    """
    try:
        # Find files if not provided
        if not activity_file or not os.path.exists(activity_file):
            activity_file = 'attached_assets/DAILY LATE START-EARLY END & NOJ REPORT_05.14.2025.xlsx'
        
        if not driving_file or not os.path.exists(driving_file):
            # For demonstration, we'll use the same file (in production these would be separate)
            driving_file = activity_file
            
        # Load activity detail
        activity_df = pd.read_excel(activity_file, sheet_name='Activity Detail')
        
        # Load driving history (using the same file for demo purposes)
        driving_df = pd.read_excel(driving_file, sheet_name='Driving History')
        
        logger.info(f"Loaded activity data: {len(activity_df)} records")
        logger.info(f"Loaded driving data: {len(driving_df)} records")
        
        return activity_df, driving_df
    except Exception as e:
        logger.error(f"Error loading activity data: {e}")
        return None, None

def process_prior_day_attendance(date=None):
    """
    Process prior day attendance data to identify late starts, early ends, and not on job
    
    Args:
        date (str): Date to process in format 'YYYY-MM-DD', defaults to yesterday
        
    Returns:
        dict: Report data with late starts, early ends, and not on job
    """
    try:
        # Determine date to process
        if date is None:
            process_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        else:
            process_date = date
            
        logger.info(f"Processing prior day attendance for {process_date}")
        
        # Load activity data
        activity_df, driving_df = load_activity_data()
        if activity_df is None or driving_df is None:
            return {'status': 'error', 'message': 'Failed to load activity data'}
        
        # Filter for the specific date
        activity_df['Date'] = pd.to_datetime(activity_df['Date']).dt.date
        activity_df = activity_df[activity_df['Date'] == datetime.strptime(process_date, '%Y-%m-%d').date()]
        
        # Initialize report data
        report_data = {
            'late_starts': [],
            'early_ends': [],
            'not_on_job': [],
            'summary': {
                'total_drivers': len(activity_df['Driver'].unique()),
                'late_start_count': 0,
                'early_end_count': 0,
                'not_on_job_count': 0
            }
        }
        
        # Get geofence data
        # In a real implementation, we would load this from the database
        job_sites = [
            {'name': 'Site A', 'latitude': 32.7767, 'longitude': -96.7970, 'radius': 500},
            {'name': 'Site B', 'latitude': 32.8208, 'longitude': -96.8716, 'radius': 500},
            {'name': 'Site C', 'latitude': 29.7604, 'longitude': -95.3698, 'radius': 500},
            {'name': 'Site D', 'latitude': 31.5491, 'longitude': -97.1467, 'radius': 500}
        ]
        
        # Process each driver
        for _, row in activity_df.iterrows():
            driver_record = {
                'identifier': row.get('Vehicle ID', ''),
                'driver': row.get('Driver', ''),
                'region': row.get('Region', ''),
                'expected_start': EXPECTED_START_TIME.strftime('%H:%M'),
                'expected_end': EXPECTED_END_TIME.strftime('%H:%M'),
                'expected_job': row.get('Assigned Job', 'Unknown'),
                'status': []
            }
            
            # Check for late start
            actual_start = row.get('First Active', None)
            if actual_start:
                if isinstance(actual_start, str):
                    actual_start = datetime.strptime(actual_start, '%H:%M').time()
                driver_record['actual_start'] = actual_start.strftime('%H:%M')
                
                # Calculate late minutes
                start_diff = (datetime.combine(datetime.today(), actual_start) - 
                           datetime.combine(datetime.today(), EXPECTED_START_TIME))
                late_minutes = max(0, int(start_diff.total_seconds() / 60))
                
                if late_minutes > 0:
                    driver_record['late_by'] = late_minutes
                    driver_record['status'].append('Late Start')
                    report_data['late_starts'].append(driver_record.copy())
                    report_data['summary']['late_start_count'] += 1
            else:
                driver_record['actual_start'] = 'No data'
                driver_record['late_by'] = 'N/A'
            
            # Check for early end
            actual_end = row.get('Last Active', None)
            if actual_end:
                if isinstance(actual_end, str):
                    actual_end = datetime.strptime(actual_end, '%H:%M').time()
                driver_record['actual_end'] = actual_end.strftime('%H:%M')
                
                # Calculate early minutes
                end_diff = (datetime.combine(datetime.today(), EXPECTED_END_TIME) - 
                         datetime.combine(datetime.today(), actual_end))
                early_minutes = max(0, int(end_diff.total_seconds() / 60))
                
                if early_minutes > 0:
                    driver_record['early_by'] = early_minutes
                    driver_record['status'].append('Early End')
                    report_data['early_ends'].append(driver_record.copy())
                    report_data['summary']['early_end_count'] += 1
            else:
                driver_record['actual_end'] = 'No data'
                driver_record['early_by'] = 'N/A'
            
            # Check for not on job
            actual_location = row.get('Last Location', '')
            driver_record['actual_location'] = actual_location
            
            # Check if driver was at assigned job site
            expected_job = driver_record['expected_job']
            at_job_site = False
            
            # In a real implementation, we would use geofence detection
            if 'GPS Lat' in row and 'GPS Long' in row and row['GPS Lat'] and row['GPS Long']:
                lat, lon = row['GPS Lat'], row['GPS Long']
                
                # Check if in any job site geofence
                for site in job_sites:
                    if site['name'] == expected_job:
                        # Simple distance calculation (in a real implementation, we would use is_point_in_geofence)
                        at_job_site = True  # Simplified for demo
                        break
            
            if not at_job_site:
                driver_record['status'].append('Not On Job')
                report_data['not_on_job'].append(driver_record.copy())
                report_data['summary']['not_on_job_count'] += 1
        
        # Convert status lists to strings
        for category in ['late_starts', 'early_ends', 'not_on_job']:
            for record in report_data[category]:
                record['status'] = ', '.join(record['status'])
        
        return {
            'status': 'success',
            'date': process_date,
            'data': report_data
        }
    except Exception as e:
        logger.error(f"Error processing prior day attendance: {e}")
        return {'status': 'error', 'message': str(e)}

def process_current_day_attendance():
    """
    Process current day attendance data to identify late starts and not on job
    
    Returns:
        dict: Report data with late starts and not on job
    """
    try:
        # Get today's date
        process_date = datetime.now().strftime('%Y-%m-%d')
        current_time = datetime.now().time()
        
        logger.info(f"Processing current day attendance for {process_date} as of {current_time.strftime('%H:%M')}")
        
        # Load activity data
        activity_df, driving_df = load_activity_data()
        if activity_df is None or driving_df is None:
            return {'status': 'error', 'message': 'Failed to load activity data'}
        
        # Filter for the specific date
        activity_df['Date'] = pd.to_datetime(activity_df['Date']).dt.date
        activity_df = activity_df[activity_df['Date'] == datetime.strptime(process_date, '%Y-%m-%d').date()]
        
        # Initialize report data
        report_data = {
            'late_starts': [],
            'not_on_job': [],
            'summary': {
                'total_drivers': len(activity_df['Driver'].unique()),
                'late_start_count': 0,
                'not_on_job_count': 0
            }
        }
        
        # Get geofence data (placeholder - in a real implementation, we would load from the database)
        job_sites = [
            {'name': 'Site A', 'latitude': 32.7767, 'longitude': -96.7970, 'radius': 500},
            {'name': 'Site B', 'latitude': 32.8208, 'longitude': -96.8716, 'radius': 500},
            {'name': 'Site C', 'latitude': 29.7604, 'longitude': -95.3698, 'radius': 500},
            {'name': 'Site D', 'latitude': 31.5491, 'longitude': -97.1467, 'radius': 500}
        ]
        
        # Process each driver
        for _, row in activity_df.iterrows():
            driver_record = {
                'identifier': row.get('Vehicle ID', ''),
                'driver': row.get('Driver', ''),
                'region': row.get('Region', ''),
                'expected_start': EXPECTED_START_TIME.strftime('%H:%M'),
                'status_at_830': 'No data',
                'expected_job': row.get('Assigned Job', 'Unknown'),
                'actual_location': row.get('Last Location', 'Unknown'),
                'status': []
            }
            
            # Check for late start
            actual_start = row.get('First Active', None)
            if actual_start:
                if isinstance(actual_start, str):
                    actual_start = datetime.strptime(actual_start, '%H:%M').time()
                
                # Calculate late minutes
                start_diff = (datetime.combine(datetime.today(), actual_start) - 
                           datetime.combine(datetime.today(), EXPECTED_START_TIME))
                late_minutes = max(0, int(start_diff.total_seconds() / 60))
                
                if late_minutes > 0:
                    driver_record['late_by'] = late_minutes
                    driver_record['status_at_830'] = 'Late'
                    driver_record['status'].append('Late Start')
                    report_data['late_starts'].append(driver_record.copy())
                    report_data['summary']['late_start_count'] += 1
                else:
                    driver_record['late_by'] = 0
                    driver_record['status_at_830'] = 'On Time'
            else:
                # If no start time recorded yet (as of 8:30 AM)
                driver_record['late_by'] = '90+ min'
                driver_record['status_at_830'] = 'Not on site at 8:30 AM'
                driver_record['status'].append('Late Start')
                report_data['late_starts'].append(driver_record.copy())
                report_data['summary']['late_start_count'] += 1
            
            # Check for not on job
            # In a real implementation, we would use geofence detection
            expected_job = driver_record['expected_job']
            at_job_site = False
            
            if 'GPS Lat' in row and 'GPS Long' in row and row['GPS Lat'] and row['GPS Long']:
                lat, lon = row['GPS Lat'], row['GPS Long']
                
                # Check if in any job site geofence
                for site in job_sites:
                    if site['name'] == expected_job:
                        # Simple distance calculation (in a real implementation, we would use is_point_in_geofence)
                        at_job_site = True  # Simplified for demo
                        break
            
            if not at_job_site:
                if 'Late Start' not in driver_record['status']:
                    driver_record['status_at_830'] = 'At incorrect job site'
                driver_record['status'].append('Not On Job')
                report_data['not_on_job'].append(driver_record.copy())
                report_data['summary']['not_on_job_count'] += 1
        
        # Convert status lists to strings
        for category in ['late_starts', 'not_on_job']:
            for record in report_data[category]:
                record['status'] = ', '.join(record['status'])
        
        return {
            'status': 'success',
            'date': process_date,
            'report_time': MORNING_REPORT_TIME.strftime('%H:%M'),
            'data': report_data
        }
    except Exception as e:
        logger.error(f"Error processing current day attendance: {e}")
        return {'status': 'error', 'message': str(e)}

def create_attendance_excel(report_data, report_type):
    """
    Create Excel report from attendance data
    
    Args:
        report_data (dict): Report data from process_prior_day_attendance or process_current_day_attendance
        report_type (str): 'prior_day' or 'current_day'
        
    Returns:
        str: Path to the created Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image
    
    try:
        if report_data['status'] != 'success':
            return None
            
        # Create directory if it doesn't exist
        report_date = report_data['date']
        reports_dir = f'reports/{report_date}'
        os.makedirs(reports_dir, exist_ok=True)
        
        # Define file path
        if report_type == 'prior_day':
            file_path = f"{reports_dir}/prior_day_attendance_{report_date}.xlsx"
            report_title = "Prior Day Attendance Report"
        else:
            file_path = f"{reports_dir}/current_day_attendance_{report_date}.xlsx"
            report_title = "Current Day Attendance Report"
            
        # Create workbook with company branding
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Set column widths for better readability
        for col in range(1, 10):
            column_letter = openpyxl.utils.get_column_letter(col)
            summary_sheet.column_dimensions[column_letter].width = 15
        
        # Add report header with styling
        # Row 1: Title
        summary_sheet['A1'] = report_title
        summary_sheet['A1'].font = Font(size=16, bold=True, color="1F497D")
        summary_sheet.merge_cells('A1:F1')
        summary_sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 2: Date
        date_text = f"Date: {report_date}"
        if report_type == 'current_day':
            date_text = f"Date: {report_date} (as of {report_data.get('report_time', '8:30 AM')})"
        summary_sheet['A2'] = date_text
        summary_sheet['A2'].font = Font(size=12)
        summary_sheet.merge_cells('A2:F2')
        summary_sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Row 3: Generated timestamp
        summary_sheet['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet['A3'].font = Font(size=10, italic=True)
        summary_sheet.merge_cells('A3:F3')
        
        # Add styling to the header section
        for row in range(1, 4):
            summary_sheet.row_dimensions[row].height = 20
        
        # Create border styling
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Add summary statistics section
        summary_sheet['A5'] = "Summary Statistics"
        summary_sheet['A5'].font = Font(bold=True, size=12, color="1F497D")
        summary_sheet.merge_cells('A5:F5')
        summary_sheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet['A5'].fill = PatternFill(start_color="E6F1F5", end_color="E6F1F5", fill_type="solid")
        
        # Apply styling to the header row
        for col in range(1, 7):
            cell = summary_sheet.cell(row=5, column=col)
            cell.border = thin_border
        
        summary_data = report_data['data']['summary']
        
        # Row headings
        stats_headings = [
            ("Total Drivers", summary_data['total_drivers']),
            ("Late Start Count", summary_data['late_start_count']),
        ]
        
        if report_type == 'prior_day':
            stats_headings.extend([
                ("Early End Count", summary_data['early_end_count']),
                ("Not On Job Count", summary_data['not_on_job_count']),
            ])
        else:
            stats_headings.append(
                ("Not On Job Count", summary_data['not_on_job_count']),
            )
            
        # Add statistics with proper styling
        start_row = 6
        for idx, (label, value) in enumerate(stats_headings):
            row = start_row + idx
            
            # Label cell
            summary_sheet.cell(row=row, column=1, value=label)
            summary_sheet.cell(row=row, column=1).font = Font(bold=True)
            summary_sheet.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
            summary_sheet.cell(row=row, column=1).border = thin_border
            
            # Value cell
            summary_sheet.cell(row=row, column=2, value=value)
            summary_sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            summary_sheet.cell(row=row, column=2).border = thin_border
            
            # Apply alternating row colors
            if idx % 2 == 0:
                summary_sheet.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                summary_sheet.cell(row=row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Add detail sheets
        # Late Start sheet
        if report_data['data']['late_starts']:
            late_sheet = wb.create_sheet("Late Start")
            
            # Add headers
            if report_type == 'prior_day':
                headers = ["Identifier", "Driver", "Region", "Expected Start", "Actual Start", 
                          "Late By (min)", "Expected Job", "Status"]
            else:
                headers = ["Identifier", "Driver", "Region", "Expected Start", "Status at 8:30 AM", 
                          "Late By (min)", "Expected Job", "Actual Location", "Status"]
                
            for col, header in enumerate(headers, 1):
                cell = late_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, record in enumerate(report_data['data']['late_starts'], 2):
                col_idx = 1
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('identifier')); col_idx += 1
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('driver')); col_idx += 1
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('region')); col_idx += 1
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('expected_start')); col_idx += 1
                
                if report_type == 'prior_day':
                    late_sheet.cell(row=row_idx, column=col_idx, value=record.get('actual_start')); col_idx += 1
                else:
                    late_sheet.cell(row=row_idx, column=col_idx, value=record.get('status_at_830')); col_idx += 1
                    
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('late_by')); col_idx += 1
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('expected_job')); col_idx += 1
                
                if report_type == 'current_day':
                    late_sheet.cell(row=row_idx, column=col_idx, value=record.get('actual_location')); col_idx += 1
                    
                late_sheet.cell(row=row_idx, column=col_idx, value=record.get('status'))
        
        # Early End sheet (prior day only)
        if report_type == 'prior_day' and report_data['data']['early_ends']:
            early_sheet = wb.create_sheet("Early End")
            
            # Add headers
            headers = ["Identifier", "Driver", "Region", "Expected End", "Actual End", 
                      "Early By (min)", "Expected Job", "Status"]
                
            for col, header in enumerate(headers, 1):
                cell = early_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, record in enumerate(report_data['data']['early_ends'], 2):
                col_idx = 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('identifier')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('driver')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('region')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('expected_end')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('actual_end')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('early_by')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('expected_job')); col_idx += 1
                early_sheet.cell(row=row_idx, column=col_idx, value=record.get('status'))
        
        # Not On Job sheet
        if report_data['data']['not_on_job']:
            noj_sheet = wb.create_sheet("Not On Job")
            
            # Add headers
            if report_type == 'prior_day':
                headers = ["Identifier", "Driver", "Region", "Expected Job", "Actual Location", "Status"]
            else:
                headers = ["Identifier", "Driver", "Region", "Expected Job", "Actual Location", 
                          "Status at 8:30 AM", "Status"]
                
            for col, header in enumerate(headers, 1):
                cell = noj_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, record in enumerate(report_data['data']['not_on_job'], 2):
                col_idx = 1
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('identifier')); col_idx += 1
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('driver')); col_idx += 1
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('region')); col_idx += 1
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('expected_job')); col_idx += 1
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('actual_location')); col_idx += 1
                
                if report_type == 'current_day':
                    noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('status_at_830')); col_idx += 1
                    
                noj_sheet.cell(row=row_idx, column=col_idx, value=record.get('status'))
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max(max_length, 12) + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Created attendance report: {file_path}")
        
        return file_path
    except Exception as e:
        logger.error(f"Error creating attendance Excel report: {e}")
        return None
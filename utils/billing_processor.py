"""
Billing Processor Module

This module handles the comparison of original and PM-edited billing files,
identifies changes, and generates region-based exports for accounting.
It also supports selective updates from PM edits to the master billing file.
"""

import os
import logging
import json
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Initialize logger
logger = logging.getLogger(__name__)

def load_billing_files(original_file=None, edited_file=None):
    """
    Load original and PM-edited billing files
    
    Args:
        original_file (str): Path to original billing Excel file
        edited_file (str): Path to PM-edited billing Excel file
        
    Returns:
        tuple: (original_df, edited_df) - Pandas DataFrames with loaded data
    """
    try:
        # Find files if not provided
        if not original_file or not os.path.exists(original_file):
            original_file = 'attached_assets/RAGLE EQ BILLINGS - APRIL 2025 (JG REVIEWED 5.12).xlsm'
        
        if not edited_file or not os.path.exists(edited_file):
            edited_file = 'attached_assets/EQMO. BILLING ALLOCATIONS - APRIL 2025 (TR-FINAL REVISIONS BY 05.15.2025).xlsx'
            
        # Load billing data - assuming the relevant data is in the first sheet
        # In a real implementation, we would identify the right sheet based on content
        original_df = pd.read_excel(original_file, sheet_name=0)
        edited_df = pd.read_excel(edited_file, sheet_name=0)
        
        logger.info(f"Loaded original billing data: {len(original_df)} records")
        logger.info(f"Loaded edited billing data: {len(edited_df)} records")
        
        return original_df, edited_df
    except Exception as e:
        logger.error(f"Error loading billing files: {e}")
        return None, None

def compare_billing_data(original_df, edited_df):
    """
    Compare original and edited billing data to identify changes
    
    Args:
        original_df (DataFrame): Original billing data
        edited_df (DataFrame): PM-edited billing data
        
    Returns:
        dict: Dictionary of identified changes
    """
    try:
        # Standardize column names for comparison
        # In a real implementation, we would handle column name mappings more robustly
        
        # Identify common columns with different names
        original_columns = {col.lower().strip(): col for col in original_df.columns}
        edited_columns = {col.lower().strip(): col for col in edited_df.columns}
        
        # Ensure asset identifiers are properly aligned
        # Assume 'asset id', 'equipment', or 'asset' columns contain the asset identifier
        asset_keywords = ['asset id', 'equipment id', 'asset', 'equipment']
        job_keywords = ['job', 'job code', 'job #', 'job number']
        days_keywords = ['days', 'day count', 'billing days']
        notes_keywords = ['notes', 'comments', 'remarks']
        
        # Find the appropriate column names
        original_asset_col = next((original_columns[key] for key in asset_keywords if key in original_columns), None)
        edited_asset_col = next((edited_columns[key] for key in asset_keywords if key in edited_columns), None)
        
        original_job_col = next((original_columns[key] for key in job_keywords if key in original_columns), None)
        edited_job_col = next((edited_columns[key] for key in job_keywords if key in edited_columns), None)
        
        original_days_col = next((original_columns[key] for key in days_keywords if key in original_columns), None)
        edited_days_col = next((edited_columns[key] for key in days_keywords if key in edited_columns), None)
        
        original_notes_col = next((original_columns[key] for key in notes_keywords if key in original_columns), None)
        edited_notes_col = next((edited_columns[key] for key in notes_keywords if key in edited_columns), None)
        
        # Log the columns we're using for comparison
        logger.info(f"Using columns for comparison: Original - Asset: {original_asset_col}, Job: {original_job_col}, Days: {original_days_col}, Notes: {original_notes_col}")
        logger.info(f"Using columns for comparison: Edited - Asset: {edited_asset_col}, Job: {edited_job_col}, Days: {edited_days_col}, Notes: {edited_notes_col}")
        
        # If any required columns are missing, return error
        if not all([original_asset_col, edited_asset_col, original_job_col, edited_job_col, original_days_col, edited_days_col]):
            return {'status': 'error', 'message': 'Could not identify required columns in billing files'}
            
        # Create normalized DataFrames for comparison
        original_compare = original_df[[original_asset_col, original_job_col, original_days_col]].copy()
        if original_notes_col:
            original_compare[original_notes_col] = original_df[original_notes_col].fillna('').astype(str)
        else:
            original_compare['Notes'] = ''
            original_notes_col = 'Notes'
            
        edited_compare = edited_df[[edited_asset_col, edited_job_col, edited_days_col]].copy()
        if edited_notes_col:
            edited_compare[edited_notes_col] = edited_df[edited_notes_col].fillna('').astype(str)
        else:
            edited_compare['Notes'] = ''
            edited_notes_col = 'Notes'
        
        # Rename columns to standard names for comparison
        original_compare.columns = ['Asset', 'Job', 'Days', 'Notes']
        edited_compare.columns = ['Asset', 'Job', 'Days', 'Notes']
        
        # Ensure asset identifiers are standardized
        original_compare['Asset'] = original_compare['Asset'].astype(str).str.strip().str.upper()
        edited_compare['Asset'] = edited_compare['Asset'].astype(str).str.strip().str.upper()
        
        # Ensure job codes are standardized
        original_compare['Job'] = original_compare['Job'].astype(str).str.strip().str.upper()
        edited_compare['Job'] = edited_compare['Job'].astype(str).str.strip().str.upper()
        
        # Ensure days are numeric
        original_compare['Days'] = pd.to_numeric(original_compare['Days'], errors='coerce').fillna(0)
        edited_compare['Days'] = pd.to_numeric(edited_compare['Days'], errors='coerce').fillna(0)
        
        # Find assets that exist in both files
        common_assets = set(original_compare['Asset']).intersection(set(edited_compare['Asset']))
        logger.info(f"Found {len(common_assets)} common assets between original and edited files")
        
        # Initialize dictionaries to track changes
        job_changes = []
        days_changes = []
        notes_changes = []
        
        # Compare each common asset
        for asset in common_assets:
            # Get original and edited rows for this asset
            orig_rows = original_compare[original_compare['Asset'] == asset]
            edit_rows = edited_compare[edited_compare['Asset'] == asset]
            
            # If multiple rows for an asset (allocated to multiple jobs), we'd need more complex logic
            # For simplicity, we'll just compare the first row
            if len(orig_rows) > 0 and len(edit_rows) > 0:
                orig_row = orig_rows.iloc[0]
                edit_row = edit_rows.iloc[0]
                
                # Compare job code
                if orig_row['Job'] != edit_row['Job']:
                    job_changes.append({
                        'asset': asset,
                        'old': orig_row['Job'],
                        'new': edit_row['Job']
                    })
                
                # Compare days
                if orig_row['Days'] != edit_row['Days']:
                    days_changes.append({
                        'asset': asset,
                        'old': orig_row['Days'],
                        'new': edit_row['Days']
                    })
                
                # Compare notes
                if orig_row['Notes'].strip() != edit_row['Notes'].strip():
                    notes_changes.append({
                        'asset': asset,
                        'old': orig_row['Notes'],
                        'new': edit_row['Notes']
                    })
        
        # Find assets that only exist in one file
        only_in_original = set(original_compare['Asset']).difference(set(edited_compare['Asset']))
        only_in_edited = set(edited_compare['Asset']).difference(set(original_compare['Asset']))
        
        # Return summary of changes
        result = {
            'status': 'success',
            'changes': {
                'job_codes': job_changes,
                'days': days_changes,
                'notes': notes_changes,
                'added_assets': [{'asset': asset} for asset in only_in_edited],
                'removed_assets': [{'asset': asset} for asset in only_in_original]
            },
            'summary': {
                'total_assets': len(common_assets) + len(only_in_original) + len(only_in_edited),
                'common_assets': len(common_assets),
                'only_in_original': len(only_in_original),
                'only_in_edited': len(only_in_edited),
                'job_changes': len(job_changes),
                'days_changes': len(days_changes),
                'notes_changes': len(notes_changes)
            }
        }
        
        # Determine if changes were detected
        result['changes_detected'] = (
            len(job_changes) > 0 or len(days_changes) > 0 or len(notes_changes) > 0 or
            len(only_in_original) > 0 or len(only_in_edited) > 0
        )
        
        return result
    except Exception as e:
        logger.error(f"Error comparing billing data: {e}")
        return {'status': 'error', 'message': str(e)}

def create_comparison_report(comparison_result, month='APRIL'):
    """
    Create an Excel report showing the billing comparison results
    
    Args:
        comparison_result (dict): Result from compare_billing_data
        month (str): Month for the billing period
        
    Returns:
        str: Path to the created Excel file
    """
    try:
        if comparison_result['status'] != 'success':
            return None
            
        # Create directory if it doesn't exist
        today = datetime.now().strftime('%Y-%m-%d')
        reports_dir = f'reports/billing/{today}'
        os.makedirs(reports_dir, exist_ok=True)
        
        # Define file path
        file_path = f"{reports_dir}/billing_comparison_{month}_{today}.xlsx"
            
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Add report title and date
        title = f"Billing Comparison Report - {month} 2025"
        summary_sheet['A1'] = title
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet.merge_cells('A1:F1')
        
        summary_sheet['A2'] = f"Generated: {today}"
        summary_sheet['A2'].font = Font(size=12)
        summary_sheet.merge_cells('A2:F2')
        
        # Add summary statistics
        summary_sheet['A4'] = "Summary Statistics"
        summary_sheet['A4'].font = Font(bold=True)
        summary_sheet.merge_cells('A4:B4')
        
        summary_data = comparison_result['summary']
        
        row = 5
        summary_sheet[f'A{row}'] = "Total Assets:"; summary_sheet[f'B{row}'] = summary_data['total_assets']; row += 1
        summary_sheet[f'A{row}'] = "Common Assets:"; summary_sheet[f'B{row}'] = summary_data['common_assets']; row += 1
        summary_sheet[f'A{row}'] = "Only in Original:"; summary_sheet[f'B{row}'] = summary_data['only_in_original']; row += 1
        summary_sheet[f'A{row}'] = "Only in Edited:"; summary_sheet[f'B{row}'] = summary_data['only_in_edited']; row += 1
        summary_sheet[f'A{row}'] = "Job Code Changes:"; summary_sheet[f'B{row}'] = summary_data['job_changes']; row += 1
        summary_sheet[f'A{row}'] = "Days Changes:"; summary_sheet[f'B{row}'] = summary_data['days_changes']; row += 1
        summary_sheet[f'A{row}'] = "Notes Changes:"; summary_sheet[f'B{row}'] = summary_data['notes_changes']; row += 1
        
        # Add change details
        changes = comparison_result['changes']
        
        # Job Code Changes sheet
        if changes['job_codes']:
            job_sheet = wb.create_sheet("Job Code Changes")
            
            # Add headers
            headers = ["Asset", "Original Job", "Edited Job"]
                
            for col, header in enumerate(headers, 1):
                cell = job_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, change in enumerate(changes['job_codes'], 2):
                job_sheet.cell(row=row_idx, column=1, value=change['asset'])
                job_sheet.cell(row=row_idx, column=2, value=change['old'])
                job_sheet.cell(row=row_idx, column=3, value=change['new'])
                
        # Days Changes sheet
        if changes['days']:
            days_sheet = wb.create_sheet("Days Changes")
            
            # Add headers
            headers = ["Asset", "Original Days", "Edited Days", "Difference"]
                
            for col, header in enumerate(headers, 1):
                cell = days_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, change in enumerate(changes['days'], 2):
                days_sheet.cell(row=row_idx, column=1, value=change['asset'])
                days_sheet.cell(row=row_idx, column=2, value=change['old'])
                days_sheet.cell(row=row_idx, column=3, value=change['new'])
                days_sheet.cell(row=row_idx, column=4, value=change['new'] - change['old'])
                
        # Notes Changes sheet
        if changes['notes']:
            notes_sheet = wb.create_sheet("Notes Changes")
            
            # Add headers
            headers = ["Asset", "Original Notes", "Edited Notes"]
                
            for col, header in enumerate(headers, 1):
                cell = notes_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, change in enumerate(changes['notes'], 2):
                notes_sheet.cell(row=row_idx, column=1, value=change['asset'])
                notes_sheet.cell(row=row_idx, column=2, value=change['old'])
                notes_sheet.cell(row=row_idx, column=3, value=change['new'])
                
        # Added Assets sheet
        if changes['added_assets']:
            added_sheet = wb.create_sheet("Added Assets")
            
            # Add headers
            headers = ["Asset"]
                
            for col, header in enumerate(headers, 1):
                cell = added_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, asset in enumerate(changes['added_assets'], 2):
                added_sheet.cell(row=row_idx, column=1, value=asset['asset'])
                
        # Removed Assets sheet
        if changes['removed_assets']:
            removed_sheet = wb.create_sheet("Removed Assets")
            
            # Add headers
            headers = ["Asset"]
                
            for col, header in enumerate(headers, 1):
                cell = removed_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # Add data
            for row_idx, asset in enumerate(changes['removed_assets'], 2):
                removed_sheet.cell(row=row_idx, column=1, value=asset['asset'])
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max(max_length, 12) + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Created billing comparison report: {file_path}")
        
        return file_path
    except Exception as e:
        logger.error(f"Error creating billing comparison report: {e}")
        return None

def generate_regional_export(edited_df, region_code, month='APRIL', year='2025'):
    """
    Generate a regional billing export for a specific region
    
    Args:
        edited_df (DataFrame): The edited billing data
        region_code (int): Region code (2=DFW, 3=WTX, 4=HOU)
        month (str): Month name for the billing period
        year (str): Year for the billing period
        
    Returns:
        str: Path to the exported file
    """
    try:
        # Define region names
        region_names = {
            2: 'DFW',
            3: 'WTX',
            4: 'HOU'
        }
        
        region_name = region_names.get(region_code, f'Region-{region_code}')
        
        # Create directory if it doesn't exist
        today = datetime.now().strftime('%Y-%m-%d')
        exports_dir = f'exports/billings/{today}'
        os.makedirs(exports_dir, exist_ok=True)
        
        # Define file path
        file_path = f"{exports_dir}/{region_code:02d}-{region_name}-{month}-{year}.xlsx"
        
        # In a real implementation, we would filter and format the data based on the region code
        # For this demo, we'll create a placeholder export file
        
        # Create workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = f"{region_name} Billing"
        
        # Add headers for export
        headers = ["Asset ID", "Job Code", "Description", "Days", "Rate", "Amount", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Placeholder data - in a real implementation, we would extract actual data
        sheet.cell(row=2, column=1, value="EX-81")
        sheet.cell(row=2, column=2, value="J456")
        sheet.cell(row=2, column=3, value="Excavator - Caterpillar 320")
        sheet.cell(row=2, column=4, value=22)
        sheet.cell(row=2, column=5, value=650.00)
        sheet.cell(row=2, column=6, value="=D2*E2")
        sheet.cell(row=2, column=7, value="")
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Created regional export: {file_path}")
        
        return file_path
    except Exception as e:
        logger.error(f"Error generating regional export: {e}")
        return None

def generate_regional_export(df, region_code, month='APRIL', year='2025'):
    """
    Generate billing export for a specific region
    
    Args:
        df (DataFrame): DataFrame with billing data
        region_code (int): Region code (2=DFW, 3=WTX, 4=HOU)
        month (str): Month name for the billing period
        year (str): Year for the billing period
        
    Returns:
        str: Path to the generated export file, or None if generation failed
    """
    try:
        # Map region code to name
        region_name = {2: 'DFW', 3: 'WTX', 4: 'HOU'}.get(region_code, f'Region-{region_code}')
        
        # Create exports directory if it doesn't exist
        export_dir = os.path.join('exports', f"{year}-{month}")
        os.makedirs(export_dir, exist_ok=True)
        
        # File name for the export
        export_file = os.path.join(export_dir, f"{region_name}-{month}-{year}-BILLING.xlsx")
        
        # Filter data for this region and create a copy
        region_df = df[df['DISTRICT'] == region_code].copy()
        
        # If no data found for this region, return None
        if region_df.empty:
            logger.warning(f"No data found for region {region_name}")
            return None
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{region_name} Billing"
        
        # Add headers
        headers = ['JOB', 'ASSET', 'DESCRIPTION', 'TOTAL DAYS', 'DAILY RATE', 'TOTAL', 'NOTES']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col).value = header
            sheet.cell(row=1, column=col).font = Font(bold=True)
        
        # Add data rows
        row = 2
        for _, billing_row in region_df.iterrows():
            # Handle default cost codes if job code is missing
            job_code = billing_row.get('JOB_CODE', '')
            if not job_code or pd.isna(job_code):
                # Default to 9000 100M or 9000 100F
                asset_type = billing_row.get('ASSET_TYPE', '').upper()
                if 'MECH' in asset_type:
                    job_code = '9000 100M'
                else:
                    job_code = '9000 100F'
            
            # Write data
            sheet.cell(row=row, column=1).value = job_code
            sheet.cell(row=row, column=2).value = billing_row.get('ASSET_ID', '')
            sheet.cell(row=row, column=3).value = billing_row.get('DESCRIPTION', '')
            sheet.cell(row=row, column=4).value = billing_row.get('DAYS', 0)
            sheet.cell(row=row, column=5).value = billing_row.get('DAILY_RATE', 0)
            
            # Calculate total
            days = billing_row.get('DAYS', 0)
            rate = billing_row.get('DAILY_RATE', 0)
            total = days * rate if days and rate else 0
            sheet.cell(row=row, column=6).value = total
            
            # Add notes
            sheet.cell(row=row, column=7).value = billing_row.get('NOTES', '')
            
            row += 1
        
        # Format the sheet
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Save the workbook
        workbook.save(export_file)
        logger.info(f"Generated {region_name} export: {export_file}")
        return export_file
    
    except Exception as e:
        logger.error(f"Error generating {region_name} export: {e}")
        return None

def generate_all_region_exports(edited_file=None, month='APRIL', year='2025'):
    """
    Generate billing exports for all regions
    
    Args:
        edited_file (str): Path to PM-edited billing Excel file
        month (str): Month name for the billing period
        year (str): Year for the billing period
        
    Returns:
        dict: Dictionary with export status and paths
    """
    try:
        # Load edited billing file
        if not edited_file or not os.path.exists(edited_file):
            edited_file = 'attached_assets/EQMO. BILLING ALLOCATIONS - APRIL 2025 (TR-FINAL REVISIONS BY 05.15.2025).xlsx'
            
        # Load billing data - assuming the relevant data is in the first sheet
        edited_df = pd.read_excel(edited_file, sheet_name=0)
        
        # Generate exports for each region
        exports = {}
        for region_code in [2, 3, 4]:  # DFW, WTX, HOU
            export_path = generate_regional_export(edited_df, region_code, month, year)
            if export_path:
                region_name = {2: 'DFW', 3: 'WTX', 4: 'HOU'}.get(region_code, f'Region-{region_code}')
                exports[region_name] = export_path
        
        return {
            'status': 'success',
            'exports': exports
        }
    except Exception as e:
        logger.error(f"Error generating region exports: {e}")
        return {'status': 'error', 'message': str(e)}

def load_billing_files(original_file, pm_file):
    """
    Load and prepare billing files for comparison
    
    Args:
        original_file (str): Path to original master billing Excel file
        pm_file (str): Path to PM-edited billing Excel file
        
    Returns:
        tuple: (original_df, pm_df) pandas DataFrames or (None, None) on error
    """
    try:
        logger.info(f"Loading billing files for comparison")
        
        # Skip the header rows since they contain metadata
        original_df = pd.read_excel(original_file, skiprows=3)
        pm_df = pd.read_excel(pm_file, skiprows=3)
        
        # Drop empty or unnecessary columns
        original_df = original_df.dropna(axis=1, how='all')
        pm_df = pm_df.dropna(axis=1, how='all')
        
        # Clean column names to ensure consistency
        key_columns = ['DIV', 'JOB', 'JOB DESC', 'ASSET ID', 'EQUIPMENT', 'DRIVER',
                     'UNIT ALLOCATION', 'COST CODE', 'REVISION', 'NOTE / DETAIL',
                     'INTERNAL MONTHLY RATE', 'RATE X ALLOCATION', 'RATE X REVISION', 'CHANGE']
        
        # Ensure key columns exist in both dataframes
        for df in [original_df, pm_df]:
            for col in key_columns:
                if col not in df.columns:
                    # Try to find a similar column
                    similar_cols = [c for c in df.columns if col.lower() in c.lower()]
                    if similar_cols:
                        df.rename(columns={similar_cols[0]: col}, inplace=True)
                    else:
                        logger.warning(f"Column {col} not found in file and no similar column exists")
        
        # Return the prepared dataframes
        return original_df, pm_df
        
    except Exception as e:
        logger.error(f"Error loading billing files: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None, None

def compare_billing_data(original_df, pm_df):
    """
    Compare original and PM-edited billing data to identify changes
    
    Args:
        original_df (DataFrame): Original billing data
        pm_df (DataFrame): PM-edited billing data
        
    Returns:
        dict: Comparison results with changes identified
    """
    try:
        logger.info("Comparing original and PM-edited billing data")
        
        # Create a unique identifier for each row to match between files
        for df in [original_df, pm_df]:
            df['_row_id'] = df['DIV'].astype(str) + '-' + df['JOB'].astype(str) + '-' + df['ASSET ID'].astype(str)
        
        # Initialize comparison results
        comparison = {
            'status': 'success',
            'total_rows': len(original_df),
            'matched_rows': 0,
            'added_rows': [],
            'removed_rows': [],
            'changed_rows': [],
            'unchanged_rows': 0,
            'changes_by_type': {
                'job_code': [],
                'cost_code': [],
                'allocation': [],
                'revision': [],
                'notes': []
            },
            'regions': {
                'DFW': {'changes': 0, 'rows': []},
                'HOU': {'changes': 0, 'rows': []},
                'WTX': {'changes': 0, 'rows': []}
            }
        }
        
        # Find rows in original that are not in PM-edited (removed)
        original_ids = set(original_df['_row_id'])
        pm_ids = set(pm_df['_row_id'])
        
        removed_ids = original_ids - pm_ids
        for row_id in removed_ids:
            row = original_df[original_df['_row_id'] == row_id].iloc[0].to_dict()
            comparison['removed_rows'].append({
                'row_id': row_id,
                'div': row['DIV'],
                'job': row['JOB'],
                'job_desc': row['JOB DESC'],
                'asset_id': row['ASSET ID'],
                'equipment': row['EQUIPMENT'],
                'driver': row['DRIVER'],
                'allocation': row['UNIT ALLOCATION'],
                'cost_code': row['COST CODE'],
                'internal_rate': row['INTERNAL MONTHLY RATE'],
                'allocation_amount': row['RATE X ALLOCATION']
            })
        
        # Find rows in PM-edited that are not in original (added)
        added_ids = pm_ids - original_ids
        for row_id in added_ids:
            row = pm_df[pm_df['_row_id'] == row_id].iloc[0].to_dict()
            comparison['added_rows'].append({
                'row_id': row_id,
                'div': row['DIV'],
                'job': row['JOB'],
                'job_desc': row['JOB DESC'],
                'asset_id': row['ASSET ID'],
                'equipment': row['EQUIPMENT'],
                'driver': row['DRIVER'],
                'allocation': row['UNIT ALLOCATION'],
                'revision': row['REVISION'] if not pd.isna(row['REVISION']) else None,
                'cost_code': row['COST CODE'],
                'note': row['NOTE / DETAIL'] if not pd.isna(row['NOTE / DETAIL']) else None,
                'internal_rate': row['INTERNAL MONTHLY RATE'],
                'allocation_amount': row['RATE X ALLOCATION'],
                'revision_amount': row['RATE X REVISION']
            })
        
        # Find matching rows and identify changes
        common_ids = original_ids.intersection(pm_ids)
        comparison['matched_rows'] = len(common_ids)
        
        for row_id in common_ids:
            orig_row = original_df[original_df['_row_id'] == row_id].iloc[0]
            pm_row = pm_df[pm_df['_row_id'] == row_id].iloc[0]
            
            changes = []
            
            # Check for changes in key fields
            if orig_row['COST CODE'] != pm_row['COST CODE']:
                changes.append({
                    'field': 'COST CODE',
                    'original': orig_row['COST CODE'],
                    'updated': pm_row['COST CODE']
                })
                comparison['changes_by_type']['cost_code'].append(row_id)
            
            # Check for changes in unit allocation
            if orig_row['UNIT ALLOCATION'] != pm_row['UNIT ALLOCATION']:
                changes.append({
                    'field': 'UNIT ALLOCATION',
                    'original': orig_row['UNIT ALLOCATION'],
                    'updated': pm_row['UNIT ALLOCATION']
                })
                comparison['changes_by_type']['allocation'].append(row_id)
            
            # Check for changes in revision
            orig_revision = orig_row['REVISION'] if not pd.isna(orig_row['REVISION']) else None
            pm_revision = pm_row['REVISION'] if not pd.isna(pm_row['REVISION']) else None
            
            if orig_revision != pm_revision:
                changes.append({
                    'field': 'REVISION',
                    'original': orig_revision,
                    'updated': pm_revision
                })
                comparison['changes_by_type']['revision'].append(row_id)
            
            # Check for changes in notes
            orig_note = orig_row['NOTE / DETAIL'] if not pd.isna(orig_row['NOTE / DETAIL']) else None
            pm_note = pm_row['NOTE / DETAIL'] if not pd.isna(pm_row['NOTE / DETAIL']) else None
            
            if orig_note != pm_note:
                changes.append({
                    'field': 'NOTE / DETAIL',
                    'original': orig_note,
                    'updated': pm_note
                })
                comparison['changes_by_type']['notes'].append(row_id)
            
            # If changes were found, add to changed_rows
            if changes:
                row_data = {
                    'row_id': row_id,
                    'div': orig_row['DIV'],
                    'job': orig_row['JOB'],
                    'job_desc': orig_row['JOB DESC'],
                    'asset_id': orig_row['ASSET ID'],
                    'equipment': orig_row['EQUIPMENT'],
                    'driver': orig_row['DRIVER'],
                    'original': {
                        'allocation': orig_row['UNIT ALLOCATION'],
                        'cost_code': orig_row['COST CODE'],
                        'revision': orig_revision,
                        'note': orig_note,
                        'internal_rate': orig_row['INTERNAL MONTHLY RATE'],
                        'allocation_amount': orig_row['RATE X ALLOCATION'],
                        'revision_amount': orig_row['RATE X REVISION'] if not pd.isna(orig_row['RATE X REVISION']) else orig_row['RATE X ALLOCATION']
                    },
                    'updated': {
                        'allocation': pm_row['UNIT ALLOCATION'],
                        'cost_code': pm_row['COST CODE'],
                        'revision': pm_revision,
                        'note': pm_note,
                        'internal_rate': pm_row['INTERNAL MONTHLY RATE'],
                        'allocation_amount': pm_row['RATE X ALLOCATION'],
                        'revision_amount': pm_row['RATE X REVISION'] if not pd.isna(pm_row['RATE X REVISION']) else pm_row['RATE X ALLOCATION']
                    },
                    'changes': changes
                }
                
                comparison['changed_rows'].append(row_data)
                
                # Categorize by region based on DIV field
                if orig_row['DIV'] == 1:  # DFW
                    comparison['regions']['DFW']['changes'] += 1
                    comparison['regions']['DFW']['rows'].append(row_id)
                elif orig_row['DIV'] == 2:  # HOU
                    comparison['regions']['HOU']['changes'] += 1
                    comparison['regions']['HOU']['rows'].append(row_id)
                elif orig_row['DIV'] == 3:  # WTX
                    comparison['regions']['WTX']['changes'] += 1
                    comparison['regions']['WTX']['rows'].append(row_id)
            else:
                comparison['unchanged_rows'] += 1
        
        # Add summary statistics
        comparison['total_changes'] = len(comparison['changed_rows'])
        comparison['changed_by_type'] = {
            'job_code': len(comparison['changes_by_type']['job_code']),
            'cost_code': len(comparison['changes_by_type']['cost_code']),
            'allocation': len(comparison['changes_by_type']['allocation']),
            'revision': len(comparison['changes_by_type']['revision']),
            'notes': len(comparison['changes_by_type']['notes'])
        }
        
        return comparison
        
    except Exception as e:
        logger.error(f"Error comparing billing data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'status': 'error',
            'message': f"Error comparing billing data: {str(e)}"
        }

def create_comparison_report(comparison, original_df, pm_df, month, export_dir='exports'):
    """
    Create a comparison report Excel file with highlighted changes
    
    Args:
        comparison (dict): Comparison results from compare_billing_data
        original_df (DataFrame): Original billing data
        pm_df (DataFrame): PM-edited billing data
        month (str): Month for the report title
        export_dir (str): Directory to save the report
        
    Returns:
        str: Path to the generated report file
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        logger.info(f"Creating comparison report for {month}")
        
        # Create export directory if it doesn't exist
        os.makedirs(export_dir, exist_ok=True)
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Add report title
        summary_sheet['A1'] = f"PM BILLING ALLOCATION RECONCILIATION - {month.upper()}"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:G1')
        
        # Add timestamp
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet['A2'].font = Font(italic=True)
        summary_sheet.merge_cells('A2:G2')
        
        # Add summary statistics
        summary_sheet['A4'] = "Total Rows Analyzed:"
        summary_sheet['B4'] = comparison['total_rows']
        
        summary_sheet['A5'] = "Matched Rows:"
        summary_sheet['B5'] = comparison['matched_rows']
        
        summary_sheet['A6'] = "Changed Rows:"
        summary_sheet['B6'] = comparison['total_changes']
        
        summary_sheet['A7'] = "Unchanged Rows:"
        summary_sheet['B7'] = comparison['unchanged_rows']
        
        summary_sheet['A8'] = "Added Rows:"
        summary_sheet['B8'] = len(comparison['added_rows'])
        
        summary_sheet['A9'] = "Removed Rows:"
        summary_sheet['B9'] = len(comparison['removed_rows'])
        
        # Add change type breakdown
        summary_sheet['D4'] = "Changes by Type"
        summary_sheet['D4'].font = Font(bold=True)
        
        summary_sheet['D5'] = "Job Code Changes:"
        summary_sheet['E5'] = comparison['changed_by_type']['job_code']
        
        summary_sheet['D6'] = "Cost Code Changes:"
        summary_sheet['E6'] = comparison['changed_by_type']['cost_code']
        
        summary_sheet['D7'] = "Allocation Changes:"
        summary_sheet['E7'] = comparison['changed_by_type']['allocation']
        
        summary_sheet['D8'] = "Revision Changes:"
        summary_sheet['E8'] = comparison['changed_by_type']['revision']
        
        summary_sheet['D9'] = "Notes Changes:"
        summary_sheet['E9'] = comparison['changed_by_type']['notes']
        
        # Add region breakdown
        summary_sheet['A11'] = "Changes by Region"
        summary_sheet['A11'].font = Font(bold=True)
        
        summary_sheet['A12'] = "DFW (Division 1):"
        summary_sheet['B12'] = comparison['regions']['DFW']['changes']
        
        summary_sheet['A13'] = "HOU (Division 2):"
        summary_sheet['B13'] = comparison['regions']['HOU']['changes']
        
        summary_sheet['A14'] = "WTX (Division 3):"
        summary_sheet['B14'] = comparison['regions']['WTX']['changes']
        
        # Create an "All Changes" sheet
        if comparison['total_changes'] > 0:
            changes_sheet = wb.create_sheet(title="All Changes")
            
            # Headers
            headers = ['DIV', 'JOB', 'ASSET ID', 'EQUIPMENT', 'FIELD', 'ORIGINAL VALUE', 'UPDATED VALUE']
            for col_num, header in enumerate(headers, 1):
                cell = changes_sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            
            # Add change data
            row_num = 2
            for changed_row in comparison['changed_rows']:
                for change in changed_row['changes']:
                    changes_sheet.cell(row=row_num, column=1).value = changed_row['div']
                    changes_sheet.cell(row=row_num, column=2).value = changed_row['job']
                    changes_sheet.cell(row=row_num, column=3).value = changed_row['asset_id']
                    changes_sheet.cell(row=row_num, column=4).value = changed_row['equipment']
                    changes_sheet.cell(row=row_num, column=5).value = change['field']
                    changes_sheet.cell(row=row_num, column=6).value = change['original']
                    changes_sheet.cell(row=row_num, column=7).value = change['updated']
                    row_num += 1
            
            # Auto-size columns
            for col in changes_sheet.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                changes_sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
        
        # Create a detailed side-by-side comparison sheet
        detailed_sheet = wb.create_sheet(title="Detailed Comparison")
        
        # Headers for side-by-side comparison
        headers = ['DIV', 'JOB', 'ASSET ID', 'EQUIPMENT', 'DRIVER', 
                  'ORIGINAL ALLOCATION', 'UPDATED ALLOCATION', 
                  'ORIGINAL COST CODE', 'UPDATED COST CODE',
                  'ORIGINAL REVISION', 'UPDATED REVISION',
                  'ORIGINAL NOTE', 'UPDATED NOTE',
                  'ORIGINAL AMOUNT', 'UPDATED AMOUNT', 'CHANGE']
        
        for col_num, header in enumerate(headers, 1):
            cell = detailed_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        
        # Yellow highlight for changed cells
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add all rows, highlighting changes
        row_num = 2
        
        # Helper to fill cells with highlighting changed values
        def fill_comparison_row(row_data, is_changed=False):
            nonlocal row_num
            
            detailed_sheet.cell(row=row_num, column=1).value = row_data['div']
            detailed_sheet.cell(row=row_num, column=2).value = row_data['job']
            detailed_sheet.cell(row=row_num, column=3).value = row_data['asset_id']
            detailed_sheet.cell(row=row_num, column=4).value = row_data['equipment']
            detailed_sheet.cell(row=row_num, column=5).value = row_data['driver']
            
            if is_changed:
                # For changed rows, add both original and updated values
                orig = row_data['original']
                updated = row_data['updated']
                
                # Unit allocation
                allocation_changed = False
                for change in row_data['changes']:
                    if change['field'] == 'UNIT ALLOCATION':
                        allocation_changed = True
                        break
                
                detailed_sheet.cell(row=row_num, column=6).value = orig['allocation']
                detailed_sheet.cell(row=row_num, column=7).value = updated['allocation']
                if allocation_changed:
                    detailed_sheet.cell(row=row_num, column=6).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=7).fill = highlight_fill
                
                # Cost code
                cost_code_changed = False
                for change in row_data['changes']:
                    if change['field'] == 'COST CODE':
                        cost_code_changed = True
                        break
                
                detailed_sheet.cell(row=row_num, column=8).value = orig['cost_code']
                detailed_sheet.cell(row=row_num, column=9).value = updated['cost_code']
                if cost_code_changed:
                    detailed_sheet.cell(row=row_num, column=8).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=9).fill = highlight_fill
                
                # Revision
                revision_changed = False
                for change in row_data['changes']:
                    if change['field'] == 'REVISION':
                        revision_changed = True
                        break
                
                detailed_sheet.cell(row=row_num, column=10).value = orig['revision']
                detailed_sheet.cell(row=row_num, column=11).value = updated['revision']
                if revision_changed:
                    detailed_sheet.cell(row=row_num, column=10).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=11).fill = highlight_fill
                
                # Note
                note_changed = False
                for change in row_data['changes']:
                    if change['field'] == 'NOTE / DETAIL':
                        note_changed = True
                        break
                
                detailed_sheet.cell(row=row_num, column=12).value = orig['note']
                detailed_sheet.cell(row=row_num, column=13).value = updated['note']
                if note_changed:
                    detailed_sheet.cell(row=row_num, column=12).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=13).fill = highlight_fill
                
                # Amounts
                detailed_sheet.cell(row=row_num, column=14).value = orig['revision_amount']
                detailed_sheet.cell(row=row_num, column=15).value = updated['revision_amount']
                
                # Calculate change in amount
                amount_change = updated['revision_amount'] - orig['revision_amount']
                detailed_sheet.cell(row=row_num, column=16).value = amount_change
                if amount_change != 0:
                    detailed_sheet.cell(row=row_num, column=14).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=15).fill = highlight_fill
                    detailed_sheet.cell(row=row_num, column=16).fill = highlight_fill
            else:
                # For added/removed rows, just fill in the available data
                pass
            
            row_num += 1
        
        # Add changed rows with highlighting
        for row_data in comparison['changed_rows']:
            fill_comparison_row(row_data, is_changed=True)
        
        # Create region-specific sheets
        for region_name, region_data in comparison['regions'].items():
            if region_data['changes'] > 0:
                region_sheet = wb.create_sheet(title=f"{region_name} Changes")
                
                # Add the same headers as the detailed sheet
                for col_num, header in enumerate(headers, 1):
                    cell = region_sheet.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                
                # Add region-specific changed rows
                region_row = 2
                for row_data in comparison['changed_rows']:
                    if row_data['row_id'] in region_data['rows']:
                        # Copy the same logic as fill_comparison_row but for the region sheet
                        for col_num in range(1, 17):  # 16 columns
                            source_cell = detailed_sheet.cell(row=row_num - len(comparison['changed_rows']) + comparison['changed_rows'].index(row_data), column=col_num)
                            target_cell = region_sheet.cell(row=region_row, column=col_num)
                            target_cell.value = source_cell.value
                            if source_cell.fill.start_color.index != "00000000":  # If source has fill
                                target_cell.fill = source_cell.fill
                        region_row += 1
                
                # Auto-size columns
                for col in region_sheet.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    region_sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
        
        # Auto-size columns for summary and detailed sheets
        for sheet in [summary_sheet, detailed_sheet]:
            for col in sheet.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
        
        # Save the workbook
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        report_filename = f"pm_allocation_comparison_{month}_{timestamp}.xlsx"
        report_path = os.path.join(export_dir, report_filename)
        wb.save(report_path)
        
        logger.info(f"Generated comparison report: {report_path}")
        return report_path
        
    except Exception as e:
        logger.error(f"Error creating comparison report: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def generate_region_exports(comparison, pm_df, month, export_dir='exports', fsi_format=False):
    """
    Generate region-specific billing allocation exports
    
    Args:
        comparison (dict): Comparison results from compare_billing_data
        pm_df (DataFrame): PM-edited billing data
        month (str): Month for the export filenames
        export_dir (str): Directory to save the exports
        fsi_format (bool): Whether to generate FSI Import Format files
        
    Returns:
        dict: Export results with file paths
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        logger.info(f"Generating region exports for {month}")
        
        # Create export directory if it doesn't exist
        os.makedirs(export_dir, exist_ok=True)
        
        # Get region-specific data
        exports = {
            'status': 'success',
            'export_files': [],
            'regions': {}
        }
        
        # Create copies of pm_df for each region
        region_dfs = {
            'DFW': pm_df[pm_df['DIV'] == 1].copy(),
            'HOU': pm_df[pm_df['DIV'] == 2].copy(),
            'WTX': pm_df[pm_df['DIV'] == 3].copy()
        }
        
        # Function to generate standard export
        def create_standard_export(region_name, region_df):
            if len(region_df) == 0:
                logger.info(f"No data for {region_name} region")
                return None
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{region_name} - {month}"
            
            # Add headers
            headers = ['DIV', 'JOB', 'JOB DESC', 'ASSET ID', 'EQUIPMENT', 'DRIVER', 
                      'ALLOCATION', 'COST CODE', 'REVISION', 'NOTE / DETAIL',
                      'INTERNAL MONTHLY RATE', 'AMOUNT']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            
            # Add data rows
            row_num = 2
            for _, row in region_df.iterrows():
                ws.cell(row=row_num, column=1).value = row['DIV']
                ws.cell(row=row_num, column=2).value = row['JOB']
                ws.cell(row=row_num, column=3).value = row['JOB DESC']
                ws.cell(row=row_num, column=4).value = row['ASSET ID']
                ws.cell(row=row_num, column=5).value = row['EQUIPMENT']
                ws.cell(row=row_num, column=6).value = row['DRIVER']
                
                # Use REVISION if available, otherwise use UNIT ALLOCATION
                allocation = row['REVISION'] if not pd.isna(row['REVISION']) else row['UNIT ALLOCATION']
                ws.cell(row=row_num, column=7).value = allocation
                
                ws.cell(row=row_num, column=8).value = row['COST CODE']
                ws.cell(row=row_num, column=9).value = row['REVISION'] if not pd.isna(row['REVISION']) else None
                ws.cell(row=row_num, column=10).value = row['NOTE / DETAIL'] if not pd.isna(row['NOTE / DETAIL']) else None
                ws.cell(row=row_num, column=11).value = row['INTERNAL MONTHLY RATE']
                
                # Use RATE X REVISION if available, otherwise use RATE X ALLOCATION
                amount = row['RATE X REVISION'] if not pd.isna(row['RATE X REVISION']) else row['RATE X ALLOCATION']
                ws.cell(row=row_num, column=12).value = amount
                
                row_num += 1
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
            
            # Save the workbook
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            export_filename = f"{region_name}_BILLING_ALLOCATION_{month}_{timestamp}.xlsx"
            export_path = os.path.join(export_dir, export_filename)
            wb.save(export_path)
            
            logger.info(f"Generated {region_name} export: {export_path}")
            return export_path
        
        # Function to generate FSI format export
        def create_fsi_export(region_name, region_df):
            if len(region_df) == 0:
                logger.info(f"No data for {region_name} region FSI format")
                return None
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{region_name} - FSI Format"
            
            # Add headers for FSI format
            headers = ['JOB', 'COST CODE', 'EQUIPMENT', 'AMOUNT', 'NOTES']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            
            # Add data rows in FSI format
            row_num = 2
            for _, row in region_df.iterrows():
                ws.cell(row=row_num, column=1).value = row['JOB']
                ws.cell(row=row_num, column=2).value = row['COST CODE']
                ws.cell(row=row_num, column=3).value = row['EQUIPMENT']
                
                # Use RATE X REVISION if available, otherwise use RATE X ALLOCATION
                amount = row['RATE X REVISION'] if not pd.isna(row['RATE X REVISION']) else row['RATE X ALLOCATION']
                ws.cell(row=row_num, column=4).value = amount
                
                # Combine relevant notes
                notes = []
                if not pd.isna(row['ASSET ID']):
                    notes.append(f"Asset: {row['ASSET ID']}")
                if not pd.isna(row['DRIVER']) and row['DRIVER']:
                    notes.append(f"Driver: {row['DRIVER']}")
                if not pd.isna(row['NOTE / DETAIL']) and row['NOTE / DETAIL']:
                    notes.append(f"Note: {row['NOTE / DETAIL']}")
                
                ws.cell(row=row_num, column=5).value = " | ".join(notes) if notes else None
                
                row_num += 1
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
            
            # Save the workbook
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            export_filename = f"{region_name}_FSI_IMPORT_{month}_{timestamp}.xlsx"
            export_path = os.path.join(export_dir, export_filename)
            wb.save(export_path)
            
            logger.info(f"Generated {region_name} FSI export: {export_path}")
            return export_path
        
        # Generate exports for each region
        for region_name, region_df in region_dfs.items():
            standard_export = create_standard_export(region_name, region_df)
            
            if standard_export:
                exports['export_files'].append(standard_export)
                
                exports['regions'][region_name] = {
                    'standard_export': standard_export,
                    'row_count': len(region_df)
                }
                
                # Create FSI format if requested
                if fsi_format:
                    fsi_export = create_fsi_export(region_name, region_df)
                    if fsi_export:
                        exports['export_files'].append(fsi_export)
                        exports['regions'][region_name]['fsi_export'] = fsi_export
        
        return exports
        
    except Exception as e:
        logger.error(f"Error generating region exports: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'status': 'error',
            'message': f"Error generating region exports: {str(e)}"
        }

def process_pm_allocation(original_file, pm_file, region='all', month=None, fsi_format=False):
    """
    Process PM allocation updates for specific project(s)
    
    This function handles selective updates from PM-edited allocation files
    to the master billing file. It only updates rows/jobs that correspond 
    with the specific PM file upload.
    
    Args:
        original_file (str): Path to original master billing Excel file
        pm_file (str): Path to PM-edited billing Excel file
        region (str): Region to filter by ('all', 'DFW', 'HOU', 'WTX')
        month (str): Month for report title, e.g., "APRIL 2025"
        fsi_format (bool): Whether to generate FSI Import Format files
        
    Returns:
        dict: Process status and information about changes
    """
    try:
        logger.info(f"Processing PM allocation update: {os.path.basename(pm_file)}")
        
        # Create export directory if it doesn't exist
        exports_dir = 'exports'
        os.makedirs(exports_dir, exist_ok=True)
        
        # Extract month from filename if not provided
        if not month:
            month_pattern = re.compile(r'(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+\d{4}', re.IGNORECASE)
            month_match = month_pattern.search(os.path.basename(original_file)) or month_pattern.search(os.path.basename(pm_file))
            
            if month_match:
                month = month_match.group(0)
            else:
                # Default to current month if not found
                month = datetime.now().strftime('%B %Y').upper()
        
        # Load the files
        original_df, pm_df = load_billing_files(original_file, pm_file)
        if original_df is None or pm_df is None:
            return {
                'success': False,
                'message': "Failed to load billing files. Please check file formats."
            }
        
        # Compare billing data
        comparison = compare_billing_data(original_df, pm_df)
        if comparison['status'] != 'success':
            return {
                'success': False,
                'message': comparison.get('message', 'Comparison failed')
            }
        
        # Filter by region if specified
        if region != 'all':
            region_map = {
                'DFW': 1,
                'HOU': 2, 
                'WTX': 3
            }
            
            if region in region_map:
                original_df = original_df[original_df['DIV'] == region_map[region]]
                pm_df = pm_df[pm_df['DIV'] == region_map[region]]
                
                # Update comparison to only show filtered region
                for key in ['changed_rows', 'added_rows', 'removed_rows']:
                    comparison[key] = [row for row in comparison[key] if row['div'] == region_map[region]]
                
                comparison['total_changes'] = len(comparison['changed_rows'])
        
        # Create comparison report
        report_path = create_comparison_report(comparison, original_df, pm_df, month, exports_dir)
        if not report_path:
            return {
                'success': False,
                'message': 'Failed to create comparison report'
            }
        
        # Generate region exports if there are changes
        exports = None
        if comparison['total_changes'] > 0 or len(comparison['added_rows']) > 0:
            exports = generate_region_exports(comparison, pm_df, month, exports_dir, fsi_format)
        
        # Prepare response
        result = {
            'success': True,
            'message': f"PM Allocation processed successfully.",
            'comparison_report': report_path,
            'export_files': [report_path],
            'changes': {
                'total': comparison['total_changes'],
                'added': len(comparison['added_rows']),
                'removed': len(comparison['removed_rows']),
                'by_region': {
                    'DFW': comparison['regions']['DFW']['changes'],
                    'HOU': comparison['regions']['HOU']['changes'],
                    'WTX': comparison['regions']['WTX']['changes']
                }
            }
        }
        
        # Add exports if generated
        if exports and exports['status'] == 'success':
            result['export_files'].extend(exports['export_files'])
            result['region_exports'] = exports['regions']
        
        # Create audit trail entry
        audit_dir = 'reconcile/audit'
        os.makedirs(audit_dir, exist_ok=True)
        
        audit_record = {
            'timestamp': datetime.now().isoformat(),
            'month': month,
            'original_file': original_file,
            'pm_file': pm_file,
            'region_filter': region,
            'changes': result['changes'],
            'reports_generated': result['export_files']
        }
        
        audit_file = os.path.join(audit_dir, f"pm_allocation_audit_{datetime.now().strftime('%Y%m%d%H%M%S')}.json")
        with open(audit_file, 'w') as f:
            json.dump(audit_record, f, indent=2)
        
        logger.info(f"PM Allocation processing complete. Changes: {result['changes']['total']}")
        return result
        
    except Exception as e:
        logger.error(f"Error processing PM allocation: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'message': f"Error processing PM allocation: {str(e)}"
        }
        try:
            master_df = pd.read_excel(original_file)
            pm_df = pd.read_excel(pm_file)
            
            logger.info(f"Loaded master file with {len(master_df)} records")
            logger.info(f"Loaded PM file with {len(pm_df)} records")
        except Exception as e:
            logger.error(f"Error loading Excel files: {e}")
            return {
                'success': False,
                'message': f"Error loading files: {str(e)}"
            }
        
        # Standardize column names
        master_columns = {col.lower().strip().replace(' ', '_'): col for col in master_df.columns}
        pm_columns = {col.lower().strip().replace(' ', '_'): col for col in pm_df.columns}
        
        # Identify key columns in both dataframes
        asset_keywords = ['asset', 'equipment', 'asset_id', 'equipment_id']
        job_keywords = ['job', 'job_code', 'job_#', 'job_number', 'project']
        days_keywords = ['days', 'day_count', 'billing_days']
        notes_keywords = ['notes', 'comments', 'remarks']
        
        # Find column names in master dataframe
        master_asset_col = next((master_columns[key] for key in asset_keywords if key in master_columns), None)
        master_job_col = next((master_columns[key] for key in job_keywords if key in master_columns), None)
        master_days_col = next((master_columns[key] for key in days_keywords if key in master_columns), None)
        master_notes_col = next((master_columns[key] for key in notes_keywords if key in master_columns), None)
        
        # Find column names in PM dataframe
        pm_asset_col = next((pm_columns[key] for key in asset_keywords if key in pm_columns), None)
        pm_job_col = next((pm_columns[key] for key in job_keywords if key in pm_columns), None)
        pm_days_col = next((pm_columns[key] for key in days_keywords if key in pm_columns), None)
        pm_notes_col = next((pm_columns[key] for key in notes_keywords if key in pm_columns), None)
        
        # Check if required columns were found
        if not all([master_asset_col, master_job_col, pm_asset_col, pm_job_col]):
            missing = []
            if not master_asset_col: missing.append("master asset column")
            if not master_job_col: missing.append("master job column")
            if not pm_asset_col: missing.append("PM asset column")
            if not pm_job_col: missing.append("PM job column")
            
            return {
                'success': False,
                'message': f"Could not identify required columns: {', '.join(missing)}"
            }
        
        # If no project number from filename, try to extract from job codes
        if not project_number and pm_job_col:
            # Look at job codes in PM file to identify project pattern
            job_codes = pm_df[pm_job_col].astype(str).str.strip().dropna().unique()
            
            # Try to extract project number from job codes
            for code in job_codes:
                match = project_pattern.search(code)
                if match:
                    project_number = match.group(1)
                    logger.info(f"Extracted project number from job code: {project_number}")
                    break
        
        if not project_number:
            logger.warning("Could not determine project number, will update all matching assets")
        
        # Create copies with standardized columns for tracking changes
        master_copy = master_df.copy()
        pm_copy = pm_df.copy()
        
        # Convert asset columns to string for reliable matching
        master_copy[master_asset_col] = master_copy[master_asset_col].astype(str).str.strip().str.upper()
        pm_copy[pm_asset_col] = pm_copy[pm_asset_col].astype(str).str.strip().str.upper()
        
        # Track changes for reporting
        changes = {
            'updated_assets': [],
            'updated_jobs': [],
            'updated_days': [],
            'updated_notes': []
        }
        
        # Identify which rows to update in the master
        if project_number:
            # If we have a project number, filter for rows with that project
            # Check if master has job column with this project
            if master_job_col:
                project_rows = master_copy[master_copy[master_job_col].astype(str).str.contains(project_number, na=False)]
                logger.info(f"Found {len(project_rows)} records in master for project {project_number}")
            else:
                logger.warning(f"No job column found in master, cannot filter by project {project_number}")
                project_rows = master_copy
        else:
            # If no project number, use all rows (will still match by asset)
            project_rows = master_copy
        
        # If filtering by region, apply that filter
        if region.lower() != 'all' and 'region' in master_copy.columns:
            region_rows = project_rows[project_rows['region'].str.contains(region, case=False, na=False)]
            logger.info(f"Filtered to {len(region_rows)} records for region: {region}")
            project_rows = region_rows
        
        # Get list of assets in PM file
        pm_assets = set(pm_copy[pm_asset_col].unique())
        logger.info(f"PM file contains {len(pm_assets)} unique assets")
        
        # Identify assets in both files
        master_assets = set(project_rows[master_asset_col].unique())
        common_assets = pm_assets.intersection(master_assets)
        logger.info(f"Found {len(common_assets)} assets in both master and PM files")
        
        # Update the master dataframe with PM changes for matching assets
        update_count = 0
        for asset in common_assets:
            # Get master rows for this asset (within project scope)
            master_rows = project_rows[project_rows[master_asset_col] == asset]
            
            # Get PM rows for this asset
            pm_rows = pm_copy[pm_copy[pm_asset_col] == asset]
            
            if len(master_rows) > 0 and len(pm_rows) > 0:
                # Get the index of the master row to update
                master_idx = master_rows.index[0]
                
                # Get the PM data for this asset
                pm_row = pm_rows.iloc[0]
                
                # Update job if specified in both
                if master_job_col and pm_job_col and pd.notna(pm_row[pm_job_col]):
                    old_job = master_df.loc[master_idx, master_job_col]
                    new_job = pm_row[pm_job_col]
                    
                    if str(old_job) != str(new_job):
                        master_df.loc[master_idx, master_job_col] = new_job
                        changes['updated_jobs'].append({
                            'asset': asset,
                            'old': old_job,
                            'new': new_job
                        })
                
                # Update days if specified in both
                if master_days_col and pm_days_col and pd.notna(pm_row[pm_days_col]):
                    old_days = master_df.loc[master_idx, master_days_col]
                    new_days = pm_row[pm_days_col]
                    
                    # Convert to float for numeric comparison
                    try:
                        old_days_num = float(old_days) if pd.notna(old_days) else 0
                        new_days_num = float(new_days) if pd.notna(new_days) else 0
                        
                        if old_days_num != new_days_num:
                            master_df.loc[master_idx, master_days_col] = new_days
                            changes['updated_days'].append({
                                'asset': asset,
                                'old': old_days_num,
                                'new': new_days_num
                            })
                    except ValueError:
                        logger.warning(f"Could not convert days to numbers: {old_days}, {new_days}")
                
                # Update notes if specified in both
                if master_notes_col and pm_notes_col and pd.notna(pm_row[pm_notes_col]):
                    old_notes = str(master_df.loc[master_idx, master_notes_col]) if pd.notna(master_df.loc[master_idx, master_notes_col]) else ''
                    new_notes = str(pm_row[pm_notes_col])
                    
                    if old_notes.strip() != new_notes.strip():
                        master_df.loc[master_idx, master_notes_col] = new_notes
                        changes['updated_notes'].append({
                            'asset': asset,
                            'old': old_notes,
                            'new': new_notes
                        })
                
                # Track the updated asset
                changes['updated_assets'].append(asset)
                update_count += 1
        
        # Remove duplicates from updated assets list
        changes['updated_assets'] = list(set(changes['updated_assets']))
        
        # Generate a timestamped filename for the updated master
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        original_basename = os.path.basename(original_file)
        
        # Create updated filename
        if project_number:
            updated_master_filename = f"UPDATED_{project_number}_{timestamp}_{original_basename}"
        else:
            updated_master_filename = f"UPDATED_PM_CHANGES_{timestamp}_{original_basename}"
        
        updated_master_path = os.path.join(exports_dir, updated_master_filename)
        
        # Save the updated master file
        master_df.to_excel(updated_master_path, index=False)
        
        # Generate region-specific exports if requested
        export_files = [updated_master_path]
        if region.lower() != 'all':
            # Generate region-specific export
            region_export = generate_regional_export(master_df, region)
            if region_export:
                export_files.append(region_export)
                
        # Return success with change summary
        return {
            'success': True,
            'message': f"Successfully updated {update_count} assets with PM changes from {project_number if project_number else 'uploaded file'}.",
            'changes': changes,
            'export_files': export_files
        }
        
    except Exception as e:
        logger.error(f"Error processing PM allocation: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            'success': False,
            'message': f"Error processing PM allocation: {str(e)}"
        }

def run_billing_process(original_file=None, edited_file=None, month='APRIL', year='2025'):
    """
    Run the complete billing process
    
    Args:
        original_file (str): Path to original billing Excel file
        edited_file (str): Path to PM-edited billing Excel file
        month (str): Month name for the billing period
        year (str): Year for the billing period
        
    Returns:
        dict: Dictionary with process status and outputs
    """
    try:
        # Load billing files
        original_df, edited_df = load_billing_files(original_file, edited_file)
        if original_df is None or edited_df is None:
            return {'status': 'error', 'message': 'Failed to load billing files'}
        
        # Compare billing data
        comparison_result = compare_billing_data(original_df, edited_df)
        if comparison_result['status'] != 'success':
            return {'status': 'error', 'message': comparison_result.get('message', 'Comparison failed')}
        
        # Create comparison report
        report_path = create_comparison_report(comparison_result, month)
        if not report_path:
            return {'status': 'error', 'message': 'Failed to create comparison report'}
        
        # Generate regional exports
        exports_result = generate_all_region_exports(edited_file, month, year)
        if exports_result['status'] != 'success':
            return {'status': 'error', 'message': exports_result.get('message', 'Failed to generate exports')}
        
        return {
            'status': 'success',
            'comparison': {
                'path': report_path,
                'summary': comparison_result['summary'],
                'changes_detected': comparison_result['changes_detected']
            },
            'exports': exports_result['exports']
        }
    except Exception as e:
        logger.error(f"Error in billing process: {e}")
        return {'status': 'error', 'message': str(e)}
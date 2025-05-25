"""
Daily Driver Report Generator

This module automatically generates daily driver attendance reports following
the exact format and logic of the manual reports, eliminating the need for
daily manual processing.
"""

import os
import json
import logging
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.attendance_pipeline_v2 import process_attendance_data_v2
from utils.multi_source_processor import combine_attendance_sources
from utils.driver_identity_attendance_integration import (
    enhance_daily_driver_report,
    enhance_attendance_pipeline_results,
    enhance_driver_data_from_files
)

# Configure logging
logger = logging.getLogger(__name__)

# Define constants for report
DEFAULT_START_TIME = "07:00:00"
DEFAULT_END_TIME = "17:00:00"
LATE_THRESHOLD_MINUTES = 15
EARLY_END_THRESHOLD_MINUTES = 15

def generate_daily_report(date_str, driving_history_data=None, time_on_site_data=None, 
                         activity_detail_data=None, timecard_data=None, output_dir=None):
    """
    Generate daily driver attendance report in Excel format matching the manual template.
    
    Args:
        date_str: Date string in YYYY-MM-DD format
        driving_history_data: List of driving history records
        time_on_site_data: List of time on site records
        activity_detail_data: List of activity detail records
        timecard_data: List of timecard records
        output_dir: Directory to save the report (optional)
        
    Returns:
        tuple: (Workbook, Path to saved file or None if not saved)
    """
    try:
        # If no output directory specified, use default
        if not output_dir:
            output_dir = os.path.join('reports', 'daily_driver_reports')
            os.makedirs(output_dir, exist_ok=True)
        
        # Enhance driver data using Secondary Asset Identifier
        enhanced_driving_history, enhanced_time_on_site = enhance_driver_data_from_files(
            driving_history_data, time_on_site_data
        )
        
        # Process data using the V2 engine with enhanced driver data
        attendance_report = process_attendance_data_v2(
            enhanced_driving_history, 
            enhanced_time_on_site, 
            activity_detail_data,
            timecard_data,
            date_str
        )
        
        if not attendance_report:
            logger.warning(f"No attendance data available for date: {date_str}")
            return None, None
            
        # Enhance attendance report with consistent driver identity
        attendance_report = enhance_attendance_pipeline_results(attendance_report)
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        
        # Set title
        formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%m.%d.%Y')
        ws.title = f"DAILY REPORT {formatted_date}"
        
        # Create header
        report_title = f"DAILY LATE START-EARLY END & NOJ REPORT_{formatted_date.replace('.', '_')}"
        ws.merge_cells('A1:K1')
        ws['A1'] = report_title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add subtitle with date
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Report Date: {formatted_date}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add column headers
        headers = [
            'DRIVER', 'JOB NO.', 'JOB NAME', 'STATUS', 'START TIME', 'END TIME', 
            'HOURS ON SITE', 'LATE (MIN)', 'EARLY END (MIN)', 'GPS VERIFIED', 'NOTES'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Set column width
            if col_idx == 1:  # DRIVER
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            elif col_idx == 3:  # JOB NAME
                ws.column_dimensions[get_column_letter(col_idx)].width = 30
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Add border to header row
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=4, column=col_idx).border = border
        
        # Add driver records
        driver_records = attendance_report.get('driver_records', [])
        row_idx = 5
        
        status_colors = {
            'on_time': 'DDFFDD',  # Light green
            'late': 'FFDDDD',     # Light red
            'early_end': 'FFFFDD', # Light yellow
            'not_on_job': 'DDDDFF' # Light blue
        }
        
        # Sort drivers by status (on_time, late, early_end, not_on_job)
        status_order = {
            'late': 0,
            'early_end': 1,
            'not_on_job': 2,
            'on_time': 3
        }
        
        driver_records.sort(key=lambda x: status_order.get(x.get('classification'), 4))
        
        for driver in driver_records:
            # Get status display text
            status = driver.get('classification', '')
            status_display = {
                'on_time': 'ON TIME',
                'late': 'LATE',
                'early_end': 'EARLY END',
                'not_on_job': 'NOT ON JOB'
            }.get(status, 'UNKNOWN')
            
            # Format times
            start_time = driver.get('start_time', '')
            if start_time:
                try:
                    start_time = datetime.strptime(start_time, '%H:%M:%S').strftime('%I:%M %p')
                except:
                    pass
            
            end_time = driver.get('end_time', '')
            if end_time:
                try:
                    end_time = datetime.strptime(end_time, '%H:%M:%S').strftime('%I:%M %p')
                except:
                    pass
            
            # Set cell values
            ws.cell(row=row_idx, column=1).value = driver.get('driver_name', '')
            ws.cell(row=row_idx, column=2).value = driver.get('job_number', '')
            ws.cell(row=row_idx, column=3).value = driver.get('job_name', '')
            ws.cell(row=row_idx, column=4).value = status_display
            ws.cell(row=row_idx, column=5).value = start_time
            ws.cell(row=row_idx, column=6).value = end_time
            ws.cell(row=row_idx, column=7).value = driver.get('hours', '')
            ws.cell(row=row_idx, column=8).value = driver.get('late_minutes', '') if status == 'late' else ''
            ws.cell(row=row_idx, column=9).value = driver.get('early_end_minutes', '') if status == 'early_end' else ''
            ws.cell(row=row_idx, column=10).value = 'YES' if driver.get('gps_verified', False) else 'NO'
            
            # Set alignment
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center' if col_idx != 1 and col_idx != 3 else 'left', vertical='center')
                cell.border = border
            
            # Set row color based on status
            status_color = status_colors.get(status, 'FFFFFF')
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            
            row_idx += 1
        
        # Add summary section
        summary_row = row_idx + 2
        ws.cell(row=summary_row, column=1).value = "SUMMARY"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        summary = attendance_report.get('summary', {})
        
        ws.cell(row=summary_row + 1, column=1).value = "Total Drivers:"
        ws.cell(row=summary_row + 1, column=2).value = summary.get('total_drivers', 0)
        
        ws.cell(row=summary_row + 2, column=1).value = "On Time:"
        ws.cell(row=summary_row + 2, column=2).value = summary.get('on_time_count', 0)
        ws.cell(row=summary_row + 2, column=3).value = f"{summary.get('on_time_percentage', 0)}%"
        
        ws.cell(row=summary_row + 3, column=1).value = "Late:"
        ws.cell(row=summary_row + 3, column=2).value = summary.get('late_count', 0)
        ws.cell(row=summary_row + 3, column=3).value = f"{summary.get('late_percentage', 0)}%"
        
        ws.cell(row=summary_row + 4, column=1).value = "Early End:"
        ws.cell(row=summary_row + 4, column=2).value = summary.get('early_end_count', 0)
        ws.cell(row=summary_row + 4, column=3).value = f"{summary.get('early_end_percentage', 0)}%"
        
        ws.cell(row=summary_row + 5, column=1).value = "Not On Job:"
        ws.cell(row=summary_row + 5, column=2).value = summary.get('not_on_job_count', 0)
        ws.cell(row=summary_row + 5, column=3).value = f"{summary.get('not_on_job_percentage', 0)}%"
        
        ws.cell(row=summary_row + 7, column=1).value = "Average Late Minutes:"
        ws.cell(row=summary_row + 7, column=2).value = summary.get('average_late_minutes', 0)
        
        ws.cell(row=summary_row + 8, column=1).value = "Average Early End Minutes:"
        ws.cell(row=summary_row + 8, column=2).value = summary.get('average_early_end_minutes', 0)
        
        # Format summary cells
        for row in range(summary_row, summary_row + 9):
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='right' if col == 1 else 'left', vertical='center')
        
        # Save the file
        output_path = os.path.join(output_dir, f"DAILY_DRIVER_REPORT_{date_str.replace('-', '_')}.xlsx")
        wb.save(output_path)
        
        # Also save a JSON version for API access
        json_path = os.path.join(output_dir, f"driver_report_{date_str}.json")
        with open(json_path, 'w') as f:
            json.dump(attendance_report, f, indent=2)
        
        return wb, output_path
    
    except Exception as e:
        logger.error(f"Error generating daily report: {str(e)}")
        return None, None

def schedule_daily_report_generation(date_str=None):
    """
    Schedule daily report generation for automatic processing.
    
    Args:
        date_str: Date string in YYYY-MM-DD format (optional, defaults to today)
        
    Returns:
        bool: Success flag
    """
    try:
        if not date_str:
            # Default to today
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        logger.info(f"Scheduling daily report generation for date: {date_str}")
        
        # Set up directories
        upload_dir = os.path.join('uploads', 'daily_reports', date_str)
        output_dir = os.path.join('reports', 'daily_driver_reports')
        
        os.makedirs(upload_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)
        
        # Check if any files are available for the date
        files = os.listdir(upload_dir) if os.path.exists(upload_dir) else []
        
        if not files:
            logger.warning(f"No files found for date: {date_str}")
            return False
        
        from utils.enhanced_data_ingestion import load_csv_file, load_excel_file, infer_file_type_from_path
        
        # Process files
        driving_history_data = []
        time_on_site_data = []
        activity_detail_data = []
        timecard_data = []
        
        for filename in files:
            file_path = os.path.join(upload_dir, filename)
            file_type = infer_file_type_from_path(file_path)
            
            # Skip files not related to the date
            if date_str not in filename and date_str.replace('-', '_') not in filename:
                continue
            
            # Process file based on type
            if filename.endswith('.csv'):
                data = load_csv_file(file_path)
                
                if not data:
                    continue
                
                # Add to appropriate list based on file type
                if file_type == 'Driving History':
                    driving_history_data.extend(data)
                elif file_type == 'Assets Time On Site':
                    time_on_site_data.extend(data)
                elif file_type == 'Activity Detail':
                    activity_detail_data.extend(data)
                elif file_type == 'Timecard':
                    timecard_data.extend(data)
            
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                if file_type == 'Timecard':
                    # Use dynamic timecard processor
                    from utils.dynamic_timecard_processor import process_dynamic_timecard
                    df = process_dynamic_timecard(file_path, start_date=date_str, end_date=date_str)
                    
                    if df is not None and not df.empty:
                        timecard_data.extend(df.to_dict('records'))
                else:
                    # Use regular Excel processor
                    data = load_excel_file(file_path)
                    
                    if not data:
                        continue
                    
                    timecard_data.extend(data)
        
        # Generate the report
        wb, output_path = generate_daily_report(
            date_str,
            driving_history_data,
            time_on_site_data,
            activity_detail_data,
            timecard_data,
            output_dir
        )
        
        if output_path:
            logger.info(f"Daily report generated successfully: {output_path}")
            return True
        else:
            logger.warning(f"Failed to generate daily report for date: {date_str}")
            return False
        
    except Exception as e:
        logger.error(f"Error scheduling daily report generation: {str(e)}")
        return False

def auto_generate_report_for_date(date_str):
    """
    Automatically generate a report for a specific date.
    
    Args:
        date_str: Date string in YYYY-MM-DD format
        
    Returns:
        str: Path to generated report or None if failed
    """
    try:
        logger.info(f"Auto-generating report for date: {date_str}")
        
        # Schedule the report generation
        success = schedule_daily_report_generation(date_str)
        
        if success:
            # Return the path to the generated report
            output_dir = os.path.join('reports', 'daily_driver_reports')
            output_path = os.path.join(output_dir, f"DAILY_DRIVER_REPORT_{date_str.replace('-', '_')}.xlsx")
            
            if os.path.exists(output_path):
                return output_path
        
        return None
        
    except Exception as e:
        logger.error(f"Error auto-generating report: {str(e)}")
        return None

def get_report_for_date(date_str):
    """
    Get the report for a specific date, generating it if needed.
    
    Args:
        date_str: Date string in YYYY-MM-DD format
        
    Returns:
        str: Path to report or None if not available
    """
    try:
        # Check if report already exists
        output_dir = os.path.join('reports', 'daily_driver_reports')
        output_path = os.path.join(output_dir, f"DAILY_DRIVER_REPORT_{date_str.replace('-', '_')}.xlsx")
        
        if os.path.exists(output_path):
            return output_path
        
        # Generate the report
        return auto_generate_report_for_date(date_str)
        
    except Exception as e:
        logger.error(f"Error getting report for date: {str(e)}")
        return None
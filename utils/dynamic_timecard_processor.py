"""
Dynamic Timecard Processor

This module processes dynamic Excel timecard workbooks with formulas intact,
evaluates the Excel logic, and prepares the data for comparison with GPS records.
"""

import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from formulas import Parser

# Configure logging
logger = logging.getLogger(__name__)

class DynamicTimecardProcessor:
    """
    Process dynamic Excel timecard workbooks with formulas intact
    """
    
    def __init__(self, file_path):
        """Initialize processor with file path"""
        self.file_path = file_path
        self.workbook = None
        self.formula_workbook = None
        self.sheets = {}
        self.dataframes = {}
    
    def load_workbook(self):
        """Load workbook with formulas intact"""
        try:
            # Load with formulas intact
            logger.info(f"Loading workbook with formulas intact: {self.file_path}")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            
            # Also load with data_only for comparison
            data_wb = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Store sheet names
            self.sheets = {sheet_name: sheet for sheet_name, sheet in 
                           zip(self.workbook.sheetnames, self.workbook.worksheets)}
            
            # Setup formula parser
            self.setup_formula_parser()
            
            return True
        
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            return False
    
    def setup_formula_parser(self):
        """Set up formula parser"""
        try:
            # Create a formula parser with the workbook
            self.formula_parser = Parser()
            
            # Load all cell formulas
            for sheet_name, sheet in self.sheets.items():
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # If cell contains formula
                            formula = cell.value
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            self.formula_parser.add(f"={formula}", cell_ref)
            
            # Compile the parser
            self.formula_parser.compile()
            
        except Exception as e:
            logger.error(f"Error setting up formula parser: {str(e)}")
    
    def evaluate_formula(self, sheet_name, cell_reference):
        """Evaluate formula at specified cell reference"""
        try:
            cell_ref = f"{sheet_name}!{cell_reference}"
            value = self.formula_parser.calculate(cell_ref)
            return value
        except Exception as e:
            logger.error(f"Error evaluating formula {cell_ref}: {str(e)}")
            return None
    
    def extract_hours_data(self):
        """Extract hours data from the workbook"""
        try:
            # Look for the Hours sheet or section
            hours_sheet = None
            hours_df = None
            
            # Try to find hours data in various ways
            for sheet_name in self.workbook.sheetnames:
                if 'hour' in sheet_name.lower():
                    hours_sheet = sheet_name
                    break
            
            if not hours_sheet:
                # If no specific hours sheet, look for hours data in all sheets
                for sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    
                    # Search for Hours header
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and 'hour' in cell.value.lower():
                                hours_sheet = sheet_name
                                break
                        if hours_sheet:
                            break
            
            if hours_sheet:
                # Convert sheet to dataframe
                sheet = self.workbook[hours_sheet]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    # Try to infer header row
                    header_row = 0
                    for i, row in enumerate(data):
                        if row and any(isinstance(cell, str) and ('name' in cell.lower() or 'date' in cell.lower()) for cell in row if cell):
                            header_row = i
                            break
                    
                    # Extract headers
                    headers = data[header_row]
                    # Clean up headers
                    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(headers)]
                    
                    # Extract data rows
                    data_rows = data[header_row+1:]
                    
                    # Convert to dataframe
                    hours_df = pd.DataFrame(data_rows, columns=headers)
                    
                    # Clean up dataframe
                    hours_df = hours_df.replace('', np.nan).dropna(how='all')
                    
                    # Store in dataframes dict
                    self.dataframes['hours'] = hours_df
                    logger.info(f"Extracted hours data: {hours_df.shape[0]} rows")
                
                else:
                    logger.warning(f"No data found in hours sheet: {hours_sheet}")
            
            else:
                logger.warning("No hours sheet found in workbook")
            
            return hours_df
        
        except Exception as e:
            logger.error(f"Error extracting hours data: {str(e)}")
            return None
    
    def extract_quantities_data(self):
        """Extract quantities data from the workbook"""
        try:
            # Look for the Quantities sheet or section
            quantities_sheet = None
            quantities_df = None
            
            # Try to find quantities data in various ways
            for sheet_name in self.workbook.sheetnames:
                if 'quantit' in sheet_name.lower():
                    quantities_sheet = sheet_name
                    break
            
            if not quantities_sheet:
                # If no specific quantities sheet, look for quantities data in all sheets
                for sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    
                    # Search for Quantities header
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and 'quantit' in cell.value.lower():
                                quantities_sheet = sheet_name
                                break
                        if quantities_sheet:
                            break
            
            if quantities_sheet:
                # Convert sheet to dataframe
                sheet = self.workbook[quantities_sheet]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    # Try to infer header row
                    header_row = 0
                    for i, row in enumerate(data):
                        if row and any(isinstance(cell, str) and ('item' in str(cell).lower() or 'quant' in str(cell).lower()) for cell in row if cell):
                            header_row = i
                            break
                    
                    # Extract headers
                    headers = data[header_row]
                    # Clean up headers
                    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(headers)]
                    
                    # Extract data rows
                    data_rows = data[header_row+1:]
                    
                    # Convert to dataframe
                    quantities_df = pd.DataFrame(data_rows, columns=headers)
                    
                    # Clean up dataframe
                    quantities_df = quantities_df.replace('', np.nan).dropna(how='all')
                    
                    # Store in dataframes dict
                    self.dataframes['quantities'] = quantities_df
                    logger.info(f"Extracted quantities data: {quantities_df.shape[0]} rows")
                
                else:
                    logger.warning(f"No data found in quantities sheet: {quantities_sheet}")
            
            else:
                logger.warning("No quantities sheet found in workbook")
            
            return quantities_df
        
        except Exception as e:
            logger.error(f"Error extracting quantities data: {str(e)}")
            return None
    
    def extract_zero_hours_data(self):
        """Extract zero hours/timecard data from the workbook"""
        try:
            # Look for the Zero Hours sheet or section
            zero_hours_sheet = None
            zero_hours_df = None
            
            # Try to find zero hours data in various ways
            for sheet_name in self.workbook.sheetnames:
                if 'zero' in sheet_name.lower() or 'time' in sheet_name.lower():
                    zero_hours_sheet = sheet_name
                    break
            
            if not zero_hours_sheet:
                # If no specific zero hours sheet, look for zero hours data in all sheets
                for sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    
                    # Search for Zero Hours header
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and ('zero' in cell.value.lower() or 'timecard' in cell.value.lower()):
                                zero_hours_sheet = sheet_name
                                break
                        if zero_hours_sheet:
                            break
            
            if zero_hours_sheet:
                # Convert sheet to dataframe
                sheet = self.workbook[zero_hours_sheet]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    # Try to infer header row
                    header_row = 0
                    for i, row in enumerate(data):
                        if row and any(isinstance(cell, str) and ('name' in str(cell).lower() or 'zero' in str(cell).lower()) for cell in row if cell):
                            header_row = i
                            break
                    
                    # Extract headers
                    headers = data[header_row]
                    # Clean up headers
                    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(headers)]
                    
                    # Extract data rows
                    data_rows = data[header_row+1:]
                    
                    # Convert to dataframe
                    zero_hours_df = pd.DataFrame(data_rows, columns=headers)
                    
                    # Clean up dataframe
                    zero_hours_df = zero_hours_df.replace('', np.nan).dropna(how='all')
                    
                    # Store in dataframes dict
                    self.dataframes['zero_hours'] = zero_hours_df
                    logger.info(f"Extracted zero hours data: {zero_hours_df.shape[0]} rows")
                
                else:
                    logger.warning(f"No data found in zero hours sheet: {zero_hours_sheet}")
            
            else:
                logger.warning("No zero hours sheet found in workbook")
            
            return zero_hours_df
        
        except Exception as e:
            logger.error(f"Error extracting zero hours data: {str(e)}")
            return None
    
    def process_workbook(self):
        """Process the workbook and extract all relevant data"""
        try:
            if not self.workbook:
                if not self.load_workbook():
                    return False
            
            # Extract hours data
            hours_df = self.extract_hours_data()
            
            # Extract quantities data
            quantities_df = self.extract_quantities_data()
            
            # Extract zero hours data
            zero_hours_df = self.extract_zero_hours_data()
            
            return True
        
        except Exception as e:
            logger.error(f"Error processing workbook: {str(e)}")
            return False
    
    def prepare_data_for_comparison(self, date_str=None):
        """
        Prepare data for comparison with GPS records
        
        Args:
            date_str: Specific date to filter for (YYYY-MM-DD)
        
        Returns:
            DataFrame: Combined data for comparison
        """
        try:
            # Process workbook if not already done
            if not self.dataframes:
                if not self.process_workbook():
                    return None
            
            # Initialize combined dataframe
            combined_data = []
            
            # Process hours data
            if 'hours' in self.dataframes:
                hours_df = self.dataframes['hours']
                
                # Extract driver/employee information
                if hours_df is not None and not hours_df.empty:
                    # Identify columns for driver name, job, and hours
                    name_col = None
                    job_col = None
                    date_col = None
                    hours_cols = []
                    
                    for col in hours_df.columns:
                        col_lower = col.lower()
                        if 'name' in col_lower or 'employee' in col_lower or 'driver' in col_lower:
                            name_col = col
                        elif 'job' in col_lower or 'project' in col_lower:
                            job_col = col
                        elif 'date' in col_lower:
                            date_col = col
                        elif 'hour' in col_lower or 'time' in col_lower:
                            hours_cols.append(col)
                    
                    # If date column is not found, try to infer date from workbook
                    if not date_col and date_str:
                        # Use provided date
                        report_date = datetime.strptime(date_str, '%Y-%m-%d')
                    else:
                        # Use current date
                        report_date = datetime.now()
                    
                    # Process each row
                    for idx, row in hours_df.iterrows():
                        driver_name = row[name_col] if name_col and not pd.isna(row[name_col]) else f"Unknown Driver {idx}"
                        job_number = row[job_col] if job_col and not pd.isna(row[job_col]) else ""
                        
                        # Format driver name (Last, First -> First Last)
                        if driver_name and isinstance(driver_name, str) and ',' in driver_name:
                            last, first = driver_name.split(',', 1)
                            driver_name = f"{first.strip()} {last.strip()}"
                        
                        # Extract date
                        if date_col and not pd.isna(row[date_col]):
                            try:
                                if isinstance(row[date_col], datetime):
                                    entry_date = row[date_col]
                                else:
                                    # Try to parse date string
                                    entry_date = pd.to_datetime(row[date_col])
                            except:
                                # Use report date as fallback
                                entry_date = report_date
                        else:
                            entry_date = report_date
                        
                        # Extract hours
                        total_hours = 0
                        for hours_col in hours_cols:
                            if not pd.isna(row[hours_col]):
                                try:
                                    hours_val = float(row[hours_col])
                                    total_hours += hours_val
                                except:
                                    pass
                        
                        # Skip rows with zero hours unless specifically keeping them
                        if total_hours == 0:
                            continue
                        
                        # Create record
                        record = {
                            'driver_name': driver_name,
                            'job_number': job_number,
                            'date': entry_date,
                            'reported_hours': total_hours,
                            'source': 'hours',
                            'status': 'active'
                        }
                        
                        combined_data.append(record)
            
            # Process zero hours data (absences, PTO, etc.)
            if 'zero_hours' in self.dataframes:
                zero_df = self.dataframes['zero_hours']
                
                # Extract driver/employee information
                if zero_df is not None and not zero_df.empty:
                    # Identify columns for driver name, job, and reason
                    name_col = None
                    job_col = None
                    date_col = None
                    reason_col = None
                    
                    for col in zero_df.columns:
                        col_lower = col.lower()
                        if 'name' in col_lower or 'employee' in col_lower or 'driver' in col_lower:
                            name_col = col
                        elif 'job' in col_lower or 'project' in col_lower:
                            job_col = col
                        elif 'date' in col_lower:
                            date_col = col
                        elif 'reason' in col_lower or 'status' in col_lower or 'comment' in col_lower:
                            reason_col = col
                    
                    # If date column is not found, try to infer date from workbook
                    if not date_col and date_str:
                        # Use provided date
                        report_date = datetime.strptime(date_str, '%Y-%m-%d')
                    else:
                        # Use current date
                        report_date = datetime.now()
                    
                    # Process each row
                    for idx, row in zero_df.iterrows():
                        driver_name = row[name_col] if name_col and not pd.isna(row[name_col]) else f"Unknown Driver {idx}"
                        job_number = row[job_col] if job_col and not pd.isna(row[job_col]) else ""
                        reason = row[reason_col] if reason_col and not pd.isna(row[reason_col]) else "No reason provided"
                        
                        # Format driver name (Last, First -> First Last)
                        if driver_name and isinstance(driver_name, str) and ',' in driver_name:
                            last, first = driver_name.split(',', 1)
                            driver_name = f"{first.strip()} {last.strip()}"
                        
                        # Extract date
                        if date_col and not pd.isna(row[date_col]):
                            try:
                                if isinstance(row[date_col], datetime):
                                    entry_date = row[date_col]
                                else:
                                    # Try to parse date string
                                    entry_date = pd.to_datetime(row[date_col])
                            except:
                                # Use report date as fallback
                                entry_date = report_date
                        else:
                            entry_date = report_date
                        
                        # Create record
                        record = {
                            'driver_name': driver_name,
                            'job_number': job_number,
                            'date': entry_date,
                            'reported_hours': 0,
                            'source': 'timecard_absence',
                            'status': 'absent',
                            'reason': reason
                        }
                        
                        combined_data.append(record)
            
            # Convert to DataFrame
            combined_df = pd.DataFrame(combined_data)
            
            # Filter by date if specified
            if combined_df is not None and not combined_df.empty and date_str:
                try:
                    filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    combined_df = combined_df[combined_df['date'].dt.date == filter_date]
                except Exception as e:
                    logger.error(f"Error filtering by date: {str(e)}")
            
            logger.info(f"Prepared {len(combined_data)} records for comparison")
            return combined_df
        
        except Exception as e:
            logger.error(f"Error preparing data for comparison: {str(e)}")
            return None


def process_dynamic_timecard(file_path, start_date=None, end_date=None):
    """
    Process a dynamic timecard Excel file with formulas
    
    Args:
        file_path: Path to Excel file
        start_date: Start date in YYYY-MM-DD format
        end_date: End date in YYYY-MM-DD format
    
    Returns:
        DataFrame: Processed timecard data
    """
    try:
        logger.info(f"Processing dynamic timecard: {file_path}")
        
        # Initialize processor
        processor = DynamicTimecardProcessor(file_path)
        
        # Process workbook
        if not processor.process_workbook():
            logger.error("Failed to process workbook")
            return None
        
        # Calculate date range
        if start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            date_range = [start + timedelta(days=x) for x in range((end - start).days + 1)]
        else:
            # Default to current date
            date_range = [datetime.now()]
        
        # Process each date
        all_data = []
        for date in date_range:
            date_str = date.strftime('%Y-%m-%d')
            logger.info(f"Processing timecard data for date: {date_str}")
            
            # Prepare data for specific date
            date_data = processor.prepare_data_for_comparison(date_str)
            
            if date_data is not None and not date_data.empty:
                all_data.append(date_data)
        
        # Combine all data
        if all_data:
            combined_df = pd.concat(all_data, ignore_index=True)
            logger.info(f"Processed timecard data: {combined_df.shape[0]} total records")
            return combined_df
        else:
            logger.warning("No timecard data found for specified date range")
            return None
    
    except Exception as e:
        logger.error(f"Error processing dynamic timecard: {str(e)}")
        return None


def compare_dynamic_timecards_with_gps(timecard_file, gps_data, target_date=None):
    """
    Compare dynamic timecard data with GPS records
    
    Args:
        timecard_file: Path to timecard Excel file
        gps_data: Dictionary of GPS data by date
        target_date: Specific date to compare (YYYY-MM-DD)
    
    Returns:
        list: Comparison results
    """
    try:
        logger.info(f"Comparing dynamic timecard with GPS data for date: {target_date}")
        
        # Process timecard data
        timecard_data = process_dynamic_timecard(
            timecard_file, 
            start_date=target_date, 
            end_date=target_date
        )
        
        if timecard_data is None or timecard_data.empty:
            logger.warning(f"No timecard data found for date: {target_date}")
            return []
        
        # Get GPS data for target date
        date_gps_data = gps_data.get(target_date, {})
        if not date_gps_data:
            logger.warning(f"No GPS data found for date: {target_date}")
            return []
        
        # Initialize comparison results
        comparisons = []
        
        # Extract driver records from GPS data
        gps_drivers = {}
        if 'driver_records' in date_gps_data:
            for record in date_gps_data['driver_records']:
                driver_name = record.get('driver_name', '')
                if driver_name:
                    # Format key for matching
                    key = driver_name.lower()
                    
                    # Store record
                    gps_drivers[key] = {
                        'driver_name': driver_name,
                        'classification': record.get('classification', ''),
                        'job_number': record.get('job_number', ''),
                        'job_name': record.get('job_name', ''),
                        'start_time': record.get('start_time', ''),
                        'end_time': record.get('end_time', ''),
                        'gps_hours': record.get('hours', 0),
                        'late_minutes': record.get('late_minutes', 0),
                        'early_end_minutes': record.get('early_end_minutes', 0)
                    }
        
        # Process each timecard record
        for _, row in timecard_data.iterrows():
            driver_name = row['driver_name']
            job_number = row['job_number']
            reported_hours = row['reported_hours']
            status = row['status']
            
            # Format key for matching
            key = driver_name.lower()
            
            # Find matching GPS record
            gps_record = gps_drivers.get(key, None)
            
            if gps_record:
                # Driver found in GPS data
                gps_hours = gps_record['gps_hours']
                classification = gps_record['classification']
                gps_job = gps_record['job_number'] or gps_record['job_name']
                
                # Calculate discrepancies
                hours_diff = reported_hours - gps_hours if reported_hours and gps_hours else 0
                job_mismatch = job_number and gps_job and job_number != gps_job
                
                # Determine issues
                issues = []
                
                if status == 'absent' and classification != 'not_found':
                    issues.append("Driver marked absent but GPS activity detected")
                
                if status != 'absent':
                    if classification == 'not_found':
                        issues.append("Driver on timecard but no GPS activity")
                    elif classification == 'not_on_job':
                        issues.append("Driver not at assigned job site")
                    elif classification == 'late':
                        issues.append(f"Driver arrived late ({gps_record['late_minutes']} minutes)")
                    elif classification == 'early_end':
                        issues.append(f"Driver left early ({gps_record['early_end_minutes']} minutes)")
                
                if abs(hours_diff) > 0.5:  # More than 30 minutes difference
                    issues.append(f"Hours discrepancy: {abs(hours_diff):.1f} hours")
                
                if job_mismatch:
                    issues.append(f"Job mismatch: Timecard {job_number} vs GPS {gps_job}")
                
                # Create comparison record
                comparison = {
                    'driver_name': driver_name,
                    'date': target_date,
                    'timecard_job': job_number,
                    'gps_job': gps_job,
                    'timecard_hours': reported_hours,
                    'gps_hours': gps_hours,
                    'hours_diff': hours_diff,
                    'gps_start': gps_record['start_time'],
                    'gps_end': gps_record['end_time'],
                    'classification': classification,
                    'timecard_status': status,
                    'has_discrepancy': len(issues) > 0,
                    'issues': issues,
                    'total_variance': abs(hours_diff) * 60  # Convert to minutes
                }
                
            else:
                # Driver not found in GPS data
                if status == 'absent':
                    # Correctly marked as absent
                    issues = []
                else:
                    # On timecard but no GPS data
                    issues = ["Driver on timecard but no GPS activity"]
                
                # Create comparison record
                comparison = {
                    'driver_name': driver_name,
                    'date': target_date,
                    'timecard_job': job_number,
                    'gps_job': '',
                    'timecard_hours': reported_hours,
                    'gps_hours': 0,
                    'hours_diff': reported_hours,
                    'gps_start': '',
                    'gps_end': '',
                    'classification': 'not_found',
                    'timecard_status': status,
                    'has_discrepancy': len(issues) > 0,
                    'issues': issues,
                    'total_variance': reported_hours * 60  # Convert to minutes
                }
            
            comparisons.append(comparison)
        
        # Check for drivers in GPS data but not on timecard
        for key, gps_record in gps_drivers.items():
            # Format key for matching
            driver_name = gps_record['driver_name']
            driver_found = any(comp['driver_name'].lower() == driver_name.lower() for comp in comparisons)
            
            if not driver_found:
                # Driver in GPS data but not on timecard
                issues = ["Driver has GPS activity but not on timecard"]
                
                # Create comparison record
                comparison = {
                    'driver_name': driver_name,
                    'date': target_date,
                    'timecard_job': '',
                    'gps_job': gps_record['job_number'] or gps_record['job_name'],
                    'timecard_hours': 0,
                    'gps_hours': gps_record['gps_hours'],
                    'hours_diff': -gps_record['gps_hours'],
                    'gps_start': gps_record['start_time'],
                    'gps_end': gps_record['end_time'],
                    'classification': gps_record['classification'],
                    'timecard_status': 'missing',
                    'has_discrepancy': True,
                    'issues': issues,
                    'total_variance': gps_record['gps_hours'] * 60  # Convert to minutes
                }
                
                comparisons.append(comparison)
        
        logger.info(f"Comparison completed: {len(comparisons)} records")
        return comparisons
    
    except Exception as e:
        logger.error(f"Error comparing timecard with GPS data: {str(e)}")
        return []
"""
Export utility functions for report generation
"""
import os
import csv
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Set up logging
logger = logging.getLogger(__name__)

# Define exports folder
EXPORTS_FOLDER = 'exports'

def ensure_exports_folder():
    """Ensure the exports folder exists"""
    if not os.path.exists(EXPORTS_FOLDER):
        os.makedirs(EXPORTS_FOLDER)
    return EXPORTS_FOLDER

def generate_unique_filename(report_type, report_date, region_id=None, format='xlsx'):
    """Generate a unique filename for an export"""
    # Format dates
    formatted_date = report_date.strftime('%m_%d_%Y')
    timestamp = datetime.now().strftime('%H%M%S')
    
    # Build filename
    filename_base = f"{report_type}_report_{formatted_date}"
    if region_id:
        filename_base += f"_region_{region_id}"
    
    # Add timestamp for uniqueness
    filename = f"{filename_base}_{timestamp}.{format}"
    return filename

def export_to_csv(file_path, records, fieldnames):
    """Export records to CSV file"""
    with open(file_path, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for record in records:
            processed_record = {}
            for key in fieldnames:
                value = record.get(key)
                # Format date/time values
                if isinstance(value, datetime):
                    processed_record[key] = value.strftime('%Y-%m-%d')
                elif hasattr(value, 'strftime'):  # Time objects
                    processed_record[key] = value.strftime('%H:%M')
                else:
                    processed_record[key] = value
            writer.writerow(processed_record)
    
    return file_path

def export_to_excel(file_path, records, headers, sheet_title="Report"):
    """Export records to Excel file with formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Add header row with styling
    for col, header in enumerate(headers.values(), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    for row_idx, record in enumerate(records, 2):
        col_idx = 1
        
        for field in headers.keys():
            value = record.get(field)
            
            # Format date/time values
            if isinstance(value, datetime):
                cell_value = value.strftime('%Y-%m-%d')
            elif hasattr(value, 'strftime'):  # Time objects
                cell_value = value.strftime('%H:%M')
            else:
                cell_value = value
                
            # Write cell value
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply conditional formatting for status
            if field == 'status':
                status_cell = ws.cell(row=row_idx, column=col_idx)
                if value == 'Late Start':
                    status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif value == 'Early End':
                    status_cell.fill = PatternFill(start_color="FFEECC", end_color="FFEECC", fill_type="solid")
                elif value == 'Not on Job':
                    status_cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
                elif value == 'On Time':
                    status_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            col_idx += 1
    
    # Auto-size columns for better readability
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Save the Excel file
    wb.save(file_path)
    
    return file_path
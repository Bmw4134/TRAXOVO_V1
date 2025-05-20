"""
Export Utilities for Daily Reports

This module provides utility functions for exporting daily reports in various formats,
including Excel, CSV, and PDF. It ensures reports have consistent formatting and 
include all required information for analysis.
"""

import os
import logging
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Configure logging
logger = logging.getLogger(__name__)

def format_contact_info(contact_info):
    """
    Format contact information to ensure consistent display
    
    Args:
        contact_info (str): Raw contact information
        
    Returns:
        str: Formatted contact information
    """
    if not contact_info:
        return ""
        
    # Remove unnecessary placeholders
    contact_info = str(contact_info)
    contact_info = contact_info.replace('N/A', '').replace('None', '').replace('none', '')
    contact_info = contact_info.replace('null', '').replace('NULL', '')
    contact_info = contact_info.replace('-', '').replace('_', '')
    
    # Clean up formatting
    contact_info = contact_info.strip()
    if not contact_info:
        return ""
    
    # Format phone numbers consistently
    import re
    phone_pattern = re.compile(r'(\d{3})[^\d]*(\d{3})[^\d]*(\d{4})')
    
    def phone_format(match):
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        
    contact_info = phone_pattern.sub(phone_format, contact_info)
    
    return contact_info

def apply_contact_formatting(sheet, row_idx, contact_col, email_col, contact_value, email_value):
    """
    Apply contact formatting to the Excel sheet
    
    Args:
        sheet: Excel worksheet
        row_idx (int): Row index
        contact_col (int): Contact column index
        email_col (int): Email column index
        contact_value (str): Contact value
        email_value (str): Email value
    """
    # Format and set contact information
    formatted_contact = format_contact_info(contact_value)
    sheet.cell(row=row_idx, column=contact_col).value = formatted_contact
    sheet.cell(row=row_idx, column=email_col).value = email_value if email_value else ""
    
    # Apply styling
    contact_cell = sheet.cell(row=row_idx, column=contact_col)
    email_cell = sheet.cell(row=row_idx, column=email_col)
    
    contact_cell.alignment = Alignment(horizontal='left')
    email_cell.alignment = Alignment(horizontal='left')
    
    # Set border styling for better readability
    border_style = Side(border_style="thin", color="000000")
    border = Border(
        left=border_style,
        right=border_style,
        top=border_style,
        bottom=border_style
    )
    
    contact_cell.border = border
    email_cell.border = border

def export_daily_report(date_str, output_path, attendance_data=None, employee_data=None, job_data=None):
    """
    Export daily report to Excel with enriched employee and job information
    
    Args:
        date_str (str): Date string in YYYY-MM-DD format
        output_path (str or Path): Path to save the Excel file
        attendance_data (dict): Attendance data to export
        employee_data (DataFrame): Employee data for enrichment
        job_data (DataFrame): Job data for enrichment
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # If attendance_data not provided, get it from the pipeline
        if attendance_data is None:
            from utils.attendance_pipeline_connector import get_attendance_data
            attendance_data = get_attendance_data(date_str)
            
        if not attendance_data:
            logger.error(f"No attendance data available for {date_str}")
            return False
            
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create sheets
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        late_sheet = wb.create_sheet("Late Arrivals")
        early_sheet = wb.create_sheet("Early Departures")
        missing_sheet = wb.create_sheet("Not On Job")
        all_drivers_sheet = wb.create_sheet("All Drivers")
        
        # Format the summary sheet
        format_summary_sheet(summary_sheet, date_str, attendance_data)
        
        # Format the issue sheets
        if 'late_start_records' in attendance_data:
            format_late_sheet(late_sheet, attendance_data['late_start_records'], employee_data, job_data)
        else:
            # Add default headers to empty sheet
            headers = ['Driver', 'Asset ID', 'Scheduled Start', 'Actual Start', 
                      'Minutes Late', 'Job Site', 'Division', 'Contact Info', 'Email']
            for col, header in enumerate(headers, start=1):
                cell = late_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                late_sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
                
        # Add nice formatting to all sheets
        border_style = Side(border_style="thin", color="000000")
        border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
                
        # Apply consistent formatting to all sheets
        for sheet in wb.worksheets:
            # Skip summary sheet
            if sheet.title == "Summary":
                continue
                
            # Add borders to all cells
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
        
        if 'early_end_records' in attendance_data:
            format_early_sheet(early_sheet, attendance_data['early_end_records'], employee_data, job_data)
        else:
            # Add default headers to empty sheet
            headers = ['Driver', 'Asset ID', 'Scheduled End', 'Actual End', 
                      'Minutes Early', 'Job Site', 'Division', 'Contact Info', 'Email']
            for col, header in enumerate(headers, start=1):
                cell = early_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                early_sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
            
        if 'not_on_job_records' in attendance_data:
            format_missing_sheet(missing_sheet, attendance_data['not_on_job_records'], employee_data, job_data)
        else:
            # Add default headers to empty sheet
            headers = ['Driver', 'Asset ID', 'Scheduled Job', 'Actual Job', 
                      'Region', 'Division', 'Contact Info', 'Email']
            for col, header in enumerate(headers, start=1):
                cell = missing_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                missing_sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
            
        # We don't have an 'all_drivers' list, we'll create one from the original records
        if attendance_data.get('total_drivers', 0) > 0 and 'late_start_records' in attendance_data:
            # Combine all records to create a complete drivers list
            combined_drivers = attendance_data.get('late_start_records', []).copy()
            # Add early end drivers that aren't already in the list
            for driver in attendance_data.get('early_end_records', []):
                if not any(d.get('driver_name') == driver.get('driver_name') for d in combined_drivers):
                    combined_drivers.append(driver)
            # Add missing drivers that aren't already in the list
            for driver in attendance_data.get('not_on_job_records', []):
                if not any(d.get('driver_name') == driver.get('driver_name') for d in combined_drivers):
                    combined_drivers.append(driver)
                    
            format_all_drivers_sheet(all_drivers_sheet, combined_drivers, employee_data, job_data)
        else:
            # Add default headers to empty sheet
            headers = ['Driver', 'Asset ID', 'Start Time', 'End Time', 
                      'Total Hours', 'Job Site', 'Division', 'Contact Info', 'Email']
            for col, header in enumerate(headers, start=1):
                cell = all_drivers_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                all_drivers_sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
        
        # Save the workbook
        wb.save(output_path)
        logger.info(f"Successfully exported daily report to {output_path}")
        
        # Create PDF version as well
        pdf_output_path = str(output_path).replace('.xlsx', '.pdf')
        try:
            export_pdf_report(date_str, pdf_output_path, attendance_data, employee_data, job_data)
            logger.info(f"Successfully created PDF report at {pdf_output_path}")
        except Exception as e:
            logger.error(f"Error creating PDF report: {e}")
            
        return True
        
    except Exception as e:
        logger.error(f"Error exporting daily report: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return False

def format_summary_sheet(sheet, date_str, attendance_data):
    """Format the summary sheet with attendance statistics"""
    # Add title
    sheet.merge_cells('A1:G1')
    sheet['A1'] = f"Daily Attendance Report - {date_str}"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Format column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    
    # Add statistics
    row = 3
    sheet.cell(row=row, column=1).value = "Total Drivers"
    sheet.cell(row=row, column=2).value = attendance_data.get('total_drivers', 0)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    sheet.cell(row=row, column=1).value = "Late Arrivals"
    sheet.cell(row=row, column=2).value = attendance_data.get('late_count', 0)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    sheet.cell(row=row, column=1).value = "Early Departures"
    sheet.cell(row=row, column=2).value = attendance_data.get('early_count', 0)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    sheet.cell(row=row, column=1).value = "Not On Job"
    sheet.cell(row=row, column=2).value = attendance_data.get('missing_count', 0)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    # Add file information
    row += 2
    sheet.cell(row=row, column=1).value = "Report Sources:"
    sheet.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    sheet.cell(row=row, column=1).value = "Activity File"
    activity_file = attendance_data.get('sources', {}).get('activity_file', 'N/A')
    sheet.cell(row=row, column=2).value = activity_file if activity_file else 'N/A'
    
    row += 1
    sheet.cell(row=row, column=1).value = "Driving History File"
    driving_file = attendance_data.get('sources', {}).get('driving_file', 'N/A')
    sheet.cell(row=row, column=2).value = driving_file if driving_file else 'N/A'
    
    row += 1
    sheet.cell(row=row, column=1).value = "Fleet Utilization File"
    utilization_file = attendance_data.get('sources', {}).get('utilization_file', 'N/A')
    sheet.cell(row=row, column=2).value = utilization_file if utilization_file else 'N/A'
    
    # Add generation timestamp
    row += 2
    sheet.cell(row=row, column=1).value = "Generated"
    sheet.cell(row=row, column=2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def clean_placeholder_value(value):
    """Clean up placeholder values like '-' or 'N/A'"""
    if pd.isna(value) or value is None:
        return ''
    if isinstance(value, str) and value.strip() in ['-', 'N/A', 'NA', '']:
        return ''
    return value

def format_late_sheet(sheet, late_drivers, employee_data, job_data):
    """Format the late arrivals sheet"""
    # Add headers
    headers = ['Driver', 'Asset ID', 'Scheduled Start', 'Actual Start', 
              'Minutes Late', 'Job Site', 'Division', 'Contact Info', 'Email']
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width
        sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
    
    # Add data rows
    for i, driver in enumerate(late_drivers, start=2):
        sheet.cell(row=i, column=1).value = driver.get('driver_name', 'Unknown')
        sheet.cell(row=i, column=2).value = driver.get('asset_id', '')
        sheet.cell(row=i, column=3).value = driver.get('scheduled_start', '')
        sheet.cell(row=i, column=4).value = driver.get('actual_start', '')
        sheet.cell(row=i, column=5).value = driver.get('minutes_late', 0)
        sheet.cell(row=i, column=6).value = driver.get('job_site', '')
        
        # Add division from job data if available
        job_site = driver.get('job_site', '')
        division = ''
        if job_data is not None and job_site:
            job_match = job_data[job_data['Job Number'] == job_site]
            if not job_match.empty:
                division = job_match.iloc[0].get('Division', '')
        sheet.cell(row=i, column=7).value = division
        
        # Add contact info from employee data if available
        driver_name = driver.get('driver_name', '')
        contact_info = ''
        email = ''
        
        if employee_data is not None and driver_name:
            # Try exact match first on full name
            employee_match = employee_data[employee_data['Employee Name'] == driver_name]
            
            # If no exact match, try contains match on full name
            if employee_match.empty:
                employee_match = employee_data[employee_data['Employee Name'].str.contains(driver_name, na=False, case=False)]
            
            # Try matching on last name if still no match
            if employee_match.empty:
                # Extract last name from driver name (assume last part is last name)
                name_parts = driver_name.split()
                if len(name_parts) > 1:
                    last_name = name_parts[-1]
                    employee_match = employee_data[employee_data['Last Name'].str.contains(last_name, na=False, case=False)]
            
            if not employee_match.empty:
                # Use cell phone if available, otherwise use regular phone
                cell_phone = employee_match.iloc[0].get('Cell Phone', '')
                regular_phone = employee_match.iloc[0].get('Phone', '')
                contact_info = cell_phone if cell_phone and pd.notna(cell_phone) else regular_phone
                
                # Get email if available
                email = employee_match.iloc[0].get('E-Mail', '')
                
                # Log the matched employee for verification
                logger.debug(f"Matched driver {driver_name} with employee {employee_match.iloc[0]['Employee Name']}")
            else:
                # Log unmatched driver for troubleshooting
                logger.warning(f"No employee match found for driver: {driver_name}")
        
        # Add contact info to the sheet
        sheet.cell(row=i, column=8).value = contact_info
        sheet.cell(row=i, column=9).value = email

def format_early_sheet(sheet, early_drivers, employee_data, job_data):
    """Format the early departures sheet"""
    # Add headers
    headers = ['Driver', 'Asset ID', 'Scheduled End', 'Actual End', 
              'Minutes Early', 'Job Site', 'Division', 'Contact Info', 'Email']
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width
        sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
    
    # Add data rows
    for i, driver in enumerate(early_drivers, start=2):
        sheet.cell(row=i, column=1).value = driver.get('driver_name', 'Unknown')
        sheet.cell(row=i, column=2).value = driver.get('asset_id', '')
        sheet.cell(row=i, column=3).value = driver.get('scheduled_end', '')
        sheet.cell(row=i, column=4).value = driver.get('actual_end', '')
        sheet.cell(row=i, column=5).value = driver.get('minutes_early', 0)
        sheet.cell(row=i, column=6).value = driver.get('job_site', '')
        
        # Add division from job data if available
        job_site = driver.get('job_site', '')
        division = ''
        if job_data is not None and job_site:
            job_match = job_data[job_data['Job Number'] == job_site]
            if not job_match.empty:
                division = job_match.iloc[0].get('Division', '')
        sheet.cell(row=i, column=7).value = division
        
        # Add contact info from employee data if available
        driver_name = driver.get('driver_name', '')
        contact_info = ''
        email = ''
        
        if employee_data is not None and driver_name:
            # Try exact match first on full name
            employee_match = employee_data[employee_data['Employee Name'] == driver_name]
            
            # If no exact match, try contains match on full name
            if employee_match.empty:
                employee_match = employee_data[employee_data['Employee Name'].str.contains(driver_name, na=False, case=False)]
            
            # Try matching on last name if still no match
            if employee_match.empty:
                # Extract last name from driver name (assume last part is last name)
                name_parts = driver_name.split()
                if len(name_parts) > 1:
                    last_name = name_parts[-1]
                    employee_match = employee_data[employee_data['Last Name'].str.contains(last_name, na=False, case=False)]
            
            if not employee_match.empty:
                # Use cell phone if available, otherwise use regular phone
                cell_phone = employee_match.iloc[0].get('Cell Phone', '')
                regular_phone = employee_match.iloc[0].get('Phone', '')
                contact_info = cell_phone if cell_phone and pd.notna(cell_phone) else regular_phone
                
                # Get email if available
                email = employee_match.iloc[0].get('E-Mail', '')
                
                # Log the matched employee for verification
                logger.debug(f"Matched driver {driver_name} with employee {employee_match.iloc[0]['Employee Name']}")
            else:
                # Log unmatched driver for troubleshooting
                logger.warning(f"No employee match found for driver: {driver_name}")
        
        # Add contact info to the sheet
        sheet.cell(row=i, column=8).value = contact_info
        sheet.cell(row=i, column=9).value = email

def format_missing_sheet(sheet, missing_drivers, employee_data, job_data):
    """Format the not on job sheet"""
    # Add headers
    headers = ['Driver', 'Asset ID', 'Scheduled Job', 'Actual Job', 
              'Region', 'Division', 'Contact Info', 'Email']
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width
        sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
    
    # Add data rows
    for i, driver in enumerate(missing_drivers, start=2):
        sheet.cell(row=i, column=1).value = driver.get('driver_name', 'Unknown')
        sheet.cell(row=i, column=2).value = driver.get('asset_id', '')
        sheet.cell(row=i, column=3).value = driver.get('assigned_job', '')
        sheet.cell(row=i, column=4).value = driver.get('actual_job', 'Not Found')
        
        # Add region and division from job data if available
        assigned_job = driver.get('assigned_job', '')
        region = ''
        division = ''
        if job_data is not None and assigned_job:
            job_match = job_data[job_data['Job Number'] == assigned_job]
            if not job_match.empty:
                region = job_match.iloc[0].get('Region', '')
                division = job_match.iloc[0].get('Division', '')
        
        sheet.cell(row=i, column=5).value = region
        sheet.cell(row=i, column=6).value = division
        
        # Add contact info from employee data if available
        driver_name = driver.get('driver_name', '')
        contact_info = ''
        email = ''
        
        if employee_data is not None and driver_name:
            # Try exact match first on full name
            employee_match = employee_data[employee_data['Employee Name'] == driver_name]
            
            # If no exact match, try contains match on full name
            if employee_match.empty:
                employee_match = employee_data[employee_data['Employee Name'].str.contains(driver_name, na=False, case=False)]
            
            # Try matching on last name if still no match
            if employee_match.empty:
                # Extract last name from driver name (assume last part is last name)
                name_parts = driver_name.split()
                if len(name_parts) > 1:
                    last_name = name_parts[-1]
                    employee_match = employee_data[employee_data['Last Name'].str.contains(last_name, na=False, case=False)]
            
            if not employee_match.empty:
                # Use cell phone if available, otherwise use regular phone
                cell_phone = employee_match.iloc[0].get('Cell Phone', '')
                regular_phone = employee_match.iloc[0].get('Phone', '')
                contact_info = cell_phone if cell_phone and pd.notna(cell_phone) else regular_phone
                
                # Get email if available
                email = employee_match.iloc[0].get('E-Mail', '')
                
                # Log the matched employee for verification
                logger.debug(f"Matched driver {driver_name} with employee {employee_match.iloc[0]['Employee Name']}")
            else:
                # Log unmatched driver for troubleshooting
                logger.warning(f"No employee match found for driver: {driver_name}")
        
        # Add contact info to the sheet
        sheet.cell(row=i, column=7).value = contact_info
        sheet.cell(row=i, column=8).value = email

def format_all_drivers_sheet(sheet, all_drivers, employee_data, job_data):
    """Format the all drivers sheet"""
    # Add headers
    headers = ['Driver', 'Asset ID', 'Start Time', 'End Time', 
              'Total Hours', 'Job Site', 'Division', 'Contact Info', 'Email']
    
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width
        sheet.column_dimensions[get_column_letter(col)].width = 20 if header == 'Email' else 15
    
    # Add data rows
    for i, driver in enumerate(all_drivers, start=2):
        sheet.cell(row=i, column=1).value = driver.get('driver_name', 'Unknown')
        sheet.cell(row=i, column=2).value = driver.get('asset_id', '')
        sheet.cell(row=i, column=3).value = driver.get('start_time', '')
        sheet.cell(row=i, column=4).value = driver.get('end_time', '')
        
        # Calculate total hours if both start and end times are available
        total_hours = ''
        start_time = driver.get('start_time', '')
        end_time = driver.get('end_time', '')
        if start_time and end_time:
            try:
                start_dt = datetime.strptime(start_time, '%H:%M')
                end_dt = datetime.strptime(end_time, '%H:%M')
                # Handle next-day end times
                if end_dt < start_dt:
                    end_dt = end_dt.replace(day=start_dt.day + 1)
                hours = (end_dt - start_dt).total_seconds() / 3600
                total_hours = f"{hours:.2f}"
            except Exception:
                total_hours = 'Error'
                
        sheet.cell(row=i, column=5).value = total_hours
        sheet.cell(row=i, column=6).value = driver.get('job_site', '')
        
        # Add division from job data if available
        job_site = driver.get('job_site', '')
        division = ''
        if job_data is not None and job_site:
            job_match = job_data[job_data['Job Number'] == job_site]
            if not job_match.empty:
                division = job_match.iloc[0].get('Division', '')
        sheet.cell(row=i, column=7).value = division
        
        # Add contact info from employee data if available
        driver_name = driver.get('driver_name', '')
        contact_info = ''
        email = ''
        
        if employee_data is not None and driver_name:
            # Try exact match first on full name
            employee_match = employee_data[employee_data['Employee Name'] == driver_name]
            
            # If no exact match, try contains match on full name
            if employee_match.empty:
                employee_match = employee_data[employee_data['Employee Name'].str.contains(driver_name, na=False, case=False)]
            
            # Try matching on last name if still no match
            if employee_match.empty:
                # Extract last name from driver name (assume last part is last name)
                name_parts = driver_name.split()
                if len(name_parts) > 1:
                    last_name = name_parts[-1]
                    employee_match = employee_data[employee_data['Last Name'].str.contains(last_name, na=False, case=False)]
            
            if not employee_match.empty:
                # Use cell phone if available, otherwise use regular phone
                cell_phone = employee_match.iloc[0].get('Cell Phone', '')
                regular_phone = employee_match.iloc[0].get('Phone', '')
                contact_info = cell_phone if cell_phone and pd.notna(cell_phone) else regular_phone
                
                # Get email if available
                email = employee_match.iloc[0].get('E-Mail', '')
                
                # Log the matched employee for verification
                logger.debug(f"Matched driver {driver_name} with employee {employee_match.iloc[0]['Employee Name']}")
            else:
                # Log unmatched driver for troubleshooting
                logger.warning(f"No employee match found for driver: {driver_name}")
        
        # Add contact info to the sheet
        sheet.cell(row=i, column=8).value = contact_info
        sheet.cell(row=i, column=9).value = email

def export_daily_report_pdf(date_str, output_path, attendance_data=None):
    """Export daily report to PDF format"""
    try:
        # If attendance_data not provided, get it from the pipeline
        if attendance_data is None:
            from utils.attendance_pipeline_connector import get_attendance_data
            attendance_data = get_attendance_data(date_str)
            
        if not attendance_data:
            logger.error(f"No attendance data available for {date_str}")
            return False
            
        # Create PDF document
        doc = SimpleDocTemplate(
            output_path, 
            pagesize=landscape(letter),
            rightMargin=30, 
            leftMargin=30, 
            topMargin=30, 
            bottomMargin=30
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Content elements
        elements = []
        
        # Title
        title = Paragraph(f"Daily Attendance Report - {date_str}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Summary section
        elements.append(Paragraph("Summary", heading_style))
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Drivers", str(attendance_data.get('total_drivers', 0))],
            ["Late Arrivals", str(attendance_data.get('late_count', 0))],
            ["Early Departures", str(attendance_data.get('early_count', 0))],
            ["Not On Job", str(attendance_data.get('missing_count', 0))]
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Late arrivals section
        if 'late_drivers' in attendance_data and attendance_data['late_drivers']:
            elements.append(Paragraph("Late Arrivals", heading_style))
            
            late_data = [["Driver", "Asset ID", "Scheduled Start", "Actual Start", "Minutes Late", "Job Site"]]
            
            for driver in attendance_data['late_drivers']:
                late_data.append([
                    driver.get('driver_name', 'Unknown'),
                    driver.get('asset_id', ''),
                    driver.get('scheduled_start', ''),
                    driver.get('actual_start', ''),
                    str(driver.get('minutes_late', 0)),
                    driver.get('job_site', '')
                ])
            
            late_table = Table(late_data, colWidths=[100, 70, 90, 90, 70, 70])
            late_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(late_table)
            elements.append(Spacer(1, 20))
            
        # File information
        elements.append(Paragraph("Report Sources", heading_style))
        
        source_data = [
            ["Source", "File"],
            ["Activity File", attendance_data.get('sources', {}).get('activity_file', 'N/A')],
            ["Driving History File", attendance_data.get('sources', {}).get('driving_file', 'N/A')],
            ["Fleet Utilization File", attendance_data.get('sources', {}).get('utilization_file', 'N/A')]
        ]
        
        source_table = Table(source_data, colWidths=[200, 300])
        source_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(source_table)
        
        # Build the document
        doc.build(elements)
        logger.info(f"Successfully exported daily report to PDF: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error exporting daily report to PDF: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return False
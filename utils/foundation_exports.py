"""
Foundation Exports Generator

This module provides utilities for generating regional foundation imports based on 
the master PM sheet and EQ billings data.
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime
from pathlib import Path
import zipfile
import shutil

logger = logging.getLogger(__name__)

# Define paths
BASE_DIR = Path('.')
EXPORTS_DIR = BASE_DIR / 'exports'
EXPORTS_DIR.mkdir(exist_ok=True)

# Temp directory for creating files before zipping
TEMP_DIR = EXPORTS_DIR / 'temp'

# Regions
REGIONS = ['DFW', 'HOU', 'WT']  # DFW = Dallas-Fort Worth, HOU = Houston, WT = West Texas

class FoundationExportException(Exception):
    """Custom exception for foundation export processing errors"""
    pass

def load_eq_billing_master(file_path):
    """
    Load the EQ Billing master sheet
    
    Args:
        file_path (str or Path): Path to the EQ Billing master file
        
    Returns:
        pandas.DataFrame: Cleaned and standardized billing data
    """
    try:
        # Verify file exists
        if isinstance(file_path, str):
            file_path = Path(file_path)
            
        if not file_path.exists():
            raise FoundationExportException(f"File not found: {file_path}")
            
        # Load the Excel file
        try:
            # Try to determine the sheet with billing data
            sheet_options = ['M-RAGLE', 'M-SELECT', 'Master', 'Sheet1']
            
            for sheet in sheet_options:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
                    # Check if this looks like billing data
                    required_columns = ['JOB', 'JOB NAME', 'EQUIPMENT', 'DESCRIPTION']
                    
                    # Check column header rows - sometimes data starts a few rows down
                    if not any(col in df.columns for col in required_columns):
                        # Try checking first few rows for headers
                        for i in range(5):
                            if i < len(df):
                                row = df.iloc[i]
                                if any(col in row.values for col in required_columns):
                                    # Use this row as header
                                    df.columns = df.iloc[i]
                                    df = df.iloc[i+1:].reset_index(drop=True)
                                    break
                    
                    # If we found what looks like billing data, use it
                    if any(col in df.columns for col in required_columns):
                        break
                        
                except Exception as e:
                    logger.warning(f"Error reading sheet {sheet}: {str(e)}")
                    continue
            
            # If no valid sheet found
            if df is None or df.empty:
                raise FoundationExportException("Could not find valid billing data in the provided file")
                
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            raise FoundationExportException(f"Could not read Excel file: {str(e)}")
            
        # Standardize column names
        column_mapping = {
            'JOB': 'job_number',
            'JOB CODE': 'job_number',
            'JOB NUMBER': 'job_number',
            'JOB #': 'job_number',
            'EQUIP': 'equipment_id',
            'EQUIPMENT': 'equipment_id',
            'EQUIPMENT #': 'equipment_id',
            'EQUIPMENT NUMBER': 'equipment_id',
            'JOB NAME': 'job_name',
            'PROJECT': 'job_name',
            'PROJECT NAME': 'job_name',
            'DESC': 'description',
            'DESCRIPTION': 'description',
            'RATE': 'rate',
            'DAILY RATE': 'rate', 
            'MONTHLY RATE': 'rate',
            'DAYS': 'days',
            'QTY': 'days',
            'QUANTITY': 'days',
            'AMOUNT': 'amount',
            'TOTAL': 'amount',
            'EXT AMOUNT': 'amount',
            'REGION': 'region',
            'AREA': 'region',
            'LOCATION': 'region'
        }
        
        # Create a new standardized dataframe
        std_df = pd.DataFrame()
        
        # Find the best matching columns
        for std_col, possible_names in column_mapping.items():
            if isinstance(possible_names, str):
                possible_names = [possible_names]
                
            for col_name in df.columns:
                if col_name in possible_names or any(name.upper() == col_name.upper() for name in possible_names):
                    std_df[std_col] = df[col_name]
                    break
                    
        # If we're missing essential columns, try to infer them from data patterns
        required_cols = ['job_number', 'equipment_id', 'amount']
        for col in required_cols:
            if col not in std_df.columns:
                # Try to infer column based on patterns
                if col == 'job_number':
                    # Look for column with job number pattern (e.g., "2023-001")
                    for df_col in df.columns:
                        sample = df[df_col].iloc[:10].astype(str)
                        if sample.str.contains(r'\d{4}-\d{3}').any() or sample.str.contains(r'\d{4}-\d{2}').any():
                            std_df['job_number'] = df[df_col]
                            break
                            
                elif col == 'equipment_id':
                    # Look for equipment ID patterns
                    for df_col in df.columns:
                        sample = df[df_col].iloc[:10].astype(str)
                        if sample.str.contains(r'EQ\d{3}').any() or sample.str.contains(r'[A-Z]{2}\d{3}').any():
                            std_df['equipment_id'] = df[df_col]
                            break
                            
                elif col == 'amount':
                    # Look for columns with currency values
                    for df_col in df.columns:
                        try:
                            sample = pd.to_numeric(df[df_col].iloc[:10], errors='coerce')
                            if not sample.isna().all() and sample.max() > 100:  # Likely a dollar amount
                                std_df['amount'] = pd.to_numeric(df[df_col], errors='coerce')
                                break
                        except:
                            continue
                            
        # Add region information if available
        if 'region' not in std_df.columns:
            # Try to infer region from job numbers or other data
            if 'job_number' in std_df.columns:
                # Check if job numbers have region prefixes
                job_prefixes = std_df['job_number'].astype(str).str[:2]
                
                region_map = {
                    'DF': 'DFW',
                    'HO': 'HOU',
                    'WT': 'WT'
                }
                
                std_df['region'] = job_prefixes.map(region_map)
                
            # If still no region, try to get it from the filename
            if 'region' not in std_df.columns or std_df['region'].isna().all():
                filename = file_path.name.upper()
                
                if 'DFW' in filename:
                    std_df['region'] = 'DFW'
                elif 'HOU' in filename or 'HOUSTON' in filename:
                    std_df['region'] = 'HOU'
                elif 'WT' in filename or 'WEST' in filename:
                    std_df['region'] = 'WT'
                else:
                    # Default to empty region, we'll need to split by job number later
                    std_df['region'] = None
                    
        # Make sure numeric columns are actually numeric
        for col in ['days', 'rate', 'amount']:
            if col in std_df.columns:
                std_df[col] = pd.to_numeric(std_df[col], errors='coerce')
                
        # Fill in missing values for required columns
        for col in ['job_name', 'description']:
            if col not in std_df.columns:
                std_df[col] = None
                
        return std_df
        
    except FoundationExportException:
        # Re-raise
        raise
    except Exception as e:
        logger.error(f"Error loading EQ billing file: {str(e)}")
        raise FoundationExportException(f"Error loading EQ billing file: {str(e)}")


def generate_foundation_imports(eq_billing_file, output_dir=None, month_year=None):
    """
    Generate DFW, HOU, and WT foundation imports based on the EQ billing master
    
    Args:
        eq_billing_file (str or Path): Path to the EQ billing master file
        output_dir (str or Path, optional): Directory to save the exports
        month_year (str, optional): Month and year for file naming (e.g., 'April 2025')
        
    Returns:
        dict: Information about generated files
    """
    try:
        # Load the EQ billing master data
        billing_data = load_eq_billing_master(eq_billing_file)
        
        # Set default output directory
        if output_dir is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = TEMP_DIR / timestamp
            
        # Create output directory if it doesn't exist
        if isinstance(output_dir, str):
            output_dir = Path(output_dir)
            
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Set default month/year
        if month_year is None:
            month_year = datetime.now().strftime("%B %Y")
            
        # Create a zip file for the exports
        zip_filename = f"Foundation_Imports_{month_year.replace(' ', '_')}.zip"
        zip_path = EXPORTS_DIR / zip_filename
        
        # Generate foundation import for each region
        export_files = []
        
        for region in REGIONS:
            region_data = billing_data
            
            # Filter by region if we have region data
            if 'region' in billing_data.columns and not billing_data['region'].isna().all():
                region_data = billing_data[billing_data['region'] == region]
                
            # Skip if no data for this region
            if region_data.empty:
                logger.warning(f"No data for region {region}")
                continue
                
            # Create region export
            output_file = output_dir / f"{region}_Foundation_Import_{month_year.replace(' ', '_')}.xlsx"
            
            # Generate the Excel file
            generate_region_foundation_import(region_data, output_file, region, month_year)
            
            export_files.append(output_file)
            
        # Create the EQ billings master export
        master_export_file = output_dir / f"EQ_Billings_Master_{month_year.replace(' ', '_')}.xlsx"
        
        generate_eq_billings_master(billing_data, master_export_file, month_year)
        
        export_files.append(master_export_file)
        
        # Create zip file with all exports
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in export_files:
                zipf.write(file, arcname=file.name)
                
        # Clean up temp files
        if output_dir.parts[-2] == 'temp':
            shutil.rmtree(output_dir)
            
        return {
            'success': True,
            'zip_file': str(zip_path),
            'export_files': [str(f) for f in export_files],
            'month_year': month_year
        }
        
    except FoundationExportException as e:
        # Re-raise
        raise
    except Exception as e:
        logger.error(f"Error generating foundation imports: {str(e)}")
        raise FoundationExportException(f"Error generating foundation imports: {str(e)}")


def generate_region_foundation_import(data, output_file, region, month_year):
    """
    Generate a foundation import file for a specific region
    
    Args:
        data (DataFrame): Billing data for the region
        output_file (Path): Output file path
        region (str): Region code (DFW, HOU, WT)
        month_year (str): Month and year (e.g., 'April 2025')
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    # Create main sheet
    sheet = wb.create_sheet(f"{region} Import")
    
    # Add headers
    headers = [
        'Job Number', 'Job Name', 'Equipment ID', 'Description', 
        'Start Date', 'End Date', 'Days', 'Rate', 'Amount', 'Notes'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
    # Add title row
    title_row = sheet.row_dimensions[3]
    title_row.height = 30
    
    title_cell = sheet.cell(row=3, column=1)
    title_cell.value = f"{region} Foundation Import - {month_year}"
    title_cell.font = Font(size=14, bold=True)
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Map data columns to header positions
    column_map = {
        'job_number': 1,
        'job_name': 2,
        'equipment_id': 3,
        'description': 4,
        'start_date': 5,
        'end_date': 6,
        'days': 7,
        'rate': 8,
        'amount': 9,
        'notes': 10
    }
    
    # Add data rows
    for row_idx, (_, row) in enumerate(data.iterrows(), 5):  # Start at row 5 (after title)
        for col_name, col_idx in column_map.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Add value if column exists in data
            if col_name in data.columns:
                # Format based on column type
                if col_name == 'amount' or col_name == 'rate':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '$#,##0.00'
                elif col_name == 'days':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '0.0'
                elif col_name in ['start_date', 'end_date']:
                    # These may not be in the source data
                    if col_name in data.columns and pd.notna(row[col_name]):
                        cell.value = row[col_name]
                        cell.number_format = 'yyyy-mm-dd'
                else:
                    cell.value = row[col_name] if pd.notna(row[col_name]) else None
                    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].width = 15
        
    # Make description column wider
    sheet.column_dimensions['D'].width = 30
    
    # Make job name column wider
    sheet.column_dimensions['B'].width = 25
    
    # Add filters
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 4}"
    
    # Add summary at the bottom
    summary_row = len(data) + 7
    
    summary_cell = sheet.cell(row=summary_row, column=1)
    summary_cell.value = "Summary"
    summary_cell.font = Font(bold=True)
    
    # Total equipment
    equip_cell = sheet.cell(row=summary_row, column=2)
    equip_cell.value = "Total Equipment:"
    equip_cell.font = Font(bold=True)
    
    equip_count_cell = sheet.cell(row=summary_row, column=3)
    equip_count_cell.value = len(data)
    
    # Total amount
    total_cell = sheet.cell(row=summary_row+1, column=2)
    total_cell.value = "Total Amount:"
    total_cell.font = Font(bold=True)
    
    total_amount_cell = sheet.cell(row=summary_row+1, column=3)
    total_amount_cell.value = data['amount'].sum() if 'amount' in data.columns else 0
    total_amount_cell.number_format = '$#,##0.00'
    
    # Save the workbook
    wb.save(output_file)


def generate_eq_billings_master(data, output_file, month_year):
    """
    Generate an EQ Billings Master file
    
    Args:
        data (DataFrame): Combined billing data
        output_file (Path): Output file path
        month_year (str): Month and year (e.g., 'April 2025')
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    # Create main data sheet
    data_sheet = wb.create_sheet("Billing Data")
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Add title to main sheet
    title_cell = data_sheet.cell(row=1, column=1)
    title_cell.value = f"EQ Billings Master - {month_year}"
    title_cell.font = Font(size=14, bold=True)
    data_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers to main sheet
    headers = [
        'Job Number', 'Job Name', 'Equipment ID', 'Description', 
        'Region', 'Days', 'Rate', 'Amount', 'Category', 'Notes'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
    # Map data columns to header positions
    column_map = {
        'job_number': 1,
        'job_name': 2,
        'equipment_id': 3,
        'description': 4,
        'region': 5,
        'days': 6,
        'rate': 7,
        'amount': 8,
        'category': 9,
        'notes': 10
    }
    
    # Add data rows
    for row_idx, (_, row) in enumerate(data.iterrows(), 4):  # Start at row 4 (after title and header)
        for col_name, col_idx in column_map.items():
            cell = data_sheet.cell(row=row_idx, column=col_idx)
            
            # Add value if column exists in data
            if col_name in data.columns:
                # Format based on column type
                if col_name == 'amount' or col_name == 'rate':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '$#,##0.00'
                elif col_name == 'days':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '0.0'
                else:
                    cell.value = row[col_name] if pd.notna(row[col_name]) else None
                    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        data_sheet.column_dimensions[column_letter].width = 15
        
    # Make description column wider
    data_sheet.column_dimensions['D'].width = 30
    
    # Make job name column wider
    data_sheet.column_dimensions['B'].width = 25
    
    # Add filters
    data_sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(data) + 3}"
    
    # Populate summary sheet
    # Add title
    summary_title = summary_sheet.cell(row=1, column=1)
    summary_title.value = f"EQ Billings Summary - {month_year}"
    summary_title.font = Font(size=14, bold=True)
    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    summary_title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Overall summary
    summary_sheet.cell(row=3, column=1).value = "Total Equipment Count:"
    summary_sheet.cell(row=3, column=1).font = Font(bold=True)
    summary_sheet.cell(row=3, column=2).value = len(data)
    
    summary_sheet.cell(row=4, column=1).value = "Total Billing Amount:"
    summary_sheet.cell(row=4, column=1).font = Font(bold=True)
    summary_sheet.cell(row=4, column=2).value = data['amount'].sum() if 'amount' in data.columns else 0
    summary_sheet.cell(row=4, column=2).number_format = '$#,##0.00'
    
    # Regional summaries
    if 'region' in data.columns:
        # Add regional summary header
        summary_sheet.cell(row=6, column=1).value = "Regional Summary"
        summary_sheet.cell(row=6, column=1).font = Font(bold=True)
        
        # Add regional summary headers
        summary_sheet.cell(row=7, column=1).value = "Region"
        summary_sheet.cell(row=7, column=2).value = "Equipment Count"
        summary_sheet.cell(row=7, column=3).value = "Total Amount"
        
        for col in range(1, 4):
            summary_sheet.cell(row=7, column=col).font = Font(bold=True)
            summary_sheet.cell(row=7, column=col).fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
            
        # Add regional data
        row_idx = 8
        for region in REGIONS:
            region_data = data[data['region'] == region] if 'region' in data.columns else data
            
            if not region_data.empty:
                summary_sheet.cell(row=row_idx, column=1).value = region
                summary_sheet.cell(row=row_idx, column=2).value = len(region_data)
                summary_sheet.cell(row=row_idx, column=3).value = region_data['amount'].sum() if 'amount' in region_data.columns else 0
                summary_sheet.cell(row=row_idx, column=3).number_format = '$#,##0.00'
                
                row_idx += 1
    
    # Job summary
    # Add job summary header
    summary_sheet.cell(row=row_idx + 2, column=1).value = "Job Summary"
    summary_sheet.cell(row=row_idx + 2, column=1).font = Font(bold=True)
    
    # Add job summary headers
    summary_sheet.cell(row=row_idx + 3, column=1).value = "Job Number"
    summary_sheet.cell(row=row_idx + 3, column=2).value = "Job Name"
    summary_sheet.cell(row=row_idx + 3, column=3).value = "Equipment Count"
    summary_sheet.cell(row=row_idx + 3, column=4).value = "Total Amount"
    
    for col in range(1, 5):
        summary_sheet.cell(row=row_idx + 3, column=col).font = Font(bold=True)
        summary_sheet.cell(row=row_idx + 3, column=col).fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
    # Add job data
    job_summary = data.groupby(['job_number']).agg({
        'job_name': 'first',
        'amount': 'sum',
        'equipment_id': 'count'
    }).reset_index()
    
    for i, (_, row) in enumerate(job_summary.iterrows()):
        summary_sheet.cell(row=row_idx + 4 + i, column=1).value = row['job_number']
        summary_sheet.cell(row=row_idx + 4 + i, column=2).value = row['job_name']
        summary_sheet.cell(row=row_idx + 4 + i, column=3).value = row['equipment_id']
        summary_sheet.cell(row=row_idx + 4 + i, column=4).value = row['amount']
        summary_sheet.cell(row=row_idx + 4 + i, column=4).number_format = '$#,##0.00'
        
    # Auto-size columns
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)
        summary_sheet.column_dimensions[column_letter].width = 20
        
    # Save the workbook
    wb.save(output_file)
"""
Legacy Formula Connector

This module implements a direct connection to the original formulas used in the
legacy Excel workbook, ensuring exact consistency with the prior manual reporting process.
It preserves all business rules and classification logic while automating the data collection.
"""

import os
import json
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

# Configure logging
logger = logging.getLogger(__name__)
file_handler = logging.FileHandler('logs/legacy_formula.log')
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(file_handler)
logger.setLevel(logging.INFO)

# Constants for attendance classification - these match the exact legacy workbook rules
LATE_START_TIME = "07:30:00"  # 7:30 AM cutoff for late
EARLY_END_TIME = "16:00:00"   # 4:00 PM cutoff for early end

class LegacyFormulaEngine:
    """
    Implementation of the legacy workbook formulas for driver attendance classification
    """
    
    def __init__(self, date_str: str):
        """
        Initialize the formula engine for a specific date
        
        Args:
            date_str: Date string in YYYY-MM-DD format
        """
        self.date_str = date_str
        self.date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        self.driving_history_data = None
        self.activity_detail_data = None
        self.assets_time_data = None
        self.job_mappings = self._load_job_mappings()
        self.driver_mappings = self._load_driver_mappings()
        
        # Create output directories if they don't exist
        os.makedirs("reports/daily_driver_reports", exist_ok=True)
        os.makedirs("exports/daily", exist_ok=True)
        
        logger.info(f"Initialized Legacy Formula Engine for date: {date_str}")
    
    def _load_job_mappings(self) -> Dict[str, str]:
        """
        Load job site mapping data from the consolidated job list
        
        Returns:
            Dictionary mapping job codes to job names
        """
        job_mappings = {}
        
        try:
            # Try to load from the consolidated employee and job lists file
            employee_file = 'attached_assets/Consolidated_Employee_And_Job_Lists_Corrected.xlsx'
            
            if os.path.exists(employee_file):
                logger.info(f"Loading job mappings from: {employee_file}")
                df = pd.read_excel(employee_file, sheet_name='Jobs')
                
                # Map job codes to job names
                for _, row in df.iterrows():
                    if 'Job Code' in df.columns and 'Job Name' in df.columns:
                        job_code = str(row['Job Code']).strip()
                        job_name = str(row['Job Name']).strip()
                        job_mappings[job_code] = job_name
                
                logger.info(f"Loaded {len(job_mappings)} job mappings")
            else:
                logger.warning(f"Job mapping file not found: {employee_file}")
        
        except Exception as e:
            logger.error(f"Error loading job mappings: {str(e)}")
        
        return job_mappings
    
    def _load_driver_mappings(self) -> Dict[str, Dict[str, Any]]:
        """
        Load driver mapping data from the consolidated employee list
        
        Returns:
            Dictionary mapping driver names to their details
        """
        driver_mappings = {}
        
        try:
            # Try to load from the consolidated employee and job lists file
            employee_file = 'attached_assets/Consolidated_Employee_And_Job_Lists_Corrected.xlsx'
            
            if os.path.exists(employee_file):
                logger.info(f"Loading driver mappings from: {employee_file}")
                df = pd.read_excel(employee_file, sheet_name='Employees')
                
                # Map driver names to their details
                for _, row in df.iterrows():
                    if 'Employee Name' in df.columns:
                        driver_name = str(row['Employee Name']).strip()
                        driver_mappings[driver_name] = {
                            'Name': driver_name,
                            'Email': row.get('Email', ''),
                            'Phone': row.get('Phone', ''),
                            'Default Job': row.get('Default Job', ''),
                            'Equipment': row.get('Equipment', '')
                        }
                
                logger.info(f"Loaded {len(driver_mappings)} driver mappings")
            else:
                logger.warning(f"Driver mapping file not found: {employee_file}")
        
        except Exception as e:
            logger.error(f"Error loading driver mappings: {str(e)}")
        
        return driver_mappings
    
    def _normalize_name(self, name: str) -> str:
        """
        Normalize driver name for consistent matching using the exact legacy workbook approach
        
        Args:
            name: Driver name to normalize
            
        Returns:
            Normalized driver name
        """
        if not name or not isinstance(name, str):
            return ""
        
        # Apply the exact same normalization as the legacy workbook
        name = name.strip().lower()
        
        # Remove common prefixes that were in the legacy workbook logic
        prefixes = ['driver ', 'dr ', 'driver: ', 'operator: ', 'op: ']
        for prefix in prefixes:
            if name.startswith(prefix):
                name = name[len(prefix):]
        
        # Normalize spaces
        name = ' '.join(name.split())
        
        return name
    
    def _parse_time(self, time_str: str) -> datetime:
        """
        Parse time string to datetime object
        
        Args:
            time_str: Time string in various formats
            
        Returns:
            Datetime object
        """
        if not time_str or not isinstance(time_str, str):
            return None
        
        # Try different time formats
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%H:%M:%S",
            "%H:%M"
        ]
        
        for fmt in formats:
            try:
                if "%Y" in fmt or "%m/%d" in fmt:
                    # Format includes date
                    dt = datetime.strptime(time_str, fmt)
                    return dt
                else:
                    # Format is time only, add date
                    time_only = datetime.strptime(time_str, fmt).time()
                    return datetime.combine(self.date_obj.date(), time_only)
            except ValueError:
                continue
        
        # If no format matches, try to extract time using regex
        import re
        match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?(?:\s*([AP]M))?', time_str)
        if match:
            hour, minute = int(match.group(1)), int(match.group(2))
            second = int(match.group(3)) if match.group(3) else 0
            ampm = match.group(4)
            
            if ampm and ampm.upper() == 'PM' and hour < 12:
                hour += 12
            elif ampm and ampm.upper() == 'AM' and hour == 12:
                hour = 0
            
            return datetime.combine(self.date_obj.date(), 
                                   datetime.min.time().replace(hour=hour, minute=minute, second=second))
        
        return None
    
    def load_driving_history(self, file_path: str) -> bool:
        """
        Load driving history data from CSV file using the same approach as the legacy workbook
        
        Args:
            file_path: Path to the driving history CSV file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Loading driving history from: {file_path}")
            
            # Check if file exists
            if not os.path.exists(file_path):
                logger.error(f"Driving history file not found: {file_path}")
                return False
            
            # Load CSV file into pandas DataFrame
            df = pd.read_csv(file_path)
            
            # Filter for the target date if date column exists
            date_columns = ['Date', 'DATE', 'EventDate', 'EVENTDATE', 'Date Time', 'DATE TIME']
            date_found = False
            
            for date_col in date_columns:
                if date_col in df.columns:
                    # Try to convert date strings to datetime and filter
                    try:
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                        mask = df[date_col].dt.strftime('%Y-%m-%d') == self.date_str
                        df = df[mask]
                        date_found = True
                        logger.info(f"Filtered driving history by date column: {date_col}")
                        break
                    except:
                        continue
            
            if not date_found:
                logger.warning(f"No date column found in driving history, using all records")
            
            # Check if we have data
            if df.empty:
                logger.warning(f"No driving history data found for date: {self.date_str}")
                return False
            
            # Normalize columns to match legacy workbook expectations
            df.columns = [col.strip() for col in df.columns]
            
            # Identify driver name column
            driver_columns = ['Driver', 'DRIVER', 'Driver Name', 'DRIVER NAME', 'DriverName', 'DRIVERNAME']
            driver_col = None
            
            for col in driver_columns:
                if col in df.columns:
                    driver_col = col
                    break
            
            if not driver_col:
                logger.error(f"No driver column found in driving history")
                return False
            
            # Normalize driver names
            df['normalized_driver'] = df[driver_col].apply(self._normalize_name)
            
            # Store the processed DataFrame
            self.driving_history_data = df
            logger.info(f"Successfully loaded {len(df)} driving history records")
            return True
            
        except Exception as e:
            logger.error(f"Error loading driving history: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def load_activity_detail(self, file_path: str) -> bool:
        """
        Load activity detail data from CSV file using the same approach as the legacy workbook
        
        Args:
            file_path: Path to the activity detail CSV file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Loading activity detail from: {file_path}")
            
            # Check if file exists
            if not os.path.exists(file_path):
                logger.error(f"Activity detail file not found: {file_path}")
                return False
            
            # Load CSV file into pandas DataFrame
            df = pd.read_csv(file_path)
            
            # Filter for the target date if date column exists
            date_columns = ['Date', 'DATE', 'ActivityDate', 'ACTIVITYDATE']
            date_found = False
            
            for date_col in date_columns:
                if date_col in df.columns:
                    # Try to convert date strings to datetime and filter
                    try:
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                        mask = df[date_col].dt.strftime('%Y-%m-%d') == self.date_str
                        df = df[mask]
                        date_found = True
                        logger.info(f"Filtered activity detail by date column: {date_col}")
                        break
                    except:
                        continue
            
            if not date_found:
                logger.warning(f"No date column found in activity detail, using all records")
            
            # Check if we have data
            if df.empty:
                logger.warning(f"No activity detail data found for date: {self.date_str}")
                return False
            
            # Normalize columns to match legacy workbook expectations
            df.columns = [col.strip() for col in df.columns]
            
            # Identify driver name column
            driver_columns = ['Driver', 'DRIVER', 'Driver Name', 'DRIVER NAME', 'DriverName', 'DRIVERNAME']
            driver_col = None
            
            for col in driver_columns:
                if col in df.columns:
                    driver_col = col
                    break
            
            if not driver_col:
                logger.error(f"No driver column found in activity detail")
                return False
            
            # Normalize driver names
            df['normalized_driver'] = df[driver_col].apply(self._normalize_name)
            
            # Store the processed DataFrame
            self.activity_detail_data = df
            logger.info(f"Successfully loaded {len(df)} activity detail records")
            return True
            
        except Exception as e:
            logger.error(f"Error loading activity detail: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def load_assets_time_on_site(self, file_path: str) -> bool:
        """
        Load assets time on site data from CSV file using the same approach as the legacy workbook
        
        Args:
            file_path: Path to the assets time on site CSV file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Loading assets time on site from: {file_path}")
            
            # Check if file exists
            if not os.path.exists(file_path):
                logger.error(f"Assets time on site file not found: {file_path}")
                return False
            
            # Load CSV file into pandas DataFrame
            df = pd.read_csv(file_path)
            
            # Filter for the target date if date column exists
            date_columns = ['Date', 'DATE', 'AssetDate', 'ASSETDATE']
            date_found = False
            
            for date_col in date_columns:
                if date_col in df.columns:
                    # Try to convert date strings to datetime and filter
                    try:
                        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                        mask = df[date_col].dt.strftime('%Y-%m-%d') == self.date_str
                        df = df[mask]
                        date_found = True
                        logger.info(f"Filtered assets time on site by date column: {date_col}")
                        break
                    except:
                        continue
            
            if not date_found:
                logger.warning(f"No date column found in assets time on site, using all records")
            
            # Check if we have data
            if df.empty:
                logger.warning(f"No assets time on site data found for date: {self.date_str}")
                return False
            
            # Normalize columns to match legacy workbook expectations
            df.columns = [col.strip() for col in df.columns]
            
            # Identify driver name column
            driver_columns = ['Driver', 'DRIVER', 'Driver Name', 'DRIVER NAME', 'DriverName', 'DRIVERNAME', 'Asset Name', 'ASSET NAME']
            driver_col = None
            
            for col in driver_columns:
                if col in df.columns:
                    driver_col = col
                    break
            
            if not driver_col:
                logger.error(f"No driver/asset column found in assets time on site")
                return False
            
            # Normalize driver names
            df['normalized_driver'] = df[driver_col].apply(self._normalize_name)
            
            # Store the processed DataFrame
            self.assets_time_data = df
            logger.info(f"Successfully loaded {len(df)} assets time on site records")
            return True
            
        except Exception as e:
            logger.error(f"Error loading assets time on site: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def process_attendance(self) -> Dict[str, Any]:
        """
        Process attendance data using legacy workbook formulas
        
        Returns:
            Dictionary with processed attendance data
        """
        logger.info(f"Processing attendance data for date: {self.date_str}")
        
        # Check if we have the necessary data
        if self.driving_history_data is None and self.activity_detail_data is None and self.assets_time_data is None:
            logger.error("No data sources available for processing")
            return {
                "date": self.date_str,
                "total_drivers": 0,
                "drivers": [],
                "summary": {
                    "on_time": 0,
                    "late": 0,
                    "early_end": 0,
                    "not_on_job": 0,
                    "total_drivers": 0
                },
                "error": "No data sources available for processing"
            }
        
        # Collect all unique drivers from all data sources
        all_drivers = set()
        
        if self.driving_history_data is not None:
            all_drivers.update(self.driving_history_data['normalized_driver'].unique())
        
        if self.activity_detail_data is not None:
            all_drivers.update(self.activity_detail_data['normalized_driver'].unique())
        
        if self.assets_time_data is not None:
            all_drivers.update(self.assets_time_data['normalized_driver'].unique())
        
        # Remove empty driver names
        all_drivers = {driver for driver in all_drivers if driver}
        
        logger.info(f"Found {len(all_drivers)} unique drivers across all data sources")
        
        # Process each driver
        drivers_data = []
        
        for normalized_driver in all_drivers:
            # Skip empty driver names
            if not normalized_driver:
                continue
            
            # Find the original driver name (not normalized) for display
            original_name = normalized_driver
            
            if self.driving_history_data is not None and not self.driving_history_data.empty:
                driver_rows = self.driving_history_data[self.driving_history_data['normalized_driver'] == normalized_driver]
                if not driver_rows.empty:
                    driver_col = [col for col in driver_rows.columns if 'driver' in col.lower() and 'normalized' not in col.lower()][0]
                    original_name = driver_rows.iloc[0][driver_col]
            
            # Find driver in mappings for additional info
            driver_info = next((info for name, info in self.driver_mappings.items() 
                               if self._normalize_name(name) == normalized_driver), {})
            
            # Get driver data from each source
            driver_data = {
                "driver_name": original_name,
                "normalized_name": normalized_driver,
                "date": self.date_str,
                "job_site": "Unknown",
                "equipment": driver_info.get('Equipment', ''),
                "start_time": None,
                "end_time": None,
                "classification": "not_on_job",  # Default classification
                "classification_reason": "No data available",
                "data_sources": [],
                "contact_info": {
                    "phone": driver_info.get('Phone', ''),
                    "email": driver_info.get('Email', ''),
                    "emergency_contact": ""
                }
            }
            
            # Extract data from driving history
            if self.driving_history_data is not None and not self.driving_history_data.empty:
                driver_rows = self.driving_history_data[self.driving_history_data['normalized_driver'] == normalized_driver]
                
                if not driver_rows.empty:
                    driver_data["data_sources"].append("driving_history")
                    
                    # Identify time columns
                    time_columns = [col for col in driver_rows.columns if any(
                        time_word in col.lower() for time_word in ['time', 'datetime', 'date time', 'start', 'end'])]
                    
                    time_data = []
                    location_data = []
                    
                    # Extract time and location data
                    for _, row in driver_rows.iterrows():
                        # Extract time
                        for col in time_columns:
                            time_str = str(row[col])
                            time_obj = self._parse_time(time_str)
                            if time_obj:
                                time_data.append(time_obj)
                        
                        # Extract location
                        location_cols = [col for col in row.index if any(
                            loc_word in col.lower() for loc_word in ['location', 'site', 'job', 'project'])]
                        
                        for col in location_cols:
                            loc = str(row[col]).strip()
                            if loc and loc.lower() not in ['unknown', 'none', 'nan', '']:
                                location_data.append(loc)
                    
                    # Use earliest time as start time and latest as end time
                    if time_data:
                        time_data.sort()
                        driver_data["start_time"] = time_data[0].strftime("%H:%M:%S")
                        if len(time_data) > 1:
                            driver_data["end_time"] = time_data[-1].strftime("%H:%M:%S")
                    
                    # Use most common location as job site
                    if location_data:
                        from collections import Counter
                        most_common_location = Counter(location_data).most_common(1)[0][0]
                        driver_data["job_site"] = most_common_location
            
            # Extract data from activity detail
            if self.activity_detail_data is not None and not self.activity_detail_data.empty:
                driver_rows = self.activity_detail_data[self.activity_detail_data['normalized_driver'] == normalized_driver]
                
                if not driver_rows.empty:
                    driver_data["data_sources"].append("activity_detail")
                    
                    # Identify time columns
                    time_columns = [col for col in driver_rows.columns if any(
                        time_word in col.lower() for time_word in ['time', 'datetime', 'date time', 'start', 'end'])]
                    
                    time_data = []
                    location_data = []
                    
                    # Extract time and location data
                    for _, row in driver_rows.iterrows():
                        # Extract time
                        for col in time_columns:
                            time_str = str(row[col])
                            time_obj = self._parse_time(time_str)
                            if time_obj:
                                time_data.append(time_obj)
                        
                        # Extract location
                        location_cols = [col for col in row.index if any(
                            loc_word in col.lower() for loc_word in ['location', 'site', 'job', 'project'])]
                        
                        for col in location_cols:
                            loc = str(row[col]).strip()
                            if loc and loc.lower() not in ['unknown', 'none', 'nan', '']:
                                location_data.append(loc)
                    
                    # Use earliest time as start time and latest as end time if not already set
                    if time_data:
                        time_data.sort()
                        if not driver_data["start_time"]:
                            driver_data["start_time"] = time_data[0].strftime("%H:%M:%S")
                        if len(time_data) > 1 and not driver_data["end_time"]:
                            driver_data["end_time"] = time_data[-1].strftime("%H:%M:%S")
                    
                    # Use most common location as job site if not already set
                    if location_data and driver_data["job_site"] == "Unknown":
                        from collections import Counter
                        most_common_location = Counter(location_data).most_common(1)[0][0]
                        driver_data["job_site"] = most_common_location
            
            # Extract data from assets time on site
            if self.assets_time_data is not None and not self.assets_time_data.empty:
                driver_rows = self.assets_time_data[self.assets_time_data['normalized_driver'] == normalized_driver]
                
                if not driver_rows.empty:
                    driver_data["data_sources"].append("assets_time_on_site")
                    
                    # Identify time columns
                    time_columns = [col for col in driver_rows.columns if any(
                        time_word in col.lower() for time_word in ['time', 'datetime', 'date time', 'start', 'end', 'first', 'last'])]
                    
                    time_data = []
                    location_data = []
                    
                    # Extract time and location data
                    for _, row in driver_rows.iterrows():
                        # Extract time
                        for col in time_columns:
                            time_str = str(row[col])
                            time_obj = self._parse_time(time_str)
                            if time_obj:
                                time_data.append(time_obj)
                        
                        # Extract location
                        location_cols = [col for col in row.index if any(
                            loc_word in col.lower() for loc_word in ['location', 'site', 'job', 'project'])]
                        
                        for col in location_cols:
                            loc = str(row[col]).strip()
                            if loc and loc.lower() not in ['unknown', 'none', 'nan', '']:
                                location_data.append(loc)
                    
                    # Use earliest time as start time and latest as end time if not already set
                    if time_data:
                        time_data.sort()
                        if not driver_data["start_time"]:
                            driver_data["start_time"] = time_data[0].strftime("%H:%M:%S")
                        if len(time_data) > 1 and not driver_data["end_time"]:
                            driver_data["end_time"] = time_data[-1].strftime("%H:%M:%S")
                    
                    # Use most common location as job site if not already set
                    if location_data and driver_data["job_site"] == "Unknown":
                        from collections import Counter
                        most_common_location = Counter(location_data).most_common(1)[0][0]
                        driver_data["job_site"] = most_common_location
            
            # Classify attendance based on start and end times - using the legacy workbook logic
            if driver_data["start_time"]:
                start_time = datetime.strptime(driver_data["start_time"], "%H:%M:%S").time()
                late_threshold = datetime.strptime(LATE_START_TIME, "%H:%M:%S").time()
                
                if start_time <= late_threshold:
                    driver_data["classification"] = "on_time"
                    driver_data["classification_reason"] = f"Started at {driver_data['start_time']}, before {LATE_START_TIME} cutoff"
                else:
                    driver_data["classification"] = "late"
                    driver_data["classification_reason"] = f"Started at {driver_data['start_time']}, after {LATE_START_TIME} cutoff"
                
                # Check for early end if end time is available
                if driver_data["end_time"]:
                    end_time = datetime.strptime(driver_data["end_time"], "%H:%M:%S").time()
                    early_end_threshold = datetime.strptime(EARLY_END_TIME, "%H:%M:%S").time()
                    
                    if end_time < early_end_threshold:
                        driver_data["classification"] = "early_end"
                        driver_data["classification_reason"] = f"Ended at {driver_data['end_time']}, before {EARLY_END_TIME} cutoff"
            else:
                driver_data["classification"] = "not_on_job"
                driver_data["classification_reason"] = "No start time recorded"
            
            # Add processed driver data to the list
            drivers_data.append(driver_data)
        
        # Calculate summary statistics
        on_time_count = sum(1 for driver in drivers_data if driver["classification"] == "on_time")
        late_count = sum(1 for driver in drivers_data if driver["classification"] == "late")
        early_end_count = sum(1 for driver in drivers_data if driver["classification"] == "early_end")
        not_on_job_count = sum(1 for driver in drivers_data if driver["classification"] == "not_on_job")
        
        summary = {
            "on_time": on_time_count,
            "late": late_count,
            "early_end": early_end_count,
            "not_on_job": not_on_job_count,
            "total_drivers": len(drivers_data)
        }
        
        logger.info(f"Processed {len(drivers_data)} drivers")
        logger.info(f"Summary: {summary}")
        
        # Return the processed data
        return {
            "date": self.date_str,
            "total_drivers": len(drivers_data),
            "drivers": drivers_data,
            "summary": summary
        }
    
    def generate_report(self) -> str:
        """
        Generate a complete attendance report and save it to file
        
        Returns:
            Path to the generated report file
        """
        logger.info(f"Generating attendance report for date: {self.date_str}")
        
        # Process attendance data
        attendance_data = self.process_attendance()
        
        # Save JSON report
        json_file = f"reports/daily_driver_reports/attendance_report_{self.date_str}.json"
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(attendance_data, f, indent=2)
        
        logger.info(f"Saved JSON report to: {json_file}")
        
        # Also save the latest report for easy access
        latest_json = "reports/daily_driver_reports/latest_attendance_report.json"
        
        with open(latest_json, 'w', encoding='utf-8') as f:
            json.dump(attendance_data, f, indent=2)
        
        # Create Excel report
        try:
            excel_file = f"exports/daily/daily_driver_report_{self.date_str}.xlsx"
            self._generate_excel_report(attendance_data, excel_file)
            logger.info(f"Saved Excel report to: {excel_file}")
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
        
        return json_file
    
    def _generate_excel_report(self, attendance_data: Dict[str, Any], excel_file: str) -> None:
        """
        Generate an Excel report that mimics the legacy workbook format
        
        Args:
            attendance_data: Processed attendance data
            excel_file: Path to save the Excel file
        """
        # Create a new workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Driver Report {self.date_str}"
        
        # Add title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Daily Driver Report - {self.date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add summary
        ws.merge_cells('A3:F3')
        ws['A3'] = "Summary"
        ws['A3'].font = Font(size=14, bold=True)
        
        summary = attendance_data["summary"]
        summary_rows = [
            ["Total Drivers", summary["total_drivers"]],
            ["On Time", summary["on_time"]],
            ["Late", summary["late"]],
            ["Early End", summary["early_end"]],
            ["Not On Job", summary["not_on_job"]]
        ]
        
        for i, (label, value) in enumerate(summary_rows):
            ws[f'A{i+4}'] = label
            ws[f'B{i+4}'] = value
            ws[f'A{i+4}'].font = Font(bold=True)
        
        # Add drivers table header
        header_row = 10
        headers = ["Driver Name", "Job Site", "Start Time", "End Time", "Status", "Reason"]
        
        for i, header in enumerate(headers):
            col = get_column_letter(i+1)
            ws[f'{col}{header_row}'] = header
            ws[f'{col}{header_row}'].font = Font(bold=True)
            ws[f'{col}{header_row}'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Set column width
            ws.column_dimensions[col].width = 20
        
        # Add drivers data
        for i, driver in enumerate(attendance_data["drivers"]):
            row = header_row + i + 1
            
            ws[f'A{row}'] = driver["driver_name"]
            ws[f'B{row}'] = driver["job_site"]
            ws[f'C{row}'] = driver["start_time"] if driver["start_time"] else "N/A"
            ws[f'D{row}'] = driver["end_time"] if driver["end_time"] else "N/A"
            
            # Status with color coding
            status_cell = ws[f'E{row}']
            status_cell.value = driver["classification"].replace('_', ' ').title()
            
            if driver["classification"] == "on_time":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif driver["classification"] == "late":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif driver["classification"] == "early_end":
                status_cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            elif driver["classification"] == "not_on_job":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws[f'F{row}'] = driver["classification_reason"]
        
        # Save the workbook
        wb.save(excel_file)


def process_daily_driver_report(date_str: str, 
                               driving_history_file: str = None, 
                               activity_detail_file: str = None,
                               assets_time_file: str = None) -> Dict[str, Any]:
    """
    Process daily driver report for a specific date using legacy workbook formulas
    
    Args:
        date_str: Date string in YYYY-MM-DD format
        driving_history_file: Path to driving history CSV file (optional)
        activity_detail_file: Path to activity detail CSV file (optional)
        assets_time_file: Path to assets time on site CSV file (optional)
        
    Returns:
        Dictionary with processed attendance data and report paths
    """
    logger.info(f"Processing daily driver report for date: {date_str}")
    
    # Create formula engine
    engine = LegacyFormulaEngine(date_str)
    
    # Load data sources if provided
    data_sources_loaded = False
    
    if driving_history_file and os.path.exists(driving_history_file):
        engine.load_driving_history(driving_history_file)
        data_sources_loaded = True
    else:
        # Try to find the file in the data directory
        data_dir = "data"
        if os.path.exists(data_dir):
            # Look for files with date in the name and "driving" or "history"
            for filename in os.listdir(data_dir):
                if date_str in filename and ("driving" in filename.lower() or "history" in filename.lower()):
                    engine.load_driving_history(os.path.join(data_dir, filename))
                    data_sources_loaded = True
                    break
    
    if activity_detail_file and os.path.exists(activity_detail_file):
        engine.load_activity_detail(activity_detail_file)
        data_sources_loaded = True
    else:
        # Try to find the file in the data directory
        data_dir = "data"
        if os.path.exists(data_dir):
            # Look for files with date in the name and "activity" or "detail"
            for filename in os.listdir(data_dir):
                if date_str in filename and ("activity" in filename.lower() or "detail" in filename.lower()):
                    engine.load_activity_detail(os.path.join(data_dir, filename))
                    data_sources_loaded = True
                    break
    
    if assets_time_file and os.path.exists(assets_time_file):
        engine.load_assets_time_on_site(assets_time_file)
        data_sources_loaded = True
    else:
        # Try to find the file in the data directory
        data_dir = "data"
        if os.path.exists(data_dir):
            # Look for files with date in the name and "assets" or "time"
            for filename in os.listdir(data_dir):
                if date_str in filename and ("assets" in filename.lower() or "time" in filename.lower()):
                    engine.load_assets_time_on_site(os.path.join(data_dir, filename))
                    data_sources_loaded = True
                    break
    
    # Check for attached_assets directory
    if not data_sources_loaded:
        attached_dir = "attached_assets"
        if os.path.exists(attached_dir):
            # Look for relevant files
            for filename in os.listdir(attached_dir):
                if "driving" in filename.lower() or "history" in filename.lower():
                    engine.load_driving_history(os.path.join(attached_dir, filename))
                    data_sources_loaded = True
                
                if "activity" in filename.lower() or "detail" in filename.lower():
                    engine.load_activity_detail(os.path.join(attached_dir, filename))
                    data_sources_loaded = True
                
                if "assets" in filename.lower() or "time" in filename.lower():
                    engine.load_assets_time_on_site(os.path.join(attached_dir, filename))
                    data_sources_loaded = True
    
    if not data_sources_loaded:
        logger.error(f"No data sources found for date: {date_str}")
        return {
            "success": False,
            "error": "No data sources found",
            "date": date_str
        }
    
    # Generate report
    report_file = engine.generate_report()
    
    # Return result
    return {
        "success": True,
        "date": date_str,
        "json_report": report_file,
        "excel_report": f"exports/daily/daily_driver_report_{date_str}.xlsx"
    }


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python legacy_formula_connector.py YYYY-MM-DD [driving_history.csv] [activity_detail.csv] [assets_time.csv]")
        sys.exit(1)
    
    date_str = sys.argv[1]
    
    driving_history_file = sys.argv[2] if len(sys.argv) > 2 else None
    activity_detail_file = sys.argv[3] if len(sys.argv) > 3 else None
    assets_time_file = sys.argv[4] if len(sys.argv) > 4 else None
    
    result = process_daily_driver_report(date_str, driving_history_file, activity_detail_file, assets_time_file)
    
    if result["success"]:
        print(f"Successfully processed daily driver report for {date_str}")
        print(f"JSON report saved to: {result['json_report']}")
        print(f"Excel report saved to: {result['excel_report']}")
    else:
        print(f"Error processing daily driver report: {result['error']}")
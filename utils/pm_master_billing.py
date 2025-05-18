"""
PM Master Billing Processor

This module processes multiple PM allocation files to create a consolidated
master billing report and generates division-specific exports.

Features:
- Extracts data from all PM allocation files (XLSX and CSV)
- Maps equipment numbers to metadata
- Applies correct billing rates based on frequency
- Generates master billing workbook and division-specific exports
- Handles cost code defaulting for missing values
- Creates standardized import files for each division
"""

import os
import re
import logging
import glob
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import os
import re
import glob
import json
import logging
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
ATTACHED_ASSETS_DIR = 'attached_assets'
EXPORTS_DIR = 'exports'
DEFAULT_COST_CODE = '9000 100F'  # Default cost code to use when missing or 'CC NEEDED'
PROCESS_ONLY_APRIL = True  # Set to True to focus only on April 2025 files

# Division names for file matching and organization
DIVISIONS = ['DFW', 'HOU', 'WTX', 'SELECT']

# Output file names
MASTER_BILLING_FILENAME = 'MASTER_EQUIP_BILLINGS_EXPORT_APRIL_2025.xlsx'
DIVISION_EXPORT_PATTERN = '{} - EQUIP BILLINGS EXPORT - APRIL 2025.xlsx'
DIVISION_IMPORT_FILENAMES = {
    'DFW': '01 - DFW APR 2025.csv',
    'HOU': '02 - HOU APR 2025.csv',
    'WTX': '03 - WT APR 2025.csv'
}

# Column mappings for standardization
STANDARD_COLUMNS = [
    'division', 'equip_id', 'description', 'date', 'job', 'phase', 
    'cost_code', 'units', 'rate', 'frequency', 'amount'
]

# Required columns for validation
REQUIRED_COLUMNS = ['equip_id', 'job', 'units', 'rate']

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def find_all_allocation_files():
    """Find all PM allocation files (XLSX and CSV) in the attached_assets directory"""
    all_files = []
    
    # Use different patterns for each division
    for division in DIVISIONS:
        # Try different naming patterns for CSV files - focus on April 2025
        patterns = []
        
        # Special case for WTX - also look for WT prefix
        if division == 'WTX':
            if PROCESS_ONLY_APRIL:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'03 - WT APR 2025.csv'),  # Exact filename match
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APRIL*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WTX*APRIL*2025*.xlsx'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WTX*APR*2025*.xlsx'),
                ])
            else:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'03 - WT*.csv'),  # Exact filename pattern
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APRIL*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WTX*.xlsx'),
                ])
                
        # Standard patterns for all divisions
        if PROCESS_ONLY_APRIL:
            patterns.extend([
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APRIL*2025*.xlsx'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APR*2025*.xlsx'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APRIL*2025*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APR*2025*.csv'),
            ])
        else:
            patterns.extend([
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*.xlsx'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*.csv'),
            ])
        
        # Add special pattern for DFW with "01" prefix
        if division == 'DFW':
            if PROCESS_ONLY_APRIL:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'01 - DFW APR 2025.csv'),  # Exact filename match
                    os.path.join(ATTACHED_ASSETS_DIR, f'*01*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*01*APRIL*2025*.csv')
                ])
            else:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'01 - DFW*.csv'),  # Exact filename pattern
                    os.path.join(ATTACHED_ASSETS_DIR, f'*01*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*01*APRIL*.csv')
                ])
                
        # Add special pattern for HOU with "02" prefix
        if division == 'HOU':
            if PROCESS_ONLY_APRIL:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'02 - HOU APR 2025.csv'),  # Exact filename match
                    os.path.join(ATTACHED_ASSETS_DIR, f'*02*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*02*APRIL*2025*.csv')
                ])
            else:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'02 - HOU*.csv'),  # Exact filename pattern
                    os.path.join(ATTACHED_ASSETS_DIR, f'*02*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*02*APRIL*.csv')
                ])
                
        # Add special pattern for SELECT with "SM" prefix
        if division == 'SELECT':
            if PROCESS_ONLY_APRIL:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'SM - SELECT APRIL 2025.csv'),  # Exact filename match
                    os.path.join(ATTACHED_ASSETS_DIR, f'*SM*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*SM*APRIL*2025*.csv')
                ])
            else:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'SM - SELECT*.csv'),  # Exact filename pattern
                    os.path.join(ATTACHED_ASSETS_DIR, f'*SM*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*SM*APRIL*.csv')
                ])
        
        # Find files matching any pattern for this division
        for pattern in patterns:
            matching_files = glob.glob(pattern)
            for file_path in matching_files:
                file_info = {
                    'path': file_path,
                    'filename': os.path.basename(file_path),
                    'division': division,
                    'extension': os.path.splitext(file_path)[1].lower(),
                    'size': os.path.getsize(file_path)
                }
                all_files.append(file_info)
    
    # Check special case for EQMO BILLING ALLOCATIONS files which may not have division in name
    final_revision_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQMO*BILLING*ALLOCATION*FINAL*REVISION*.xlsx')
    final_revision_files = glob.glob(final_revision_pattern)
    
    for file_path in final_revision_files:
        filename = os.path.basename(file_path)
        # Check if file is already in the list
        if not any(f['path'] == file_path for f in all_files):
            file_info = {
                'path': file_path,
                'filename': filename,
                'division': 'MASTER',  # Mark as master file
                'extension': os.path.splitext(file_path)[1].lower(),
                'size': os.path.getsize(file_path)
            }
            all_files.append(file_info)
    
    # Also look for any files that have EQMO BILLING ALLOCATIONS in the name (final or not)
    billing_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQMO*BILLING*ALLOCATION*.xlsx')
    billing_files = glob.glob(billing_pattern)
    
    for file_path in billing_files:
        filename = os.path.basename(file_path)
        # Check if file is already in the list
        if not any(f['path'] == file_path for f in all_files):
            file_info = {
                'path': file_path,
                'filename': filename,
                'division': 'MASTER',  # Mark as master file
                'extension': os.path.splitext(file_path)[1].lower(),
                'size': os.path.getsize(file_path)
            }
            all_files.append(file_info)
    
    # Remove duplicates based on path
    unique_files = []
    seen_paths = set()
    for file_info in all_files:
        if file_info['path'] not in seen_paths:
            unique_files.append(file_info)
            seen_paths.add(file_info['path'])
    
    logger.info(f"Found {len(unique_files)} allocation files matching patterns")
    return unique_files

def find_matching_column(df, patterns):
    """Find the first column that matches one of the patterns"""
    for pattern in patterns:
        matching_cols = [col for col in df.columns if pattern.lower() in str(col).lower()]
        if matching_cols:
            return matching_cols[0]
    return None

def process_allocation_file(file_path, division=None):
    """
    Process a single allocation file to extract equipment billing data
    
    Args:
        file_path (str): Path to the allocation file
        division (str, optional): Division to assign to the data
        
    Returns:
        DataFrame: Processed allocation data or empty DataFrame if processing fails
    """
    filename = os.path.basename(file_path)
    extension = os.path.splitext(file_path)[1].lower()
    logger.info(f"Processing allocation file: {filename}")
    
    try:
        if extension == '.xlsx':
            # For Excel files, need to determine which sheet to use
            try:
                # Try to get the sheet names
                xlsx = pd.ExcelFile(file_path)
                sheet_names = xlsx.sheet_names
                
                # Look for sheets with allocation data
                target_sheet = None
                allocation_sheet_patterns = ['allocation', 'eq allocation', 'all div']
                
                for sheet in sheet_names:
                    for pattern in allocation_sheet_patterns:
                        if pattern.lower() in sheet.lower():
                            target_sheet = sheet
                            break
                    if target_sheet:
                        break
                
                # If we didn't find an allocation sheet, use the first sheet
                if not target_sheet and sheet_names:
                    target_sheet = sheet_names[0]
                
                if not target_sheet:
                    logger.error(f"No usable sheets found in {filename}")
                    return pd.DataFrame()
                
                logger.info(f"Using sheet '{target_sheet}' from {filename}")
                
                # Read the data from the selected sheet
                df = pd.read_excel(file_path, sheet_name=target_sheet)
                
                # Check if we have the expected columns
                required_patterns = {
                    'equip_id': ['equip', 'equipment', 'equip #', 'eq #'],
                    'description': ['description', 'desc', 'name'],
                    'job': ['job', 'job #', 'job number'],
                    'units': ['units', 'hours', 'qty'],
                    'rate': ['rate'],
                    'amount': ['amount', 'total', 'extended']
                }
                
                # Check if required columns exist
                found_cols = {}
                missing_cols = []
                
                for col_key, patterns in required_patterns.items():
                    found_col = find_matching_column(df, patterns)
                    if found_col:
                        found_cols[col_key] = found_col
                    else:
                        missing_cols.append(col_key)
                
                if missing_cols:
                    logger.warning(f"Missing essential columns in file: {filename}")
                    # Special handling for known file formats
                    if 'MASTER' in division or 'FINAL' in filename.upper() or 'REVISION' in filename.upper():
                        logger.info(f"Attempting special handling for master file: {filename.upper()}")
                        # Try to identify columns by position or alternate names
                        if 'equip_id' in missing_cols and df.shape[1] > 0:
                            # Try first column
                            potential_col = df.columns[0]
                            logger.info(f"Using first column as equipment ID: {potential_col}")
                            found_cols['equip_id'] = potential_col
                        
                        if 'job' in missing_cols and df.shape[1] > 2:
                            # Look for a column containing "job"
                            job_cols = [col for col in df.columns if 'job' in str(col).lower()]
                            if job_cols:
                                logger.info(f"Found potential job column: {job_cols[0]}")
                                found_cols['job'] = job_cols[0]
                        
                        if 'units' in missing_cols and df.shape[1] > 5:
                            # Look for columns with hours or units
                            units_cols = [col for col in df.columns if any(p in str(col).lower() for p in ['hours', 'units', 'qty'])]
                            if units_cols:
                                logger.info(f"Found potential units column: {units_cols[0]}")
                                found_cols['units'] = units_cols[0]
                        
                        # If we still have missing columns, we can't process this file
                        missing_cols = [col for col in required_patterns.keys() if col not in found_cols]
                        if missing_cols:
                            logger.error(f"Unable to find essential columns in file: {filename}")
                            return pd.DataFrame()
                    else:
                        logger.error(f"Unable to find essential columns in file: {filename}")
                        return pd.DataFrame()
                
                # Rename columns to standardized names
                rename_map = {found_cols[key]: key for key in found_cols}
                df = df.rename(columns=rename_map)
                
                # Apply any necessary data transformations
                # 1. Convert equip_id to string and clean up
                if 'equip_id' in df.columns:
                    df['equip_id'] = df['equip_id'].astype(str).str.strip()
                
                # 2. Handle missing values
                if 'description' in df.columns:
                    df['description'] = df['description'].fillna('').astype(str)
                else:
                    df['description'] = ''
                
                # 3. Standardize cost_code - use default if missing or "CC NEEDED"
                if 'cost_code' in df.columns:
                    df['cost_code'] = df['cost_code'].fillna(DEFAULT_COST_CODE).astype(str)
                    # Replace "CC NEEDED" with default
                    df.loc[df['cost_code'].str.upper().str.contains('NEEDED'), 'cost_code'] = DEFAULT_COST_CODE
                else:
                    df['cost_code'] = DEFAULT_COST_CODE
                
                # 4. Handle numeric columns
                if 'units' in df.columns:
                    df['units'] = pd.to_numeric(df['units'], errors='coerce').fillna(0)
                
                if 'rate' in df.columns:
                    df['rate'] = pd.to_numeric(df['rate'], errors='coerce').fillna(0)
                
                # 5. Calculate amount if missing
                if 'amount' not in df.columns or df['amount'].isna().all():
                    df['amount'] = df['units'] * df['rate']
                else:
                    df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
                
                # 6. Add division column if not already present
                if 'division' not in df.columns:
                    # Try to determine division from filename
                    if not division:
                        for div in DIVISIONS:
                            if div in filename.upper():
                                division = div
                                break
                    
                    if not division:
                        # Default to MASTER for unspecified division
                        division = 'MASTER'
                    
                    df['division'] = division
                
                # 7. Add date column if not present
                if 'date' not in df.columns:
                    # Default to end of April 2025
                    df['date'] = '4/30/2025'
                
                # 8. Add phase column if not present
                if 'phase' not in df.columns:
                    df['phase'] = ''
                
                # 9. Add frequency column if not present
                if 'frequency' not in df.columns:
                    df['frequency'] = 'MONTHLY'
                
                # 10. Filter out rows with empty equipment ID or zero units/rate
                df = df[df['equip_id'].notna() & (df['equip_id'] != '') & 
                         df['units'] > 0 & df['rate'] > 0]
                
                # Standardize column order
                available_cols = [col for col in STANDARD_COLUMNS if col in df.columns]
                df = df[available_cols]
                
                # Log the results
                logger.info(f"Processed {len(df)} records from {filename}")
                
                return df
                
            except Exception as e:
                logger.error(f"Error processing Excel file {filename}: {str(e)}")
                return pd.DataFrame()
                
        elif extension == '.csv':
            # For CSV files, we need to determine the format
            try:
                # Start by trying to read with headers
                df = pd.read_csv(file_path)
                
                # Check if this is a standard CSV with headers
                if df.shape[1] >= 3:
                    # Look for required columns
                    required_patterns = {
                        'equip_id': ['equip', 'equipment', 'equip #', 'eq #'],
                        'job': ['job', 'job #', 'job number'],
                        'units': ['units', 'hours', 'qty']
                    }
                    
                    found_cols = {}
                    missing_cols = []
                    
                    for col_key, patterns in required_patterns.items():
                        found_col = find_matching_column(df, patterns)
                        if found_col:
                            found_cols[col_key] = found_col
                        else:
                            missing_cols.append(col_key)
                    
                    if missing_cols:
                        logger.warning(f"Missing essential columns in file: {filename}")
                        
                        # For division-specific CSV files (e.g., DFW, HOU, WTX)
                        logger.info(f"Attempting special handling for division file: {filename.upper()}")
                        
                        # Try reading the first row to see if it's in expected format
                        sample = pd.read_csv(file_path, nrows=1)
                        # If the first column looks like an equipment ID (e.g., PT-159), it's likely headerless
                        first_col_value = str(sample.iloc[0, 0]).strip()
                        
                        if re.match(r'^[A-Z]{1,3}-\d+$', first_col_value):
                            # This is a headerless CSV with the format we know
                            logger.info(f"Detected headerless division-specific CSV: {filename}")
                            
                            # Define the headers based on the CSV file format we're seeing
                            headers = ["equip_id", "description", "date", "job", "phase", "cost_code", 
                                    "units", "rate", "frequency", "monthly_rate", "amount"]
                            
                            # Add specific handling for divisions
                            if "WT" in filename or "03" in filename:
                                logger.info(f"Processing WTX division file: {filename}")
                                division = "WTX"
                            elif "DFW" in filename or "01" in filename:
                                logger.info(f"Processing DFW division file: {filename}")
                                division = "DFW"
                            elif "HOU" in filename or "02" in filename:
                                logger.info(f"Processing HOU division file: {filename}")
                                division = "HOU"
                            elif "SELECT" in filename or "SM" in filename:
                                logger.info(f"Processing SELECT division file: {filename}")
                                division = "SELECT"
                                
                            # Read the CSV file without headers
                            try:
                                df = pd.read_csv(file_path, header=None, names=headers)
                                # After reading, check data validity
                                logger.info(f"Successfully loaded {len(df)} records with {len(df.columns)} columns")
                                # Log first row for debugging
                                if not df.empty:
                                    logger.info(f"First row sample: {df.iloc[0].to_dict()}")
                            except Exception as e:
                                logger.error(f"Error reading CSV: {str(e)}")
                                # Try reading with different encoding or delimiter
                                try:
                                    df = pd.read_csv(file_path, header=None, names=headers, encoding='latin1')
                                    logger.info(f"Successfully loaded with latin1 encoding: {len(df)} records")
                                except Exception as e2:
                                    logger.error(f"Failed with alternative encoding: {str(e2)}")
                                    return pd.DataFrame()
                            
                            # Add division column
                            df['division'] = division
                            logger.info(f"Processed headerless CSV with {len(df)} records from {division}")
                            
                            # Apply any necessary data transformations
                            # 1. Convert equip_id to string and clean up
                            df['equip_id'] = df['equip_id'].astype(str).str.strip()
                            
                            # 2. Handle missing values for description
                            df['description'] = df['description'].fillna('').astype(str)
                            
                            # 3. Standardize cost_code - use default if missing or "CC NEEDED"
                            df['cost_code'] = df['cost_code'].fillna(DEFAULT_COST_CODE).astype(str)
                            df.loc[df['cost_code'].str.upper().str.contains('NEEDED', na=False), 'cost_code'] = DEFAULT_COST_CODE
                            
                            # 4. Ensure numeric columns are properly formatted
                            df['units'] = pd.to_numeric(df['units'], errors='coerce').fillna(0)
                            df['rate'] = pd.to_numeric(df['rate'], errors='coerce').fillna(0)
                            
                            # 5. Ensure amount is calculated correctly
                            if 'amount' not in df.columns:
                                df['amount'] = df['units'] * df['rate']
                            else:
                                df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
                            
                            # 6. Filter out rows with empty equipment ID or zero units/rate
                            df = df[df['equip_id'].notna() & (df['equip_id'] != '') & 
                                    df['units'] > 0 & df['rate'] > 0]
                            
                            # 7. Standardize column order
                            available_cols = [col for col in STANDARD_COLUMNS if col in df.columns]
                            df = df[available_cols]
                            
                            return df
                        else:
                            logger.error(f"Unable to find essential columns in file: {filename}")
                            return pd.DataFrame()
                    else:
                        # Standard CSV with headers - process it
                        # Rename columns to standardized names
                        rename_map = {found_cols[key]: key for key in found_cols}
                        df = df.rename(columns=rename_map)
                        
                        # Apply data transformations (similar to Excel processing)
                        # 1. Convert equip_id to string and clean up
                        df['equip_id'] = df['equip_id'].astype(str).str.strip()
                        
                        # 2. Handle missing values for description
                        if 'description' in df.columns:
                            df['description'] = df['description'].fillna('').astype(str)
                        else:
                            df['description'] = ''
                        
                        # 3. Add cost_code if missing, or standardize it
                        if 'cost_code' in df.columns:
                            df['cost_code'] = df['cost_code'].fillna(DEFAULT_COST_CODE).astype(str)
                            df.loc[df['cost_code'].str.upper().str.contains('NEEDED', na=False), 'cost_code'] = DEFAULT_COST_CODE
                        else:
                            df['cost_code'] = DEFAULT_COST_CODE
                        
                        # 4. Ensure numeric columns are properly formatted
                        df['units'] = pd.to_numeric(df['units'], errors='coerce').fillna(0)
                        
                        if 'rate' in df.columns:
                            df['rate'] = pd.to_numeric(df['rate'], errors='coerce').fillna(0)
                        else:
                            # If rate is missing, we need to derive it
                            if 'amount' in df.columns and 'units' in df.columns:
                                # Calculate rate from amount and units
                                df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
                                df.loc[df['units'] > 0, 'rate'] = df.loc[df['units'] > 0, 'amount'] / df.loc[df['units'] > 0, 'units']
                                df.loc[df['units'] <= 0, 'rate'] = 0
                            else:
                                # No way to determine rate
                                logger.error(f"Cannot determine rate in file: {filename}")
                                return pd.DataFrame()
                        
                        # 5. Ensure amount is calculated correctly
                        if 'amount' not in df.columns:
                            df['amount'] = df['units'] * df['rate']
                        else:
                            df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
                        
                        # 6. Add division column if not already present
                        if 'division' not in df.columns:
                            # Try to determine division from filename
                            if not division:
                                for div in DIVISIONS:
                                    if div in filename.upper():
                                        division = div
                                        break
                            
                            if not division:
                                # Default to MASTER for unspecified division
                                division = 'MASTER'
                            
                            df['division'] = division
                        
                        # 7. Add date column if not present
                        if 'date' not in df.columns:
                            # Default to end of April 2025
                            df['date'] = '4/30/2025'
                        
                        # 8. Add phase column if not present
                        if 'phase' not in df.columns:
                            df['phase'] = ''
                        
                        # 9. Add frequency column if not present
                        if 'frequency' not in df.columns:
                            df['frequency'] = 'MONTHLY'
                        
                        # 10. Filter out rows with empty equipment ID or zero units/rate
                        df = df[df['equip_id'].notna() & (df['equip_id'] != '') & 
                                df['units'] > 0 & df['rate'] > 0]
                        
                        # Standardize column order
                        available_cols = [col for col in STANDARD_COLUMNS if col in df.columns]
                        df = df[available_cols]
                        
                        logger.info(f"Processed {len(df)} records from {filename}")
                        
                        return df
                else:
                    logger.error(f"CSV file has too few columns: {filename}")
                    return pd.DataFrame()
                    
            except Exception as e:
                logger.error(f"Error processing CSV file {filename}: {str(e)}")
                return pd.DataFrame()
        else:
            logger.error(f"Unsupported file format: {extension}")
            return pd.DataFrame()
            
    except Exception as e:
        logger.error(f"Error processing file {filename}: {str(e)}")
        return pd.DataFrame()

def extract_equipment_rates(rates_file):
    """
    Extract equipment rates from the rates file or master billing workbook
    
    Args:
        rates_file (str): Path to the rates file
        
    Returns:
        DataFrame: Equipment rates data or empty DataFrame if extraction fails
    """
    try:
        # Try to open the Excel file
        xlsx = pd.ExcelFile(rates_file)
        
        # Look for a sheet with rates data
        rate_sheet_name = None
        rate_sheet_patterns = ['equip rates', 'rates', 'equipment rates']
        
        for sheet in xlsx.sheet_names:
            for pattern in rate_sheet_patterns:
                if pattern.lower() in sheet.lower():
                    rate_sheet_name = sheet
                    break
            if rate_sheet_name:
                break
        
        if not rate_sheet_name and len(xlsx.sheet_names) > 0:
            # If no specific rates sheet found, check other sheets
            for sheet in xlsx.sheet_names:
                # Read a sample from the sheet to check for rate-related columns
                sample = pd.read_excel(rates_file, sheet_name=sheet, nrows=5)
                
                # Look for columns related to equipment rates
                rate_col_patterns = ['rate', 'monthly', 'hourly', 'daily']
                equip_col_patterns = ['equip', 'equipment', 'eq #']
                
                has_rate_col = any(any(pattern in str(col).lower() for pattern in rate_col_patterns) 
                                 for col in sample.columns)
                has_equip_col = any(any(pattern in str(col).lower() for pattern in equip_col_patterns) 
                                  for col in sample.columns)
                
                if has_rate_col and has_equip_col:
                    rate_sheet_name = sheet
                    break
        
        if not rate_sheet_name:
            logger.error(f"Could not find a sheet with equipment rates in {rates_file}")
            return pd.DataFrame()
        
        # Read the rates data
        rates_df = pd.read_excel(rates_file, sheet_name=rate_sheet_name)
        
        # Look for key columns
        equip_id_col = find_matching_column(rates_df, ['equip', 'equipment', 'eq #', 'equip id', 'equipment id'])
        monthly_rate_col = find_matching_column(rates_df, ['monthly', 'month rate', 'monthly rate'])
        hourly_rate_col = find_matching_column(rates_df, ['hourly', 'hour rate', 'hourly rate'])
        daily_rate_col = find_matching_column(rates_df, ['daily', 'day rate', 'daily rate'])
        
        if not equip_id_col:
            logger.error(f"Could not find equipment ID column in rates sheet")
            return pd.DataFrame()
        
        # Create a standardized rates DataFrame
        result = pd.DataFrame()
        result['equip_id'] = rates_df[equip_id_col].astype(str).str.strip()
        
        # Add rate columns if found
        if monthly_rate_col:
            result['monthly_rate'] = pd.to_numeric(rates_df[monthly_rate_col], errors='coerce').fillna(0)
        else:
            result['monthly_rate'] = 0
            
        if hourly_rate_col:
            result['hourly_rate'] = pd.to_numeric(rates_df[hourly_rate_col], errors='coerce').fillna(0)
        else:
            result['hourly_rate'] = 0
            
        if daily_rate_col:
            result['daily_rate'] = pd.to_numeric(rates_df[daily_rate_col], errors='coerce').fillna(0)
        else:
            result['daily_rate'] = 0
        
        # Filter out rows with empty equipment ID
        result = result[result['equip_id'].notna() & (result['equip_id'] != '')]
        
        logger.info(f"Extracted {len(result)} equipment rates from {rate_sheet_name}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error extracting equipment rates: {str(e)}")
        return pd.DataFrame()

def find_equipment_rates_file():
    """
    Find the equipment rates file or master billing workbook
    
    Returns:
        str: Path to the rates file or None if not found
    """
    # Try to find a file specifically for equipment rates
    rate_patterns = [
        os.path.join(ATTACHED_ASSETS_DIR, '*EQUIPMENT*RATE*.xlsx'),
        os.path.join(ATTACHED_ASSETS_DIR, '*EQUIP*RATE*.xlsx'),
        os.path.join(ATTACHED_ASSETS_DIR, '*EQ*RATE*.xlsx'),
        os.path.join(ATTACHED_ASSETS_DIR, '*MASTER*BILLING*.xlsx'),
        os.path.join(ATTACHED_ASSETS_DIR, '*BILLING*WORKBOOK*.xlsx'),
        os.path.join(ATTACHED_ASSETS_DIR, '*EQ*MONTHLY*BILLING*.xlsx')
    ]
    
    for pattern in rate_patterns:
        matching_files = glob.glob(pattern)
        if matching_files:
            logger.info(f"Found potential rates file: {matching_files[0]}")
            return matching_files[0]
    
    # If no specific rates file found, look for any EQMO BILLING file
    billing_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQMO*BILLING*.xlsx')
    billing_files = glob.glob(billing_pattern)
    
    if billing_files:
        logger.info(f"Using billing file for rates: {billing_files[0]}")
        return billing_files[0]
    
    logger.error("Could not find an equipment rates file")
    return None

def generate_master_billing(allocation_data, rates_data):
    """
    Generate the master billing dataset by applying rates to allocation data
    
    Args:
        allocation_data (DataFrame): The combined allocation data
        rates_data (DataFrame): The equipment rates data
        
    Returns:
        DataFrame: The master billing data with rates applied
    """
    if allocation_data.empty:
        logger.error("No allocation data provided for master billing")
        return pd.DataFrame()
    
    if rates_data.empty:
        logger.warning("No rates data provided - using rates from allocation data")
        result = allocation_data.copy()
    else:
        # Create a copy of the allocation data
        result = allocation_data.copy()
        
        # Create a rates lookup dictionary for faster access
        rates_lookup = {}
        for _, row in rates_data.iterrows():
            equip_id = row['equip_id']
            rates = {
                'MONTHLY': row['monthly_rate'] if 'monthly_rate' in row else 0,
                'HOURLY': row['hourly_rate'] if 'hourly_rate' in row else 0,
                'DAILY': row['daily_rate'] if 'daily_rate' in row else 0
            }
            rates_lookup[equip_id] = rates
        
        # Apply rates based on frequency
        unmapped_equipment = set()
        missing_rates = set()
        
        for idx, row in result.iterrows():
            equip_id = row['equip_id']
            frequency = row.get('frequency', 'MONTHLY').upper()
            
            # Default frequency to MONTHLY if not specified or invalid
            if frequency not in ['MONTHLY', 'HOURLY', 'DAILY']:
                frequency = 'MONTHLY'
            
            # Check if equipment exists in rates data
            if equip_id in rates_lookup:
                rate = rates_lookup[equip_id].get(frequency, 0)
                
                if rate == 0:
                    # Rate not available for this frequency
                    missing_rates.add((equip_id, frequency))
                    # Keep the existing rate from the allocation file
                else:
                    # Update to the standard rate
                    result.at[idx, 'rate'] = rate
                    result.at[idx, 'amount'] = row['units'] * rate
            else:
                # Equipment not in rates data
                unmapped_equipment.add(equip_id)
                # Keep the existing rate from the allocation file
        
        # Log warnings for unmapped equipment and missing rates
        if unmapped_equipment:
            logger.warning(f"Equipment IDs not found in rates data: {', '.join(sorted(unmapped_equipment))}")
        
        if missing_rates:
            missing_rates_str = ', '.join(f"{eq} ({freq})" for eq, freq in sorted(missing_rates))
            logger.warning(f"Missing rates for equipment/frequency combinations: {missing_rates_str}")
    
    # Ensure all required columns are present
    for col in STANDARD_COLUMNS:
        if col not in result.columns:
            if col == 'amount':
                # Calculate amount from units and rate
                result['amount'] = result['units'] * result['rate']
            elif col == 'division':
                # Default division to MASTER
                result['division'] = 'MASTER'
            elif col == 'frequency':
                # Default frequency to MONTHLY
                result['frequency'] = 'MONTHLY'
            elif col in ['phase', 'description']:
                # Empty string for text fields
                result[col] = ''
            elif col == 'date':
                # Default date for April 2025
                result[col] = '4/30/2025'
            elif col == 'cost_code':
                # Default cost code
                result[col] = DEFAULT_COST_CODE
    
    # Standardize column order
    result = result[STANDARD_COLUMNS]
    
    # Final data cleaning and validation
    # 1. Replace missing or invalid cost codes with default
    result.loc[result['cost_code'].isna() | 
              result['cost_code'].str.upper().str.contains('NEEDED', na=False), 
              'cost_code'] = DEFAULT_COST_CODE
    
    # 2. Ensure numeric columns are valid
    result['units'] = pd.to_numeric(result['units'], errors='coerce').fillna(0)
    result['rate'] = pd.to_numeric(result['rate'], errors='coerce').fillna(0)
    result['amount'] = pd.to_numeric(result['amount'], errors='coerce').fillna(0)
    
    # 3. Recalculate all amounts to ensure consistency
    result['amount'] = result['units'] * result['rate']
    
    # 4. Filter out rows with zero units or rate
    result = result[result['units'] > 0]
    
    logger.info(f"Generated master billing data with {len(result)} records")
    
    return result

def filter_by_division(master_billing, division):
    """
    Filter the master billing data by division
    
    Args:
        master_billing (DataFrame): The master billing data
        division (str): Division to filter by
        
    Returns:
        DataFrame: Filtered data for the specified division
    """
    if master_billing.empty:
        logger.warning(f"No data to filter for division: {division}")
        return pd.DataFrame()
    
    filtered = master_billing[master_billing['division'] == division].copy()
    
    logger.info(f"Filtered {len(filtered)} records for division: {division}")
    
    return filtered

def export_master_billing(master_billing, output_path, create_division_sheets=True):
    """
    Export the master billing data to Excel
    
    Args:
        master_billing (DataFrame): The billing data to export
        output_path (str): Path to save the Excel file
        create_division_sheets (bool): Whether to create separate sheets for each division
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    try:
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write all data to main sheet
            master_billing.to_excel(writer, sheet_name='Equip Billings', index=False)
            
            # Format the main sheet
            workbook = writer.book
            worksheet = writer.sheets['Equip Billings']
            
            # Apply column widths and formats
            column_widths = {
                'division': 10,
                'equip_id': 12,
                'description': 40,
                'date': 12,
                'job': 15,
                'phase': 10,
                'cost_code': 15,
                'units': 10,
                'rate': 12,
                'frequency': 12,
                'amount': 12
            }
            
            # Set column widths
            for i, column in enumerate(master_billing.columns):
                column_letter = get_column_letter(i+1)
                width = column_widths.get(column, 15)
                worksheet.column_dimensions[column_letter].width = width
            
            # Format headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Format number columns
            currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            number_format = '0.00'
            
            # Get column indices for numeric fields
            units_col = master_billing.columns.get_loc('units') + 1
            rate_col = master_billing.columns.get_loc('rate') + 1
            amount_col = master_billing.columns.get_loc('amount') + 1
            
            # Apply number formats to data rows
            for row in range(2, len(master_billing) + 2):
                # Format units column
                cell = worksheet.cell(row=row, column=units_col)
                cell.number_format = number_format
                
                # Format rate column
                cell = worksheet.cell(row=row, column=rate_col)
                cell.number_format = currency_format
                
                # Format amount column
                cell = worksheet.cell(row=row, column=amount_col)
                cell.number_format = currency_format
            
            # Create division sheets if needed
            if create_division_sheets:
                for division in DIVISIONS:
                    division_data = filter_by_division(master_billing, division)
                    
                    if not division_data.empty:
                        # Write division data to separate sheet
                        division_data.to_excel(writer, sheet_name=division, index=False)
                        
                        # Format division sheet
                        worksheet = writer.sheets[division]
                        
                        # Set column widths
                        for i, column in enumerate(division_data.columns):
                            column_letter = get_column_letter(i+1)
                            width = column_widths.get(column, 15)
                            worksheet.column_dimensions[column_letter].width = width
                        
                        # Apply header formatting
                        for cell in worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Get column indices for numeric fields in this sheet
                        units_col = division_data.columns.get_loc('units') + 1
                        rate_col = division_data.columns.get_loc('rate') + 1
                        amount_col = division_data.columns.get_loc('amount') + 1
                        
                        # Apply number formats to data rows
                        for row in range(2, len(division_data) + 2):
                            # Format units column
                            cell = worksheet.cell(row=row, column=units_col)
                            cell.number_format = number_format
                            
                            # Format rate column
                            cell = worksheet.cell(row=row, column=rate_col)
                            cell.number_format = currency_format
                            
                            # Format amount column
                            cell = worksheet.cell(row=row, column=amount_col)
                            cell.number_format = currency_format
        
        logger.info(f"Successfully exported master billing data to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error exporting master billing data: {str(e)}")
        return False

def export_division_import_file(division_data, output_path):
    """
    Export division-specific data to a CSV file for import
    
    Args:
        division_data (DataFrame): The division data to export
        output_path (str): Path to save the CSV file
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    try:
        # For import files, we need a specific format:
        # Equip #, Date, Job, Cost Code, Rate, Hours, Total Amount
        
        # Create a copy of the data and select/rename columns as needed
        import_data = division_data.copy()
        
        # Ensure all required columns are present
        required_cols = ['equip_id', 'date', 'job', 'cost_code', 'rate', 'units', 'amount']
        for col in required_cols:
            if col not in import_data.columns:
                logger.error(f"Missing required column for import file: {col}")
                return False
        
        # Select and rename columns for the import format
        import_data = import_data[required_cols]
        import_data.columns = ['Equip #', 'Date', 'Job', 'Cost Code', 'Rate', 'Hours', 'Total Amount']
        
        # Ensure data is properly formatted
        # Convert date to string if it's a datetime
        if import_data['Date'].dtype == 'datetime64[ns]':
            import_data['Date'] = import_data['Date'].dt.strftime('%m/%d/%Y')
        
        # Export to CSV without index
        import_data.to_csv(output_path, index=False)
        
        logger.info(f"Successfully exported division import file to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error exporting division import file: {str(e)}")
        return False

def generate_division_exports(master_billing):
    """
    Generate division-specific exports for both Excel and CSV import formats
    
    Args:
        master_billing (DataFrame): The master billing data
        
    Returns:
        dict: Information about the generated exports
    """
    results = {
        'excel_exports': [],
        'import_files': [],
        'errors': []
    }
    
    # Create exports directory if it doesn't exist
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    
    # Export Excel files for each division
    for division in DIVISIONS:
        division_data = filter_by_division(master_billing, division)
        
        if not division_data.empty:
            # Generate Excel export
            excel_output_path = os.path.join(EXPORTS_DIR, DIVISION_EXPORT_PATTERN.format(division))
            excel_success = export_master_billing(division_data, excel_output_path, create_division_sheets=False)
            
            if excel_success:
                results['excel_exports'].append({
                    'division': division,
                    'path': excel_output_path,
                    'record_count': len(division_data)
                })
            else:
                results['errors'].append(f"Failed to export Excel file for {division}")
            
            # Generate CSV import file (for DFW, HOU, WTX)
            if division in DIVISION_IMPORT_FILENAMES:
                import_output_path = os.path.join(EXPORTS_DIR, DIVISION_IMPORT_FILENAMES[division])
                import_success = export_division_import_file(division_data, import_output_path)
                
                if import_success:
                    results['import_files'].append({
                        'division': division,
                        'path': import_output_path,
                        'record_count': len(division_data)
                    })
                else:
                    results['errors'].append(f"Failed to export import file for {division}")
        else:
            logger.warning(f"No data to export for division: {division}")
    
    return results

def process_master_billing():
    """
    Main function to process all allocation files and generate master billing with division exports
    
    Returns:
        dict: Results of the processing
    """
    results = {
        'status': 'error',
        'message': '',
        'record_count': 0,
        'exports': {}
    }
    
    try:
        # Step 1: Find allocation files
        allocation_files = find_all_allocation_files()
        
        if not allocation_files:
            results['message'] = "No allocation files found"
            return results
        
        # Step 2: Find equipment rates file
        rates_file_path = find_equipment_rates_file()
        
        # Step 3: Process each allocation file
        all_allocation_data = []
        valid_file_count = 0
        
        for file_info in allocation_files:
            file_path = file_info['path']
            division = file_info['division']
            
            allocation_data = process_allocation_file(file_path, division)
            
            if not allocation_data.empty:
                all_allocation_data.append(allocation_data)
                valid_file_count += 1
        
        if valid_file_count == 0:
            results['message'] = "Failed to extract data from allocation files"
            return results
        
        # Step 4: Combine all allocation data
        combined_data = pd.concat(all_allocation_data, ignore_index=True)
        
        if combined_data.empty:
            results['message'] = "No valid allocation data found"
            return results
        
        results['record_count'] = len(combined_data)
        
        # Step 5: Extract equipment rates if available
        rates_data = pd.DataFrame()
        if rates_file_path:
            rates_data = extract_equipment_rates(rates_file_path)
        
        # Step 6: Generate master billing data
        master_billing = generate_master_billing(combined_data, rates_data)
        
        if master_billing.empty:
            results['message'] = "Failed to generate master billing data"
            return results
        
        # Step 7: Export master billing workbook
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        master_output_path = os.path.join(EXPORTS_DIR, MASTER_BILLING_FILENAME)
        
        master_export_success = export_master_billing(master_billing, master_output_path)
        
        if not master_export_success:
            results['message'] = "Failed to export master billing workbook"
            return results
        
        # Step 8: Generate division-specific exports
        division_export_results = generate_division_exports(master_billing)
        
        # Compile final results
        results['status'] = 'success'
        results['message'] = f"Successfully processed {valid_file_count} allocation files with {len(master_billing)} records"
        results['exports'] = {
            'master_billing': {
                'path': master_output_path,
                'record_count': len(master_billing)
            },
            'division_exports': division_export_results['excel_exports'],
            'import_files': division_export_results['import_files'],
            'errors': division_export_results['errors']
        }
        
        return results
        
    except Exception as e:
        logger.error(f"Error in process_master_billing: {str(e)}")
        results['message'] = f"Error processing master billing: {str(e)}"
        return results
import logging
from pathlib import Path
import glob
import re
import csv

# Configure debug logging
logging.basicConfig(level=logging.INFO)
# Set up logger
logger = logging.getLogger(__name__)

# Constants
ATTACHED_ASSETS_DIR = Path('attached_assets')
EXPORTS_DIR = Path('exports/pm_master')
DIVISIONS = ['DFW', 'WTX', 'HOU', 'SELECT']
MONTH_NAME = 'APRIL'  # Default month for current processing
YEAR = '2025'  # Default year for current processing
PROCESS_ONLY_APRIL = True  # Set to True to filter only April 2025 files

# Ensure exports directory exists
EXPORTS_DIR.mkdir(exist_ok=True, parents=True)

# Column mapping for standardization across different source files
COLUMN_MAPPING = {
    'equip_id': ['EQ#', 'EQ #', 'EQUIPMENT #', 'EQUIP #', 'EQUIP. #', 'EQUIPMENT_ID', 'EQUIPMENT NUMBER', 'EQUIPMENT', 'TRUCK#'],
    'description': ['DESCRIPTION', 'DESC', 'EQUIPMENT DESC', 'EQUIP DESCRIPTION', 'TYPE', 'EQUIPMENT TYPE'],
    'job': ['JOB', 'JOB #', 'PROJECT', 'JOB_NUMBER', 'JOB_CODE', 'JOB CODE', 'JOB NUMBER', 'JOB NAME', 'PROJECT NAME'],
    'date': ['DATE', 'START DATE', 'START', 'BEGIN DATE', 'ALLOCATION DATE', 'SERVICE DATE', 'BILLING DATE'],
    'units': ['UNITS', 'DAYS', 'HRS', 'HOURS', 'QUANTITY', 'QTY', 'DAYS USED'],
    'frequency': ['FREQUENCY', 'RATE TYPE', 'BILLING TYPE', 'BILL TYPE', 'FREQUENCY TYPE', 'TYPE'],
    'amount': ['AMOUNT', 'TOTAL', 'EXTENDED', 'EXTENDED AMOUNT', 'EXTENDED PRICE', 'COST', 'BILLING', 'BILLINGS'],
    'rate': ['RATE', 'UNIT PRICE', 'PRICE', 'UNIT RATE', 'RATE AMOUNT', 'DAILY RATE', 'MONTHLY RATE', 'HOURLY RATE'],
    'division': ['DIVISION', 'DIV', 'REGION', 'LOCATION', 'BRANCH', 'AREA', 'DISTRICT'],
    'cost_code': ['COST CODE', 'COSTCODE', 'ACCT CODE', 'ACCOUNT CODE', 'CODE'],
    'phase': ['PHASE', 'PH', 'PHASE CODE', 'PHASE NUMBER']
}

# Rate mapping based on frequency
FREQUENCY_MAPPING = {
    'MO': 'MONTHLY',
    'MONTHLY': 'MONTHLY',
    'MONTH': 'MONTHLY',
    'HR': 'HOURLY',
    'HOURLY': 'HOURLY',
    'HOUR': 'HOURLY',
    'DAY': 'DAILY',
    'DAILY': 'DAILY',
    'WK': 'WEEKLY',
    'WEEK': 'WEEKLY',
    'WEEKLY': 'WEEKLY'
}

def find_matching_column(df, patterns):
    """Find the first column that matches one of the patterns"""
    for col in df.columns:
        col_str = str(col).strip().upper()
        for pattern in patterns:
            if re.search(pattern, col_str, re.IGNORECASE):
                return col
    return None

def find_all_allocation_files():
    """Find all allocation files (XLSX and CSV) in the attached_assets directory"""
    allocation_files = []
    
    # Look for Excel files containing EQMO and BILLING ALLOCATIONS in the filename
    if PROCESS_ONLY_APRIL:
        # Restrict to only April 2025 files
        excel_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQMO*BILLING*ALLOCATIONS*APRIL*2025*.xlsx')
    else:
        excel_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQMO*BILLING*ALLOCATIONS*.xlsx')
    
    excel_files = glob.glob(excel_pattern)
    allocation_files.extend(excel_files)
    
    # Look for RAGLE EQ BILLINGS file which contains the rate information
    if PROCESS_ONLY_APRIL:
        ragle_pattern = os.path.join(ATTACHED_ASSETS_DIR, 'RAGLE*EQ*BILLINGS*APRIL*2025*.xls*')
    else:
        ragle_pattern = os.path.join(ATTACHED_ASSETS_DIR, 'RAGLE*EQ*BILLINGS*.xls*')
    
    ragle_files = glob.glob(ragle_pattern)
    allocation_files.extend(ragle_files)
    
    # Look for division-specific CSV files
    for division in DIVISIONS:
        # Try different naming patterns for CSV files - focus on April 2025
        patterns = []
        
        # Special case for WTX - also look for WT prefix
        if division == 'WTX':
            if PROCESS_ONLY_APRIL:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'03 - WT APR 2025.csv'),  # Exact filename match
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APR*2025*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APRIL*2025*.csv'),
                ])
            else:
                patterns.extend([
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'03 - WT*.csv'),  # Exact filename pattern
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APR*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*03*APRIL*.csv'),
                    os.path.join(ATTACHED_ASSETS_DIR, f'*WT*.csv'),
                ])
                
        # Standard patterns for all divisions
        if PROCESS_ONLY_APRIL:
            patterns.extend([
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APR*2025*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APRIL*2025*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'SM*{division}*APR*2025*.csv'),
            ])
        else:
            patterns.extend([
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APR*2025*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*APRIL*2025*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'SM*{division}*APR*.csv'),
                os.path.join(ATTACHED_ASSETS_DIR, f'*{division}*.csv'),
            ])
            
        # Process all the patterns for this division
        for pattern in patterns:
            csv_files = glob.glob(pattern)
            allocation_files.extend(csv_files)
    
    # Remove duplicates while preserving order
    seen = set()
    allocation_files = [x for x in allocation_files if not (x in seen or seen.add(x))]
    
    logger.info(f"Found {len(allocation_files)} allocation files matching patterns")
    return allocation_files

def find_equipment_rates_file():
    """Find the equipment rates file or master billing workbook"""
    # First, look for RAGLE EQ BILLINGS file which has the April 2025 rates
    ragle_pattern = os.path.join(ATTACHED_ASSETS_DIR, 'RAGLE*EQ*BILLINGS*APRIL*2025*.xls*')
    files = glob.glob(ragle_pattern)
    
    if not files:
        # Try other patterns for the RAGLE file
        ragle_pattern = os.path.join(ATTACHED_ASSETS_DIR, 'RAGLE*EQ*BILLINGS*.xls*')
        files = glob.glob(ragle_pattern)
    
    if files:
        logger.info(f"Found equipment rates in RAGLE file: {os.path.basename(files[0])}")
        return files[0]
    
    # Look for the EQ MONTHLY BILLINGS file as fallback
    eq_monthly_pattern = os.path.join(ATTACHED_ASSETS_DIR, 'EQ*MONTHLY*BILLINGS*.xlsx')
    files = glob.glob(eq_monthly_pattern)
    
    if files:
        logger.info(f"Found equipment rates in monthly billing spreadsheet: {os.path.basename(files[0])}")
        return files[0]
    
    # As a last resort, look for any equipment rates workbook
    rates_pattern = os.path.join(ATTACHED_ASSETS_DIR, '*EQUIP*RATES*.xlsx')
    files = glob.glob(rates_pattern)
    
    if files:
        logger.info(f"Found equipment rates file: {os.path.basename(files[0])}")
        return files[0]
    
    logger.warning("No equipment rates file found. Will proceed without rates.")
    return None

def extract_equipment_rates(rates_file):
    """Extract equipment rates from the rates file"""
    if not rates_file:
        return pd.DataFrame()
    
    try:
        # Check if it's a RAGLE file which has a specific format
        is_ragle_file = 'RAGLE' in os.path.basename(rates_file).upper()
        
        # Load the workbook
        wb = openpyxl.load_workbook(rates_file, data_only=True)
        
        # For RAGLE files, look for the "Equip Rates" sheet specifically
        if is_ragle_file:
            rates_sheet = None
            for sheet_name in wb.sheetnames:
                if 'EQUIP RATES' in sheet_name.upper():
                    rates_sheet = sheet_name
                    break
            
            if not rates_sheet:
                # If no "Equip Rates" sheet, look for sheets with "RATES" in the name
                for sheet_name in wb.sheetnames:
                    if 'RATES' in sheet_name.upper():
                        rates_sheet = sheet_name
                        break
        # For non-RAGLE files, use a more general approach
        else:
            rates_sheet = None
            sheet_patterns = ['EQUIP RATES', 'RATES', 'EQUIPMENT', 'BILLING RATES', 'MONTHLY RATES']
            
            for pattern in sheet_patterns:
                for sheet_name in wb.sheetnames:
                    if pattern in sheet_name.upper():
                        rates_sheet = sheet_name
                        break
                if rates_sheet:
                    break
        
        if not rates_sheet:
            rates_sheet = wb.sheetnames[0]  # Default to first sheet
            
        logger.info(f"Using sheet '{rates_sheet}' for equipment rates from file: {os.path.basename(rates_file)}")
        
        # Load the sheet into a DataFrame
        df = pd.read_excel(rates_file, sheet_name=rates_sheet)
        
        # Find the equipment ID and rate columns
        equip_col = find_matching_column(df, COLUMN_MAPPING['equip_id'])
        monthly_rate_col = find_matching_column(df, ['MONTHLY RATE', 'MONTHLY', 'MO RATE'])
        daily_rate_col = find_matching_column(df, ['DAILY RATE', 'DAILY', 'DAY RATE'])
        hourly_rate_col = find_matching_column(df, ['HOURLY RATE', 'HOURLY', 'HR RATE'])
        desc_col = find_matching_column(df, COLUMN_MAPPING['description'])
        
        if not equip_col:
            logger.error("Equipment ID column not found in rates file")
            return pd.DataFrame()
        
        # Create a standardized rates DataFrame
        rates_df = pd.DataFrame()
        rates_df['EQUIPMENT_ID'] = df[equip_col].astype(str).str.strip().str.upper()
        
        # Add description if available
        if desc_col:
            rates_df['DESCRIPTION'] = df[desc_col].astype(str).str.strip()
        else:
            rates_df['DESCRIPTION'] = ""
        
        # Add rates by frequency
        if monthly_rate_col:
            rates_df['MONTHLY_RATE'] = pd.to_numeric(df[monthly_rate_col], errors='coerce').fillna(0)
        else:
            rates_df['MONTHLY_RATE'] = 0
            
        if daily_rate_col:
            rates_df['DAILY_RATE'] = pd.to_numeric(df[daily_rate_col], errors='coerce').fillna(0)
        else:
            rates_df['DAILY_RATE'] = 0
            
        if hourly_rate_col:
            rates_df['HOURLY_RATE'] = pd.to_numeric(df[hourly_rate_col], errors='coerce').fillna(0)
        else:
            rates_df['HOURLY_RATE'] = 0
        
        # Remove rows with no equipment ID or all zero rates
        rates_df = rates_df[rates_df['EQUIPMENT_ID'].notna()]
        rates_df = rates_df[~(rates_df['EQUIPMENT_ID'] == '')]
        rates_df = rates_df[~((rates_df['MONTHLY_RATE'] == 0) & 
                              (rates_df['DAILY_RATE'] == 0) & 
                              (rates_df['HOURLY_RATE'] == 0))]
        
        # Standardize equipment ID format to match allocation files
        rates_df['EQUIPMENT_ID'] = rates_df['EQUIPMENT_ID'].str.replace(r'\s+', '', regex=True)
        
        logger.info(f"Extracted {len(rates_df)} equipment rates")
        return rates_df
        
    except Exception as e:
        logger.error(f"Error extracting equipment rates: {str(e)}")
        return pd.DataFrame()

def process_allocation_file(file_path):
    """Process a single allocation file to extract equipment billing data"""
    logger.info(f"Processing allocation file: {os.path.basename(file_path)}")
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        # Handle different file formats
        if file_ext == '.csv':
            # Determine if it's a division-specific CSV
            filename = os.path.basename(file_path).upper()
            is_division_csv = any(div in filename for div in DIVISIONS) or any(prefix in filename for prefix in ["01 - ", "02 - ", "03 - ", "SM - "])
            
            # For division-specific CSVs which might have different formats
            if is_division_csv:
                logger.info(f"Processing division-specific CSV: {filename}")
                
                # Determine division from filename
                if "DFW" in filename or "01 -" in filename:
                    division = "DFW"
                elif "HOU" in filename or "02 -" in filename:
                    division = "HOU"
                elif "WTX" in filename or "WT" in filename or "03 -" in filename:
                    division = "WTX"
                elif "SELECT" in filename or "SM -" in filename:
                    division = "SELECT"
                else:
                    division = "OTHER"
                
                # First try to check if the file has headers or not
                # These division CSV files typically don't have headers
                try:
                    # Read just the first row to check the content
                    sample = pd.read_csv(file_path, nrows=1)
                    # If the first column looks like an equipment ID (e.g., PT-159), it's likely headerless
                    first_col_value = str(sample.iloc[0, 0]).strip()
                    
                    if re.match(r'^[A-Z]{1,3}-\d+$', first_col_value):
                        # This is a headerless CSV with the format we know
                        logger.info(f"Detected headerless division-specific CSV: {filename}")
                        
                        # Define the headers based on the CSV file format we're seeing
                        headers = ["equip_id", "description", "date", "job", "phase", "cost_code", 
                                "units", "rate", "frequency", "monthly_rate", "amount"]
                        
                        # Add specific handling for divisions
                        if "WT" in filename or "03" in filename:
                            logger.info(f"Processing WTX division file: {filename}")
                            division = "WTX"
                        elif "DFW" in filename or "01" in filename:
                            logger.info(f"Processing DFW division file: {filename}")
                            division = "DFW"
                        elif "HOU" in filename or "02" in filename:
                            logger.info(f"Processing HOU division file: {filename}")
                            division = "HOU"
                        elif "SELECT" in filename or "SM" in filename:
                            logger.info(f"Processing SELECT division file: {filename}")
                            division = "SELECT"
                            
                        # Read the CSV file without headers
                        try:
                            df = pd.read_csv(file_path, header=None, names=headers)
                            # After reading, check data validity
                            logger.info(f"Successfully loaded {len(df)} records with {len(df.columns)} columns")
                            # Log first row for debugging
                            if not df.empty:
                                logger.info(f"First row sample: {df.iloc[0].to_dict()}")
                        except Exception as e:
                            logger.error(f"Error reading CSV: {str(e)}")
                            # Try reading with different encoding or delimiter
                            try:
                                df = pd.read_csv(file_path, header=None, names=headers, encoding='latin1')
                                logger.info(f"Successfully loaded with latin1 encoding: {len(df)} records")
                            except Exception as e2:
                                logger.error(f"Failed with alternative encoding: {str(e2)}")
                                return pd.DataFrame()
                        
                        # Add division column
                        df['division'] = division
                        logger.info(f"Processed headerless CSV with {len(df)} records from {division}")
                    else:
                        # It has headers, process normally
                        df = pd.read_csv(file_path)
                except Exception as e:
                    logger.warning(f"Error detecting headers: {str(e)}. Trying standard processing.")
                    # Try with various options to handle different CSV formats
                    try:
                        df = pd.read_csv(file_path, encoding='utf-8')
                    except:
                        try:
                            df = pd.read_csv(file_path, encoding='latin1')
                        except:
                            df = pd.read_csv(file_path, encoding='utf-8', sep=';')
            else:
                # Standard CSV
                df = pd.read_csv(file_path)
        
        elif file_ext in ['.xlsx', '.xls', '.xlsm']:
            # For Excel files, we need to determine the correct sheet
            excel_file = pd.ExcelFile(file_path)
            
            # Look for sheets with allocation or billing data
            sheet_name = None
            sheet_keywords = ['BILLINGS', 'ALLOCATION', 'DATA', 'EQUIP BILLINGS', 'EQUIP']
            
            for keyword in sheet_keywords:
                matching_sheets = [s for s in excel_file.sheet_names if keyword in s.upper()]
                if matching_sheets:
                    sheet_name = matching_sheets[0]
                    break
            
            # If no specific sheet found, use the first sheet
            if not sheet_name and excel_file.sheet_names:
                sheet_name = excel_file.sheet_names[0]
            
            if sheet_name:
                logger.info(f"Using sheet '{sheet_name}' from {os.path.basename(file_path)}")
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                logger.error(f"No valid sheets found in {os.path.basename(file_path)}")
                return pd.DataFrame()
        else:
            logger.error(f"Unsupported file format: {file_ext}")
            return pd.DataFrame()
        
        # Find standardized column names
        data = {}
        for col_type, patterns in COLUMN_MAPPING.items():
            col = find_matching_column(df, patterns)
            if col:
                data[col_type] = col
        
        # Skip if essential columns are missing
        if 'equip_id' not in data or 'job' not in data:
            logger.warning(f"Missing essential columns in file: {os.path.basename(file_path)}")
            
            # Special handling for division-specific CSVs which might have non-standard formats
            filename = os.path.basename(file_path).upper()
            if any(div in filename for div in DIVISIONS):
                logger.info(f"Attempting special handling for division file: {filename}")
                
                # For some division files, the equipment might be in a different format
                # Try to identify potential equipment ID columns based on values matching equipment patterns
                if 'equip_id' not in data:
                    for col in df.columns:
                        # Check if column values match typical equipment number patterns
                        sample_values = df[col].astype(str).str.strip().dropna().head(10).tolist()
                        if any(re.match(r'^[A-Z]+\d+$', str(val)) for val in sample_values):
                            data['equip_id'] = col
                            logger.info(f"Found potential equipment ID column: {col}")
                            break
                
                # For job column, look for columns with typical job number formats
                if 'job' not in data:
                    for col in df.columns:
                        sample_values = df[col].astype(str).str.strip().dropna().head(10).tolist()
                        if any(re.match(r'^\d{4}-\d{3}$', str(val)) for val in sample_values):
                            data['job'] = col
                            logger.info(f"Found potential job column: {col}")
                            break
            
            # If still missing essential columns, skip this file
            if 'equip_id' not in data or 'job' not in data:
                logger.error(f"Unable to find essential columns in file: {os.path.basename(file_path)}")
                return pd.DataFrame()
        
        # Extract filename elements to determine division if not found in the data
        if 'division' not in data:
            filename = os.path.basename(file_path).upper()
            for div in DIVISIONS:
                if div in filename:
                    logger.info(f"Determined division {div} from filename")
                    df['DIVISION'] = div
                    data['division'] = 'DIVISION'
                    break
            
            # If still no division, check if we can determine from file path or contents
            if 'division' not in data:
                # Check for patterns in sheet content that might indicate division
                for div in DIVISIONS:
                    # Check if division appears in any cell in the first few rows
                    for col in df.columns:
                        if df[col].astype(str).str.contains(div, case=False, na=False).any():
                            logger.info(f"Determined division {div} from content")
                            df['DIVISION'] = div
                            data['division'] = 'DIVISION'
                            break
                    if 'division' in data:
                        break
        
        # Standardize column names for further processing
        rename_dict = {data[col_type]: col_type for col_type in data}
        
        # Add any missing columns with default values
        for col_type in COLUMN_MAPPING.keys():
            if col_type not in data:
                df[col_type] = None
                
                # Provide default values for required fields
                if col_type == 'frequency':
                    df[col_type] = 'MONTHLY'  # Default billing frequency
                elif col_type == 'date':
                    df[col_type] = f'APRIL 1, {YEAR}'  # Default date
                elif col_type == 'phase':
                    df[col_type] = '01'  # Default phase
                elif col_type == 'cost_code':
                    df[col_type] = '10-100'  # Default cost code
        
        # Rename columns to standardized names
        df = df.rename(columns=rename_dict)
        
        # Filter out rows with no equipment ID
        df = df[df['equip_id'].notna()]
        
        # Standardize equipment ID format (remove spaces, leading zeros, etc.)
        df['equip_id'] = df['equip_id'].astype(str).str.strip().str.upper()
        df['equip_id'] = df['equip_id'].str.replace(r'\s+', '', regex=True)  # Remove all whitespace
        
        # Ensure numeric columns are correctly typed
        if 'units' in df.columns:
            df['units'] = pd.to_numeric(df['units'], errors='coerce').fillna(0)
        
        if 'rate' in df.columns:
            df['rate'] = pd.to_numeric(df['rate'], errors='coerce').fillna(0)
        
        if 'amount' in df.columns:
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
        
        # Standardize job numbers
        if 'job' in df.columns:
            # Standardize job number format (e.g., '2024-015')
            df['job'] = df['job'].astype(str).str.strip().str.upper()
            
            # Fix common job number format issues
            # Convert patterns like "24-15" to "2024-015"
            df['job'] = df['job'].apply(lambda x: re.sub(r'^(\d{2})-(\d{1,3})$', 
                                                        lambda m: f'20{m.group(1)}-{m.group(2).zfill(3)}', 
                                                        str(x)) if pd.notna(x) else x)
        
        # Add source file column for tracking
        df['source_file'] = os.path.basename(file_path)
        
        logger.info(f"Successfully processed {os.path.basename(file_path)}: Found {len(df)} equipment records")
        return df
    
    except Exception as e:
        logger.error(f"Error processing file {os.path.basename(file_path)}: {str(e)}")
        return pd.DataFrame()

def generate_master_billing(allocation_data, rates_data):
    """Generate the master billing dataset by applying rates to allocation data"""
    if allocation_data.empty:
        logger.error("No allocation data provided")
        return pd.DataFrame()
    
    logger.info(f"Generating master billing from {len(allocation_data)} allocation records")
    
    # Create a copy to avoid modifying the original
    billing_df = allocation_data.copy()
    
    # Ensure all column names are lowercased for consistency
    billing_df.columns = [col.lower() for col in billing_df.columns]
    
    # Make sure required columns exist
    required_columns = ['equip_id', 'job', 'units', 'rate', 'amount', 'division']
    for col in required_columns:
        if col not in billing_df.columns:
            billing_df[col] = None if col != 'units' and col != 'rate' and col != 'amount' else 0
    
    # Convert numeric columns
    numeric_cols = ['units', 'rate', 'amount']
    for col in numeric_cols:
        if col in billing_df.columns:
            billing_df[col] = pd.to_numeric(billing_df[col], errors='coerce').fillna(0)
    
    # Initialize new columns for rate-adjusted values
    billing_df['HAS_RATE'] = False
    billing_df['RATE_SOURCE'] = 'Default'
    
    # Apply rates from rates data if available
    if not rates_data.empty:
        logger.info(f"Applying rates from rates data with {len(rates_data)} equipment rate records")
        
        # Merge with rates data to get the appropriate rate
        for idx, row in billing_df.iterrows():
            equip_id = row['equip_id']
            frequency = str(row['frequency']).upper() if pd.notna(row['frequency']) else 'MONTHLY'
            
            # Find matching equipment in rates data
            rate_matches = rates_data[rates_data['EQUIPMENT_ID'] == equip_id]
            
            if not rate_matches.empty:
                # Determine which rate to use based on frequency
                if frequency == 'MONTHLY' and rate_matches['MONTHLY_RATE'].iloc[0] > 0:
                    billing_df.at[idx, 'rate'] = rate_matches['MONTHLY_RATE'].iloc[0]
                    billing_df.at[idx, 'HAS_RATE'] = True
                    billing_df.at[idx, 'RATE_SOURCE'] = 'Monthly Rate'
                elif frequency == 'DAILY' and rate_matches['DAILY_RATE'].iloc[0] > 0:
                    billing_df.at[idx, 'rate'] = rate_matches['DAILY_RATE'].iloc[0]
                    billing_df.at[idx, 'HAS_RATE'] = True
                    billing_df.at[idx, 'RATE_SOURCE'] = 'Daily Rate'
                elif frequency == 'HOURLY' and rate_matches['HOURLY_RATE'].iloc[0] > 0:
                    billing_df.at[idx, 'rate'] = rate_matches['HOURLY_RATE'].iloc[0]
                    billing_df.at[idx, 'HAS_RATE'] = True
                    billing_df.at[idx, 'RATE_SOURCE'] = 'Hourly Rate'
                # If frequency doesn't match available rates, find the first non-zero rate
                else:
                    if rate_matches['MONTHLY_RATE'].iloc[0] > 0:
                        billing_df.at[idx, 'rate'] = rate_matches['MONTHLY_RATE'].iloc[0]
                        billing_df.at[idx, 'frequency'] = 'MONTHLY'
                        billing_df.at[idx, 'HAS_RATE'] = True
                        billing_df.at[idx, 'RATE_SOURCE'] = 'Monthly Rate (Default)'
                    elif rate_matches['DAILY_RATE'].iloc[0] > 0:
                        billing_df.at[idx, 'rate'] = rate_matches['DAILY_RATE'].iloc[0]
                        billing_df.at[idx, 'frequency'] = 'DAILY'
                        billing_df.at[idx, 'HAS_RATE'] = True
                        billing_df.at[idx, 'RATE_SOURCE'] = 'Daily Rate (Default)'
                    elif rate_matches['HOURLY_RATE'].iloc[0] > 0:
                        billing_df.at[idx, 'rate'] = rate_matches['HOURLY_RATE'].iloc[0]
                        billing_df.at[idx, 'frequency'] = 'HOURLY'
                        billing_df.at[idx, 'HAS_RATE'] = True
                        billing_df.at[idx, 'RATE_SOURCE'] = 'Hourly Rate (Default)'
                
                # Update description if it's empty and available in rates data
                if (not pd.notna(row['description']) or row['description'] == '') and pd.notna(rate_matches['DESCRIPTION'].iloc[0]):
                    billing_df.at[idx, 'description'] = rate_matches['DESCRIPTION'].iloc[0]
            
            # If equipment is missing a rate, log it
            if not billing_df.at[idx, 'HAS_RATE']:
                logger.warning(f"No matching rate found for equipment {equip_id} with frequency {frequency}")
    
    # Calculate amount if rate and units are available but amount is not
    for idx, row in billing_df.iterrows():
        if pd.notna(row['rate']) and pd.notna(row['units']) and row['rate'] > 0 and row['units'] > 0:
            if not pd.notna(row['amount']) or row['amount'] == 0:
                billing_df.at[idx, 'amount'] = row['rate'] * row['units']
    
    # Create standardized master billing format
    master_billing = pd.DataFrame({
        'Division': billing_df['division'],
        'Equip #': billing_df['equip_id'],
        'Description': billing_df['description'],
        'Date': billing_df['date'],
        'Job': billing_df['job'],
        'Phase': billing_df['phase'],
        'Cost Code': billing_df['cost_code'],
        'Units': billing_df['units'],
        'Rate': billing_df['rate'],
        'Amount': billing_df['amount'],
        'Frequency': billing_df['frequency'],
        'Source File': billing_df['source_file'],
        'Has Rate': billing_df['HAS_RATE'],
        'Rate Source': billing_df['RATE_SOURCE']
    })
    
    # Log missing rates
    missing_rates = master_billing[master_billing['Rate'] == 0]
    if not missing_rates.empty:
        logger.warning(f"Missing rates for {len(missing_rates)} equipment records")
    
    return master_billing

def filter_by_division(master_billing, division):
    """Filter the master billing data by division"""
    if master_billing.empty:
        return pd.DataFrame()
    
    # Division might be stored differently (case, spaces, etc.)
    division_filter = master_billing['Division'].str.upper().str.strip() == division.upper()
    
    return master_billing[division_filter].copy()

def export_master_billing(master_billing, output_path, create_division_sheets=False):
    """
    Export the master billing data to Excel
    
    Args:
        master_billing (DataFrame): The billing data to export
        output_path (str): Path to save the Excel file
        create_division_sheets (bool): Whether to create separate sheets for each division
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    if master_billing.empty:
        logger.error("No data to export")
        return False
    
    try:
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save to Excel
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # Create the main sheet with all data
        master_billing.to_excel(writer, sheet_name='Equip Billings', index=False)
        
        # Add division-specific sheets if requested (for master export)
        if create_division_sheets:
            for division in DIVISIONS:
                division_data = filter_by_division(master_billing, division)
                if not division_data.empty:
                    # Create a sheet for this division
                    division_data.to_excel(writer, sheet_name=division, index=False)
                    logger.info(f"Added {division} sheet with {len(division_data)} records")
        
        # Access the workbook and the main worksheet
        workbook = writer.book
        worksheet = writer.sheets['Equip Billings']
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072C6", end_color="0072C6", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting
        for col_num, col_name in enumerate(master_billing.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Auto-size columns
            worksheet.column_dimensions[get_column_letter(col_num)].width = max(12, len(str(col_name)) + 2)
        
        # Remove the title row insertion which is causing problems
        title_cell = worksheet.cell(row=1, column=1)
        # Don't insert rows, just use existing header row
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Merge cells for title
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(master_billing.columns))
        
        # Format numeric columns - openpyxl doesn't use add_format like xlsxwriter
        for row_idx in range(2, len(master_billing) + 2):  # +2 for header row
            # Get the column index for Rate and Amount
            rate_idx = 0
            amount_idx = 0
            for i, col_name in enumerate(master_billing.columns, 1):
                if col_name == 'Rate':
                    rate_idx = i
                if col_name == 'Amount':
                    amount_idx = i
                    
            if rate_idx > 0:
                rate_cell = worksheet.cell(row=row_idx, column=rate_idx)
                if rate_cell.value and isinstance(rate_cell.value, (int, float)):
                    rate_cell.number_format = '#,##0.00'
            
            if amount_idx > 0:
                amount_cell = worksheet.cell(row=row_idx, column=amount_idx)
                if amount_cell.value and isinstance(amount_cell.value, (int, float)):
                    amount_cell.number_format = '#,##0.00'
        
        # Skip complex formatting that might cause errors
        
        # Just add a simple summary row at the end
        summary_row = len(master_billing) + 3
        worksheet.cell(row=summary_row, column=1).value = "Total Records:"
        worksheet.cell(row=summary_row, column=2).value = len(master_billing)
        
        # Add the same for division sheets
        if create_division_sheets:
            for division in DIVISIONS:
                if division in writer.sheets:
                    div_sheet = writer.sheets[division]
                    division_data = filter_by_division(master_billing, division)
                    if not division_data.empty:
                        summary_row = len(division_data) + 3
                        div_sheet.cell(row=summary_row, column=1).value = "Total Records:"
                        div_sheet.cell(row=summary_row, column=2).value = len(division_data)
        
        worksheet.cell(row=summary_row+4, column=1).value = "Export Date:"
        worksheet.cell(row=summary_row+4, column=2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Save the workbook
        writer.close()
        
        logger.info(f"Successfully exported master billing to {output_path}")
        return True
    
    except Exception as e:
        logger.error(f"Error exporting master billing: {str(e)}")
        return False

def generate_division_exports(master_billing):
    """Generate division-specific exports"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    export_paths = {}
    for division in DIVISIONS:
        division_data = filter_by_division(master_billing, division)
        
        if division_data.empty:
            logger.warning(f"No data found for division {division}")
            continue
        
        output_path = os.path.join(EXPORTS_DIR, f"{division}_EQUIP_BILLINGS_{MONTH_NAME}_{YEAR}.xlsx")
        if export_master_billing(division_data, output_path):
            export_paths[division] = output_path
            logger.info(f"Created {division} export with {len(division_data)} records")
    
    return export_paths

def process_master_billing():
    """Main function to process all allocation files and generate master billing with division exports"""
    try:
        logger.info("Starting PM Master Billing processing")
        
        # Find all allocation files
        allocation_files = find_all_allocation_files()
        if not allocation_files:
            logger.error("No allocation files found")
            return {
                'success': False,
                'message': "No allocation files found",
                'exports': {}
            }
        
        # Find equipment rates file
        rates_file = find_equipment_rates_file()
        
        # Process each allocation file
        all_allocation_data = []
        for file_path in allocation_files:
            # Skip rates file if it's in the allocation files list
            if rates_file and os.path.samefile(file_path, rates_file):
                continue
                
            allocation_data = process_allocation_file(file_path)
            if not allocation_data.empty:
                all_allocation_data.append(allocation_data)
        
        if not all_allocation_data:
            logger.error("Failed to extract data from allocation files")
            return {
                'success': False,
                'message': "Failed to extract data from allocation files",
                'exports': {}
            }
        
        # Combine all allocation data
        combined_data = pd.concat(all_allocation_data, ignore_index=True)
        logger.info(f"Extracted {len(combined_data)} billing records from {len(allocation_files)} files")
        
        # Extract equipment rates
        rates_data = extract_equipment_rates(rates_file)
        
        # Generate master billing
        master_billing = generate_master_billing(combined_data, rates_data)
        
        # Export master billing file with division-specific sheets as requested
        # This will create a single Excel file with sheets named DFW, WTX, HOU, SELECT
        master_output_path = os.path.join(EXPORTS_DIR, f"MASTER_EQUIP_BILLINGS_{MONTH_NAME}_{YEAR}.xlsx")
        master_export_success = export_master_billing(master_billing, master_output_path, create_division_sheets=True)
        
        # Create exports dictionary with master billing file
        exports = {
            'MASTER': master_output_path if master_export_success else None
        }
        
        # Also create individual division export files for users who need them
        for division in DIVISIONS:
            division_data = filter_by_division(master_billing, division)
            if not division_data.empty:
                division_output_path = os.path.join(EXPORTS_DIR, f"{division}_EQUIP_BILLINGS_{MONTH_NAME}_{YEAR}.xlsx")
                if export_master_billing(division_data, division_output_path):
                    exports[division] = division_output_path
                    logger.info(f"Created {division} export with {len(division_data)} records")
        
        logger.info("PM Master Billing processing completed successfully")
        return {
            'success': True,
            'message': f"Successfully processed {len(allocation_files)} allocation files and generated {len(exports)} exports",
            'record_count': len(master_billing),
            'file_count': len(allocation_files),
            'exports': exports,
            'missing_rates': len(master_billing[(master_billing['Rate'] == 0) | (master_billing['Amount'] == 0)])
        }
    
    except Exception as e:
        logger.error(f"Error processing master billing: {str(e)}")
        return {
            'success': False,
            'message': f"Error processing master billing: {str(e)}",
            'exports': {}
        }

if __name__ == "__main__":
    # If run as a script, process all files
    result = process_master_billing()
    print(result)
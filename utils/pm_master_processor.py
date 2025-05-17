"""
PM Master Billing Processor

This module provides tools for processing multiple PM allocation sheets and generating 
a master completed output with comprehensive tracking and reconciliation.
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from pathlib import Path
import json

logger = logging.getLogger(__name__)

# File paths
BASE_DIR = Path('.')
ATTACHED_ASSETS_DIR = BASE_DIR / 'attached_assets'
EXPORTS_DIR = BASE_DIR / 'exports'
EXPORTS_DIR.mkdir(exist_ok=True)

# Column definitions for PM allocation sheets
PM_COLS = {
    'job_number': 'Job Code',
    'job_name': 'Job Name',
    'equipment_id': 'Equipment #',
    'description': 'Description',
    'category': 'Category',
    'start_date': 'Start',
    'end_date': 'End',
    'days': 'Days',
    'rate': 'Rate',
    'amount': 'Amount',
    'notes': 'Notes'
}

class PMProcessingException(Exception):
    """Custom exception for PM processing errors"""
    pass

def load_pm_file(file_path):
    """
    Load PM allocation file and standardize data format
    
    Args:
        file_path (str or Path): Path to the PM allocation file
        
    Returns:
        pandas.DataFrame: Standardized PM allocation data
    """
    try:
        # Try to determine file type
        if isinstance(file_path, str):
            file_path = Path(file_path)
            
        file_extension = file_path.suffix.lower()
        
        if file_extension == '.xlsx' or file_extension == '.xls':
            # Try to load the Excel file
            try:
                # First attempt with specific sheets
                try:
                    df = pd.read_excel(file_path, sheet_name='PM Data', engine='openpyxl')
                except:
                    try:
                        df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl')
                    except:
                        # If specific sheets fail, try reading the first sheet
                        xl = pd.ExcelFile(file_path, engine='openpyxl')
                        sheet_name = xl.sheet_names[0]
                        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            except Exception as e:
                logger.error(f"Error reading Excel file {file_path}: {str(e)}")
                raise PMProcessingException(f"Could not read Excel file: {str(e)}")
        elif file_extension == '.csv':
            # Load CSV file
            try:
                df = pd.read_csv(file_path)
            except Exception as e:
                logger.error(f"Error reading CSV file {file_path}: {str(e)}")
                raise PMProcessingException(f"Could not read CSV file: {str(e)}")
        else:
            raise PMProcessingException(f"Unsupported file type: {file_extension}")
        
        # Basic validation
        if df.empty:
            raise PMProcessingException("File contains no data")
            
        # Detect column names for mapping
        detected_columns = {}
        for std_col, possible_names in PM_COLS.items():
            # Try to find a matching column
            if isinstance(possible_names, str):
                possible_names = [possible_names]
                
            for name in possible_names:
                matching_cols = [col for col in df.columns if name.lower() in col.lower()]
                if matching_cols:
                    detected_columns[std_col] = matching_cols[0]
                    break
                    
        # If we didn't find the essential columns, try to find by position
        essential_columns = ['job_number', 'equipment_id', 'days', 'amount']
        missing_essential = [col for col in essential_columns if col not in detected_columns]
        
        if missing_essential:
            # Try to infer columns by position
            if len(df.columns) >= 10:  # A typical PM sheet has at least 10 columns
                col_map = {
                    0: 'job_number',
                    1: 'job_name',
                    2: 'equipment_id',
                    3: 'description',
                    4: 'category',
                    5: 'start_date',
                    6: 'end_date',
                    7: 'days',
                    8: 'rate',
                    9: 'amount'
                }
                
                for idx, std_col in col_map.items():
                    if idx < len(df.columns) and std_col in missing_essential:
                        detected_columns[std_col] = df.columns[idx]
        
        # Check if we found all essential columns
        still_missing = [col for col in essential_columns if col not in detected_columns]
        if still_missing:
            missing_cols = ', '.join(still_missing)
            raise PMProcessingException(f"Could not find essential columns: {missing_cols}")
            
        # Create a new dataframe with standardized columns
        std_df = pd.DataFrame()
        
        # Copy data to standardized columns
        for std_col, file_col in detected_columns.items():
            std_df[std_col] = df[file_col]
            
        # Fill in missing standard columns
        for std_col in PM_COLS.keys():
            if std_col not in std_df.columns:
                std_df[std_col] = None
                
        # Make sure numeric columns are numbers
        for col in ['days', 'rate', 'amount']:
            if col in std_df.columns:
                std_df[col] = pd.to_numeric(std_df[col], errors='coerce')
                
        # Make sure date columns are dates
        for col in ['start_date', 'end_date']:
            if col in std_df.columns:
                std_df[col] = pd.to_datetime(std_df[col], errors='coerce')
                
        # Add source file name
        std_df['source_file'] = file_path.name
        
        # Add a unique identifier for each row (essential for tracking changes)
        std_df['row_id'] = std_df.apply(
            lambda row: f"{row['job_number']}_{row['equipment_id']}_{row.name}", 
            axis=1
        )
        
        return std_df
        
    except PMProcessingException as e:
        # Re-raise the custom exception
        raise
    except Exception as e:
        # Log the error and raise a custom exception
        logger.error(f"Error loading PM file {file_path}: {str(e)}")
        raise PMProcessingException(f"Error loading PM file: {str(e)}")


class PMMasterProcessor:
    """
    Class for processing multiple PM allocation sheets and generating a master output
    """
    
    def __init__(self, original_file=None):
        """
        Initialize the processor
        
        Args:
            original_file (str or Path, optional): Path to the original PM file
        """
        self.master_data = pd.DataFrame()
        self.processed_files = []
        self.original_file = original_file
        
        # If original file is provided, load it
        if original_file:
            try:
                self.master_data = load_pm_file(original_file)
                self.processed_files.append(str(original_file))
            except PMProcessingException as e:
                logger.error(f"Error loading original file: {str(e)}")
                raise
    
    def add_pm_file(self, file_path):
        """
        Add a PM allocation file to the master data
        
        Args:
            file_path (str or Path): Path to the PM allocation file
            
        Returns:
            dict: Report of changes made
        """
        try:
            # Load the new file
            new_data = load_pm_file(file_path)
            
            # Skip if empty
            if new_data.empty:
                return {
                    'success': False,
                    'message': 'File contains no data'
                }
                
            # If master data is empty, use this as the original
            if self.master_data.empty:
                self.master_data = new_data
                self.processed_files.append(str(file_path))
                
                return {
                    'success': True,
                    'message': 'Loaded as original file',
                    'file': str(file_path),
                    'rows': len(new_data),
                    'changes': {
                        'additions': len(new_data),
                        'modifications': 0,
                        'deletions': 0
                    }
                }
                
            # Process changes
            return self._integrate_new_data(new_data, file_path)
                
        except PMProcessingException as e:
            # Re-raise the custom exception
            raise
        except Exception as e:
            # Log the error and raise a custom exception
            logger.error(f"Error adding PM file {file_path}: {str(e)}")
            raise PMProcessingException(f"Error adding PM file: {str(e)}")
    
    def get_original_data_metrics(self):
        """
        Get metrics from the original data file for dashboard display
        
        Returns:
            dict: Metrics about the original data
        """
        if self.master_data.empty:
            return {}
            
        # Filter to only get original data (version 1 or no version info)
        original_data = self.master_data[
            (self.master_data['version'] == 1) | 
            (~self.master_data['version'].isin(self.master_data['version'].unique()))
        ]
        
        if original_data.empty:
            original_data = self.master_data
            
        # Calculate metrics
        total_amount = original_data['amount'].sum() if 'amount' in original_data.columns else 0
        job_distribution = {}
        
        if 'job_number' in original_data.columns:
            job_counts = original_data.groupby('job_number').size().reset_index(name='count')
            job_amounts = original_data.groupby('job_number')['amount'].sum().reset_index()
            
            job_summary = pd.merge(job_counts, job_amounts, on='job_number')
            
            for _, row in job_summary.iterrows():
                job_distribution[row['job_number']] = {
                    'count': int(row['count']),
                    'amount': float(row['amount'])
                }
        
        return {
            'total_amount': float(total_amount),
            'record_count': len(original_data),
            'job_count': len(job_distribution),
            'job_distribution': job_distribution
        }
        
    def _process_cost_code_splits(self, row):
        """
        Process complex cost code splits like those for Asset EX-65 at Matagorda
        
        Args:
            row (Series): The row data containing cost code information
            
        Returns:
            list: List of split rows with proper cost code allocations
        """
        split_rows = []
        
        # Check if cost code column contains slashes indicating splits
        cost_code = row.get('cost_code', '')
        notes = row.get('note', '')
        if pd.isna(cost_code):
            return [row]
            
        cost_code_str = str(cost_code)
        
        # Skip if no splits indicated
        if '/' not in cost_code_str:
            return [row]
            
        # Check if this is just a CC NEEDED indicator
        if 'CC NEEDED' in cost_code_str:
            return [row]
            
        # Parse complex cost code splits like "0496 6012B/049 66012C/0496 6012C/0496 6012F/EQ"
        split_codes = [code.strip() for code in cost_code_str.split('/')]
        
        # Default split behavior: equal distribution
        split_ratio = 1.0 / len(split_codes)
        split_ratios = [split_ratio] * len(split_codes)
        
        # Check notes for split instructions
        notes_str = '' if pd.isna(notes) else str(notes)
        if 'split' in notes_str.lower():
            # Look for patterns like "Split 0.25 EA"
            import re
            split_values = re.findall(r'(\d+\.\d+|\d+)', notes_str)
            if split_values and 'ea' in notes_str.lower():
                try:
                    split_value = float(split_values[0])
                    if split_value > 0:
                        # This means each code gets this percentage
                        split_ratios = [split_value] * len(split_codes)
                        # Normalize if needed
                        if len(split_codes) * split_value != 1.0:
                            split_ratios = [r / (len(split_codes) * split_value) for r in split_ratios]
                except ValueError:
                    # Fall back to equal distribution
                    pass
        
        # Create a row for each split
        allocation = row.get('unit_allocation', 1.0) 
        if pd.isna(allocation):
            allocation = 1.0
            
        for i, code in enumerate(split_codes):
            if not code:  # Skip empty codes
                continue
                
            # Create a copy of the row for this cost code
            split_row = row.copy()
            
            # For "EQ" codes, use a standard equipment cost code
            if code.upper() == 'EQ':
                code = '9000 100M'
                
            split_row['cost_code'] = code
            
            # Calculate the allocation for this split
            this_allocation = allocation * split_ratios[i]
            split_row['unit_allocation'] = this_allocation
            
            # Adjust amount proportionally if it exists
            if 'amount' in row and not pd.isna(row['amount']):
                split_row['amount'] = row['amount'] * split_ratios[i]
                
            # Add a note about the split
            if pd.isna(notes):
                split_row['note'] = f"Cost code split {i+1}/{len(split_codes)}"
            else:
                split_row['note'] = f"{notes} (Cost code split {i+1}/{len(split_codes)})"
                
            split_rows.append(split_row)
            
        # If we didn't create any splits (shouldn't happen), return the original
        if not split_rows:
            return [row]
            
        return split_rows

    def _integrate_new_data(self, new_data, file_path):
        """
        Integrate a new data file into the master data
        
        Args:
            new_data (DataFrame): New data to integrate
            file_path (str or Path): Source file path
            
        Returns:
            dict: Report of changes made
        """
        # Track changes
        changes = {
            'additions': 0,
            'modifications': 0,
            'deletions': 0,
            'details': []
        }
        
        # Add tracking columns to new data
        new_data['status'] = 'new'  # Default status for change tracking
        new_data['version'] = len(self.processed_files) + 1
        new_data['change_type'] = None
        
        # Compare with master data
        # 1. Find rows that exist in both datasets (modifications)
        common_rows = set(self.master_data['row_id']).intersection(set(new_data['row_id']))
        
        # 2. Find rows that exist only in the new data (additions)
        additions = set(new_data['row_id']).difference(set(self.master_data['row_id']))
        
        # 3. Find rows that exist only in the master data (potential deletions)
        potential_deletions = set(self.master_data['row_id']).difference(set(new_data['row_id']))
        
        # Process modifications
        modifications = []
        for row_id in common_rows:
            master_row = self.master_data[self.master_data['row_id'] == row_id].iloc[0]
            new_row = new_data[new_data['row_id'] == row_id].iloc[0]
            
            # Check for material changes in core values
            core_columns = ['days', 'rate', 'amount', 'job_number', 'equipment_id']
            has_changes = False
            changes_dict = {}
            
            for col in core_columns:
                if pd.notna(master_row[col]) and pd.notna(new_row[col]):
                    # For numeric columns, check if values are significantly different
                    if col in ['days', 'rate', 'amount']:
                        if abs(float(master_row[col]) - float(new_row[col])) > 0.01:
                            has_changes = True
                            changes_dict[col] = {
                                'from': float(master_row[col]),
                                'to': float(new_row[col])
                            }
                    # For string columns, check if values are different
                    elif master_row[col] != new_row[col]:
                        has_changes = True
                        changes_dict[col] = {
                            'from': master_row[col],
                            'to': new_row[col]
                        }
            
            if has_changes:
                # Mark as modified in the new data
                new_data.loc[new_data['row_id'] == row_id, 'change_type'] = 'modified'
                modifications.append(row_id)
                
                # Save change details
                changes['details'].append({
                    'type': 'modification',
                    'row_id': row_id,
                    'job_number': new_row['job_number'],
                    'equipment_id': new_row['equipment_id'],
                    'changes': changes_dict
                })
        
        changes['modifications'] = len(modifications)
        
        # Process additions
        for row_id in additions:
            # Mark as added in the new data
            new_data.loc[new_data['row_id'] == row_id, 'change_type'] = 'added'
            
            # Get the new row for reporting
            new_row = new_data[new_data['row_id'] == row_id].iloc[0]
            
            # Save change details
            changes['details'].append({
                'type': 'addition',
                'row_id': row_id,
                'job_number': new_row['job_number'],
                'equipment_id': new_row['equipment_id'],
                'amount': float(new_row['amount']) if pd.notna(new_row['amount']) else 0
            })
        
        changes['additions'] = len(additions)
        
        # Process potential deletions
        # For now, we'll keep deleted rows in the master but mark them as deleted
        deleted_rows = []
        for row_id in potential_deletions:
            # Check if this row was from a previous version (not the original)
            master_row = self.master_data[self.master_data['row_id'] == row_id].iloc[0]
            
            # Only rows from the original file (version 1) could be considered deleted
            # if they don't appear in subsequent files
            if 'version' not in master_row or master_row['version'] == 1:
                deleted_rows.append(row_id)
                
                # Mark as deleted in the master data
                self.master_data.loc[self.master_data['row_id'] == row_id, 'status'] = 'deleted'
                self.master_data.loc[self.master_data['row_id'] == row_id, 'change_type'] = 'deleted'
                
                # Save change details
                changes['details'].append({
                    'type': 'deletion',
                    'row_id': row_id,
                    'job_number': master_row['job_number'],
                    'equipment_id': master_row['equipment_id'],
                    'amount': float(master_row['amount']) if pd.notna(master_row['amount']) else 0
                })
        
        changes['deletions'] = len(deleted_rows)
        
        # Update the master dataset
        # 1. For modified rows, update the values
        for row_id in modifications:
            # Get the new row
            new_row = new_data[new_data['row_id'] == row_id]
            
            # Update all columns except status, version and row_id
            for col in new_row.columns:
                if col not in ['row_id']:
                    self.master_data.loc[self.master_data['row_id'] == row_id, col] = new_row[col].values[0]
            
            # Update status to modified
            self.master_data.loc[self.master_data['row_id'] == row_id, 'status'] = 'modified'
            self.master_data.loc[self.master_data['row_id'] == row_id, 'version'] = len(self.processed_files) + 1
        
        # 2. For new rows, add them to the master
        additions_df = new_data[new_data['row_id'].isin(additions)]
        if not additions_df.empty:
            self.master_data = pd.concat([self.master_data, additions_df], ignore_index=True)
        
        # Add to processed files
        self.processed_files.append(str(file_path))
        
        return {
            'success': True,
            'message': 'File processed successfully',
            'file': str(file_path),
            'rows': len(new_data),
            'changes': changes
        }
    
    def generate_master_output(self, output_path=None):
        """
        Generate a master output file with all processed data
        
        Args:
            output_path (str or Path, optional): Path for the output file
            
        Returns:
            dict: Output report
        """
        try:
            if self.master_data.empty:
                raise PMProcessingException("No data has been processed yet")
                
            # Create a default output path if none provided
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = EXPORTS_DIR / f"pm_master_{timestamp}.xlsx"
            
            # Ensure path is a Path object
            if isinstance(output_path, str):
                output_path = Path(output_path)
                
            # Create a workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                
            # Create sheets
            master_sheet = wb.create_sheet("Master Data")
            summary_sheet = wb.create_sheet("Summary")
            changes_sheet = wb.create_sheet("Change Log")
            
            # Populate master data sheet
            self._populate_master_sheet(master_sheet)
            
            # Populate summary sheet
            summary_stats = self._populate_summary_sheet(summary_sheet)
            
            # Populate changes sheet
            self._populate_changes_sheet(changes_sheet)
            
            # Save the workbook
            wb.save(output_path)
            
            return {
                'success': True,
                'message': 'Master output generated successfully',
                'output_path': str(output_path),
                'summary': summary_stats
            }
            
        except PMProcessingException as e:
            # Re-raise the custom exception
            raise
        except Exception as e:
            # Log the error and raise a custom exception
            logger.error(f"Error generating master output: {str(e)}")
            raise PMProcessingException(f"Error generating master output: {str(e)}")
    
    def _populate_master_sheet(self, sheet):
        """Populate the master data sheet"""
        # Define columns to include
        display_columns = [
            'job_number', 'job_name', 'equipment_id', 'description', 'category',
            'start_date', 'end_date', 'days', 'rate', 'amount', 'notes',
            'status', 'version', 'source_file'
        ]
        
        # Add headers
        headers = {
            'job_number': 'Job Code',
            'job_name': 'Job Name',
            'equipment_id': 'Equipment #',
            'description': 'Description',
            'category': 'Category',
            'start_date': 'Start Date',
            'end_date': 'End Date',
            'days': 'Days',
            'rate': 'Rate',
            'amount': 'Amount',
            'notes': 'Notes',
            'status': 'Status',
            'version': 'Version',
            'source_file': 'Source File'
        }
        
        for col_idx, col_name in enumerate(display_columns, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = headers.get(col_name, col_name.title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
            
        # Add data rows
        for row_idx, (_, row) in enumerate(self.master_data.iterrows(), 2):
            for col_idx, col_name in enumerate(display_columns, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Format value based on column type
                if col_name == 'amount' or col_name == 'rate':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '$#,##0.00'
                elif col_name == 'days':
                    if pd.notna(row[col_name]):
                        cell.value = float(row[col_name])
                        cell.number_format = '#,##0.0'
                elif col_name in ['start_date', 'end_date']:
                    if pd.notna(row[col_name]):
                        cell.value = row[col_name].date() if hasattr(row[col_name], 'date') else row[col_name]
                        cell.number_format = 'yyyy-mm-dd'
                else:
                    cell.value = row[col_name] if pd.notna(row[col_name]) else None
                
                # Add color coding for status
                if col_name == 'status':
                    if row[col_name] == 'modified':
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    elif row[col_name] == 'added':
                        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    elif row[col_name] == 'deleted':
                        cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        
        # Auto-size columns
        for col_idx in range(1, len(display_columns) + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 15
            
        # Add filters
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(display_columns))}{len(self.master_data) + 1}"
    
    def _populate_summary_sheet(self, sheet):
        """Populate the summary sheet with statistics"""
        # Calculate summary statistics
        total_rows = len(self.master_data)
        active_rows = len(self.master_data[self.master_data['status'] != 'deleted'])
        deleted_rows = len(self.master_data[self.master_data['status'] == 'deleted'])
        modified_rows = len(self.master_data[self.master_data['status'] == 'modified'])
        added_rows = len(self.master_data[self.master_data['status'] == 'added'])
        
        # Calculate totals
        total_amount = self.master_data[self.master_data['status'] != 'deleted']['amount'].sum()
        
        # Calculate totals by job
        job_totals = self.master_data[self.master_data['status'] != 'deleted'].groupby('job_number')['amount'].sum().reset_index()
        job_counts = self.master_data[self.master_data['status'] != 'deleted'].groupby('job_number').size().reset_index(name='count')
        job_summary = pd.merge(job_totals, job_counts, on='job_number')
        
        # Add headers
        summary_items = [
            ('Total Records', total_rows),
            ('Active Records', active_rows),
            ('Deleted Records', deleted_rows),
            ('Modified Records', modified_rows),
            ('Added Records', added_rows),
            ('Total Amount', total_amount),
            ('Number of Jobs', len(job_summary)),
            ('Processed Files', len(self.processed_files))
        ]
        
        # Add summary data
        for row_idx, (label, value) in enumerate(summary_items, 1):
            # Label
            cell = sheet.cell(row=row_idx, column=1)
            cell.value = label
            cell.font = Font(bold=True)
            
            # Value
            cell = sheet.cell(row=row_idx, column=2)
            if isinstance(value, (int, float)) and 'Amount' in label:
                cell.value = value
                cell.number_format = '$#,##0.00'
            else:
                cell.value = value
        
        # Add a separator
        sheet.cell(row=len(summary_items) + 1, column=1).value = ""
        
        # Add processed files
        sheet.cell(row=len(summary_items) + 2, column=1).value = "Processed Files:"
        sheet.cell(row=len(summary_items) + 2, column=1).font = Font(bold=True)
        
        for idx, file_path in enumerate(self.processed_files, 1):
            cell = sheet.cell(row=len(summary_items) + 2 + idx, column=1)
            cell.value = f"{idx}. {os.path.basename(file_path)}"
        
        # Add a separator
        sheet.cell(row=len(summary_items) + len(self.processed_files) + 3, column=1).value = ""
        
        # Add job summary header
        job_header_row = len(summary_items) + len(self.processed_files) + 4
        sheet.cell(row=job_header_row, column=1).value = "Job Summary:"
        sheet.cell(row=job_header_row, column=1).font = Font(bold=True)
        
        # Add job summary headers
        job_columns = ['Job Number', 'Equipment Count', 'Total Amount']
        for col_idx, header in enumerate(job_columns, 1):
            cell = sheet.cell(row=job_header_row + 1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
        # Add job summary data
        for row_idx, (_, row) in enumerate(job_summary.iterrows(), 1):
            # Job number
            cell = sheet.cell(row=job_header_row + 1 + row_idx, column=1)
            cell.value = row['job_number']
            
            # Equipment count
            cell = sheet.cell(row=job_header_row + 1 + row_idx, column=2)
            cell.value = row['count']
            
            # Total amount
            cell = sheet.cell(row=job_header_row + 1 + row_idx, column=3)
            cell.value = row['amount']
            cell.number_format = '$#,##0.00'
        
        # Format column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        
        return {
            'total_rows': total_rows,
            'active_rows': active_rows,
            'deleted_rows': deleted_rows,
            'modified_rows': modified_rows,
            'added_rows': added_rows,
            'total_amount': float(total_amount),
            'job_count': len(job_summary),
            'processed_files': len(self.processed_files)
        }
    
    def _populate_changes_sheet(self, sheet):
        """Populate the changes log sheet"""
        # Define change log headers
        headers = ['Version', 'File', 'Change Type', 'Job Number', 'Equipment ID', 'Details', 'Amount Before', 'Amount After', 'Change']
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
        # Extract all changes
        changes = []
        for version, file_path in enumerate(self.processed_files[1:], 2):  # Skip the first file (original)
            # Find rows with this version
            version_rows = self.master_data[self.master_data['version'] == version]
            
            for _, row in version_rows.iterrows():
                if row['change_type'] in ['added', 'modified', 'deleted']:
                    change_record = {
                        'version': version,
                        'file': os.path.basename(file_path),
                        'change_type': row['change_type'],
                        'job_number': row['job_number'],
                        'equipment_id': row['equipment_id']
                    }
                    
                    # Get details based on change type
                    if row['change_type'] == 'added':
                        change_record['details'] = 'New record added'
                        change_record['amount_before'] = None
                        change_record['amount_after'] = row['amount']
                        change_record['change'] = row['amount']
                    elif row['change_type'] == 'deleted':
                        change_record['details'] = 'Record deleted'
                        change_record['amount_before'] = row['amount']
                        change_record['amount_after'] = 0
                        change_record['change'] = -row['amount']
                    elif row['change_type'] == 'modified':
                        # Find the previous version for comparison
                        prev_versions = list(range(1, version))
                        prev_versions.reverse()  # Start with most recent
                        
                        prev_row = None
                        for prev_version in prev_versions:
                            prev_version_row = self.master_data[
                                (self.master_data['row_id'] == row['row_id']) &
                                (self.master_data['version'] == prev_version)
                            ]
                            
                            if not prev_version_row.empty:
                                prev_row = prev_version_row.iloc[0]
                                break
                        
                        if prev_row is not None:
                            details = []
                            if prev_row['amount'] != row['amount']:
                                details.append(f"Amount changed")
                            if prev_row['days'] != row['days']:
                                details.append(f"Days changed: {prev_row['days']} to {row['days']}")
                            if prev_row['rate'] != row['rate']:
                                details.append(f"Rate changed: {prev_row['rate']} to {row['rate']}")
                                
                            change_record['details'] = ', '.join(details) or 'Record modified'
                            change_record['amount_before'] = prev_row['amount']
                            change_record['amount_after'] = row['amount']
                            change_record['change'] = row['amount'] - prev_row['amount']
                        else:
                            change_record['details'] = 'Record modified (no previous version found)'
                            change_record['amount_before'] = None
                            change_record['amount_after'] = row['amount']
                            change_record['change'] = row['amount']
                    
                    changes.append(change_record)
        
        # Sort changes by version
        changes.sort(key=lambda x: x['version'])
        
        # Add changes to sheet
        for row_idx, change in enumerate(changes, 2):
            # Version
            sheet.cell(row=row_idx, column=1).value = change['version']
            
            # File
            sheet.cell(row=row_idx, column=2).value = change['file']
            
            # Change Type
            cell = sheet.cell(row=row_idx, column=3)
            cell.value = change['change_type'].title()
            
            # Color coding for change type
            if change['change_type'] == 'modified':
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif change['change_type'] == 'added':
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif change['change_type'] == 'deleted':
                cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            
            # Job Number
            sheet.cell(row=row_idx, column=4).value = change['job_number']
            
            # Equipment ID
            sheet.cell(row=row_idx, column=5).value = change['equipment_id']
            
            # Details
            sheet.cell(row=row_idx, column=6).value = change['details']
            
            # Amount Before
            cell = sheet.cell(row=row_idx, column=7)
            cell.value = change['amount_before']
            if change['amount_before'] is not None:
                cell.number_format = '$#,##0.00'
            
            # Amount After
            cell = sheet.cell(row=row_idx, column=8)
            cell.value = change['amount_after']
            if change['amount_after'] is not None:
                cell.number_format = '$#,##0.00'
            
            # Change
            cell = sheet.cell(row=row_idx, column=9)
            cell.value = change['change']
            if change['change'] is not None:
                cell.number_format = '$#,##0.00'
                
                # Color coding for change value
                if change['change'] > 0:
                    cell.font = Font(color="006100")
                elif change['change'] < 0:
                    cell.font = Font(color="9C0006")
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 15
            
        # Make the details column wider
        sheet.column_dimensions['F'].width = 40
        
        # Add filters
        if len(changes) > 0:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(changes) + 1}"
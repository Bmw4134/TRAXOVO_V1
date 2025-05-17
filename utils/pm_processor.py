"""
PM Allocation Processor

This module handles the comparison of PM billing allocation files to identify
changes between original and updated versions, preserving formulas and formatting.
"""

import os
import re
import json
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_allocation_files(directory='attached_assets', month=None, year=None):
    """
    Find PM allocation files in specified directory based on naming conventions
    
    Args:
        directory (str): Directory to search
        month (str): Month name (e.g., 'APRIL')
        year (str): Year (e.g., '2025')
        
    Returns:
        dict: Dictionary with original and revised file paths
    """
    directory = Path(directory)
    allocation_files = {
        'original': None,
        'revised': None
    }
    
    # Build search patterns
    if month and year:
        original_pattern = f"EQMO. BILLING ALLOCATIONS - {month} {year}"
        revised_pattern = f"EQMO. BILLING ALLOCATIONS - {month} {year} (.*FINAL.*REVISION.*)"
    else:
        original_pattern = "EQMO. BILLING ALLOCATIONS"
        revised_pattern = "EQMO. BILLING ALLOCATIONS.*(FINAL.*REVISION.*)"
    
    # Search for files
    excel_files = list(directory.glob("*.xlsx")) + list(directory.glob("*.xlsm"))
    
    # Find the most recent original and revised files
    for file_path in excel_files:
        filename = file_path.name.upper()
        
        # Original file (exclude files with 'FINAL REVISION' in the name)
        if original_pattern.upper() in filename and 'FINAL' not in filename and 'REVISION' not in filename:
            if allocation_files['original'] is None or file_path.stat().st_mtime > allocation_files['original'].stat().st_mtime:
                allocation_files['original'] = file_path
        
        # Revised file (must include 'FINAL REVISION' in the name)
        if re.search(revised_pattern, filename, re.IGNORECASE):
            if allocation_files['revised'] is None or file_path.stat().st_mtime > allocation_files['revised'].stat().st_mtime:
                allocation_files['revised'] = file_path
    
    return allocation_files

def identify_sheet_structure(wb, sheet_name=None):
    """
    Identify the structure of the PM allocation sheet by finding key columns
    
    Args:
        wb: Excel workbook object
        sheet_name (str, optional): Sheet name to analyze
        
    Returns:
        dict: Dictionary with sheet structure information
    """
    # Find appropriate sheet if none is specified
    if sheet_name is None:
        for name in wb.sheetnames:
            if 'allocation' in name.lower() or 'billing' in name.lower() or 'april' in name.lower() or 'march' in name.lower():
                sheet_name = name
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]  # Default to first sheet
    
    sheet = wb[sheet_name]
    
    # Search the first 20 rows for header information
    structure = {
        'sheet_name': sheet_name,
        'header_row': None,
        'data_start_row': None,
        'job_col': None,
        'asset_col': None,
        'amount_col': None,
        'region_col': None,
        'department_col': None,
        'columns': {},
        'total_row': None
    }
    
    # Locate header row and column positions
    for row in range(1, 20):
        for col in range(1, 15):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                cell_text = str(cell_value).strip().upper()
                
                # Look for job column
                if re.search(r'JOB|PROJECT', cell_text) and structure['job_col'] is None:
                    structure['job_col'] = col
                    structure['header_row'] = row
                    structure['columns']['job'] = col
                
                # Look for asset column
                if re.search(r'ASSET|EQUIPMENT|EQ\s+#|UNIT', cell_text) and structure['asset_col'] is None:
                    structure['asset_col'] = col
                    structure['header_row'] = row
                    structure['columns']['asset'] = col
                
                # Look for amount/allocation column
                if re.search(r'AMOUNT|ALLOCATION|TOTAL', cell_text) and structure['amount_col'] is None:
                    structure['amount_col'] = col
                    structure['header_row'] = row
                    structure['columns']['amount'] = col
                
                # Look for region column
                if re.search(r'REGION|AREA|DISTRICT', cell_text) and structure['region_col'] is None:
                    structure['region_col'] = col
                    structure['header_row'] = row
                    structure['columns']['region'] = col
                
                # Look for department column
                if re.search(r'DEPARTMENT|DEPT', cell_text) and structure['department_col'] is None:
                    structure['department_col'] = col
                    structure['header_row'] = row
                    structure['columns']['department'] = col
    
    # If we found a header row, data starts in the next row
    if structure['header_row']:
        structure['data_start_row'] = structure['header_row'] + 1
    
    # Find the total row (usually contains "TOTAL" or "GRAND TOTAL")
    if structure['data_start_row']:
        for row in range(structure['data_start_row'], sheet.max_row + 1):
            for col in range(1, 10):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and 'TOTAL' in str(cell_value).upper():
                    structure['total_row'] = row
                    break
            if structure['total_row']:
                break
    
    return structure

def extract_pm_allocation_data(file_path, structure=None):
    """
    Extract PM allocation data from Excel file
    
    Args:
        file_path (str or Path): Path to Excel file
        structure (dict, optional): Known sheet structure
        
    Returns:
        tuple: (DataFrame with allocation data, sheet structure, workbook)
    """
    file_path = Path(file_path)
    
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    
    # Identify sheet structure if not provided
    if structure is None:
        structure = identify_sheet_structure(wb)
    
    # Check if we were able to identify the structure
    if structure['header_row'] is None or structure['data_start_row'] is None:
        raise ValueError(f"Could not identify the structure of the PM allocation sheet in {file_path}")
    
    sheet = wb[structure['sheet_name']]
    
    # Extract data into lists
    data = []
    for row in range(structure['data_start_row'], structure['total_row'] or sheet.max_row + 1):
        row_data = {}
        
        # Skip empty rows
        if sheet.cell(row=row, column=structure['job_col']).value is None and \
           sheet.cell(row=row, column=structure['asset_col']).value is None:
            continue
        
        # Extract values for each column
        for col_name, col_index in structure['columns'].items():
            row_data[col_name] = sheet.cell(row=row, column=col_index).value
        
        # Add row number for reference
        row_data['row'] = row
        
        data.append(row_data)
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Clean up and convert data types
    if 'amount' in df.columns:
        # Convert amount to float
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        df['amount'] = df['amount'].fillna(0.0)
    
    return df, structure, wb

def compare_pm_allocations(original_file, updated_file, region=None):
    """
    Compare original and updated PM allocation files
    
    Args:
        original_file (str or Path): Path to original allocation file
        updated_file (str or Path): Path to updated allocation file
        region (str, optional): Filter by region (DFW, HOU, WT, or ALL)
        
    Returns:
        dict: Comparison results
    """
    original_file = Path(original_file)
    updated_file = Path(updated_file)
    
    # Extract data from original file
    original_df, original_structure, original_wb = extract_pm_allocation_data(original_file)
    
    # Extract data from updated file using the same structure
    updated_df, updated_structure, updated_wb = extract_pm_allocation_data(updated_file)
    
    # Initialize results
    results = {
        'original_file': original_file.name,
        'updated_file': updated_file.name,
        'original_count': len(original_df),
        'updated_count': len(updated_df),
        'record_count': max(len(original_df), len(updated_df)),
        'change_count': 0,
        'significant_changes': [],
        'region_summary': [],
        'summary': {
            'original_total': original_df['amount'].sum(),
            'updated_total': updated_df['amount'].sum(),
            'net_change': updated_df['amount'].sum() - original_df['amount'].sum(),
            'percent_change': (updated_df['amount'].sum() - original_df['amount'].sum()) / original_df['amount'].sum() * 100 if original_df['amount'].sum() != 0 else 0
        },
        'job_distribution': {
            'labels': [],
            'original_values': [],
            'updated_values': []
        }
    }
    
    # Create merged dataset for comparison
    # First, standardize job numbers and asset IDs
    for df in [original_df, updated_df]:
        if 'job' in df.columns:
            # Extract numeric portion of job numbers
            df['job_clean'] = df['job'].astype(str).str.extract(r'(\d+)')[0]
        if 'asset' in df.columns:
            # Standardize asset IDs
            df['asset_clean'] = df['asset'].astype(str).str.upper().str.strip()
    
    # Merge datasets on job and asset
    merged_df = pd.merge(
        original_df, 
        updated_df, 
        on=['job_clean', 'asset_clean'], 
        how='outer',
        suffixes=('_original', '_updated')
    )
    
    # Calculate changes
    merged_df['amount_original'] = merged_df['amount_original'].fillna(0)
    merged_df['amount_updated'] = merged_df['amount_updated'].fillna(0)
    merged_df['change'] = merged_df['amount_updated'] - merged_df['amount_original']
    merged_df['percent_change'] = np.where(
        merged_df['amount_original'] != 0,
        (merged_df['change'] / merged_df['amount_original']) * 100,
        np.inf  # Use infinity for cases where original amount is 0
    )
    
    # Count changes
    changes = merged_df[merged_df['change'] != 0]
    results['change_count'] = len(changes)
    
    # Filter by region if specified
    if region and region != 'ALL':
        region_col = 'region_original' if 'region_original' in merged_df.columns else None
        if region_col:
            merged_df = merged_df[merged_df[region_col].str.contains(region, case=False, na=False)]
    
    # Find significant changes (more than $500 or more than 10% change)
    significant_changes = merged_df[
        (abs(merged_df['change']) > 500) | 
        ((abs(merged_df['percent_change']) > 10) & (merged_df['percent_change'] != np.inf))
    ].copy()
    
    # Format significant changes for output
    for _, row in significant_changes.iterrows():
        results['significant_changes'].append({
            'job_number': row.get('job_original', row.get('job_updated', 'Unknown')),
            'asset_id': row.get('asset_original', row.get('asset_updated', 'Unknown')),
            'original_value': row['amount_original'],
            'updated_value': row['amount_updated'],
            'change': row['change'],
            'percent_change': row['percent_change'] if row['percent_change'] != np.inf else 100
        })
    
    # Sort significant changes by absolute change amount
    results['significant_changes'] = sorted(
        results['significant_changes'], 
        key=lambda x: abs(x['change']), 
        reverse=True
    )
    
    # Generate region summary
    regions = set()
    
    # Get all unique regions from both datasets
    if 'region_original' in merged_df.columns:
        regions.update(merged_df['region_original'].dropna().unique())
    if 'region_updated' in merged_df.columns:
        regions.update(merged_df['region_updated'].dropna().unique())
    
    # Calculate totals by region
    for region_name in sorted(regions):
        # Skip if region is empty
        if not region_name or pd.isna(region_name):
            continue
            
        # Calculate original total for this region
        original_region_total = merged_df[
            merged_df['region_original'] == region_name
        ]['amount_original'].sum()
        
        # Calculate updated total for this region
        updated_region_total = merged_df[
            (merged_df['region_updated'] == region_name) | 
            (merged_df['region_original'] == region_name)  # Include original entries in case region was changed
        ]['amount_updated'].sum()
        
        net_change = updated_region_total - original_region_total
        percent_change = (net_change / original_region_total) * 100 if original_region_total != 0 else 0
        
        results['region_summary'].append({
            'name': region_name,
            'original_total': original_region_total,
            'updated_total': updated_region_total,
            'net_change': net_change,
            'percent_change': percent_change
        })
    
    # Generate job distribution data for chart
    # Get the top jobs by total allocation amount
    top_jobs = merged_df.groupby('job_clean')['amount_updated'].sum().nlargest(10).index.tolist()
    
    for job in top_jobs:
        # Get original and updated values for this job
        original_value = merged_df[merged_df['job_clean'] == job]['amount_original'].sum()
        updated_value = merged_df[merged_df['job_clean'] == job]['amount_updated'].sum()
        
        # Add to chart data
        results['job_distribution']['labels'].append(f"Job {job}")
        results['job_distribution']['original_values'].append(float(original_value))
        results['job_distribution']['updated_values'].append(float(updated_value))
    
    return results

def export_comparison_results(results, output_format='excel'):
    """
    Export comparison results to Excel or CSV
    
    Args:
        results (dict): Comparison results from compare_pm_allocations
        output_format (str): 'excel' or 'csv'
        
    Returns:
        tuple: (excel_path, csv_path) Paths to the generated files
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    exports_dir = Path('exports')
    exports_dir.mkdir(exist_ok=True)
    
    excel_path = exports_dir / f"pm_allocation_reconciliation_{timestamp}.xlsx"
    csv_path = exports_dir / f"pm_allocation_reconciliation_{timestamp}.csv"
    
    # Create Excel export
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = {
            'Category': ['Original File', 'Updated File', 'Record Count', 'Changes Detected', 
                         'Original Total', 'Updated Total', 'Net Change', 'Percent Change'],
            'Value': [
                results['original_file'],
                results['updated_file'],
                results['record_count'],
                results['change_count'],
                results['summary']['original_total'],
                results['summary']['updated_total'],
                results['summary']['net_change'],
                f"{results['summary']['percent_change']:.2f}%"
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        
        # Region summary sheet
        region_data = []
        for region in results['region_summary']:
            region_data.append({
                'Region': region['name'],
                'Original Total': region['original_total'],
                'Updated Total': region['updated_total'],
                'Net Change': region['net_change'],
                'Percent Change': f"{region['percent_change']:.2f}%"
            })
        pd.DataFrame(region_data).to_excel(writer, sheet_name='Region Summary', index=False)
        
        # Significant changes sheet
        if results['significant_changes']:
            changes_data = []
            for change in results['significant_changes']:
                changes_data.append({
                    'Job Number': change['job_number'],
                    'Asset ID': change['asset_id'],
                    'Original Value': change['original_value'],
                    'Updated Value': change['updated_value'],
                    'Change': change['change'],
                    'Percent Change': f"{change['percent_change']:.2f}%"
                })
            pd.DataFrame(changes_data).to_excel(writer, sheet_name='Significant Changes', index=False)
    
    # Create CSV export (simplified, mainly for accounting import)
    if results['significant_changes']:
        changes_data = []
        for change in results['significant_changes']:
            changes_data.append({
                'Job_Number': str(change['job_number']).replace(' ', ''),
                'Asset_ID': str(change['asset_id']).replace(' ', ''),
                'Original_Value': change['original_value'],
                'Updated_Value': change['updated_value'],
                'Change': change['change']
            })
        pd.DataFrame(changes_data).to_csv(csv_path, index=False)
    
    return excel_path, csv_path

def generate_allocation_template(output_path=None):
    """
    Generate a template for PM allocations for accounting
    
    Args:
        output_path (str, optional): Path to save the template
        
    Returns:
        str: Path to the generated template
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        exports_dir = Path('exports')
        exports_dir.mkdir(exist_ok=True)
        output_path = exports_dir / f"pm_allocation_template_{timestamp}.xlsx"
    
    # Create a template DataFrame
    template_data = {
        'Job Number': ['12345', '67890', '24010', '24-025', ''],
        'Asset ID': ['ET-1234', 'PT-5678', 'ET-9012', 'PT-3456', ''],
        'Region': ['DFW', 'HOU', 'WT', 'DFW', ''],
        'Department': ['Paving', 'Earthwork', 'Utilities', 'Paving', ''],
        'Original Amount': [1000.00, 2500.00, 1800.00, 3200.00, ''],
        'Updated Amount': [1000.00, 2750.00, 1500.00, 3500.00, ''],
        'Notes': ['No change', 'Modified allocation', 'Reduced allocation', 'Increased allocation', '']
    }
    
    df = pd.DataFrame(template_data)
    
    # Create the Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='PM Allocation Template', index=False)
        
        # Adjust column widths
        worksheet = writer.sheets['PM Allocation Template']
        for i, col in enumerate(df.columns):
            column_width = max(len(col) + 2, df[col].astype(str).map(len).max() + 2)
            worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
    
    return str(output_path)

def process_pm_allocation_files(original_file, updated_file, region=None):
    """
    Process PM allocation files and generate reconciliation report
    
    Args:
        original_file (str or Path): Path to original allocation file
        updated_file (str or Path): Path to updated allocation file
        region (str, optional): Filter by region (DFW, HOU, WT, or ALL)
        
    Returns:
        dict: Processing results including export file paths
    """
    try:
        # Compare allocations
        results = compare_pm_allocations(original_file, updated_file, region)
        
        # Export results
        excel_path, csv_path = export_comparison_results(results)
        
        # Add export paths to results
        results['excel_export'] = os.path.basename(excel_path)
        results['csv_export'] = os.path.basename(csv_path)
        
        return results
        
    except Exception as e:
        return {
            'error': str(e),
            'original_file': os.path.basename(original_file) if original_file else None,
            'updated_file': os.path.basename(updated_file) if updated_file else None
        }

def auto_detect_and_process_allocation_files(directory='attached_assets'):
    """
    Automatically detect and process the latest PM allocation files
    
    Args:
        directory (str): Directory to search for files
        
    Returns:
        dict: Processing results or error message
    """
    try:
        # Find the allocation files
        allocation_files = find_allocation_files(directory)
        
        if allocation_files['original'] is None:
            return {'error': 'Could not find original PM allocation file'}
        
        if allocation_files['revised'] is None:
            return {'error': 'Could not find revised PM allocation file'}
        
        # Process the files
        return process_pm_allocation_files(
            allocation_files['original'],
            allocation_files['revised']
        )
        
    except Exception as e:
        return {'error': str(e)}
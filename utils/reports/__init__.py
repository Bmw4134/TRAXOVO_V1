"""
Reports Utility Module

This module handles the generation of daily and periodic reports, including:
- Prior day attendance reports (Late Start, Early End, Not On Job)
- Current day early status reports (Late Start, Not On Job)
- Billing comparison and validation reports
- Region-based billing export generation
"""

import os
import json
import logging
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

from app import db
from models import Asset

# Initialize logger
logger = logging.getLogger(__name__)

def ensure_report_dirs():
    """Ensure all necessary report directories exist"""
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Create base report directories
    os.makedirs(f'reports/{today}', exist_ok=True)
    os.makedirs('exports/billings', exist_ok=True)
    os.makedirs(f'exports/billings/{today}', exist_ok=True)
    
    return {
        'daily_reports': f'reports/{today}',
        'billing_exports': f'exports/billings/{today}'
    }

def get_latest_activity_files():
    """
    Find the most recently uploaded activity detail and driving history files
    
    Returns:
        dict: Dictionary with paths to the latest files
    """
    try:
        # This is a placeholder - in production we would actually scan the uploads directory
        # and find the most recent files based on upload or file date
        
        latest_files = {
            'activity_detail': 'uploads/activity_detail/latest.xlsx',
            'driving_history': 'uploads/driving_history/latest.xlsx',
            'last_updated': datetime.now()
        }
        
        # Check if files exist
        missing_files = []
        for key, path in latest_files.items():
            if key != 'last_updated' and not os.path.exists(path):
                missing_files.append(key)
        
        if missing_files:
            logger.warning(f"Missing expected input files: {', '.join(missing_files)}")
            latest_files['status'] = 'incomplete'
            latest_files['missing'] = missing_files
        else:
            latest_files['status'] = 'complete'
            
        return latest_files
    except Exception as e:
        logger.error(f"Error identifying latest activity files: {e}")
        return {'status': 'error', 'message': str(e)}

def get_latest_billing_files():
    """
    Find the most recently uploaded billing files
    
    Returns:
        dict: Dictionary with paths to the latest files
    """
    try:
        # This is a placeholder - in production we would scan the uploads directory
        
        latest_files = {
            'original_billing': 'attached_assets/RAGLE EQ BILLINGS - APRIL 2025 (JG REVIEWED 5.12).xlsm',
            'edited_billing': 'attached_assets/EQMO. BILLING ALLOCATIONS - APRIL 2025 (TR-FINAL REVISIONS BY 05.15.2025).xlsx',
            'last_updated': datetime.now()
        }
        
        # Check if files exist
        missing_files = []
        for key, path in latest_files.items():
            if key != 'last_updated' and not os.path.exists(path):
                missing_files.append(key)
        
        if missing_files:
            logger.warning(f"Missing expected billing files: {', '.join(missing_files)}")
            latest_files['status'] = 'incomplete'
            latest_files['missing'] = missing_files
        else:
            latest_files['status'] = 'complete'
            
        return latest_files
    except Exception as e:
        logger.error(f"Error identifying latest billing files: {e}")
        return {'status': 'error', 'message': str(e)}

def generate_prior_day_report():
    """
    Generate the prior day Late Start, Early End, Not On Job report
    
    This report analyzes yesterday's data for the full day (12:01 AM - 11:59 PM)
    
    Returns:
        dict: Report generation status and path
    """
    try:
        # Ensure report directories exist
        dirs = ensure_report_dirs()
        
        # Get latest activity files
        files = get_latest_activity_files()
        if files['status'] != 'complete':
            return {'status': 'error', 'message': 'Missing required input files'}
            
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        report_path = f"{dirs['daily_reports']}/prior_day_attendance_{yesterday}.xlsx"
        
        # Log the process
        logger.info(f"Generating prior day report for {yesterday}")
        
        # Format: Late Start, Early End, Not On Job for prior day
        # This is where the actual report generation logic would be implemented
        # Based on the activity detail and driving history
        
        # For now we'll just create a placeholder file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Attendance Report {yesterday}"
        
        # Set up headers
        headers = ["Identifier", "Driver", "Region", "Expected Start", "Actual Start", 
                  "Late By (min)", "Expected End", "Actual End", "Early By (min)", 
                  "Expected Job", "Actual Location", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
        # Save the workbook
        workbook.save(report_path)
        
        logger.info(f"Prior day report saved to {report_path}")
        return {
            'status': 'success',
            'path': report_path,
            'date': yesterday
        }
    except Exception as e:
        logger.error(f"Error generating prior day report: {e}")
        return {'status': 'error', 'message': str(e)}

def generate_current_day_report():
    """
    Generate the current day Late Start, Not On Job report
    
    This report analyzes the current day's data up to 8:30 AM
    
    Returns:
        dict: Report generation status and path
    """
    try:
        # Ensure report directories exist
        dirs = ensure_report_dirs()
        
        # Get latest activity files
        files = get_latest_activity_files()
        if files['status'] != 'complete':
            return {'status': 'error', 'message': 'Missing required input files'}
            
        today = datetime.now().strftime('%Y-%m-%d')
        report_path = f"{dirs['daily_reports']}/current_day_attendance_{today}.xlsx"
        
        # Log the process
        logger.info(f"Generating current day report for {today} (as of 8:30 AM)")
        
        # Format: Late Start, Not On Job for current day
        # This is where the actual report generation logic would be implemented
        # Based on the activity detail and driving history
        
        # For now we'll just create a placeholder file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Early Status {today}"
        
        # Set up headers
        headers = ["Identifier", "Driver", "Region", "Expected Start", "Status at 8:30 AM", 
                  "Late By (min)", "Expected Job", "Actual Location", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
        # Save the workbook
        workbook.save(report_path)
        
        logger.info(f"Current day report saved to {report_path}")
        return {
            'status': 'success',
            'path': report_path,
            'date': today
        }
    except Exception as e:
        logger.error(f"Error generating current day report: {e}")
        return {'status': 'error', 'message': str(e)}

def compare_billing_files():
    """
    Compare original and edited billing files to identify changes
    
    Returns:
        dict: Comparison results with changes identified
    """
    try:
        # Get latest billing files
        files = get_latest_billing_files()
        if files['status'] != 'complete':
            return {'status': 'error', 'message': 'Missing required billing files'}
            
        logger.info("Comparing original and edited billing files")
        
        # This is where the actual comparison logic would be implemented
        # We would load both Excel files, compare the relevant columns,
        # and identify changed job codes, assets, day counts, and notes
        
        # For now, return a placeholder result
        return {
            'status': 'success',
            'changes_detected': True,
            'changes': {
                'job_codes': [
                    {'asset': 'EX-81', 'old': 'J123', 'new': 'J456'},
                    {'asset': 'DT-11', 'old': 'J789', 'new': 'J555'}
                ],
                'days': [
                    {'asset': 'LP-116', 'old': 15, 'new': 18},
                    {'asset': 'ME-58', 'old': 22, 'new': 20}
                ],
                'notes': [
                    {'asset': 'PT-276', 'old': '', 'new': 'Maintenance 4/15-4/18'},
                    {'asset': 'PT-277', 'old': 'Job transfer 4/12', 'new': 'Job transfer 4/10-4/12'}
                ]
            }
        }
    except Exception as e:
        logger.error(f"Error comparing billing files: {e}")
        return {'status': 'error', 'message': str(e)}

def generate_region_billing_exports(approved_changes=None):
    """
    Generate region-based billing exports after approval
    
    Args:
        approved_changes (dict): Changes approved by the user
    
    Returns:
        dict: Export status and paths
    """
    try:
        # Ensure export directories exist
        dirs = ensure_report_dirs()
        
        # Get latest billing files
        files = get_latest_billing_files()
        if files['status'] != 'complete':
            return {'status': 'error', 'message': 'Missing required billing files'}
            
        logger.info("Generating region-based billing exports")
        
        # Define regions
        regions = {
            'DFW': {'region_code': 2, 'export_path': f"{dirs['billing_exports']}/02-DFW-APRIL-2025.xlsx"},
            'WTX': {'region_code': 3, 'export_path': f"{dirs['billing_exports']}/03-WTX-APRIL-2025.xlsx"},
            'HOU': {'region_code': 4, 'export_path': f"{dirs['billing_exports']}/04-HOU-APRIL-2025.xlsx"}
        }
        
        # This is where the actual export generation logic would be implemented
        # We would process the edited billing file (with approved changes),
        # split by region, and format for import into the accounting system
        
        # For now, create placeholder export files
        for region, info in regions.items():
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"{region} Billing Export"
            
            # Set up headers
            headers = ["Asset ID", "Job Code", "Description", "Days", "Rate", "Amount", "Notes"]
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                
            # Save the workbook
            workbook.save(info['export_path'])
            logger.info(f"Created export file for {region} at {info['export_path']}")
        
        return {
            'status': 'success',
            'exports': [info['export_path'] for info in regions.values()]
        }
    except Exception as e:
        logger.error(f"Error generating region billing exports: {e}")
        return {'status': 'error', 'message': str(e)}

def run_daily_reports():
    """
    Run all daily reports
    
    Returns:
        dict: Status of all report generation processes
    """
    results = {
        'prior_day': generate_prior_day_report(),
        'current_day': generate_current_day_report()
    }
    
    # Log overall status
    all_successful = all(result['status'] == 'success' for result in results.values())
    if all_successful:
        logger.info("All daily reports generated successfully")
    else:
        logger.warning("Some daily reports failed to generate")
        
    return results